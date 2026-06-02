# =============
# ✅ IMPORTS 
# ============
from flask import Flask, render_template, request, jsonify, session, url_for, redirect, flash, send_from_directory, send_file, make_response, render_template_string
from flask_cors import CORS
import smtplib
import random
import json
import os
import shutil
import time
from datetime import timedelta, datetime, date
import uuid
import re
import ssl
import csv
import io
import math
import base64
from decimal import Decimal
from collections import defaultdict
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from werkzeug.utils import secure_filename
import requests  # type: ignore[import]
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors  # type: ignore[import]
from reportlab.lib.pagesizes import A4  # type: ignore[import]
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore[import]
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore[import]
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT  # type: ignore[import]
from reportlab.lib.units import inch, mm  # type: ignore[import]
from reportlab.pdfbase import pdfmetrics  # type: ignore[import]
from reportlab.pdfbase.ttfonts import TTFont  # type: ignore[import]
from reportlab.pdfbase.pdfmetrics import registerFontFamily  # type: ignore[import]
from dotenv import load_dotenv  # env loader
from sqlalchemy import create_engine

# PDF
from flask import make_response, request

#email
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText


from io import BytesIO


# from reportlab.pdfbase.ttfonts import TTFo
# ===================================================
# IMPORTS
# ===================================================
import string
from collections import defaultdict

import psycopg2
from psycopg2 import pool as psycopg2_pool
from psycopg2.extras import RealDictCursor
import atexit
import traceback
import socket
from urllib.parse import urlparse, unquote, quote_plus


import hashlib


DB_POOL = None
_DB_PARAMS_CACHE = None


# Load env before any DB connection (was previously below ~line 395).
_BASE_DIR_EARLY = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(_BASE_DIR_EARLY, ".env"), override=True)
load_dotenv(os.path.join(_BASE_DIR_EARLY, "env"), override=True)


def _env_truthy(name, default=True):
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() not in {"0", "false", "no", "off", ""}


def _resolve_ipv4_hostaddr(host, port):
    """Resolve host to IPv4 address so deployments without IPv6 can connect."""
    try:
        infos = socket.getaddrinfo(host, int(port), socket.AF_INET, socket.SOCK_STREAM)
        if infos:
            return infos[0][4][0]
    except Exception:
        return None
    return None


def _supabase_project_ref_from_host(host):
    """
    Extract project ref from direct Supabase host:
    db.<project-ref>.supabase.co -> <project-ref>
    """
    m = re.match(r"^db\.([a-z0-9]+)\.supabase\.co$", (host or "").strip(), flags=re.IGNORECASE)
    return m.group(1) if m else None


def _alternate_supabase_pooler_host(host):
    """
    Toggle pooler host prefix between aws-0 and aws-1 for same region.
    Some projects are provisioned on one prefix only.
    """
    h = (host or "").strip().lower()
    m = re.match(r"^aws-(0|1)-([a-z0-9-]+)\.pooler\.supabase\.com$", h)
    if not m:
        return None
    alt_prefix = "1" if m.group(1) == "0" else "0"
    return f"aws-{alt_prefix}-{m.group(2)}.pooler.supabase.com"


class _PooledConnection:
    """Proxy connection that returns underlying conn to pool on close()."""

    def __init__(self, conn, pool_obj):
        self._conn = conn
        self._pool = pool_obj

    def __getattr__(self, name):
        return getattr(self._conn, name)

    def close(self):
        if self._conn is None:
            return
        try:
            # Keep pooled connections clean if caller forgot to commit/rollback.
            if getattr(self._conn, "status", None) != psycopg2.extensions.STATUS_READY:
                self._conn.rollback()
        except Exception:
            pass
        try:
            self._pool.putconn(self._conn)
        finally:
            self._conn = None


def _effective_connect_timeout(host, raw_timeout=None):
    """Remote DB hosts need a longer connect window than local Postgres."""
    if raw_timeout is None:
        raw_timeout = int(os.getenv("DB_CONNECT_TIMEOUT") or 15)
    else:
        raw_timeout = int(raw_timeout)
    h = (host or "").strip().lower()
    if h not in {"localhost", "127.0.0.1", "::1"} and raw_timeout < 15:
        return 15
    return raw_timeout


def _apply_db_keepalive(params):
    """Keep pooled connections alive through NAT/firewalls (optional via env)."""
    if not _env_truthy("DB_KEEPALIVE", True):
        return params
    params = dict(params)
    params.setdefault("keepalives", 1)
    params.setdefault("keepalives_idle", int(os.getenv("DB_KEEPALIVES_IDLE") or 30))
    params.setdefault("keepalives_interval", int(os.getenv("DB_KEEPALIVES_INTERVAL") or 10))
    params.setdefault("keepalives_count", int(os.getenv("DB_KEEPALIVES_COUNT") or 5))
    return params


def _db_conn_params():
    global _DB_PARAMS_CACHE
    if _DB_PARAMS_CACHE is not None:
        return dict(_DB_PARAMS_CACHE)

    # Deployment-friendly DSN support (Vercel/PythonAnywhere/custom envs)
    # Prefer pooler/prisma-style URLs first; direct DB URLs can be IPv6-only on some hosts.
    dsn_env_key = None
    for k in (
        "DB_DSN",
        "POSTGRES_PRISMA_URL",
        "SUPABASE_TRANSACTION_POOLER_URL",
        "SUPABASE_SESSION_POOLER_URL",
        "SUPABASE_POOLER_URL",
        "DATABASE_URL",
        "NEON_DATABASE_URL",
        "POSTGRES_URL",
        "SUPABASE_DB_URL",
    ):
        if os.getenv(k):
            dsn_env_key = k
            break
    dsn = (os.getenv(dsn_env_key) if dsn_env_key else "")
    dsn = (dsn or "").strip()

    db_host = (os.getenv("DB_HOST") or os.getenv("host") or "localhost").strip()
    db_name = (os.getenv("DB_NAME") or os.getenv("dbname") or "POS_Billing").strip()
    db_user = (os.getenv("DB_USER") or os.getenv("user") or "postgres").strip()
    db_pass = (os.getenv("DB_PASSWORD") or os.getenv("password") or "Pos@123").strip()
    db_port = int((os.getenv("DB_PORT") or os.getenv("port") or 5432))
    db_sslmode = (os.getenv("DB_SSLMODE") or ("require" if "supabase.co" in (db_host or "") else "prefer")).strip()
    db_connect_timeout = _effective_connect_timeout(db_host)
    force_ipv4 = _env_truthy("DB_FORCE_IPV4", True)
    pooler_host = (os.getenv("SUPABASE_POOLER_HOST") or "").strip()
    pooler_port = int(os.getenv("SUPABASE_POOLER_PORT") or 6543)
    supabase_region = (os.getenv("SUPABASE_REGION") or "").strip()
    deployed_region = (os.getenv("SUPABASE_DEFAULT_POOLER_REGION") or "ap-south-1").strip()
    is_deployed_runtime = bool(
        os.getenv("PYTHONANYWHERE_SITE")
        or os.getenv("PA_SITE")
        or os.getenv("WEBSITE_HOSTNAME")
        or os.getenv("RENDER")
        or os.getenv("VERCEL")
    )
    # Runtime-aware DSN selection so localhost and deployed can coexist.
    if is_deployed_runtime and os.getenv("DEPLOY_DB_DSN"):
        dsn_env_key = "DEPLOY_DB_DSN"
        dsn = (os.getenv("DEPLOY_DB_DSN") or "").strip()
    elif (not is_deployed_runtime) and os.getenv("LOCAL_DB_DSN"):
        dsn_env_key = "LOCAL_DB_DSN"
        dsn = (os.getenv("LOCAL_DB_DSN") or "").strip()
    elif (
        not is_deployed_runtime
        and dsn_env_key == "DB_DSN"
        and _env_truthy("LOCAL_PREFER_HOST_CONFIG", True)
        and not os.getenv("LOCAL_DB_DSN")
    ):
        # Keep local dev on host/user/password config unless LOCAL_DB_DSN is set.
        dsn = ""
        dsn_env_key = None

    # psycopg2 expects postgres:// or postgresql:// (not sqlalchemy dialect suffixes)
    if dsn.startswith("postgresql+psycopg2://"):
        dsn = "postgresql://" + dsn[len("postgresql+psycopg2://") :]
    elif dsn.startswith("postgres+psycopg2://"):
        dsn = "postgres://" + dsn[len("postgres+psycopg2://") :]

    # Auto-convert direct Supabase host to pooler host when region is provided.
    # We avoid guessing regions because wrong poolers cause "Tenant or user not found".
    project_ref = _supabase_project_ref_from_host(db_host)
    effective_region = supabase_region or (deployed_region if is_deployed_runtime else "")
    if not pooler_host and project_ref and effective_region:
        pooler_host = f"aws-0-{effective_region}.pooler.supabase.com"
        pooler_port = int(os.getenv("SUPABASE_POOLER_PORT") or 6543)
        print(f"Using derived Supabase pooler host: {pooler_host}:{pooler_port}")

    if pooler_host:
        db_host = pooler_host
        db_port = pooler_port
        # Supabase pooler expects tenant-suffixed username (e.g. postgres.<project_ref>).
        if project_ref and db_user and "." not in db_user:
            db_user = f"{db_user}.{project_ref}"

    if dsn:
        if dsn_env_key:
            print(f"DB DSN source: {dsn_env_key}")
        # Some providers use postgres://; psycopg2 accepts both, but normalize anyway.
        if dsn.startswith("postgres://"):
            dsn = "postgresql://" + dsn[len("postgres://") :]

        parsed = urlparse(dsn)
        host = (parsed.hostname or db_host or "").strip()
        port = int(parsed.port or db_port)
        database = ((parsed.path or "").lstrip("/") or db_name).strip()
        user = (unquote(parsed.username) if parsed.username is not None else db_user).strip()
        password = (unquote(parsed.password) if parsed.password is not None else db_pass).strip()

        db_search_path = os.getenv("DB_SEARCH_PATH")
        if db_search_path is None:
            db_search_path = "pos,public"
        db_search_path = db_search_path.strip()

        params = {
            "host": host,
            "database": database,
            "user": user,
            "password": password,
            "port": port,
            "sslmode": db_sslmode,
            "connect_timeout": _effective_connect_timeout(host, db_connect_timeout),
        }
        if db_search_path:
            params["options"] = f"-c search_path={db_search_path}"

        explicit_hostaddr = (os.getenv("DB_HOSTADDR") or "").strip()
        if explicit_hostaddr:
            params["hostaddr"] = explicit_hostaddr
            print(f"Using DB_HOSTADDR override: {explicit_hostaddr}")
        elif force_ipv4 and host and host not in {"localhost", "127.0.0.1"}:
            hostaddr = _resolve_ipv4_hostaddr(host, port)
            if hostaddr:
                params["hostaddr"] = hostaddr
                print(f"Using IPv4 DB hostaddr for {host}: {hostaddr}")
        _DB_PARAMS_CACHE = _apply_db_keepalive(params)
        return dict(_DB_PARAMS_CACHE)

    db_search_path = os.getenv("DB_SEARCH_PATH")
    if db_search_path is None:
        db_search_path = "pos,public"
    db_search_path = db_search_path.strip()

    params = {
        "host": db_host,
        "database": db_name,
        "user": db_user,
        "password": db_pass,
        "port": db_port,
        "sslmode": db_sslmode,
        "connect_timeout": db_connect_timeout,
    }
    if db_search_path:
        params["options"] = f"-c search_path={db_search_path}"

    explicit_hostaddr = (os.getenv("DB_HOSTADDR") or "").strip()
    if explicit_hostaddr:
        params["hostaddr"] = explicit_hostaddr
        print(f"Using DB_HOSTADDR override: {explicit_hostaddr}")
    elif force_ipv4 and db_host and db_host not in {"localhost", "127.0.0.1"}:
        hostaddr = _resolve_ipv4_hostaddr(db_host, db_port)
        if hostaddr:
            params["hostaddr"] = hostaddr
            print(f"Using IPv4 DB hostaddr for {db_host}: {hostaddr}")
    if (
        "supabase.co" in (params.get("host") or "")
        and (params.get("host") or "").startswith("db.")
        and "hostaddr" not in params
        and not os.getenv("SUPABASE_POOLER_HOST")
        and not os.getenv("SUPABASE_REGION")
        and not os.getenv("DB_DSN")
    ):
        print(
            "Supabase direct host detected without IPv4 override. "
            "Set DB_DSN to Supabase pooler URL (recommended), "
            "or set SUPABASE_POOLER_HOST/SUPABASE_REGION."
        )
    _DB_PARAMS_CACHE = _apply_db_keepalive(params)
    return dict(_DB_PARAMS_CACHE)


def _discard_pooled_connection(pool, conn):
    """Return a broken connection to the pool and close it."""
    if conn is None:
        return
    try:
        pool.putconn(conn, close=True)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def _checkout_pooled_connection(pool, attempts=3):
    """Get a live connection from the pool (validates with SELECT 1)."""
    last_err = None
    for _ in range(attempts):
        conn = None
        try:
            conn = pool.getconn()
            if getattr(conn, "closed", 0):
                _discard_pooled_connection(pool, conn)
                conn = None
                continue
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
            conn.rollback()
            return conn
        except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
            last_err = e
            _discard_pooled_connection(pool, conn)
            conn = None
        except Exception as e:
            last_err = e
            _discard_pooled_connection(pool, conn)
            conn = None
    if last_err is not None:
        raise last_err
    raise psycopg2.OperationalError("Could not obtain a healthy database connection from pool")


def _direct_db_connect(params):
    """Connect outside the pool, with one alternate Supabase pooler host retry."""
    try:
        return psycopg2.connect(**params)
    except Exception as e2:
        msg = str(e2)
        host = str(params.get("host") or "")
        alt_host = _alternate_supabase_pooler_host(host)
        if alt_host and "Tenant or user not found" in msg:
            alt_params = dict(params)
            alt_params["host"] = alt_host
            alt_params.pop("hostaddr", None)
            print(f"Retrying DB connect with alternate pooler host: {alt_host}")
            return psycopg2.connect(**alt_params)
        raise


def _init_db_pool():
    """Initialize global postgres pool once (lazy)."""
    global DB_POOL
    if DB_POOL is not None:
        return DB_POOL
    min_conn = int(os.getenv("DB_POOL_MINCONN") or 1)
    max_conn = int(os.getenv("DB_POOL_MAXCONN") or 20)
    DB_POOL = psycopg2_pool.ThreadedConnectionPool(min_conn, max_conn, **_db_conn_params())
    return DB_POOL


def _close_db_pool():
    global DB_POOL
    if DB_POOL is not None:
        try:
            DB_POOL.closeall()
        finally:
            DB_POOL = None


atexit.register(_close_db_pool)


def get_db_connection():
    """Get DB connection from global pool; fallback to direct connect with retry."""
    params = _db_conn_params()
    retries = max(1, int(os.getenv("DB_CONNECT_RETRIES") or 2))
    retry_delay = float(os.getenv("DB_CONNECT_RETRY_DELAY") or 0.5)
    last_err = None

    for attempt in range(retries):
        try:
            p = _init_db_pool()
            conn = _checkout_pooled_connection(p)
            return _PooledConnection(conn, p)
        except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
            last_err = e
            _close_db_pool()
            if attempt + 1 < retries:
                time.sleep(retry_delay)
        except Exception as e:
            last_err = e
            _close_db_pool()
            break

    for attempt in range(retries):
        try:
            return _direct_db_connect(params)
        except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
            last_err = e
            if attempt + 1 < retries:
                time.sleep(retry_delay)
        except Exception as e:
            last_err = e
            break

    print(f"Database connection failed: {last_err}")
    print(traceback.format_exc())
    raise last_err


def fetch_all(query, params=None):
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            return cur.fetchall()
    finally:
        conn.close()


def fetch_one(query, params=None):
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            return cur.fetchone()
    finally:
        conn.close()


def execute_query(query, params=None, commit=True, return_id=False):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(query, params)
            if commit:
                conn.commit()
            if return_id:
                return cur.fetchone()[0] if cur.description else None
    finally:
        conn.close()


# Base directory for building absolute paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

import object_storage  # noqa: E402 — optional S3 uploads (boto3); reads env at runtime


# def _migrate_dnr_data_pos_to_public(cur):
#     """
#     Merge DNR rows from schema `pos` into `public` when both exist (search_path used to
#     create tables in `pos` while pgAdmin queries `public` by default).

#     Copies any `dnr_id` (and child rows) present in pos but missing in public — not only
#     when public is empty, so a partial public table does not block migration.
#     """
#     try:
#         cur.execute(
#             """
#             SELECT EXISTS (
#                 SELECT 1 FROM information_schema.tables
#                 WHERE table_schema = 'pos' AND table_name = 'deliverynote_returns'
#             )
#             """
#         )
#         if not cur.fetchone()[0]:
#             return
#         cur.execute("SELECT COUNT(*) FROM pos.deliverynote_returns")
#         if (cur.fetchone() or [0])[0] == 0:
#             return
#         cur.execute(
#             """
#             INSERT INTO public.deliverynote_returns
#             SELECT p.*
#             FROM pos.deliverynote_returns p
#             WHERE NOT EXISTS (
#                 SELECT 1 FROM public.deliverynote_returns x WHERE x.dnr_id = p.dnr_id
#             )
#             """
#         )
#         # Child rows: omit SERIAL id so we never collide with ids already used in public.
#         cur.execute(
#             """
#             INSERT INTO public.deliverynote_return_items (
#                 dnr_id, product_id, product_name, uom, invoiced_qty, returned_qty, serial_no, return_reason
#             )
#             SELECT pi.dnr_id, pi.product_id, pi.product_name, pi.uom, pi.invoiced_qty,
#                    pi.returned_qty, pi.serial_no, pi.return_reason
#             FROM pos.deliverynote_return_items pi
#             WHERE EXISTS (
#                 SELECT 1 FROM public.deliverynote_returns h WHERE h.dnr_id = pi.dnr_id
#             )
#             AND NOT EXISTS (
#                 SELECT 1 FROM public.deliverynote_return_items x
#                 WHERE x.dnr_id = pi.dnr_id
#                   AND x.product_id IS NOT DISTINCT FROM pi.product_id
#                   AND COALESCE(x.serial_no, '') = COALESCE(pi.serial_no, '')
#             )
#             """
#         )
#         cur.execute(
#             """
#             INSERT INTO public.deliverynote_return_history (
#                 dnr_id, action, description, created_by, created_at
#             )
#             SELECT h.dnr_id, h.action, h.description, h.created_by, h.created_at
#             FROM pos.deliverynote_return_history h
#             WHERE EXISTS (
#                 SELECT 1 FROM public.deliverynote_returns d WHERE d.dnr_id = h.dnr_id
#             )
#             AND NOT EXISTS (
#                 SELECT 1 FROM public.deliverynote_return_history x
#                 WHERE x.dnr_id = h.dnr_id AND x.created_at = h.created_at
#             )
#             """
#         )
#         cur.execute(
#             """
#             INSERT INTO public.deliverynote_return_comments (
#                 dnr_id, comment, created_by, created_at
#             )
#             SELECT c.dnr_id, c.comment, c.created_by, c.created_at
#             FROM pos.deliverynote_return_comments c
#             WHERE EXISTS (
#                 SELECT 1 FROM public.deliverynote_returns d WHERE d.dnr_id = c.dnr_id
#             )
#             AND NOT EXISTS (
#                 SELECT 1 FROM public.deliverynote_return_comments x
#                 WHERE x.dnr_id = c.dnr_id AND x.created_at = c.created_at
#             )
#             """
#         )
#         cur.execute(
#             """
#             INSERT INTO public.deliverynote_return_attachments (dnr_id, file_name, file_path, uploaded_at)
#             SELECT a.dnr_id, a.file_name, a.file_path, a.uploaded_at
#             FROM pos.deliverynote_return_attachments a
#             WHERE EXISTS (
#                 SELECT 1 FROM public.deliverynote_returns d WHERE d.dnr_id = a.dnr_id
#             )
#             AND NOT EXISTS (
#                 SELECT 1 FROM public.deliverynote_return_attachments x
#                 WHERE x.dnr_id = a.dnr_id
#                   AND COALESCE(x.file_name, '') = COALESCE(a.file_name, '')
#             )
#             """
#         )
#         for tbl in (
#             "public.deliverynote_return_items",
#             "public.deliverynote_return_history",
#             "public.deliverynote_return_comments",
#             "public.deliverynote_return_attachments",
#         ):
#             try:
#                 cur.execute(
#                     f"""
#                     SELECT setval(
#                         pg_get_serial_sequence('{tbl}', 'id'),
#                         COALESCE((SELECT MAX(id) FROM {tbl}), 1),
#                         true
#                     )
#                     """
#                 )
#             except Exception as se:
#                 print(f"deliverynote_returns setval {tbl}: {se}")
#         print("deliverynote_returns: merged missing rows from pos → public (if any)")
#     except Exception as e:
#         print(f"deliverynote_returns pos→public migration skipped: {e}")


# def _disable_dnr_rls_if_possible(cur):
#     """Supabase: tables with RLS enabled and no policies block INSERT/SELECT for app roles."""
#     for tbl in (
#         "public.deliverynote_returns",
#         "public.deliverynote_return_items",
#         "public.deliverynote_return_history",
#         "public.deliverynote_return_comments",
#         "public.deliverynote_return_attachments",
#     ):
#         try:
#             cur.execute(f"ALTER TABLE {tbl} DISABLE ROW LEVEL SECURITY")
#         except Exception as e:
#             print(f"DNR RLS disable skipped for {tbl}: {e}")


# def _ensure_deliverynote_returns_schema():
#     """Create DNR tables if missing (same DDL as sql/deliverynote_returns_schema.sql)."""
#     path = os.path.join(BASE_DIR, "sql", "deliverynote_returns_schema.sql")
#     if not os.path.isfile(path):
#         print("deliverynote_returns schema: sql file not found, skipped")
#         return
#     try:
#         with open(path, encoding="utf-8") as f:
#             raw = f.read()
#     except OSError as e:
#         print(f"deliverynote_returns schema: read failed: {e}")
#         return
#     lines = [ln for ln in raw.splitlines() if not ln.strip().startswith("--")]
#     cleaned = "\n".join(lines)
#     parts = [p.strip() for p in cleaned.split(";") if p.strip()]
#     if not parts:
#         return
#     conn = get_db_connection()
#     try:
#         with conn.cursor() as cur:
#             for stmt in parts:
#                 cur.execute(stmt + ";")
#             _disable_dnr_rls_if_possible(cur)
#             _migrate_dnr_data_pos_to_public(cur)
#             cur.execute("SELECT COUNT(*) FROM public.deliverynote_returns")
#             _dnr_pub_n = cur.fetchone()[0]
#             cur.execute(
#                 "SELECT current_database(), current_user, current_setting('search_path', true)"
#             )
#             _dbn, _usr, _sp = cur.fetchone()
#         conn.commit()
#         print(
#             f"deliverynote_returns schema: ensured "
#             f"(public.deliverynote_returns row count = {_dnr_pub_n})"
#         )
#         print(
#             f"deliverynote_returns connection: database={_dbn!r} role={_usr!r} search_path={_sp!r}"
#         )
#         print(
#             "If pgAdmin shows 0 rows but this count > 0, pgAdmin is connected to a "
#             "different server or database than Flask — match host + database name from .env."
#         )
#         _env_db = (os.getenv("dbname") or os.getenv("DBNAME") or "").strip()
#         _env_host = (os.getenv("host") or os.getenv("HOST") or "").strip()
#         if _env_db or _env_host:
#             print(f"deliverynote_returns .env hint: host={_env_host!r} dbname={_env_db!r}")
#     except Exception as e:
#         try:
#             conn.rollback()
#         except Exception:
#             pass
#         print(f"deliverynote_returns schema ensure failed: {e}")
#     finally:
#         try:
#             conn.close()
#         except Exception:
#             pass


# # Pre-warm DB pool on startup to reduce first-login latency.
# try:
#     _init_db_pool()
# except Exception as e:
#     print(f"DB pool warmup skipped: {e}")

# try:
#     _ensure_deliverynote_returns_schema()
# except Exception as e:
#     print(f"deliverynote_returns schema ensure skipped: {e}")




# =========================================
# ✅ SQLALCHEMY ENGINE (optional)
# - Used for Supabase/Postgres connectivity
# - Driven by lowercase keys in .env: user/password/host/port/dbname
# - Exposes `engine` for scripts like `main.py`
# =========================================
SQLALCHEMY_ENGINE = None
engine = None

_USER = os.getenv("user")
_PASSWORD = os.getenv("password")
_HOST = os.getenv("host")
_PORT = os.getenv("port")
_DBNAME = os.getenv("dbname")

if _USER and _PASSWORD and _HOST and _PORT and _DBNAME and _PASSWORD != "[YOUR-PASSWORD]":
    db_search_path = os.getenv("DB_SEARCH_PATH")
    if db_search_path is None:
        db_search_path = "pos,public"
    db_search_path = db_search_path.strip()

    options_query = ""
    if db_search_path:
        options_query = f"&options={quote_plus(f'-c search_path={db_search_path}') }"

    _DATABASE_URL = (
        f"postgresql+psycopg2://{_USER}:{_PASSWORD}@{_HOST}:{_PORT}/{_DBNAME}?sslmode=require{options_query}"
    )
    try:
        SQLALCHEMY_ENGINE = create_engine(_DATABASE_URL)
        engine = SQLALCHEMY_ENGINE
    except Exception as e:
        # Keep app importable even if DB env is not configured yet.
        print(f"Failed to create SQLAlchemy engine: {e}")

def get_departments_from_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT code, name, branch, description
        FROM departments
        ORDER BY code ASC
    """)

    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [
        {
            "code": r[0],
            "name": r[1],  # UI expects name
            "branch": r[2],
            "description": r[3],
        }
        for r in rows
    ]
def get_roles_from_db():
    conn = get_db_connection()
    cur = conn.cursor()
 
    cur.execute("""
        SELECT
            r.role_id,
            COALESCE(d.name, r.department_name),
            COALESCE(d.code, ''),
            r.role_name,
            r.description,
            r.permissions,
            r.branch
        FROM roles r
        LEFT JOIN departments d
            ON LOWER(TRIM(COALESCE(r.department_name, ''))) = LOWER(TRIM(COALESCE(d.name, '')))
        ORDER BY r.role_id DESC
    """)
 
    rows = cur.fetchall()
 
    roles = []
    for r in rows:
        roles.append({
            "id": r[0],
            "department": r[1],
            "department_code": r[2],
            "role": r[3],
            "description": r[4],
            "permissions": r[5] or {},
            "branch": r[6] or "",
        })
 
    cur.close()
    conn.close()
    return roles


def _resolve_department_name(cur, department_ref):
    """Map department code or display name to canonical departments.name (FK target for roles.department_name)."""
    department_ref = (department_ref or "").strip()
    if not department_ref:
        return None
    cur.execute(
        """
        SELECT name FROM departments
        WHERE LOWER(TRIM(code)) = LOWER(TRIM(%s))
           OR LOWER(TRIM(name)) = LOWER(TRIM(%s))
        LIMIT 1
        """,
        (department_ref, department_ref),
    )
    row = cur.fetchone()
    return (row[0] or "").strip() if row else None


# =========================================
# PDF Font Setup (Supports ₹ Indian Rupee Symbol)
# =========================================
FONT_DIR = os.path.join(BASE_DIR, "static", "fonts")

pdfmetrics.registerFont(
    TTFont("DejaVuSans", os.path.join(FONT_DIR, "DejaVuSans.ttf"))
)

pdfmetrics.registerFont(
    TTFont("DejaVuSans-Bold", os.path.join(FONT_DIR, "DejaVuSans-Bold.ttf"))
)

registerFontFamily(
    "DejaVuSans",
    normal="DejaVuSans",
    bold="DejaVuSans-Bold",
    italic="DejaVuSans",
    boldItalic="DejaVuSans-Bold"


)

# =========================================
# ✅ EMAIL SENDER (SMTP / UNIVERSAL)
# =========================================
def send_email_universal(to_email, subject, body, from_email, password, smtp_server=None, port=None):
    """Send email using configured SMTP server."""
    smtp_server = smtp_server or os.getenv("SMTP_SERVER", "smtp.gmail.com")
    port = int(port or os.getenv("SMTP_PORT", "587"))
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls(context=context)
            server.login(from_email, password)
            server.sendmail(from_email, to_email, msg.as_string())
        return True

    except Exception as e:
        print("❌ Email send error:", e)
        return False


# =========================================
# ✅ FLASK APP SETUP
# =========================================
app = Flask(__name__)
CORS(app)


# =========================================
# ✅ SESSION SETTINGS & GLOBAL CONSTANTS
# =========================================
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-key")
app.permanent_session_lifetime = timedelta(days=7)     # for Remember Me
# Inactivity timeout for normal sessions when "Remember Me" is not checked (in seconds)
# 15 minutes = 900 seconds — after this, user is logged out and redirected to login
INACTIVITY_TIMEOUT = 3600


# =========================================
# ✅ FILE PATH CONSTANTS (JSON FILES)
# =========================================
def _init_upload_paths():
    """Single project uploads/ root; fall back to TMPDIR on read-only FS (e.g. Vercel serverless)."""
    root_upload = os.path.join(app.root_path, "uploads")
    try:
        os.makedirs(root_upload, exist_ok=True)
        return root_upload
    except OSError as e:
        if getattr(e, "errno", None) not in (30, 13, 1):  # EROFS, EACCES, EPERM
            raise
    root_upload = os.path.join(os.environ.get("TMPDIR", "/tmp"), "pos_app_data", "uploads")
    os.makedirs(root_upload, exist_ok=True)
    return root_upload


def _writable_upload_subdir(*parts):
    """Path under UPLOAD_FOLDER for feature-specific files; mkdir when FS allows (Vercel-safe)."""
    base = os.path.join(UPLOAD_FOLDER, *parts) if parts else UPLOAD_FOLDER
    try:
        os.makedirs(base, exist_ok=True)
    except OSError:
        pass
    return base


def _merge_legacy_upload_tree(src_dir, dest_dir):
    """Move files from a legacy folder into uploads/ (skip if destination already exists)."""
    if not src_dir or not os.path.isdir(src_dir):
        return
    src_norm = os.path.normpath(src_dir)
    dest_norm = os.path.normpath(dest_dir)
    if src_norm == dest_norm:
        return
    for root, _dirs, files in os.walk(src_dir):
        rel = os.path.relpath(root, src_dir)
        target_root = dest_dir if rel in (".", "") else os.path.join(dest_dir, rel)
        try:
            os.makedirs(target_root, exist_ok=True)
        except OSError:
            continue
        for name in files:
            sp = os.path.join(root, name)
            dp = os.path.join(target_root, name)
            if os.path.isfile(dp):
                continue
            try:
                shutil.move(sp, dp)
            except OSError:
                try:
                    shutil.copy2(sp, dp)
                except OSError:
                    pass


def _migrate_legacy_upload_dirs():
    """Consolidate static/uploads and root attachments/ into project uploads/."""
    legacy_static = os.path.join(app.root_path, "static", "uploads")
    legacy_attach = os.path.join(app.root_path, "attachments")
    _merge_legacy_upload_tree(legacy_static, UPLOAD_FOLDER)
    _merge_legacy_upload_tree(legacy_attach, QUOTATION_ATTACHMENTS_FOLDER)
    legacy_ir = os.path.join(UPLOAD_FOLDER, "invoice_return")
    if legacy_ir != INVOICE_RETURN_ATTACHMENTS_FOLDER:
        _merge_legacy_upload_tree(legacy_ir, INVOICE_RETURN_ATTACHMENTS_FOLDER)


def _resolve_stored_file_path(file_path):
    """Resolve DB-stored paths after uploads/ consolidation (supports legacy locations)."""
    if not file_path:
        return ""
    p = str(file_path).strip()
    if object_storage.is_remote_url(p):
        return p
    if os.path.isfile(p):
        return p
    rel = p.replace("\\", "/").lstrip("/")
    search_dirs = [
        UPLOAD_FOLDER,
        INVOICE_RETURN_ATTACHMENTS_FOLDER,
        INVOICE_RETURN_UPLOAD_FOLDER,
        INVOICE_ATTACHMENTS_FOLDER,
        QUOTATION_ATTACHMENTS_FOLDER,
        PURCHASE_ATTACHMENTS_FOLDER,
        STOCK_ATTACHMENTS_FOLDER,
        CREDIT_NOTE_ATTACHMENTS_FOLDER,
        SUPPLIER_ATTACHMENTS_FOLDER,
        DELIVERY_NOTE_ATTACHMENTS_FOLDER,
        DELIVERY_NOTE_RETURN_ATTACHMENTS_FOLDER,
        PRODUCT_IMAGES_FOLDER,
        IMPORT_UPLOAD_FOLDER,
        os.path.join(app.root_path, "static", "uploads"),
        os.path.join(app.root_path, "static", "uploads", "invoice_return"),
        os.path.join(app.root_path, "attachments"),
    ]
    for d in search_dirs:
        if not d:
            continue
        if rel:
            candidate = os.path.join(d, *rel.split("/"))
            if os.path.isfile(candidate):
                return candidate
        base = os.path.basename(rel)
        if base:
            candidate = os.path.join(d, base)
            if os.path.isfile(candidate):
                return candidate
    return p


ROLE_FILE = os.path.join(app.root_path, "roles.json")
DEPARTMENT_FILE = os.path.join(app.root_path, "departments.json")
UPLOAD_FOLDER = _init_upload_paths()

INVOICE_RETURN_ATTACHMENTS_FOLDER = _writable_upload_subdir("invoice_return_attachments")
INVOICE_RETURN_UPLOAD_FOLDER = INVOICE_RETURN_ATTACHMENTS_FOLDER
QUOTATION_ATTACHMENTS_FOLDER = _writable_upload_subdir("quotation_attachments")
INVOICE_ATTACHMENTS_FOLDER = _writable_upload_subdir("invoice_attachments")
PRODUCT_IMAGES_FOLDER = _writable_upload_subdir("product_images")
IMPORT_UPLOAD_FOLDER = _writable_upload_subdir("imports")
PURCHASE_ATTACHMENTS_FOLDER = _writable_upload_subdir("purchase_attachments")
STOCK_ATTACHMENTS_FOLDER = _writable_upload_subdir("stock_attachments")
CREDIT_NOTE_ATTACHMENTS_FOLDER = _writable_upload_subdir("creditnote_attachments")
SUPPLIER_ATTACHMENTS_FOLDER = _writable_upload_subdir("supplier_attachments")
DELIVERY_NOTE_ATTACHMENTS_FOLDER = _writable_upload_subdir("deliverynote_attachments")
DELIVERY_NOTE_RETURN_ATTACHMENTS_FOLDER = _writable_upload_subdir("deliverynote_return_attachments")

_DN_POD_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}


def _upload_basename(original_filename, secure=False):
    """Original file name only (no path); optional werkzeug secure_filename."""
    name = (original_filename or "").strip()
    name = name.replace("\\", "/").split("/")[-1]
    name = name.replace("\x00", "").strip()
    if secure:
        name = secure_filename(name) or "attachment"
    if not name or name in (".", ".."):
        name = "attachment"
    return name


def _upload_relative_path(entity_id, original_filename, secure=False):
    """Relative key/path: {entity_id}/{original filename} under a module folder."""
    eid = (entity_id or "").strip().replace("\\", "/").split("/")[-1]
    base = _upload_basename(original_filename, secure=secure)
    if not eid:
        return base
    return f"{eid}/{base}"


def _supplier_attachment_relative_path(supplier_id, original_filename, category=None):
    """S3/local key under supplier_attachments/: {id}/[category/]{original filename}."""
    eid = (supplier_id or "").strip().replace("\\", "/").split("/")[-1].upper()
    base = _upload_basename(original_filename)
    cat = (category or "").strip().lower().replace("\\", "/").split("/")[-1]
    if cat and cat not in (".", ".."):
        return f"{eid}/{cat}/{base}"
    if not eid:
        return base
    return f"{eid}/{base}"


def _local_upload_dest(local_root, relative_path):
    parts = str(relative_path or "attachment").replace("\\", "/").split("/")
    dest = os.path.join(local_root, *parts)
    parent = os.path.dirname(dest)
    if parent:
        os.makedirs(parent, exist_ok=True)
    return dest


def _upload_file_size_bytes(file_storage):
    """Return uploaded file size in bytes; reset read position for subsequent save/upload."""
    if not file_storage:
        return 0
    stream = getattr(file_storage, "stream", None) or file_storage
    try:
        pos = stream.tell()
        stream.seek(0, os.SEEK_END)
        size = int(stream.tell() or 0)
        stream.seek(pos)
        return size
    except Exception:
        try:
            data = file_storage.read()
            try:
                file_storage.seek(0)
            except Exception:
                pass
            return len(data) if data else 0
        except Exception:
            return 0


def _persist_module_upload(module_key, local_root, file_storage, relative_path):
    """Save to S3 (returns public URL) or local disk (returns relative path). Returns (stored_path, size)."""
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    up = object_storage.try_upload_stream(module_key, relative_path, file_storage)
    if up:
        return up
    try:
        file_storage.seek(0)
    except Exception:
        pass
    dest = _local_upload_dest(local_root, relative_path)
    file_storage.save(dest)
    return relative_path, os.path.getsize(dest)


def _remove_stored_upload(stored_path, local_root):
    p = str(stored_path or "").strip()
    if not p:
        return
    if object_storage.is_remote_url(p):
        object_storage.delete_object_by_public_url(p)
        return
    resolved = _resolve_stored_file_path(p)
    if resolved and os.path.isfile(resolved) and not object_storage.is_remote_url(resolved):
        try:
            os.remove(resolved)
            return
        except OSError:
            pass
    dest = _local_upload_dest(local_root, p)
    if os.path.isfile(dest):
        try:
            os.remove(dest)
        except OSError:
            pass


def _purge_prior_same_name_files(cur, table, entity_col, entity_id, name_col, file_name, id_col, keep_id, path_col, local_root):
    """After a new upload, delete older rows with the same display name and their stored files."""
    if keep_id is None:
        return
    cur.execute(
        f"""
        DELETE FROM {table}
        WHERE {entity_col} = %s AND {name_col} = %s AND {id_col} != %s
        RETURNING {path_col}
        """,
        (entity_id, file_name, keep_id),
    )
    for row in cur.fetchall() or []:
        _remove_stored_upload(row[0], local_root)


_migrate_legacy_upload_dirs()

PRODUCT_FILE = os.path.join(app.root_path, "product.json")
CATEGORY_FILE = os.path.join(app.root_path, "product_categories.json")
TAX_CODE_FILE = os.path.join(app.root_path, "product_tax_codes.json")
UOM_FILE = os.path.join(app.root_path, "product_uoms.json")
WAREHOUSE_FILE = os.path.join(app.root_path, "product_warehouses.json")
SIZE_FILE = os.path.join(app.root_path, "product_sizes.json")
COLOR_FILE = os.path.join(app.root_path, "product_colors.json")
SUPPLIER_FILE = os.path.join(app.root_path, "product_suppliers.json")
CUSTOMER_FILE = os.path.join(app.root_path, "customer.json")
QUOTATION_FILE = os.path.join(app.root_path, "quotation.json")
COMMENTS_FILE = os.path.join(app.root_path, "comments.json")
BILLS_FILE = os.path.join(app.root_path, "bills.json")
HOLD_FILE = os.path.join(app.root_path, "Hold-Billing.json")
SALES_ORDERS_FILE = os.path.join(app.root_path, "sales_orders.json")
DELIVERY_NOTE_FILE = os.path.join(app.root_path, "deliverynotes.json")


# =========================================
# ✅ EMAIL CONFIG (from environment)
# =========================================
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS", "")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD", "")

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SENDER_EMAIL = os.getenv("SENDER_EMAIL", EMAIL_ADDRESS)
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD", EMAIL_PASSWORD)

OTP_EXPIRY_MINUTES = int(os.getenv("OTP_EXPIRY_MINUTES", "1"))

# =========================================
# ✅ FORGOT PASSWORD + LOCKOUT CONFIG
# =========================================
_raw_base_url = os.getenv("APP_BASE_URL", "http://127.0.0.1:5000")
# Support multiple base URLs in APP_BASE_URL (comma-separated); use the first as primary
BASE_URL = _raw_base_url.split(",")[0].strip() if _raw_base_url else "http://127.0.0.1:5000"
RESET_SEND_COUNT = {}
MAX_RESET_SENDS = 5
LOCKOUT_THRESHOLD = 5
LOCKOUT_DURATION = 120  # seconds
RESET_TOKENS = {}
RESET_LOCK = {}
RESET_TOKEN_EXPIRY = 600

# =========================================
# ✅ SIGNUP OTP RATE LIMITING
# =========================================
# BUG_008: Too many OTP resend attempts should return HTTP 429 instead of 200.
# We keep a simple in‑memory counter of recent OTP sends per email.
OTP_SEND_COUNT = {}  # { email: [timestamps...] }
MAX_OTP_SENDS = 5    # max OTPs within the window
OTP_WINDOW_SECONDS = 5 * 60  # 5 minutes


# =========================================
# ✅ REGEX VALIDATION RULES
# =========================================
EMAIL_REGEX = re.compile(
    r"^[A-Za-z0-9._%+-]{3,40}@(gmail\.com|yahoo\.com|yahoo\.co\.in|outlook\.com|hotmail\.com|thestackly\.com|stackly\.in)$",
    re.IGNORECASE
)

MAX_EMAIL_LENGTH = 40
PHONE_REGEX = re.compile(r"^[0-9]{10}$")
NAME_REGEX = re.compile(r"^[A-Za-z\s]{3,20}$")


# =========================================
# ✅ QUOTATION / EMAIL OTP + RATE LIMIT CONFIG
# =========================================
OTP_EXPIRY_MINUTES = 1
MAX_ATTACHMENTS_PER_QUOTATION = 5
MAX_FILE_SIZE_MB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024
ALLOWED_EXTENSIONS = {"pdf", "doc", "docx", "xls", "xlsx", "jpg", "jpeg", "png"}
RATE_LIMIT_CONFIG = {
    "max_otp_attempts": 5,                # Max OTP attempts
    "otp_cooldown_minutes": 30,          # Cooldown after max attempts
    "max_emails_per_quotation": 3,       # Max emails per quotation
    "max_emails_per_recipient": 2,       # Max emails to same person
    "max_daily_emails_per_customer": 5,  # Max emails per day
    "min_time_between_emails_minutes": 5,  # Wait time between sends
    "requires_approval_after": 2,        # Require approval after 2 emails
}

# In-memory storage for rate limiting (quotation emails)
email_attempts = defaultdict(list)
otp_attempts = defaultdict(list)
otp_resend_attempts = defaultdict(list)
otp_blocked = defaultdict(dict)

# Shared email message containers for quotation emails
msg = MIMEMultipart("mixed")  # For both HTML and attachment
msg_alternative = MIMEMultipart("alternative")  # For HTML + plain text
msg.attach(msg_alternative)

# Simple in-memory OTP storage for quotation actions
otp_storage = {}


# =========================================
# ✅ CONTENT NEGOTIATION (HTML vs JSON for Postman)
# =========================================
def wants_json():
    """Check if client wants JSON response (API/Postman).
    True when: Accept: application/json, ?format=json, request.is_json,
    or Content-Type: application/json (many clients send this on GET in Postman)."""
    accept = (request.headers.get("Accept") or "").lower()
    if "application/json" in accept:
        return True
    if request.args.get("format") == "json":
        return True
    if request.is_json:
        return True
    ct = (request.headers.get("Content-Type") or "").lower()
    if "application/json" in ct:
        return True
    return False


# =========================================
# ✅ GLOBAL BEFORE_REQUEST
#   - AUTO SESSION TIMEOUT CHECK
# =========================================
@app.before_request
def auto_session_timeout():
    allowed_paths = [
        "/",                     # root
        "/login",
        "/signup",
        "/forgot-password",
        "/check-your-mail",
        "/check-email",
        "/reset-password",
        "/send_otp",
        "/verify_otp",
        "/send-reset-link",
        "/static",
    ]

    # allow static files
    if request.path.startswith("/static/"):
        return

    # allow the routes above
    if request.path in allowed_paths:
        return

    # for all other pages → check session timeout
    # Skip session check for some public pages above.
    # Special handling for JSON/API clients: return JSON 401 instead of redirecting to /login.
    is_api = request.path.startswith("/api/")

    if wants_json():
        # JSON clients (e.g. Postman) should get a JSON error, not an HTML redirect.
        if not check_session_timeout():
            return jsonify({"success": False, "message": "session_expired"}), 401
    else:
        # Normal browser HTML navigation: redirect to login on timeout for non-API routes.
        if not is_api and not check_session_timeout():
            return redirect(url_for("login", message="session_expired"))




# =========================================
# ✅ SESSION DEFAULT HELPERS
# =========================================
def ensure_role():
    if "user" in session and "role" not in session:
        session["role"] = "user"


# =========================================
# ✅ JSON HELPERS — Users storage shape
# =========================================
# Persisted user records: full branch-user fields + password, never "id".
# DEFAULT_BRANCH_USER_PASSWORD applies when admin-created users have no password yet.
_USER_PHONE_COUNTRY_PREFIXES = tuple(
    sorted(
        ["+91", "+971", "+974", "+966", "+94", "+880", "+977", "+1", "+44", "+61"],
        key=len,
        reverse=True,
    )
)
DEFAULT_BRANCH_USER_PASSWORD = os.getenv("DEFAULT_BRANCH_USER_PASSWORD", "Stackly@123")


def _infer_country_and_contact_from_phone(phone: str):
    """Split +… phone into (country_code, contact_number) when prefix matches a known code."""
    phone = (phone or "").strip()
    if not phone:
        return "", ""
    if not phone.startswith("+"):
        return "", re.sub(r"\D", "", phone)
    digits_only = "".join(c for c in phone[1:] if c.isdigit())
    for prefix in _USER_PHONE_COUNTRY_PREFIXES:
        p_digits = prefix[1:]
        if digits_only.startswith(p_digits):
            rest = digits_only[len(p_digits) :]
            return prefix, rest
    return "", digits_only


def normalize_user_record_for_storage(u: dict) -> dict:
    """Normalize one user dict for DB sync: drop id, ensure password + full field set."""
    if not isinstance(u, dict):
        return {}
    out = {k: v for k, v in u.items() if k != "id"}
    pwd = (out.get("password") or "").strip() if out.get("password") is not None else ""
    if not pwd:
        out["password"] = DEFAULT_BRANCH_USER_PASSWORD
    else:
        out["password"] = pwd
    name = (out.get("name") or "").strip()
    fn = (out.get("first_name") or "").strip()
    ln = (out.get("last_name") or "").strip()
    if not fn and name:
        parts = name.split(None, 1)
        fn = parts[0]
        ln = parts[1] if len(parts) > 1 else (ln or "")
    out["first_name"] = fn
    out["last_name"] = ln or ""
    out["name"] = name or f"{fn} {ln}".strip()
    cc = (out.get("country_code") or "").strip()
    cn = (out.get("contact_number") or "").strip()
    phone = (out.get("phone") or "").strip()
    if cc and cn:
        out["country_code"] = cc
        out["contact_number"] = cn
        out["phone"] = phone or f"{cc}{cn}"
    elif phone:
        icc, icn = _infer_country_and_contact_from_phone(phone)
        out["phone"] = phone
        out["country_code"] = icc
        out["contact_number"] = icn
    else:
        out["phone"] = phone
        out["country_code"] = cc
        out["contact_number"] = cn
    out["email"] = (out.get("email") or "").strip()
    out["role"] = (out.get("role") or "").strip() or "User"
    for key in ("branch", "department", "reporting_to", "available_branches", "employee_id"):
        val = out.get(key)
        out[key] = (val or "").strip() if val is not None else ""
    return out


def user_public_dict(u: dict) -> dict:
    """User object safe for JSON responses (no password or id)."""
    if not isinstance(u, dict):
        return {}
    return {k: v for k, v in u.items() if k not in ("password", "id")}


def load_users():
    """DB-backed users loader kept for compatibility with existing call sites."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT name, phone, first_name, last_name, email,
                   country_code, contact_number, branch, department,
                   role, reporting_to, available_branches, employee_id, password
            FROM users
            ORDER BY user_id DESC
            """
        )
        rows = cur.fetchall()
        out = []
        for r in rows:
            out.append(
                {
                    "name": r[0] or "",
                    "phone": r[1] or "",
                    "first_name": r[2] or "",
                    "last_name": r[3] or "",
                    "email": r[4] or "",
                    "country_code": r[5] or "",
                    "contact_number": r[6] or "",
                    "branch": r[7] or "",
                    "department": r[8] or "",
                    "role": r[9] or "User",
                    "reporting_to": r[10] or "",
                    "available_branches": str(r[11]) if r[11] is not None else "",
                    "employee_id": r[12] or "",
                    "password": r[13] or "",
                }
            )
        return out
    finally:
        cur.close()
        conn.close()




def save_users(data):
    """DB-backed users saver kept for compatibility with existing call sites."""
    if isinstance(data, dict):
        data = list(data.values())
    normalized = []
    for item in data or []:
        if isinstance(item, dict):
            norm = normalize_user_record_for_storage(item)
            item.clear()
            item.update(norm)
            normalized.append(norm)

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for u in normalized:
            email = (u.get("email") or "").strip()
            if not email:
                continue
            cur.execute(
                """
                UPDATE users
                SET name = %s,
                    phone = %s,
                    first_name = %s,
                    last_name = %s,
                    country_code = %s,
                    contact_number = %s,
                    branch = %s,
                    department = %s,
                    role = %s,
                    reporting_to = %s,
                    available_branches = %s,
                    employee_id = %s,
                    password = %s
                WHERE LOWER(email) = LOWER(%s)
                """,
                (
                    u.get("name") or "",
                    u.get("phone") or "",
                    u.get("first_name") or "",
                    u.get("last_name") or "",
                    u.get("country_code") or "",
                    u.get("contact_number") or "",
                    u.get("branch") or "",
                    u.get("department") or "",
                    u.get("role") or "User",
                    u.get("reporting_to") or "",
                    u.get("available_branches") or None,
                    u.get("employee_id") or "",
                    u.get("password") or "",
                    email,
                ),
            )
            if cur.rowcount == 0:
                cur.execute(
                    """
                    INSERT INTO users (
                        name, phone, first_name, last_name, email,
                        country_code, contact_number, branch, department,
                        role, reporting_to, available_branches, employee_id, password
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        u.get("name") or "",
                        u.get("phone") or "",
                        u.get("first_name") or "",
                        u.get("last_name") or "",
                        email,
                        u.get("country_code") or "",
                        u.get("contact_number") or "",
                        u.get("branch") or "",
                        u.get("department") or "",
                        u.get("role") or "User",
                        u.get("reporting_to") or "",
                        u.get("available_branches") or None,
                        u.get("employee_id") or "",
                        u.get("password") or "",
                    ),
                )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_failed_attempts():
    """DB-backed failed attempts map for Manage Users compatibility."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        try:
            cur.execute("SELECT LOWER(email), COALESCE(failed_attempts, 0) FROM users")
        except Exception:
            # Column may not exist yet; keep API compatible with zero values.
            return {}
        rows = cur.fetchall()
        out = {}
        for email, count in rows:
            key = (email or "").strip().lower()
            if key:
                out[key] = {"count": int(count or 0)}
        return out
    finally:
        cur.close()
        conn.close()


def save_failed_attempts(data):
    """No-op after DB migration (kept to avoid breaking call sites)."""
    return


# =========================================
# ✅ EMAIL WRAPPER HELPER
# =========================================
def send_email(to_email, subject, body):
    """Wrapper around universal email sender (supports Gmail, Outlook, Yahoo, etc.)."""
    try:
        return send_email_universal(
            to_email=to_email,
            subject=subject,
            body=body,
            from_email=EMAIL_ADDRESS,
            password=EMAIL_PASSWORD,
        )
    except Exception as e:
        print("Email send error:", e)
        return False


# =========================================
# ✅ OTP HELPERS
# =========================================
def load_otps():
    """
    DB-backed OTP loader for compatibility with existing call sites.
    Returns dict: { email: {otp, verified, timestamp} }.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS email_otp_store (
                email VARCHAR(255) PRIMARY KEY,
                otp VARCHAR(10) NOT NULL,
                otp_expiry TIMESTAMP,
                verified BOOLEAN DEFAULT FALSE
            )
            """
        )
        conn.commit()
        cur.execute(
            """
            SELECT LOWER(email), otp, otp_expiry, COALESCE(verified, FALSE)
            FROM email_otp_store
            """
        )
        rows = cur.fetchall()
        out = {}
        for email, otp, expiry, verified in rows:
            key = (email or "").strip().lower()
            if not key or not (otp or "").strip():
                continue
            ts = time.time()
            if expiry is not None:
                try:
                    ts = float(expiry.timestamp()) - 300.0
                except Exception:
                    ts = time.time()
            out[key] = {
                "otp": str(otp).strip(),
                "verified": bool(verified),
                "timestamp": ts,
            }
        return out
    finally:
        cur.close()
        conn.close()


def save_otps(otps: dict):
    """No-op after DB migration (kept to avoid breaking call sites)."""
    return


def generate_otp():
    return str(random.randint(100000, 999999))



def save_otp_in_db(email, otp):
    """Store/overwrite OTP in DB for this email."""
    email = (email or "").strip().lower()
    ts = time.time()
    try:
        expiry_dt = datetime.fromtimestamp(float(ts) + 300)
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS email_otp_store (
                email VARCHAR(255) PRIMARY KEY,
                otp VARCHAR(10) NOT NULL,
                otp_expiry TIMESTAMP,
                verified BOOLEAN DEFAULT FALSE
            )
            """
        )
        cur.execute(
            """
            INSERT INTO email_otp_store (email, otp, otp_expiry, verified)
            VALUES (%s, %s, %s, FALSE)
            ON CONFLICT (email)
            DO UPDATE SET otp = EXCLUDED.otp,
                          otp_expiry = EXCLUDED.otp_expiry,
                          verified = FALSE
            """,
            (email, otp, expiry_dt),
        )
        cur.execute(
            """
            UPDATE users
            SET email_otp = %s,
                otp_expiry = %s
            WHERE LOWER(email) = LOWER(%s)
            """,
            (otp, expiry_dt, email),
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as db_err:
        print("save_otp_in_db mirror warning:", db_err)


def verify_otp_in_db(email, otp, expiry_seconds=300):
    """
    Check OTP for email with expiry.
    If valid, mark as verified and return True.
    """
    email = (email or "").strip().lower()
    otp = (otp or "").strip()
    if not email or not otp:
        return False

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS email_otp_store (
                email VARCHAR(255) PRIMARY KEY,
                otp VARCHAR(10) NOT NULL,
                otp_expiry TIMESTAMP,
                verified BOOLEAN DEFAULT FALSE
            )
            """
        )
        cur.execute(
            """
            SELECT otp, otp_expiry
            FROM email_otp_store
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
            """,
            (email,),
        )
        row = cur.fetchone()
        if not row:
            return False
        db_otp, expiry = row
        if (db_otp or "").strip() != otp:
            return False
        if not expiry:
            return False
        expiry_naive = expiry.replace(tzinfo=None) if getattr(expiry, "tzinfo", None) else expiry
        if datetime.now() > expiry_naive:
            return False
        cur.execute(
            "UPDATE email_otp_store SET verified = TRUE WHERE LOWER(email)=LOWER(%s)",
            (email,),
        )
        conn.commit()
        return True
    finally:
        cur.close()
        conn.close()


def is_email_otp_verified(email: str) -> bool:
    """Used during signup to ensure email OTP exists and is still valid in DB."""
    email = (email or "").strip().lower()
    if not email:
        return False
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS email_otp_store (
                email VARCHAR(255) PRIMARY KEY,
                otp VARCHAR(10) NOT NULL,
                otp_expiry TIMESTAMP,
                verified BOOLEAN DEFAULT FALSE
            )
            """
        )
        cur.execute(
            """
            SELECT otp, otp_expiry, COALESCE(verified, FALSE)
            FROM email_otp_store
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
            """,
            (email,),
        )
        row = cur.fetchone()
        if not row:
            return False
        db_otp, expiry, verified = row
        if not (db_otp or "").strip() or not expiry:
            return False
        if not bool(verified):
            return False
        expiry_naive = expiry.replace(tzinfo=None) if getattr(expiry, "tzinfo", None) else expiry
        return datetime.now() <= expiry_naive
    finally:
        cur.close()
        conn.close()


def send_otp_email(to_email, otp):
    """Send the signup OTP using the universal email helper."""
    subject = "Your OTP Verification Code - Stackly POS"
    body = (
        f"Hi,\n\n"
        f"Your OTP for Stackly POS signup is: {otp}\n"
        f"It is valid for 5 minutes.\n\n"
        f"If you did not request this, you can ignore this email.\n\n"
        f"- Stackly Team"
    )
    send_email(to_email, subject, body)


# =========================================
# ✅ DEPARTMENT HELPERS
# =========================================
def normalize_department_for_storage(d):
    """Departments.json stores code, name, branch, description — never id."""
    if not isinstance(d, dict):
        return {}
    return {k: v for k, v in d.items() if k != "id"}


def department_for_api(d):
    """Department dict safe for JSON (no id)."""
    if not isinstance(d, dict):
        return {}
    return {k: v for k, v in d.items() if k != "id"}


def _dept_code_key(d):
    return (d.get("code") or "").strip().lower()


def find_department_by_code(departments, code_ref):
    """Find a department by code (case-insensitive)."""
    if not code_ref or not isinstance(departments, list):
        return None
    cref = str(code_ref).strip().lower()
    if not cref:
        return None
    for d in departments:
        if isinstance(d, dict) and _dept_code_key(d) == cref:
            return d
    return None


def load_departments():
    if not os.path.exists(DEPARTMENT_FILE):
        return []
    try:
        with open(DEPARTMENT_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if not isinstance(data, list):
                return []
            return [normalize_department_for_storage(d) if isinstance(d, dict) else d for d in data]
    except Exception:
        return []


def save_departments(departments):
    """Persist departments without id field."""
    if not isinstance(departments, list):
        departments = []
    normalized = []
    for d in departments:
        if isinstance(d, dict):
            norm = normalize_department_for_storage(d)
            d.clear()
            d.update(norm)
            normalized.append(d)
    with open(DEPARTMENT_FILE, "w", encoding="utf-8") as f:
        json.dump(normalized, f, indent=2, ensure_ascii=False)


def load_products():
    """Load all products from PostgreSQL."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM products")
        rows = cur.fetchall()
        col_names = [desc[0] for desc in cur.description]

        products = []
        for row in rows:
            product = dict(zip(col_names, row))

            # Convert numeric types used by frontend/API consumers
            if product.get("unit_price") is not None:
                product["unit_price"] = float(product["unit_price"])
            if product.get("discount") is not None:
                product["discount"] = float(product["discount"])
            if product.get("price") is not None:
                product["price"] = float(product["price"])
            if product.get("tax_percent") is not None:
                product["tax_percent"] = float(product["tax_percent"])

            # Map DB fields to frontend keys
            product["type"] = product.get("product_type") or ""
            product["category"] = product.get("category_name") or ""

            # Prefer unit_price for UI price; fallback to price column if present
            ui_price = product.get("unit_price")
            if ui_price is None:
                ui_price = product.get("price")
            product["price"] = float(ui_price or 0.0)

            # Normalize status for filters/UI
            status_raw = product.get("status", "")
            if status_raw:
                product["status"] = str(status_raw).strip().capitalize()
            else:
                product["status"] = ""

            # Frontend compatibility defaults
            product.setdefault("stock_level", 0)
            product.setdefault("description", "")
            product.setdefault("sub_category", "")
            product.setdefault("quantity", 0)
            product.setdefault("reorder_level", 0)
            product.setdefault("weight", "")
            product.setdefault("specifications", "")
            product.setdefault("related_products", "")
            product.setdefault("product_usage", "")
            product.setdefault("image", "")

            products.append(product)
        def _product_sort_key(prod):
            pid = str(prod.get("product_id") or "").strip()
            m = re.search(r"(\d+)$", pid)
            if m:
                return (1, int(m.group(1)), pid)
            return (0, -1, pid)

        # Show most recently added product IDs first (e.g. P140 above P139).
        products.sort(key=_product_sort_key, reverse=True)
        return products
    except Exception as e:
        print(f"Error in load_products: {e}")
        return []
    finally:
        if "cur" in locals():
            cur.close()
        if "conn" in locals():
            conn.close()


def save_products(products):
    """
    Upsert products without wiping the table.
    This avoids FK breakage (e.g. enquiry_product -> products.product_id).
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        for p in products:
            tax_code = p.get("tax_code", "")
            tax_percent = p.get("tax_percent")
            if tax_percent in (None, "") and tax_code:
                m = re.search(r"\((\d+(?:\.\d+)?)%\)", str(tax_code))
                if m:
                    tax_percent = float(m.group(1))

            cur.execute(
                """
                INSERT INTO products (
                    product_id, product_name, product_type, category_name,
                    unit_price, discount, description, sub_category,
                    quantity, stock_level, reorder_level,
                    weight, specifications, related_products,
                    status, product_usage, image,
                    tax_code, tax_percent, tax_description,
                    uom_name, uom_items, uom_description,
                    warehouse_name, warehouse_location, warehouse_manager,
                    warehouse_contact, warehouse_notes,
                    size, color,
                    supplier_name, supplier_contact, supplier_phone,
                    supplier_email, supplier_address
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s
                )
                ON CONFLICT (product_id) DO UPDATE SET
                    product_name = EXCLUDED.product_name,
                    product_type = EXCLUDED.product_type,
                    category_name = EXCLUDED.category_name,
                    unit_price = EXCLUDED.unit_price,
                    discount = EXCLUDED.discount,
                    description = EXCLUDED.description,
                    sub_category = EXCLUDED.sub_category,
                    quantity = EXCLUDED.quantity,
                    stock_level = EXCLUDED.stock_level,
                    reorder_level = EXCLUDED.reorder_level,
                    weight = EXCLUDED.weight,
                    specifications = EXCLUDED.specifications,
                    related_products = EXCLUDED.related_products,
                    status = EXCLUDED.status,
                    product_usage = EXCLUDED.product_usage,
                    image = EXCLUDED.image,
                    tax_code = EXCLUDED.tax_code,
                    tax_percent = EXCLUDED.tax_percent,
                    tax_description = EXCLUDED.tax_description,
                    uom_name = EXCLUDED.uom_name,
                    uom_items = EXCLUDED.uom_items,
                    uom_description = EXCLUDED.uom_description,
                    warehouse_name = EXCLUDED.warehouse_name,
                    warehouse_location = EXCLUDED.warehouse_location,
                    warehouse_manager = EXCLUDED.warehouse_manager,
                    warehouse_contact = EXCLUDED.warehouse_contact,
                    warehouse_notes = EXCLUDED.warehouse_notes,
                    size = EXCLUDED.size,
                    color = EXCLUDED.color,
                    supplier_name = EXCLUDED.supplier_name,
                    supplier_contact = EXCLUDED.supplier_contact,
                    supplier_phone = EXCLUDED.supplier_phone,
                    supplier_email = EXCLUDED.supplier_email,
                    supplier_address = EXCLUDED.supplier_address
                """,
                (
                    p.get("product_id"),
                    p.get("product_name"),
                    p.get("type", "") or p.get("product_type", ""),
                    p.get("category", "") or p.get("category_name", ""),
                    p.get("unit_price", p.get("price", 0.0)),
                    p.get("discount", 0.0),
                    p.get("description", ""),
                    p.get("sub_category", ""),
                    p.get("quantity", 0),
                    p.get("stock_level", 0),
                    p.get("reorder_level", 0),
                    p.get("weight", ""),
                    p.get("specifications", ""),
                    p.get("related_products", ""),
                    p.get("status", "Active"),
                    p.get("product_usage", ""),
                    p.get("image", ""),
                    tax_code,
                    tax_percent,
                    p.get("tax_description", ""),
                    p.get("uom", "") or p.get("uom_name", ""),
                    p.get("uom_items", 0),
                    p.get("uom_description", ""),
                    p.get("warehouse", "") or p.get("warehouse_name", ""),
                    p.get("warehouse_location", ""),
                    p.get("warehouse_manager", ""),
                    p.get("warehouse_contact", ""),
                    p.get("warehouse_notes", ""),
                    p.get("size", ""),
                    p.get("color", ""),
                    p.get("supplier", "") or p.get("supplier_name", ""),
                    p.get("supplier_contact", ""),
                    p.get("supplier_phone", ""),
                    p.get("supplier_email", ""),
                    p.get("supplier_address", ""),
                ),
            )
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()


def generate_product_id():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT product_id FROM products ORDER BY product_id DESC LIMIT 1")
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return "P101"

    match = re.search(r"(\d+)$", str(row[0]))
    if match:
        return f"P{int(match.group(1)) + 1}"
    return "P101"


@app.route('/api/products/new-id', methods=['GET'])
def get_new_product_id():
    """Returns the next auto-generated product ID"""
    product_id = generate_product_id()
    return jsonify({"productId": product_id})


# =========================================
# ✅ PRODUCT CATEGORY HELPERS
# =========================================
def load_product_categories():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT product_type, name FROM product_categories ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"product_type": r[0] or "", "name": r[1]} for r in rows]


def save_product_categories(categories):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_categories")
        for item in categories or []:
            if not isinstance(item, dict):
                continue
            cur.execute(
                "INSERT INTO product_categories (product_type, name) VALUES (%s, %s)",
                ((item.get("product_type") or "").strip(), (item.get("name") or "").strip()),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


# =========================================
# ✅ PRODUCT MASTER DROPDOWN HELPERS
#    (Tax codes, UOM, Warehouse, Size, Color, Supplier)
# =========================================
def _load_simple_list(path, cleaner):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            return []
        cleaned = []
        for item in data:
            if not isinstance(item, dict):
                continue
            cleaned_item = cleaner(item)
            if cleaned_item:
                cleaned.append(cleaned_item)
        return cleaned
    except json.JSONDecodeError:
        return []


def _save_simple_list(path, items):
    if not isinstance(items, list):
        items = []
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)


def load_tax_codes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT code, percent, description FROM product_tax_codes ORDER BY code")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"code": r[0], "percent": float(r[1]), "description": r[2] or ""} for r in rows]


def save_tax_codes(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_tax_codes")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute(
                "INSERT INTO product_tax_codes (code, percent, description) VALUES (%s, %s, %s)",
                ((item.get("code") or "").strip(), item.get("percent", 0), item.get("description", "")),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_uoms():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT name, items, description FROM product_uoms ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"name": r[0], "items": r[1] if r[1] is not None else 0, "description": r[2] or ""} for r in rows]


def save_uoms(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_uoms")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute(
                "INSERT INTO product_uoms (name, items, description) VALUES (%s, %s, %s)",
                ((item.get("name") or "").strip(), item.get("items", 0), item.get("description", "")),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_warehouses():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT name, location, manager, contact, notes FROM product_warehouses ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {"name": r[0], "location": r[1] or "", "manager": r[2] or "", "contact": r[3] or "", "notes": r[4] or ""}
        for r in rows
    ]


def save_warehouses(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_warehouses")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute(
                "INSERT INTO product_warehouses (name, location, manager, contact, notes) VALUES (%s, %s, %s, %s, %s)",
                (
                    (item.get("name") or "").strip(),
                    item.get("location", ""),
                    item.get("manager", ""),
                    item.get("contact", ""),
                    item.get("notes", ""),
                ),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_sizes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT name FROM product_sizes ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"name": r[0]} for r in rows]


def save_sizes(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_sizes")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute("INSERT INTO product_sizes (name) VALUES (%s)", ((item.get("name") or "").strip(),))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_colors():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT name FROM product_colors ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [{"name": r[0]} for r in rows]


def save_colors(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_colors")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute("INSERT INTO product_colors (name) VALUES (%s)", ((item.get("name") or "").strip(),))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def load_suppliers():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT name, contact, phone, email, address FROM product_suppliers ORDER BY name")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {"name": r[0], "contact": r[1] or "", "phone": r[2] or "", "email": r[3] or "", "address": r[4] or ""}
        for r in rows
    ]


def save_suppliers(items):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM product_suppliers")
        for item in items or []:
            if not isinstance(item, dict):
                continue
            cur.execute(
                "INSERT INTO product_suppliers (name, contact, phone, email, address) VALUES (%s, %s, %s, %s, %s)",
                (
                    (item.get("name") or "").strip(),
                    item.get("contact", ""),
                    item.get("phone", ""),
                    item.get("email", ""),
                    item.get("address", ""),
                ),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


# =========================================
# ✅ ROLE HELPERS
# =========================================
def save_roles(roles):
    with open(ROLE_FILE, "w", encoding="utf-8") as f:
        json.dump(roles, f, indent=2, ensure_ascii=False)


# =========================================
# ✅ GLOBAL AFTER_REQUEST HEADERS
# =========================================
@app.after_request
def set_security_headers(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, private"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# =========================================
# MODULE INDEX (matches sidebar: Dashboard → Masters → CRM)
# =========================================
# 1. ROOT & AUTH         — /, /login, /signup, /logout, forgot-password, reset-password, OTP
# 2. DASHBOARD           — /dashboard
# 3. MASTERS — Manage Users   — /manage-users, /create-user, /update-user, /delete-user, /api/users
# 4. MASTERS — Department & Roles — /department-roles, /api/departments, /api/roles
# 5. MASTERS — Products  — /products, /products/create, /import, /api/products
# 6. MASTERS — Customer  — /customer, /import-customer, /addnew-customer, /api/customer, /api/customers
# 7. CRM — Enquiry List  — /enquiry-list, /api/enquiries (REST for Postman)
# 8. CRM — New Enquiry   — /new-enquiry, /save-enquiry, /add-product, /api/enquiry/…
# 9. UTILITY             — /profile, /search, /logout
# =========================================

# =========================================
# 1. ROOT & AUTH
# =========================================
@app.route("/")
def root():
    return redirect(url_for("login"))


# =========================================
# 2. DASHBOARD
# =========================================
@app.route("/dashboard")
def dashboard():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "dashboard.html",
        page="dashboard",
        title="Dashboard",
        user_email=user_email,
        user_name=user_name,
    )


# =========================================
# 1. ROOT & AUTH — Auth pages (GET)
# =========================================
@app.route("/signup")
def home():
    return render_template("signup.html")


@app.route("/login")
def login():
    message = request.args.get("message", "")

    # BUG_009: When the client explicitly asks for JSON (eg. Postman),
    # treat GET on /login as an invalid HTTP method for the JSON API.
    if wants_json():
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Method not allowed. Use POST /login for authentication.",
                }
            ),
            405,
        )

    return render_template("index.html", message=message)


@app.route("/forgot-password")
def forgot_password():
    return render_template("forgot-password.html")


@app.route("/check-your-mail")
def check_your_mail_page():
    """Legacy URL — password reset is handled on /forgot-password."""
    return redirect(url_for("forgot_password"))


# =========================================
# 3. MASTERS — Manage Users
# =========================================
def _db_fetch_users_ordered(include_id: bool = False):
    """Return users in the same order as Manage Users table (latest first)."""
    conn = get_db_connection()
    cur = conn.cursor()
    if include_id:
        cur.execute(
            """
            SELECT user_id, name, email, phone, role, first_name, last_name, country_code,
                   contact_number, branch, department, reporting_to, available_branches, employee_id
            FROM users
            ORDER BY user_id DESC
            """
        )
        rows = cur.fetchall()
        users = []
        for r in rows:
            users.append(
                {
                    "user_id": r[0],
                    "name": r[1],
                    "email": r[2],
                    "phone": r[3],
                    "role": r[4],
                    "first_name": r[5],
                    "last_name": r[6],
                    "country_code": r[7],
                    "contact_number": r[8],
                    "branch": r[9],
                    "department": r[10],
                    "reporting_to": r[11],
                    "available_branches": str(r[12]) if r[12] is not None else "",
                    "employee_id": r[13],
                }
            )
    else:
        cur.execute("SELECT user_id, name, email, phone, role FROM users ORDER BY user_id DESC")
        rows = cur.fetchall()
        users = [{"user_id": r[0], "name": r[1], "email": r[2], "phone": r[3], "role": r[4]} for r in rows]
    cur.close()
    conn.close()
    return users


def _db_get_user_by_email(email: str):
    """Get single DB user row by email (case-insensitive). Includes branch/department for RBAC."""
    conn = get_db_connection()
    cur = conn.cursor()
    row = None
    try:
        cur.execute(
            """
            SELECT user_id, name, email, phone, role, branch, department
            FROM users
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
            """,
            ((email or "").strip(),),
        )
        row = cur.fetchone()
    except Exception:
        try:
            cur.execute(
                """
                SELECT user_id, name, email, phone, role
                FROM users
                WHERE LOWER(email) = LOWER(%s)
                LIMIT 1
                """,
                ((email or "").strip(),),
            )
            row = cur.fetchone()
        except Exception:
            row = None
    finally:
        cur.close()
        conn.close()
    if not row:
        return None
    if len(row) >= 7:
        return {
            "user_id": row[0],
            "name": row[1],
            "email": row[2],
            "phone": row[3],
            "role": row[4],
            "branch": row[5] or "",
            "department": row[6] or "",
        }
    return {
        "user_id": row[0],
        "name": row[1],
        "email": row[2],
        "phone": row[3],
        "role": row[4],
        "branch": "",
        "department": "",
    }


# --- RBAC: session + roles.json (matrix) + platform Admin / Super Admin ---
RBAC_MODULES = (
    "department_roles",
    "products",
    "customer",
    "new_enquiry",
    "quotation",
    "sales",
    "delivery",
    "invoice",
)


def _rbac_empty_perm():
    return {"full_access": False, "view": False, "create": False, "edit": False, "delete": False}


def _rbac_full_perm():
    return {"full_access": True, "view": True, "create": True, "edit": True, "delete": True}


def _rbac_admin_perm():
    """Admin policy: create/view/edit allowed, delete denied."""
    return {"full_access": False, "view": True, "create": True, "edit": True, "delete": False}


def normalize_menu_permissions(raw):
    """Normalize roles.json permission block (nested or flat checkbox keys from create-role UI)."""
    if not isinstance(raw, dict):
        return _rbac_empty_perm()
    if any(k in raw for k in ("full_access", "view", "create", "edit", "delete")):
        return {
            "full_access": bool(raw.get("full_access")),
            "view": bool(raw.get("view")),
            "create": bool(raw.get("create")),
            "edit": bool(raw.get("edit")),
            "delete": bool(raw.get("delete")),
        }
    fa = fv = fc = fe = fd = False
    for k, v in raw.items():
        if not v:
            continue
        ks = str(k).lower()
        if ks.endswith("_full") or ks == "full_access":
            fa = True
        elif ks.endswith("_view"):
            fv = True
        elif ks.endswith("_create"):
            fc = True
        elif ks.endswith("_edit"):
            fe = True
        elif ks.endswith("_delete"):
            fd = True
    if fa:
        return _rbac_full_perm()
    return {"full_access": False, "view": fv, "create": fc, "edit": fe, "delete": fd}


def get_current_user_profile():
    """
    Prefer PostgreSQL session + DB row (login uses DB).
    Fixes RBAC when user exists only in DB.
    """
    email = session.get("user")
    if not email:
        return None
    role = session.get("role")
    name = "User"
    department = session.get("department")
    branch = session.get("branch")

    dbu = _db_get_user_by_email(email)
    if dbu:
        name = dbu.get("name") or "User"
        role = dbu.get("role") or "User"
        if department is None:
            department = dbu.get("department")
        if branch is None:
            branch = dbu.get("branch")
    else:
        role = "User" 
    return {
        "email": email,
        "name": name,
        "role": str(role or "").strip() or "User",
        "department": (department or "").strip(),
        "branch": (branch or "").strip() or "Main Branch",
    }


   
def get_effective_permissions_for_session():
    """Effective menu permissions: platform Admin/Super Admin = full; else roles.json matrix."""
    empty = {m: _rbac_empty_perm() for m in RBAC_MODULES}
    if not session.get("user"):
        return {"is_platform_admin": False, **empty}

    prof = get_current_user_profile()
    if not prof:
        return {"is_platform_admin": False, **empty}

    rn = (prof.get("role") or "").strip().lower().replace(" ", "").replace("_", "")
    if rn in ("superadmin", "admin"):
        full = {m: _rbac_full_perm() for m in RBAC_MODULES}
        full["is_platform_admin"] = True
        return full

    roles = get_roles_from_db()
    # roles = get_roles_from_db()
    dept = (prof.get("department") or "").strip().lower()
    branch = (prof.get("branch") or "Main Branch").strip().lower()
    role_name = (prof.get("role") or "").strip().lower()
    matched = None
    for r in roles:
        if not isinstance(r, dict):
            continue
        rd = (r.get("department") or "").strip().lower()
        rb = (r.get("branch") or "").strip().lower()
        rr = (r.get("role") or "").strip().lower()
        if rd == dept and rb == branch and rr == role_name:
            matched = r
            break

    out = {"is_platform_admin": False}
    perms = (matched or {}).get("permissions") or {}
    for m in RBAC_MODULES:
        out[m] = normalize_menu_permissions(perms.get(m) or {})
    return out


@app.context_processor
def inject_rbac():
    try:
        if session.get("user"):
            return {"rbac": get_effective_permissions_for_session()}
    except Exception:
        pass
    return {"rbac": {}}


@app.context_processor
def inject_profile_display_name():
    """
    Inject consistent profile name/email for the top-right dropdown.

    Name resolution matches routes: PostgreSQL users row.
    """
    email = session.get("user")
    if not email:
        return {"profile_user_name": "User", "profile_user_email": ""}

    try:
        return {
            "profile_user_name": _get_logged_in_user_name(),
            "profile_user_email": email,
        }
    except Exception:
        return {"profile_user_name": "User", "profile_user_email": email or ""}


def _db_sync_users_id_sequence(cur):
    """Fix out-of-sync users.id sequence (common after manual imports)."""
    cur.execute(
        """
        SELECT setval(
            pg_get_serial_sequence('users', 'user_id')
            COALESCE((SELECT MAX(user_id) FROM users), 0) + 1,
            false
        )
        """
    )


# def _db_sync_customers_id_sequence(cur):
#     """Fix out-of-sync customers.id sequence (avoids duplicate key on customers_pkey)."""
#     try:
#         cur.execute(
#             """
#             SELECT setval(
#                 pg_get_serial_sequence('customers', 'id')::regclass,
#                 COALESCE((SELECT MAX(id) FROM customers), 0) + 1,
#                 false
#             )
#             WHERE pg_get_serial_sequence('customers', 'id') IS NOT NULL
#             """
#         )
#     except Exception as ex:
#         print(f"_db_sync_customers_id_sequence: {ex}")


@app.route("/manage-users")
def manage_users():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired"}), 401
        return redirect(url_for("login", message="session_expired"))

    users = _db_fetch_users_ordered(include_id=False)

    user_name = "User"
    user_role = "User"

    current_email = (user_email or "").strip().lower()

    for u in users:
        if not isinstance(u, dict):
            continue

        u_email = (u.get("email") or "").strip().lower()
        if u_email == current_email:
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    print("DEBUG manage_users: email =", user_email, "role =", user_role)

    if wants_json():
        q = (request.args.get("q") or "").strip().lower()
        try:
            page = max(1, int(request.args.get("page") or 1))
        except (TypeError, ValueError):
            page = 1
        try:
            page_size = int(request.args.get("page_size") or 10)
        except (TypeError, ValueError):
            page_size = 10
        page_size = min(max(page_size, 1), 100)

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            where_sql = ""
            params = []
            if q:
                where_sql = """
                    WHERE LOWER(name) LIKE %s
                       OR LOWER(email) LIKE %s
                       OR LOWER(phone) LIKE %s
                       OR LOWER(role) LIKE %s
                """
                like = f"%{q}%"
                params = [like, like, like, like]

            cur.execute(
                f"""
                SELECT COUNT(*)
                FROM users
                {where_sql}
                """,
                tuple(params),
            )
            total = int((cur.fetchone() or [0])[0] or 0)
            total_pages = max(1, (total + page_size - 1) // page_size)
            page = max(1, min(page, total_pages))
            offset = (page - 1) * page_size

            cur.execute(
                f"""
                SELECT
                    user_id,
                    name,
                    email,
                    phone,
                    role,
                    first_name,
                    last_name,
                    branch,
                    department,
                    reporting_to,
                    available_branches,
                    employee_id
                FROM users
                {where_sql}
                ORDER BY user_id DESC
                LIMIT %s OFFSET %s
                """,
                tuple(params + [page_size, offset]),
            )
            rows = cur.fetchall()
        finally:
            cur.close()
            conn.close()

        otp_data = load_otps()
        failed_attempts_data = load_failed_attempts()
        now_ts = time.time()

        page_users = [
            (
                lambda email_key, otp_entry, failed_entry: {
                    "user_id": r[0],
                    "name": r[1],
                    "email": r[2],
                    "phone": r[3],
                    "role": r[4],
                    "first_name": r[5] or "",
                    "last_name": r[6] or "",
                    "branch": r[7] or "",
                    "department": r[8] or "",
                    "reporting_to": r[9] or "",
                    "available_branches": str(r[10]) if r[10] is not None else "",
                    "employee_id": r[11] or "",
                    "email_otp": (otp_entry.get("otp") or "") if isinstance(otp_entry, dict) else "",
                    "otp_expiry": (
                        datetime.fromtimestamp(float(otp_entry.get("timestamp", 0)) + 300).isoformat(sep=" ")
                        if isinstance(otp_entry, dict)
                        and otp_entry.get("timestamp")
                        and (float(otp_entry.get("timestamp", 0)) + 300) >= now_ts
                        else ""
                    ),
                    "failed_attempts": (
                        int(failed_entry.get("count", 0))
                        if isinstance(failed_entry, dict)
                        else int(failed_entry or 0)
                    ),
                }
            )(
                ((r[2] or "").strip().lower()),
                otp_data.get((r[2] or "").strip().lower(), {}),
                failed_attempts_data.get((r[2] or "").strip().lower(), {}),
            )
            for r in rows
        ]

        return jsonify({
            "success": True,
            "users": page_users,
            "total": total,
            "page": page,
            "total_pages": total_pages,
            "current_user": {"email": user_email, "name": user_name, "role": user_role},
        }), 200

    return render_template(
        "manage-users.html",
        users=users,
        title="Manage Users - Stackly",
        page="manage_users",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )
# =========================================
# 4. MASTERS — Department & Roles
# =========================================
@app.route("/department-roles")
def department_roles():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
        return redirect(url_for("login", message="session_expired"))

    # departments = load_departments()
    departments = get_departments_from_db()
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    if wants_json():
        return jsonify(
            {
                "success": True,
                "departments": departments,
                "total": len(departments),
                "current_user": {"email": user_email, "name": user_name, "role": user_role},
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "department-roles.html",
        title="Department & Roles - Stackly",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        departments=departments,
    )



@app.route("/department-roles/create", methods=["GET", "POST"])
def create_department():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    branches_list = [
        {"id": "main_branch", "name": "Main Branch"},
        {"id": "branch_1", "name": "Branch 1"},
        {"id": "branch_2", "name": "Branch 2"},
    ]

    roles = get_roles_from_db()
    print("ROLES COUNT:", len(roles))

    if request.method == "POST":
        # -----------------------------------------
        #  ROLE-BASED ACCESS CHECK
        #  Only Super Admin and Admin can create departments
        # -----------------------------------------
        normalized_role = user_role.replace(" ", "").replace("_", "").lower()
        if normalized_role not in ["superadmin", "admin"]:
            error = "User cannot create new departments."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                branches=branches_list,
                roles=roles,
            )

        code = (request.form.get("code") or "").strip()
        name = (request.form.get("department_name") or "").strip()
        branch = (request.form.get("branch") or "").strip()
        desc = (request.form.get("description") or "").strip()

        if not code:
            error = "Department code is required."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                form={"code": code, "department_name": name, "branch": branch, "description": desc},
                branches=[],
                roles=roles,
            )

        if not name:
            error = "Department name is required."
            return render_template(
                "create-department.html",
                title="Create Department - Stackly",
                page="department_roles",
                section="masters",
                user_email=user_email,
                user_name=user_name,
                user_role=user_role,
                error=error,
                form={"code": code, "department_name": name, "branch": branch, "description": desc},
                branches=[],
                roles=roles,
            )

        departments = get_departments_from_db()

        # Check for duplicates (case-insensitive) - either code OR name should be unique
        for d in departments:
            existing_code = (d.get("code") or "").strip().lower()
            existing_name = (d.get("name") or "").strip().lower()
            new_code = code.lower()
            new_name = name.lower()
            
            if existing_code == new_code:
                error = "Department code already exists. Please use a different code."
                return render_template(
                    "create-department.html",
                    title="Create Department - Stackly",
                    page="department_roles",
                    section="masters",
                    user_email=user_email,
                    user_name=user_name,
                    user_role=user_role,
                    error=error,
                    form={"code": code, "department_name": name, "branch": branch, "description": desc},
                    branches=branches_list,
                roles=roles,
                )
            
            if existing_name == new_name:
                error = "Department name already exists. Please use a different name."
                return render_template(
                    "create-department.html",
                    title="Create Department - Stackly",
                    page="department_roles",
                    section="masters",
                    user_email=user_email,
                    user_name=user_name,
                    user_role=user_role,
                    error=error,
                    form={"code": code, "department_name": name, "branch": branch, "description": desc},
                    branches=branches_list,
                    roles=roles,
                )

        new_dept = {
            "code": code,
            "name": name,
            "branch": branch,
            "description": desc,
        }
        
        # 🔥 ADD DB INSERT
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO departments (code, name, branch, description)
            VALUES (%s, %s, %s, %s)
        """, (
            code,
            name,
            branch,
            desc
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Department has been created successfully", "success")
        return redirect(url_for("department_roles"))

    return render_template(
        "create-department.html",
        title="Create Department - Stackly",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        branches=branches_list,
        roles=roles,
    )

# =========================================
# 4. MASTERS — Department & Roles — Edit/Delete (UI)
# =========================================
@app.route("/department-roles/edit", methods=["POST"])
def edit_department():
    # Session check (same as Edit Product / other edit modules)
    user_email = session.get("user")
    if not user_email:
        return jsonify(success=False, error="Session expired. Please login first."), 401

    try:
        data = request.get_json(silent=True) or {}
        # Original code identifies the row (before rename); legacy clients may send "id" with the old code
        original_code = (data.get("original_code") or data.get("id") or "").strip()
        code = (data.get("code") or "").strip()
        name = (data.get("name") or "").strip()
        description = data.get("description")

        if not original_code:
            return jsonify(success=False, error="Missing department identifier (original code)"), 400

        departments = get_departments_from_db()
        if not isinstance(departments, list):
            departments = []

        current = find_department_by_code(departments, original_code)
        if not current:
            return jsonify(success=False, error="Department not found"), 404

        # Check for duplicates (case-insensitive) - exclude current department
        new_code = code.lower()
        new_name = name.lower()

        for dept in departments:
            if dept is current:
                continue
            existing_code = (dept.get("code") or "").strip().lower()
            existing_name = (dept.get("name") or "").strip().lower()
            if existing_code == new_code:
                return jsonify(success=False, error="Department code already exists. Please use a different code."), 409
            if existing_name == new_name:
                return jsonify(success=False, error="Department name already exists. Please use a different name."), 409

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            UPDATE departments
            SET code = %s,
                name = %s,
                description = %s
            WHERE LOWER(code) = LOWER(%s)
        """, (
            code,
            name,
            description,
            original_code
        ))
        conn.commit()
        cur.close()
        conn.close()

        return jsonify(success=True)
   
    except Exception as e:
        print("EDIT ERROR:", e)
        return jsonify(success=False, error=str(e)), 500

@app.route("/department-roles/delete", methods=["POST"])
def delete_department():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "error": "session_expired"}), 401

    data = request.get_json(silent=True) or {}
    code_ref = (data.get("code") or data.get("id") or "").strip()

    if not code_ref:
        return jsonify({"success": False, "error": "missing_code"}), 400
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM departments WHERE LOWER(code) = LOWER(%s)", (code_ref,))

    if cur.rowcount == 0:
        cur.close()
        conn.close()
        return jsonify({"success": False, "error": "not_found"}), 404

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

    

# =========================================
# =========================================
# 4. MASTERS — Department & Roles — APIs
# =========================================
@app.route("/api/me/permissions", methods=["GET"])
def api_me_permissions():
    """JSON: effective RBAC matrix for the logged-in user (session + roles.json)."""
    if not session.get("user"):
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    return jsonify(
        {
            "success": True,
            "permissions": get_effective_permissions_for_session(),
            "profile": get_current_user_profile(),
        }
    ), 200


@app.route("/api/departments", methods=["GET"])
def api_departments():
    """Get all departments - supports JSON response for Postman"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    q = (request.args.get("q") or "").strip().lower()
    raw_page = request.args.get("page")
    raw_page_no = request.args.get("page_no")
    raw_page_size = request.args.get("page_size")
    use_pagination = bool(q or raw_page_size or raw_page_no or (raw_page and str(raw_page).isdigit()))

    if not use_pagination:
        departments = get_departments_from_db()
        prof = get_current_user_profile() or {}
        user_name = prof.get("name") or "User"
        user_role = prof.get("role") or "User"
        perms = get_effective_permissions_for_session()
        return jsonify(
            {
                "success": True,
                "departments": [department_for_api(d) for d in departments if isinstance(d, dict)],
                "total": len(departments),
                "page": 1,
                "total_pages": 1,
                "current_user": {
                    "email": user_email,
                    "name": user_name,
                    "role": user_role,
                },
                "permissions": perms,
            }
        ), 200

    try:
        page = max(1, int(raw_page_no or raw_page or 1))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(request.args.get("page_size") or 10)
    except (TypeError, ValueError):
        page_size = 10
    page_size = min(max(page_size, 1), 100)

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        where_sql = ""
        params = []
        if q:
            like = f"%{q}%"
            where_sql = """
                WHERE LOWER(code) LIKE %s
                   OR LOWER(name) LIKE %s
                   OR LOWER(description) LIKE %s
            """
            params = [like, like, like]

        cur.execute(f"SELECT COUNT(*) FROM departments {where_sql}", tuple(params))
        total = int((cur.fetchone() or [0])[0] or 0)
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        offset = (page - 1) * page_size

        cur.execute(
            f"""
            SELECT code, name, branch, description
            FROM departments
            {where_sql}
            ORDER BY code DESC
            LIMIT %s OFFSET %s
            """,
            tuple(params + [page_size, offset]),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    departments = [
        {"code": r[0], "name": r[1], "branch": r[2], "description": r[3]}
        for r in rows
    ]
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"
    perms = get_effective_permissions_for_session()

    return jsonify(
        {
            "success": True,
            "departments": [department_for_api(d) for d in departments if isinstance(d, dict)],
            "total": total,
            "page": page,
            "total_pages": total_pages,
            "current_user": {
                "email": user_email,
                "name": user_name,
                "role": user_role,
            },
            "permissions": perms,
        }
    ), 200


@app.route("/api/departments/<path:dept_ref>", methods=["GET"])
def api_get_department(dept_ref):
    """Get single department by code (URL path, case-insensitive)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    departments = get_departments_from_db()
    department = find_department_by_code(departments, dept_ref)
    
    if not department:
        return jsonify({"success": False, "message": "Department not found"}), 404
    
    return jsonify({
        "success": True,
        "department": department_for_api(department)
    }), 200



@app.route("/api/departments", methods=["POST"])
def api_create_department():
    """Create new department"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "User cannot create new departments."
        }), 403
    
    data = request.get_json() or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or data.get("department_name") or "").strip()
    branch = (data.get("branch") or "").strip()
    description = (data.get("description") or "").strip()
    
    # Validation
    if not code:
        return jsonify({"success": False, "message": "Department code is required."}), 400
    
    if not name:
        return jsonify({"success": False, "message": "Department name is required."}), 400
    
    departments = get_departments_from_db()
    
    # Check for duplicates (case-insensitive)
    for d in departments:
        existing_code = (d.get("code") or "").strip().lower()
        existing_name = (d.get("name") or "").strip().lower()
        new_code = code.lower()
        new_name = name.lower()
        
        if existing_code == new_code:
            return jsonify({
                "success": False,
                "message": "Department code already exists. Please use a different code."
            }), 409
        
        if existing_name == new_name:
            return jsonify({
                "success": False,
                "message": "Department name already exists. Please use a different name."
            }), 409
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO departments (code, name, branch, description)
        VALUES (%s, %s, %s, %s)
        """,
        (code, name, branch, description),
    )
    conn.commit()
    cur.close()
    conn.close()

    new_dept = {"code": code, "name": name, "branch": branch, "description": description}
    
    return jsonify({
        "success": True,
        "message": "Department created successfully",
        "department": department_for_api(new_dept)
    }), 201



@app.route("/api/departments/<path:dept_ref>", methods=["PUT"])
def api_update_department(dept_ref):
    """Update existing department"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can edit departments."
        }), 403
    
    data = request.get_json() or {}
    code = (data.get("code") or "").strip()
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    
    if not dept_ref or not str(dept_ref).strip():
        return jsonify({"success": False, "message": "Department code is required in the URL."}), 400
    
    departments = get_departments_from_db()
    current = find_department_by_code(departments, dept_ref)
    if not current:
        return jsonify({"success": False, "message": "Department not found"}), 404
    
    # Check for duplicates (case-insensitive) - exclude current department
    merged_code = (code or current.get("code") or "").strip().lower()
    merged_name = (name or current.get("name") or "").strip().lower()

    for dept in departments:
        if dept is current:
            continue

        existing_code = (dept.get("code") or "").strip().lower()
        existing_name = (dept.get("name") or "").strip().lower()

        if existing_code == merged_code:
            return jsonify({
                "success": False,
                "message": "Department code already exists. Please use a different code."
            }), 409

        if existing_name == merged_name:
            return jsonify({
                "success": False,
                "message": "Department name already exists. Please use a different name."
            }), 409

    merged_code = (code or current.get("code") or "").strip()
    merged_name = (name or current.get("name") or "").strip()
    merged_branch = (data.get("branch") or current.get("branch") or "").strip()
    merged_description = description if description is not None else (current.get("description") or "")

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE departments
        SET code = %s,
            name = %s,
            branch = %s,
            description = %s
        WHERE LOWER(code) = LOWER(%s)
        """,
        (merged_code, merged_name, merged_branch, merged_description, dept_ref),
    )
    conn.commit()
    cur.close()
    conn.close()

    current["code"] = merged_code
    current["name"] = merged_name
    current["branch"] = merged_branch
    current["description"] = merged_description

    return jsonify({
        "success": True,
        "message": "Department updated successfully",
        "department": department_for_api(current)
    }), 200


@app.route("/api/departments/<path:dept_ref>", methods=["DELETE"])
def api_delete_department(dept_ref):
    """Delete department by code (URL path segment)."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    # Check user role
    users = load_users()
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_role = (u.get("role") or "User").strip()
            break
    
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can delete departments."
        }), 403
    
    if not dept_ref or not str(dept_ref).strip():
        return jsonify({"success": False, "message": "Department code is required in the URL."}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM departments WHERE LOWER(code) = LOWER(%s)", (dept_ref,))
    deleted = cur.rowcount
    conn.commit()
    cur.close()
    conn.close()

    if deleted == 0:
        return jsonify({"success": False, "message": "Department not found"}), 404

    return jsonify({
        "success": True,
        "message": "Department deleted successfully"
    }), 200


# =========================================
# 4. MASTERS — Department & Roles — Role UI
# =========================================
@app.route("/department-role/create/new")
def department_new():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    return render_template(
        "create-role.html",
        page="department_roles",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role
    )


# =========================================
# 4. MASTERS — Department & Roles — Save/Edit Role (UI)
# =========================================
@app.route("/save_role", methods=["POST"])
def save_role():
    data = request.get_json() or {}

    try:
        # -----------------------------------------
        #  ROLE-BASED ACCESS CHECK
        #  Only Super Admin and Admin can create roles
        # -----------------------------------------
        user_email = session.get("user")
        if not user_email:
            return jsonify({"status": "error", "message": "session_expired"}), 401

        users = load_users()
        user_role = "User"
        for u in users:
            if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
                user_role = (u.get("role") or "User").strip().lower()
                break

        # Normalize role for comparison
        normalized_role = user_role.replace(" ", "").replace("_", "").lower()
        if normalized_role not in ["superadmin", "admin"]:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "User cannot create roles.",
                    }
                ),
                403,
            )

        # -----------------------------------------
        #  VALIDATION: Description max 50 characters
        # -----------------------------------------
        description = (data.get("description") or "").strip()
        if description and len(description) > 50:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Description must not exceed 50 characters.",
                    }
                ),
                400,
            )

        conn = get_db_connection()
        cur = conn.cursor()
        dept_name = _resolve_department_name(cur, data.get("department"))
        if not dept_name:
            cur.close()
            conn.close()
            return jsonify(
                {"status": "error", "message": "Department not found. Select a valid department from the list."}
            ), 400

        roles = get_roles_from_db()

        # -----------------------------------------
        #  DUPLICATE CHECK
        #  - Combination of department + branch + role must be unique
        # -----------------------------------------
        new_dept = dept_name.lower()
        new_branch = (data.get("branch") or "").strip().lower()
        new_role = (data.get("role") or "").strip().lower()

        for r in roles:
            dept = (r.get("department") or "").strip().lower()
            branch = (r.get("branch") or "").strip().lower()
            role = (r.get("role") or "").strip().lower()

            if dept == new_dept and branch == new_branch and role == new_role:
                # Duplicate found → do NOT save
                cur.close()
                conn.close()
                return (
                    jsonify(
                        {
                            "status": "error",
                            "message": "This combination of Department, Branch and Role already exists.",
                        }
                    ),
                    409,
                )

        branch = (data.get("branch") or "").strip()
        role = (data.get("role") or "").strip()

        if not branch or not role:
            cur.close()
            conn.close()
            return jsonify({
                "status": "error",
                "message": "Department, Branch and Role are required"
            }), 400

        # No duplicate → append and save (department_name must be departments.name for fk_department)
        cur.execute("""
            INSERT INTO roles (department_name, branch, role_name, description, permissions)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            dept_name,
            data.get("branch"),
            data.get("role"),
            data.get("description"),
            json.dumps(data.get("permissions", {}))
        ))

        conn.commit()
        cur.close()
        conn.close()
        # save_roles(roles)             # ✅ use same saver
        return jsonify({"status": "success"})
    except Exception as e:
        print("❌ data save error:", e)
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/department-roles/create/edit", methods=["POST"])
def edit_role():
    data = request.get_json() or {}
    try:
        role_id = int(data.get("role_id"))
    except (TypeError, ValueError):
        return jsonify(success=False, error="Missing or invalid role id"), 400

    new_role = (data.get("role") or "").strip()
    description = (data.get("description") or "").strip()
    new_department = (data.get("department") or "").strip()

    if not new_role:
        return jsonify(success=False, error="Missing role data"), 400

    if description and len(description) > 50:
        return jsonify(success=False, error="Description must not exceed 50 characters."), 400

    if not new_department:
        return jsonify(success=False, error="Department is required"), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        dept_name = _resolve_department_name(cur, new_department)
        if not dept_name:
            return jsonify(
                success=False,
                error="Department not found. Select a valid department from the list.",
            ), 400

        cur.execute(
            """
            SELECT 1 FROM roles
            WHERE role_id != %s
              AND LOWER(TRIM(role_name)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(COALESCE(department_name, ''))) = LOWER(TRIM(%s))
            """,
            (role_id, new_role, dept_name),
        )
        if cur.fetchone():
            return jsonify(
                success=False,
                error="This combination of Role and Department already exists.",
            ), 409

        cur.execute(
            """
            UPDATE roles
            SET role_name = %s,
                description = %s,
                department_name = %s,
                updated_at = NOW()
            WHERE role_id = %s
            """,
            (new_role, description, dept_name, role_id),
        )

        if cur.rowcount == 0:
            conn.rollback()
            return jsonify(success=False, error="Role not found"), 404

        conn.commit()
        return jsonify(success=True)
    except Exception as e:
        conn.rollback()
        print("EDIT ERROR:", e)
        return jsonify(success=False, error=str(e)), 500
    finally:
        cur.close()
        conn.close()


@app.route("/department-roles/create/delete", methods=["POST"])
def delete_role():
    data = request.get_json(silent=True) or {}
    print("DELETE DATA:", data)

    description = data.get("description")

    if not description:
        return jsonify(success=False, error="missing_description"), 400

    roles = get_roles_from_db()
    before = len(roles)

    roles = [
        r for r in roles
        if r.get("description", "").strip().lower()
        != description.strip().lower()
    ]

    if len(roles) == before:
        return jsonify(success=False, error="not_found"), 404

    # save_roles(roles)
    return jsonify(success=True)


# =========================================
# 4. MASTERS — Department & Roles — APIs (continued)
# =========================================
@app.route("/api/roles", methods=["GET"])
def api_roles():
    """Get all roles - supports JSON response for Postman"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"
    roles = get_roles_from_db()
    
    return jsonify({
        "success": True,
        "roles": roles,
        "total": len(roles),
        "current_user": {
            "email": user_email,
            "name": user_name,
            "role": user_role
        }
    }), 200



@app.route("/api/roles/<int:role_index>", methods=["GET"])
def api_get_role(role_index):
    """Get single role by index (0-based)"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    roles = get_roles_from_db()
    
    if role_index < 0 or role_index >= len(roles):
        return jsonify({"success": False, "message": "Role index out of range"}), 404
    
    return jsonify({
        "success": True,
        "role": roles[role_index],
        "index": role_index
    }), 200


@app.route("/api/roles", methods=["POST"])
def api_create_role():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired"}), 401

    # 🔐 ROLE CHECK (optional but recommended)
    prof = get_current_user_profile() or {}
    user_role = (prof.get("role") or "").lower().replace(" ", "")

    if user_role not in ["admin", "superadmin"]:
        return jsonify({"success": False, "message": "No permission"}), 403

    data = request.get_json() or {}

    # 🔥 CORRECT MAPPING
    department_code = (data.get("department") or "").strip()
    role_name = (data.get("role") or "").strip()
    description = (data.get("description") or "").strip()

    # ✅ Validation
    if not department_code:
        return jsonify(success=False, message="Department required"), 400

    if not role_name:
        return jsonify(success=False, message="Role required"), 400

    if not description:
        return jsonify(success=False, message="Description required"), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        dept_name = _resolve_department_name(cur, department_code)
        if not dept_name:
            return jsonify(
                success=False,
                message="Department not found. Select a valid department from the list.",
            ), 400

        # 🔥 DUPLICATE CHECK (IMPORTANT) — compare using canonical name stored in DB
        cur.execute("""
            SELECT 1 FROM roles
            WHERE LOWER(TRIM(role_name)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(COALESCE(department_name, ''))) = LOWER(TRIM(%s))
        """, (role_name, dept_name))

        if cur.fetchone():
            return jsonify(success=False, message="Role already exists"), 409

        branch = (data.get("branch") or "").strip()
        perms = data.get("permissions")
        if not isinstance(perms, dict):
            perms = {}
        # ✅ INSERT: roles.department_name FK references departments.name; permissions from Create Role grid → jsonb
        cur.execute("""
            INSERT INTO roles (role_name, department_name, branch, description, created_at, permissions)
            VALUES (%s, %s, %s, %s, NOW(), %s::jsonb)
        """, (
            role_name,
            dept_name,
            branch or None,
            description,
            json.dumps(perms),
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()
        print("ERROR:", e)
        return jsonify(success=False, message=str(e)), 500

    finally:
        cur.close()
        conn.close()

    return jsonify(success=True, message="Role created successfully")

@app.route("/api/roles/<int:role_index>", methods=["PUT"])
def api_update_role(role_index):

    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired"}), 401

    prof = get_current_user_profile() or {}
    user_role = prof.get("role", "User")

    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can edit roles."
        }), 403

    data = request.get_json() or {}

    new_role_name = (data.get("role") or "").strip()
    description = (data.get("description") or "").strip()
    new_department_code = (data.get("department") or "").strip()
    permissions = data.get("permissions", {})

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # ✅ correct select
        cur.execute("""
            SELECT role_id
            FROM roles
            ORDER BY role_id
        """)
        rows = cur.fetchall()

        if role_index >= len(rows):
            return jsonify({"success": False, "message": "Role not found"}), 404

        role_id = rows[role_index][0]

        dept_name_update = None
        if new_department_code:
            dept_name_update = _resolve_department_name(cur, new_department_code)
            if not dept_name_update:
                return jsonify({"success": False, "message": "Department not found"}), 400

        # ✅ correct update (FK: department_name must be departments.name)
        cur.execute("""
            UPDATE roles
            SET 
                role_name = COALESCE(%s, role_name),
                description = COALESCE(%s, description),
                department_name = COALESCE(%s, department_name)
            WHERE role_id = %s
        """, (
            new_role_name if new_role_name else None,
            description if description else None,
            dept_name_update if new_department_code else None,
            role_id
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()   # 🔥 VERY IMPORTANT
        print("ERROR:", e)
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cur.close()
        conn.close()

    return jsonify({
        "success": True,
        "message": "Role updated successfully"
    }), 200

@app.route("/api/roles/<int:role_index>", methods=["DELETE"])
def api_delete_role(role_index):
    """Delete role by index"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    
    prof = get_current_user_profile() or {}

    user_role = (prof.get("role") or "User") \
        .strip() \
        .replace(" ", "") \
        .replace("_", "") \
        .lower()

    print("DELETE ROLE CHECK:", user_role)

    if user_role not in ["superadmin", "admin"]:

    
        return jsonify({
            "success": False,
            "message": "Only Super Admin or Admin can delete roles."
        }), 403

    # ✅ SAME SOURCE AS UI
    roles = get_roles_from_db()

    if role_index < 0 or role_index >= len(roles):
        return jsonify({"success": False, "message": "Role not found"}), 404

    role_id = roles[role_index]["id"]   # 🔥 KEY FIX

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM roles WHERE role_id = %s", (role_id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Role deleted successfully"
    }), 200
    


# =========================================
# 9. UTILITY — Profile
# =========================================
@app.route("/profile")
def profile():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()

    user_name = "User"
    mobile = ""

    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").strip().lower() == user_email.lower():
            user_name = u.get("name", "User")
            mobile = u.get("phone", "")
            break

    return render_template(
        "profile.html",
        user_email=user_email,
        user_name=user_name,
        mobile=mobile,
        page="profile",
    )


# =========================================
# 1. ROOT & AUTH — Check Email (AJAX)
# =========================================
@app.route("/check-email", methods=["POST"])
def check_email():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify({"status": "error", "message": "Email is required."}), 400

    users = load_users()
    exists = False

    if isinstance(users, list):
        for u in users:
            if isinstance(u, dict) and (u.get("email") or "").strip().lower() == email:
                exists = True
                break
    elif isinstance(users, dict) and email in users:
        exists = True

    if not exists:
        return jsonify({"status": "error", "message": "Email not registered."}), 404

    return jsonify({"status": "ok"}), 200



# =========================================
# 1. ROOT & AUTH — Forgot Password (AJAX)
# =========================================
@app.route("/send-reset-link", methods=["POST"])
def send_reset_link():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify({"status": "error", "message": "Email is required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT 1
            FROM users
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
            """,
            (email,),
        )
        exists = cur.fetchone() is not None
        cur.close()
        conn.close()

        if not exists:
            return jsonify({"status": "error", "message": "Email not registered."}), 400

        now = time.time()
        locked_until = RESET_LOCK.get(email)
        if locked_until and now < locked_until:
            remaining = int(locked_until - now)
            return jsonify({
                "status": "error",
                "message": f"This email is locked. Try again after {remaining} seconds."
            }), 429

        count = RESET_SEND_COUNT.get(email, 0)
        if count >= MAX_RESET_SENDS:
            RESET_LOCK[email] = now + LOCKOUT_DURATION
            RESET_SEND_COUNT[email] = 0
            return jsonify({
                "status": "error",
                "message": "Reset link already sent 5 times. Email is locked for 2 minutes."
            }), 429

        RESET_SEND_COUNT[email] = count + 1

        token = str(uuid.uuid4())
        RESET_TOKENS[token] = email
        reset_link = url_for("reset_password_page", token=token, _external=True)

        subject = "Reset Your Password - Stackly POS"
        body = (
            "Hi,\n\n"
            "Click the link below to reset your password:\n\n"
            f"{reset_link}\n\n"
            "If you did not request this, please ignore this email.\n\n"
            "- Stackly Team"
        )

        send_email(email, subject, body)
        return jsonify({"status": "ok"}), 200

    except Exception as e:
        print("DEBUG send-reset-link error:", e)
        return jsonify({"status": "error", "message": "Server error while sending reset link."}), 500


# =========================================
# 1. ROOT & AUTH — Reset Password
# =========================================
@app.route("/reset-password")
def reset_password_page():
    token = request.args.get("token")
    if not token or token not in RESET_TOKENS:
        return "Invalid or expired reset link.", 400
    return render_template("reset-password.html", token=token)


@app.route("/reset-password", methods=["POST"])
def reset_password_submit():
    conn = None
    cur = None
    try:
        data = request.get_json() or {}
        token = data.get("token")
        new_password = (data.get("password") or "").strip()

        if not token or not new_password:
            return jsonify({"status": "error", "message": "Token and password are required."}), 400

        email = RESET_TOKENS.get(token)
        if not email:
            return jsonify({"status": "error", "message": "Invalid or expired token."}), 400

        email_key = email.strip()
        hashed_password = hashlib.sha256(new_password.encode()).hexdigest()

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            UPDATE users
            SET password = %s
            WHERE LOWER(email) = LOWER(%s)
            """,
            (hashed_password, email_key),
        )
        updated = cur.rowcount > 0
        conn.commit()

        if not updated:
            return jsonify({"status": "error", "message": "User not found."}), 400

        RESET_TOKENS.pop(token, None)

        print("✅ Password reset for:", email_key.strip().lower())
        return jsonify({"status": "ok"}), 200

    except Exception as e:
        if conn:
            conn.rollback()
        print("❌ Reset password error:", e)
        return jsonify({"status": "error", "message": "Server error while updating password."}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()




# =========================================
# 3. MASTERS — Manage Users — Create User
# =========================================
@app.route("/create-user", methods=["GET", "POST"])
def create_user():
    user_email = session.get("user")
    if not user_email:
        # Check if JSON request
        if request.is_json or request.content_type == "application/json":
            return jsonify({"success": False, "message": "Session expired"}), 401
        return redirect(url_for("login", message="session_expired"))

    # ✅ DB USER FETCH
    

    # ✅ GET CURRENT USER FROM DB
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT name, role FROM users WHERE LOWER(email) = LOWER(%s)
    """, (user_email,))

    current_user = cursor.fetchone()

    cursor.close()
    conn.close()

    if current_user:
        user_name, user_role = current_user
    else:
        user_name = "User"
        user_role = "User"
    if request.method == "GET" and wants_json():
        return jsonify({
            "success": True,
            "page": "create-user",
            "current_user": {
                "email": user_email,
                "name": user_name,
                "role": user_role,
            },
        }), 200

    if request.method == "GET":
        return render_template(
            "create-user.html",
            title="Create User - Stackly",
            page="manage_users",
            section="masters",
            user_email=user_email,
            user_name=user_name,
            user_role=user_role,
            departments=get_departments_from_db(),
            roles=get_roles_from_db(),
        )

    # -----------------------------------------
    #  ROLE-BASED ACCESS CHECK
    #  Only Super Admin and Admin can create branch users
    # -----------------------------------------
    normalized_role = user_role.replace(" ", "").replace("_", "").lower()
    if normalized_role not in ["superadmin", "admin"]:
        error_message = "Create new branch user is restricted for your credentials."
        if request.is_json or request.content_type == "application/json":
            return jsonify({"success": False, "message": error_message}), 403
        else:
            flash(error_message, "error")
            return redirect(url_for("create_user"))

    # Determine if request is JSON or form data
    is_json_request = request.is_json or request.content_type == "application/json"
    
    if is_json_request:
        data = request.get_json(silent=True) or {}
        first_name = (data.get("first_name") or "").strip()
        last_name = (data.get("last_name") or "").strip()
        email = (data.get("email") or "").strip()
        country_code = (data.get("country_code") or "").strip()
        contact_number = (data.get("contact_number") or "").strip()
        branch = (data.get("branch") or "").strip()
        department = (data.get("department") or "").strip()
        role = (data.get("role") or "").strip()
        reporting_to = (data.get("reporting_to") or "").strip()
        available_branches = (data.get("available_branches") or "").strip()
        employee_id = (data.get("employee_id") or "").strip()
        new_password = (data.get("password") or "").strip()
    else:
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        country_code = request.form.get("country_code", "").strip()
        contact_number = request.form.get("contact_number", "").strip()
        branch = request.form.get("branch", "").strip()
        department = request.form.get("department", "").strip()
        role = request.form.get("role", "").strip()
        reporting_to = request.form.get("reporting_to", "").strip()
        available_branches = request.form.get("available_branches", "").strip()
        employee_id = request.form.get("employee_id", "").strip()
        new_password = request.form.get("password", "").strip()

    # Validation errors list
    errors = []

    # Validate First Name
    if not first_name:
        errors.append("First Name is required")
    elif len(first_name) < 3:
        errors.append("First Name must be at least 3 characters")
    elif not NAME_REGEX.match(first_name):
        errors.append("First Name should contain only letters and spaces (3-20 characters)")

    # Validate Last Name
    if not last_name:
        errors.append("Last Name is required")
    elif len(last_name) < 1:
        errors.append("Last Name must be at least 1 character")
    elif not re.match(r"^[A-Za-z\s]{1,30}$", last_name):
        errors.append("Last Name should contain only letters and spaces (1-30 characters)")

    # Validate Email
    if not email:
        errors.append("Email is required")
    elif len(email) > MAX_EMAIL_LENGTH:
        errors.append(f"Email is too long (max {MAX_EMAIL_LENGTH} characters)")
    elif not EMAIL_REGEX.match(email):
        errors.append("Enter a valid email address")

    # Validate Country Code
    valid_country_codes = ["+91", "+971", "+974", "+966", "+94", "+880", "+977", "+1", "+44", "+61"]
    phone_rules = {
        "+91": 10,   # India
        "+971": 9,   # United Arab Emirates
        "+974": 8,   # Qatar
        "+966": 9,   # Saudi Arabia
        "+94": 9,    # Sri Lanka
        "+880": 10,  # Bangladesh
        "+977": 10,  # Nepal
        "+1": 10,    # United States
        "+44": 10,   # United Kingdom (mobile)
        "+61": 9     # Australia
    }
    
    is_valid_country_code = country_code in valid_country_codes
    
    if not country_code:
        errors.append("Country code is required")
    elif not is_valid_country_code:
        errors.append(f"Invalid country code. Valid codes are: {', '.join(valid_country_codes)}")

    # Validate Contact Number
    if not contact_number:
        errors.append("Contact Number is required")
    elif not re.match(r"^\d+$", contact_number):
        errors.append("Contact Number must contain digits only")
    elif is_valid_country_code and country_code in phone_rules:
        # Only validate length if country code is valid
        required_phone_len = phone_rules[country_code]
        if len(contact_number) != required_phone_len:
            errors.append(f"Contact Number must be exactly {required_phone_len} digits for {country_code}")

    # Validate Branch
    if not branch:
        errors.append("Branch is required")

    # Validate Department
    if not department:
        errors.append("Department is required")

    # Validate Role
    if not role:
        errors.append("Role is required")

    # Validate Reporting To
    if not reporting_to:
        errors.append("Reporting To is required")
    elif len(reporting_to) < 3:
        errors.append("Reporting To must be at least 3 characters")
    elif not re.match(r"^[A-Za-z.\-\s]{3,40}$", reporting_to):
        errors.append("Reporting To may contain letters, dots, hyphens and spaces (3-40 characters)")

    # Validate Available Branches (digits only per memory)
    if not available_branches:
        errors.append("Available Branches is required")
    elif not re.match(r"^\d+$", available_branches):
        errors.append("Available Branches must contain only digits")

    # Validate Employee ID
    if not employee_id:
        errors.append("Employee ID is required")
    elif not re.match(r"^[A-Za-z0-9\-]{1,20}$", employee_id):
        errors.append("Employee ID may have letters, numbers and '-' (max 20 characters)")
    # 🔥 DB DUPLICATE CHECK
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT email, contact_number, employee_id FROM users")
    db_users = cursor.fetchall()

    for u in db_users:
        db_email = (u[0] or "").lower()
        db_contact = u[1]
        db_emp_id = u[2]

        if email.lower() == db_email:
            errors.append("Email already exists")

        if contact_number == db_contact:
            errors.append("Contact number already exists")

        if employee_id == db_emp_id:
            errors.append("Employee ID already exists")

    cursor.close()
    conn.close()
   
    # Return errors if any
    if errors:
        if is_json_request:
            return jsonify({"success": False, "message": "; ".join(errors), "errors": errors}), 400
        else:
            for error in errors:
                flash(error, "error")
            return redirect(url_for("create_user"))

    # Create new user (no persisted id; password required in file — default if not supplied)
    full_name = (first_name + " " + last_name).strip()
    full_phone = f"{country_code}{contact_number}" if country_code and contact_number else contact_number
    stored_password = new_password or DEFAULT_BRANCH_USER_PASSWORD






    new_user = {
        "name": full_name,
        "phone": full_phone,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "country_code": country_code,
        "contact_number": contact_number,
        "branch": branch,
        "department": department,
        "role": role,
        "reporting_to": reporting_to,
        "available_branches": available_branches,
        "employee_id": employee_id,
        "password": stored_password,
    }

    # data = request.get_json() if request.is_json else request.form

    conn = get_db_connection()
    cursor = conn.cursor()

    print("DEBUG NAME:", full_name)
    print("DEBUG PHONE:", full_phone)

    insert_sql = """
    INSERT INTO users (
        name, phone, first_name, last_name, email,
        country_code, contact_number, branch, department,
        role, reporting_to, available_branches, employee_id, password
    )
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """
    insert_vals = (
        full_name,
        full_phone,
        first_name,
        last_name,
        email,
        country_code,
        contact_number,
        branch,
        department,
        role,
        reporting_to,
        int(available_branches),
        employee_id,
        stored_password,
    )
    try:
        cursor.execute(insert_sql, insert_vals)
        conn.commit()
    except psycopg2.errors.UniqueViolation as e:
        # If users.id sequence is behind, sync once and retry.
        conn.rollback()
        if "users_pkey" in str(e):
            _db_sync_users_id_sequence(cursor)
            conn.commit()
            cursor.execute(insert_sql, insert_vals)
            conn.commit()
        else:
            raise
    except Exception as e:
        conn.rollback()
        if is_json_request:
            return jsonify({"success": False, "message": f"Failed to create user: {e}"}), 500
        flash("Failed to create user. Please try again.", "error")
        return redirect(url_for("create_user"))
    finally:
        cursor.close()
        conn.close()




    # Return appropriate response
    if is_json_request:
        return jsonify({
            "success": True,
            "message": "User created successfully",
            "user": {
                "name": new_user["name"],
                "phone": new_user["phone"],
                "first_name": new_user["first_name"],
                "last_name": new_user["last_name"],
                "email": new_user["email"],
                "country_code": new_user["country_code"],
                "contact_number": new_user["contact_number"],
                "branch": new_user["branch"],
                "department": new_user["department"],
                "role": new_user["role"],
                "reporting_to": new_user["reporting_to"],
                "available_branches": new_user["available_branches"],
                "employee_id": new_user["employee_id"],
            },
        }), 201
    else:
        flash("User created successfully", "success")
        return redirect(url_for("manage_users"))

def normalize_role(role: str) -> str:
    return (role or "").strip().lower().replace(" ", "").replace("_", "")# =========================================

# 3. MASTERS — Manage Users — Update User
# =========================================
@app.route("/update-user/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired"}), 401

    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    email = (data.get("email") or "").strip()
    phone = (data.get("phone") or "").strip()
    role = (data.get("role") or "").strip()
    country_code = (data.get("country_code") or "").strip()
    contact_number = (data.get("contact_number") or "").strip()
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    branch = (data.get("branch") or "").strip()
    department = (data.get("department") or "").strip()
    reporting_to = (data.get("reporting_to") or "").strip()
    available_branches = (data.get("available_branches") or "").strip()
    employee_id = (data.get("employee_id") or "").strip()

    if (
        not name
        or not email
        or not phone
        or not role
        or not first_name
        or not last_name
        or not branch
        or not department
        or not reporting_to
        or not available_branches
        or not employee_id
    ):
        return jsonify({"success": False, "message": "All fields required"}), 400

    # Keep phone parts in sync for DB rows that were created with only full phone.
    if not country_code or not contact_number:
        country_code, contact_number = _infer_country_and_contact_from_phone(phone)

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # ✅ role check
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s)",
            (user_email,),
        )
        row = cur.fetchone()

        if not row:
            return jsonify({"success": False, "message": "Current user not found"}), 403

        if normalize_role(row[0]) not in ["superadmin", "admin"]:
            return jsonify({"success": False, "message": "No permission"}), 403

        # ✅ UPDATE using user_id
        cur.execute("""
            UPDATE users
            SET name=%s,
                email=%s,
                phone=%s,
                role=%s,
                first_name=%s,
                last_name=%s,
                country_code=%s,
                contact_number=%s,
                branch=%s,
                department=%s,
                reporting_to=%s,
                available_branches=%s,
                employee_id=%s
            WHERE user_id=%s
        """, (
            name,
            email,
            phone,
            role,
            first_name,
            last_name,
            country_code,
            contact_number,
            branch,
            department,
            reporting_to,
            available_branches,
            employee_id,
            user_id,
        ))

        conn.commit()

        return jsonify({"success": True, "message": "User updated"}), 200

    except Exception as e:
        conn.rollback()
        print("❌ Update error:", e)
        return jsonify({"success": False, "message": "Update failed"}), 500

    finally:
        cur.close()
        conn.close()


# =========================================
# 5. MASTERS — Products
# =========================================
@app.route("/products")
def products():
    user_email = session.get("user")
    if not user_email:
        if wants_json():
            return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
        return redirect(url_for("login"))

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    if wants_json():
        products_list = load_products()
        return jsonify(
            {
                "success": True,
                "products": products_list,
                "total": len(products_list),
                "current_user": {"email": user_email, "name": user_name, "role": user_role},
                "permissions": get_effective_permissions_for_session(),
            }
        ), 200

    return render_template(
        "products.html",
        title="Product Master - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )


# =========================================
# ✅ API: PRODUCT CATEGORIES (PERSISTENT)
# =========================================
@app.route("/api/product-categories", methods=["GET", "POST"])
def api_product_categories():
    """
    GET  /api/product-categories?type=Electronics
         → { success, categories: ["Headphones", ...] }
    POST /api/product-categories
         JSON: { "name": "Headphones", "product_type": "Electronics" }
    """
    if request.method == "GET":
        product_type = (request.args.get("type") or "").strip()
        all_cats = load_product_categories()

        if product_type:
            # When a product type is specified, return categories
            # saved for that type PLUS any "global" categories
            # that were saved without a product_type.
            pt_norm = product_type.strip().lower()
            names = []
            for c in all_cats:
                ptype = (c.get("product_type") or "").strip().lower()
                if ptype in ("", pt_norm):
                    names.append(c.get("name") or "")
        else:
            # If no product_type is specified, return all names
            names = [c.get("name") or "" for c in all_cats]

        # remove duplicates (case-insensitive) while preserving order
        seen = set()
        unique = []
        for n in names:
            name_clean = (n or "").strip()
            if not name_clean:
                continue
            key = name_clean.lower()
            if key in seen:
                continue
            seen.add(key)
            unique.append(name_clean)

        return jsonify({"success": True, "categories": unique}), 200

    # POST → save new category
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    product_type = (data.get("product_type") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Category name is required."}), 400
    if len(name) < 3 or len(name) > 50:
        return jsonify({"success": False, "message": "Category Name Should contain atleast 3 characters."}), 400
    # Only alphabets and spaces allowed
    if not re.fullmatch(r"[A-Za-z\s]+", name):
        return jsonify({"success": False, "message": "Category name can contain only letters and spaces."}), 400

    # optional product_type – but we still store it for filtering
    all_cats = load_product_categories()
    name_lower = name.lower()
    type_lower = product_type.lower()

    for c in all_cats:
        if (c.get("name") or "").strip().lower() == name_lower and \
           (c.get("product_type") or "").strip().lower() == type_lower:
            return jsonify({"success": False, "message": "Category already exists for this product type."}), 409

    all_cats.append({"product_type": product_type, "name": name})
    save_product_categories(all_cats)

    return jsonify({"success": True, "message": "Category saved successfully."}), 201


# =========================================
# ✅ API: PRODUCT MASTER DROPDOWNS
#    Tax codes, UOM, Warehouse, Size, Color, Supplier
# =========================================
def _require_login_json():
    user_email = session.get("user")
    if not user_email:
        return None, jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    return user_email, None, None


@app.route("/api/product-tax-codes", methods=["GET", "POST"])
def api_product_tax_codes():
    if request.method == "GET":
        codes = load_tax_codes()
        return jsonify({"success": True, "items": codes}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    code = (data.get("code") or "").strip()
    description = (data.get("description") or "").strip()
    try:
        percent = float(data.get("percent", 0))
    except (TypeError, ValueError):
        percent = 0

    if not code:
        return jsonify({"success": False, "message": "Tax name is required."}), 400

    if percent < 1 or percent > 100:
        return jsonify({"success": False, "message": "Tax percentage must be between 1 and 100."}), 400

    # Extract pure tax name (before "(xx%)") for validation & duplicate check
    base_name = code.split("(")[0].strip()

    # Tax name: only alphabets and spaces, at least 3 characters
    if len(base_name) < 3:
        return jsonify({"success": False, "message": "Tax Name should contain atleast 3 characters."}), 400
    if not re.fullmatch(r"[A-Za-z\s]+", base_name):
        return jsonify({"success": False, "message": "Tax name can contain only letters and spaces."}), 400

    items = load_tax_codes()
    key = base_name.lower()
    for item in items:
        existing = (item.get("code") or "").strip()
        existing_base = existing.split("(")[0].strip().lower()
        if existing_base == key:
            # same tax name already present → treat as duplicate name
            return jsonify({"success": False, "message": "This tax name already exists."}), 409

    items.append({"code": code, "percent": percent, "description": description})
    save_tax_codes(items)
    return jsonify({"success": True, "message": "Tax code saved."}), 201


@app.route("/api/product-uoms", methods=["GET", "POST"])
def api_product_uoms():
    if request.method == "GET":
        uoms = load_uoms()
        return jsonify({"success": True, "items": uoms}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    try:
        items_val = int(data.get("items", 0))
    except (TypeError, ValueError):
        items_val = 0
    description = (data.get("description") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "UOM name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "UOM Name should contain atleast 3 characters."}), 400
    if len(name) > 50:
        return jsonify({"success": False, "message": "UOM Name should not exceed 50 characters."}), 400

    items = load_uoms()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate UOM name not allowed
            return jsonify({"success": False, "message": "UOM name already exists."}), 409

    items.append({"name": name, "items": items_val, "description": description})
    save_uoms(items)
    return jsonify({"success": True, "message": "UOM saved."}), 201


@app.route("/api/product-warehouses", methods=["GET", "POST"])
def api_product_warehouses():
    if request.method == "GET":
        warehouses = load_warehouses()
        return jsonify({"success": True, "items": warehouses}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    location = (data.get("location") or "").strip()
    manager = (data.get("manager") or "").strip()
    contact = (data.get("contact") or "").strip()
    notes = (data.get("notes") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Warehouse name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Warehouse Name should contain atleast 3 characters."}), 400
    if len(notes) > 50:
        return jsonify({"success": False, "message": "Notes must be 50 characters or less."}), 400

    items = load_warehouses()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate warehouse name not allowed
            return jsonify({"success": False, "message": "Warehouse name already exists."}), 409

    items.append(
        {
            "name": name,
            "location": location,
            "manager": manager,
            "contact": contact,
            "notes": notes,
        }
    )
    save_warehouses(items)
    return jsonify({"success": True, "message": "Warehouse saved."}), 201


@app.route("/api/product-sizes", methods=["GET", "POST"])
def api_product_sizes():
    if request.method == "GET":
        sizes = load_sizes()
        return jsonify({"success": True, "items": sizes}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Size name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Size Name should contain atleast 3 characters."}), 400

    items = load_sizes()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            return jsonify({"success": True, "message": "Size already exists."}), 200

    items.append({"name": name})
    save_sizes(items)
    return jsonify({"success": True, "message": "Size saved."}), 201


@app.route("/api/product-colors", methods=["GET", "POST"])
def api_product_colors():
    if request.method == "GET":
        colors = load_colors()
        return jsonify({"success": True, "items": colors}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Color name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Color Name should contain atleast 3 characters."}), 400

    items = load_colors()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            return jsonify({"success": True, "message": "Color already exists."}), 200

    items.append({"name": name})
    save_colors(items)
    return jsonify({"success": True, "message": "Color saved."}), 201


@app.route("/api/product-suppliers", methods=["GET", "POST"])
def api_product_suppliers():
    if request.method == "GET":
        suppliers = load_suppliers()
        return jsonify({"success": True, "items": suppliers}), 200

    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    contact = (data.get("contact") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    address = (data.get("address") or "").strip()

    if not name:
        return jsonify({"success": False, "message": "Supplier name is required."}), 400
    if len(name) < 3:
        return jsonify({"success": False, "message": "Supplier Name should contain atleast 3 characters."}), 400
    if not re.fullmatch(r"[A-Za-z\s]+", name):
        return jsonify({"success": False, "message": "Supplier Name should contain atleast 3 characters."}), 400

    # Contact person: apply same rules as Supplier Name
    if not contact or len(contact.strip()) < 3 or not re.fullmatch(r"[A-Za-z\s]+", contact.strip()):
        return jsonify({"success": False, "message": "Contact Person Name should contain atleast 3 characters."}), 400

    items = load_suppliers()
    key = name.strip().lower()
    for item in items:
        if (item.get("name") or "").strip().lower() == key:
            # Duplicate supplier name not allowed
            return jsonify({"success": False, "message": "Supplier name already exists."}), 409

    items.append(
        {
            "name": name,
            "contact": contact,
            "phone": phone,
            "email": email,
            "address": address,
        }
    )
    save_suppliers(items)
    return jsonify({"success": True, "message": "Supplier saved."}), 201

# =========================
# API: GET SINGLE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["GET"])
def api_get_product(product_id):
    """
    GET /api/products/<product_id>
    
    Returns a single product by ID
    Supports both JSON and HTML responses
    """
    # BUG_001 / BUG_006: Require login for product APIs and return JSON 401
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(
            """
            SELECT
                product_id,
                product_name,
                product_type,
                category_name,
                status,
                stock_level,
                unit_price,
                description,
                sub_category,
                tax_code,
                supplier_name
            FROM products
            WHERE product_id = %s
            LIMIT 1
            """,
            (str(product_id),),
        )
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    p = None
    if row:
        p = dict(row)
        p["type"] = p.get("product_type") or ""
        p["category"] = p.get("category_name") or ""
        p["price"] = float(p.get("unit_price") or 0.0)
        p["stock_level"] = int(p.get("stock_level") or 0)
    
    if not p:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404
    
    response_data = {
        "success": True,
        "data": p,
        "message": "Product retrieved successfully"
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200

# =========================
# API: CREATE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products", methods=["POST"])
def api_create_product():
    """
    POST /api/products
    Content-Type: application/json
    
    Request Body (JSON):
    {
        "product_name": "Product Name",
        "type": "Physical",
        "category": "Electronics",
        "status": "Active",
        "stock_level": 100,
        "price": 99.99,
        ... (other optional fields)
    }
    
    Returns created product with generated ID
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json() or {}
    
    # Validation
    product_name = (data.get("product_name") or "").strip()
    if not product_name:
        error_response = {
            "success": False,
            "message": "Product name is required",
            "error": "Validation failed"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    # Generate product ID
    product_id = generate_product_id()
    
    # Build product object
    product = {
        "product_id": str(product_id),
        "product_name": product_name,
        "type": (data.get("type") or "").strip(),
        "category": (data.get("category") or "").strip(),
        "status": (data.get("status") or "Active").strip(),
        "stock_level": int(data.get("stock_level", 0)),
        "price": float(data.get("price", 0.0)),
        "description": (data.get("description") or "").strip(),
        "sub_category": (data.get("sub_category") or "").strip(),
        "unit_price": (data.get("unit_price") or "").strip(),
        "discount": (data.get("discount") or "").strip(),
        "tax_code": (data.get("tax_code") or "").strip(),
        "quantity": (data.get("quantity") or "").strip(),
        "uom": (data.get("uom") or "").strip(),
        "reorder_level": (data.get("reorder_level") or "").strip(),
        "warehouse": (data.get("warehouse") or "").strip(),
        "size": (data.get("size") or "").strip(),
        "color": (data.get("color") or "").strip(),
        "weight": (data.get("weight") or "").strip(),
        "specifications": (data.get("specifications") or "").strip(),
        "related_products": (data.get("related_products") or "").strip(),
        "supplier": (data.get("supplier") or "").strip(),
        "product_usage": (data.get("product_usage") or "").strip(),
        "image": (data.get("image") or "").strip(),
    }
    
    # Save product directly (faster than load+rewrite all products)
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO products (
                product_id, product_name, product_type, category_name,
                status, stock_level, unit_price,
                description, sub_category, tax_code, supplier_name,
                product_usage, image
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s
            )
            """,
            (
                product["product_id"],
                product["product_name"],
                product.get("type", ""),
                product.get("category", ""),
                product.get("status", "Active"),
                int(product.get("stock_level", 0) or 0),
                float(product.get("price", 0.0) or 0.0),
                product.get("description", ""),
                product.get("sub_category", ""),
                product.get("tax_code", ""),
                product.get("supplier", ""),
                product.get("product_usage", ""),
                product.get("image", ""),
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
    
    response_data = {
        "success": True,
        "message": "Product created successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 201
    else:
        return jsonify(response_data), 201


# =========================
# API: UPDATE PRODUCT (PUT - Full Update) (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["PUT"])
def api_update_product(product_id):
    """
    PUT /api/products/<product_id>
    Content-Type: application/json
    
    Full update - replaces entire product
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json(silent=True) or {}

    # BUG_005: Validate payload and reject obviously invalid values
    errors = []
    product_name = (data.get("product_name") or "").strip()
    if not product_name:
        errors.append("Product name is required.")

    if "stock_level" in data:
        try:
            stock_val = int(data.get("stock_level"))
            if stock_val < 0:
                errors.append("Stock level must be 0 or greater.")
        except (TypeError, ValueError):
            errors.append("Stock level must be a whole number.")

    if "price" in data:
        try:
            price_val = float(data.get("price"))
            if price_val <= 0:
                errors.append("Price must be greater than 0.")
        except (TypeError, ValueError):
            errors.append("Price must be a valid number.")

    if errors:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Validation failed",
                    "errors": errors,
                }
            ),
            400,
        )
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute(
            """
            SELECT product_id, product_name, product_type, category_name, status, stock_level,
                   unit_price, description, sub_category, tax_code, supplier_name
            FROM products
            WHERE product_id = %s
            LIMIT 1
            """,
            (str(product_id),),
        )
        current = cur.fetchone()
        if not current:
            error_response = {
                "success": False,
                "message": "Product not found",
                "error": f"Product with ID '{product_id}' does not exist"
            }
            return jsonify(error_response), 404

        product = dict(current)
        product["product_name"] = (data.get("product_name") or product.get("product_name") or "").strip()
        product["type"] = (data.get("type") or product.get("product_type") or "").strip()
        product["category"] = (data.get("category") or product.get("category_name") or "").strip()
        product["status"] = (data.get("status") or product.get("status") or "Active").strip()
        product["supplier"] = (data.get("supplier") or product.get("supplier_name") or "").strip()
        try:
            product["stock_level"] = int(data.get("stock_level", product.get("stock_level", 0)))
        except Exception:
            product["stock_level"] = 0
        try:
            product["price"] = float(data.get("price", product.get("unit_price", 0)))
        except Exception:
            product["price"] = 0.0
        if "description" in data:
            product["description"] = (data.get("description") or "").strip()
        if "sub_category" in data:
            product["sub_category"] = (data.get("sub_category") or "").strip()
        if "tax_code" in data:
            product["tax_code"] = (data.get("tax_code") or "").strip()
        if "supplier" in data:
            product["supplier"] = (data.get("supplier") or "").strip()

        cur.execute(
            """
            SELECT product_id, product_name, product_type, category_name, status, stock_level, unit_price
            FROM products
            WHERE product_id <> %s
            """,
            (str(product_id),),
        )
        products = []
        for r in cur.fetchall() or []:
            d = dict(r)
            d["type"] = d.get("product_type") or ""
            d["category"] = d.get("category_name") or ""
            d["price"] = float(d.get("unit_price") or 0.0)
            products.append(d)

    finally:
        cur.close()
        conn.close()

    if not product:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        return jsonify(error_response), 404

    def _to_int(value, default=0):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def _to_float(value, default=0.0):
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    # -------------------- DUPLICATE VALIDATION (exclude current product) -----------------------------
    # Normalize values for comparison (case-insensitive for text, exact for numbers)
    updated_product_name = (product.get("product_name") or "").strip().lower()
    updated_type = (product.get("type") or "").strip().lower()
    updated_category = (product.get("category") or "").strip().lower()
    updated_status = (product.get("status") or "").strip().lower()
    updated_stock_level = _to_int(product.get("stock_level", 0), 0)
    updated_price = _to_float(product.get("price", 0.0), 0.0)
    
    # Check 1: Duplicate product name (case-insensitive, exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        if existing_name == updated_product_name:
            error_response = {
                "success": False,
                "message": f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409
    
    # Check 2: Duplicate combination (exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        existing_type = (existing.get("type") or "").strip().lower()
        existing_category = (existing.get("category") or "").strip().lower()
        existing_status = (existing.get("status") or "").strip().lower()
        existing_stock_level = _to_int(existing.get("stock_level", 0), 0)
        existing_price = _to_float(existing.get("price", 0.0), 0.0)
        
        # Compare all fields (case-insensitive for text, exact for numbers)
        if (existing_name == updated_product_name and
            existing_type == updated_type and
            existing_category == updated_category and
            existing_status == updated_status and
            existing_stock_level == updated_stock_level and
            abs(existing_price - updated_price) < 0.01):  # Float comparison with tolerance
            error_response = {
                "success": False,
                "message": f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {updated_stock_level}, Price: {updated_price}) already exists."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE products SET
                product_name = %s,
                product_type = %s,
                category_name = %s,
                status = %s,
                stock_level = %s,
                unit_price = %s,
                description = %s,
                sub_category = %s,
                tax_code = %s,
                supplier_name = %s
            WHERE product_id = %s
            """,
            (
                product.get("product_name", ""),
                product.get("type", ""),
                product.get("category", ""),
                product.get("status", "Active"),
                int(product.get("stock_level", 0) or 0),
                float(product.get("price", 0.0) or 0.0),
                product.get("description", ""),
                product.get("sub_category", ""),
                product.get("tax_code", ""),
                product.get("supplier", ""),
                str(product_id),
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
    
    response_data = {
        "success": True,
        "message": "Product updated successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200


# =========================
# API: PARTIAL UPDATE PRODUCT (PATCH) (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["PATCH"])
def api_patch_product(product_id):
    """
    PATCH /api/products/<product_id>
    Content-Type: application/json
    
    Partial update - only updates provided fields
    """
    # BUG_001 / BUG_006: Require login for product APIs
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        error_response = {
            "success": False,
            "message": "Content-Type must be application/json",
            "error": "Invalid request format"
        }
        if wants_json():
            return jsonify(error_response), 400
        else:
            return jsonify(error_response), 400
    
    data = request.get_json(silent=True) or {}

    # BUG_005: Validate any fields that are provided
    errors = []
    if "product_name" in data:
        name = (data.get("product_name") or "").strip()
        if not name:
            errors.append("Product name cannot be blank.")
    if "stock_level" in data:
        try:
            stock_val = int(data.get("stock_level"))
            if stock_val < 0:
                errors.append("Stock level must be 0 or greater.")
        except (TypeError, ValueError):
            errors.append("Stock level must be a whole number.")
    if "price" in data:
        try:
            price_val = float(data.get("price"))
            if price_val <= 0:
                errors.append("Price must be greater than 0.")
        except (TypeError, ValueError):
            errors.append("Price must be a valid number.")

    if errors:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Validation failed",
                    "errors": errors,
                }
            ),
            400,
        )
    products = load_products()
    updated = False
    product = None

    for p in products:
        if str(p.get("product_id")) == str(product_id):
            # Partial update - only update provided fields
            if "product_name" in data:
                p["product_name"] = (data.get("product_name") or "").strip()
            if "type" in data:
                p["type"] = (data.get("type") or "").strip()
            if "category" in data:
                p["category"] = (data.get("category") or "").strip()
            if "status" in data:
                p["status"] = (data.get("status") or "Active").strip()
            if "stock_level" in data:
                try:
                    p["stock_level"] = int(data.get("stock_level", 0))
                except:
                    p["stock_level"] = 0
            if "price" in data:
                try:
                    p["price"] = float(data.get("price", 0))
                except:
                    p["price"] = 0.0
            if "description" in data:
                p["description"] = (data.get("description") or "").strip()
            if "sub_category" in data:
                p["sub_category"] = (data.get("sub_category") or "").strip()
            if "tax_code" in data:
                p["tax_code"] = (data.get("tax_code") or "").strip()
            if "supplier" in data:
                p["supplier"] = (data.get("supplier") or "").strip()

            product = p
            updated = True
            break

    if not updated:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        if wants_json():
            return jsonify(error_response), 404
        else:
            return jsonify(error_response), 404

    # -------------------- DUPLICATE VALIDATION (exclude current product) -----------------------------
    # Normalize values for comparison (case-insensitive for text, exact for numbers)
    updated_product_name = (product.get("product_name") or "").strip().lower()
    updated_type = (product.get("type") or "").strip().lower()
    updated_category = (product.get("category") or "").strip().lower()
    updated_status = (product.get("status") or "").strip().lower()
    updated_stock_level = product.get("stock_level", 0)
    updated_price = product.get("price", 0.0)
    
    # Check 1: Duplicate product name (case-insensitive, exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        if existing_name == updated_product_name:
            error_response = {
                "success": False,
                "message": f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409
    
    # Check 2: Duplicate combination (exclude current product)
    for existing in products:
        if str(existing.get("product_id")) == str(product_id):
            continue  # Skip the product being updated
        existing_name = (existing.get("product_name") or "").strip().lower()
        existing_type = (existing.get("type") or "").strip().lower()
        existing_category = (existing.get("category") or "").strip().lower()
        existing_status = (existing.get("status") or "").strip().lower()
        existing_stock_level = existing.get("stock_level", 0)
        existing_price = existing.get("price", 0.0)
        
        # Compare all fields (case-insensitive for text, exact for numbers)
        if (existing_name == updated_product_name and
            existing_type == updated_type and
            existing_category == updated_category and
            existing_status == updated_status and
            existing_stock_level == updated_stock_level and
            abs(existing_price - updated_price) < 0.01):  # Float comparison with tolerance
            error_response = {
                "success": False,
                "message": f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {updated_stock_level}, Price: {updated_price}) already exists."
            }
            if wants_json():
                return jsonify(error_response), 409
            else:
                return jsonify(error_response), 409

    save_products(products)
    
    response_data = {
        "success": True,
        "message": "Product updated successfully",
        "data": product
    }
    
    if wants_json():
        return jsonify(response_data), 200
    else:
        return jsonify(response_data), 200


# =========================================
# 5. MASTERS — Products (continued)
# =========================================
@app.route("/products/create")
def create_new_product_page():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "create-new-product.html",
        title="Create Product - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


@app.route("/download-template")
def download_template():
    # 1. Create empty dataframe
    df = pd.DataFrame(columns=[
        "Product ID",
        "Product Name",
        "Type",
        "Category",
        "Status",
        "Stock Level",
        "Price"
    ])

    # 2. Save to Excel in memory
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Products")
    output.seek(0)

    # 3. Load workbook to add validations
    wb = load_workbook(output)
    ws = wb.active

    # ---------- DROPDOWN VALIDATIONS ----------
    type_validation = DataValidation(
        type="list",
        formula1='"Physical,Digital"',
        allow_blank=False
    )

    category_validation = DataValidation(
        type="list",
        formula1='"Electronics,Clothing,Food,Furniture"',
        allow_blank=False
    )

    status_validation = DataValidation(
        type="list",
        formula1='"Active,Inactive"',
        allow_blank=False
    )

    # ---------- NUMBER VALIDATIONS ----------
    stock_validation = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Stock Level",
        error="Stock Level must be a whole number (0 or greater)."
    )

    price_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Price",
        error="Price must be a number greater than 0."
    )

    product_id_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Product ID",
        error="Product ID must contain ONLY numbers (no letters or special characters)."
    )

    product_name_validation = DataValidation(
        type="custom",
        # Custom validation formulas in openpyxl should NOT include a leading "="
        formula1=(
            'AND(B2<>"",'
            'NOT(OR('
            'ISNUMBER(SEARCH("0",B2)),'
            'ISNUMBER(SEARCH("1",B2)),'
            'ISNUMBER(SEARCH("2",B2)),'
            'ISNUMBER(SEARCH("3",B2)),'
            'ISNUMBER(SEARCH("4",B2)),'
            'ISNUMBER(SEARCH("5",B2)),'
            'ISNUMBER(SEARCH("6",B2)),'
            'ISNUMBER(SEARCH("7",B2)),'
            'ISNUMBER(SEARCH("8",B2)),'
            'ISNUMBER(SEARCH("9",B2))'
            '))'
            ')'
        ),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Product Name",
        error="Only alphabets and spaces are allowed. Numbers are not permitted."
    )

    # Add validations to worksheet
    ws.add_data_validation(product_id_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(category_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(stock_validation)
    ws.add_data_validation(price_validation)
    ws.add_data_validation(product_name_validation)

    # Apply validation to columns (row 2 to 1000)
    product_id_validation.add("A2:A1000")
    product_name_validation.add("B2:B1000")
    type_validation.add("C2:C1000")
    category_validation.add("D2:D1000")
    status_validation.add("E2:E1000")
    stock_validation.add("F2:F1000")
    price_validation.add("G2:G1000")

    # 4. Save workbook again
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # 5. Download file
    return send_file(
        final_output,
        as_attachment=True,
        download_name="Product_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# public.customers columns for import template (exclude id, created_at).
# Base headers match existing import validation; remaining DB fields after City, no duplicates.
CUSTOMER_IMPORT_TEMPLATE_BASE_COLUMNS = [
    "Customer ID",
    "Name",
    "Company",
    "Email",
    "Status",
    "Credit Limit",
    "City",
]
CUSTOMER_IMPORT_TEMPLATE_COLUMNS_AFTER_CITY = [
    "Customer Type",
    "phone",
    "first_name",
    "last_name",
    "sales_rep",
    "billing_address",
    "shipping_address",
    "street",
    "state",
    "zip_code",
    "country",
    "payment_terms",
    "credit_term",
    "gstin",
]


def _customer_import_template_column_headers():
    """Return ordered template headers: base columns plus extras not already represented."""
    headers = []
    seen = set()
    for h in CUSTOMER_IMPORT_TEMPLATE_BASE_COLUMNS:
        key = h.strip().lower()
        if key not in seen:
            seen.add(key)
            headers.append(h)
    # Status column maps to customer_status in DB — treat as covered
    seen.add("customer_status")
    for h in CUSTOMER_IMPORT_TEMPLATE_COLUMNS_AFTER_CITY:
        key = h.strip().lower()
        if key not in seen:
            seen.add(key)
            headers.append(h)
    return headers


def _customer_import_all_required_columns():
    """Headers that must exist in upload/import files (base + extra template columns)."""
    return list(CUSTOMER_IMPORT_TEMPLATE_BASE_COLUMNS) + list(
        CUSTOMER_IMPORT_TEMPLATE_COLUMNS_AFTER_CITY
    )


def _validate_customer_import_extra_fields(row):
    """
    Mandatory validation for template columns after City (snake_case headers).
    Returns a list of error strings (empty if valid).
    """
    errs = []

    def blank(v):
        if v is None:
            return True
        try:
            if pd.isna(v):
                return True
        except (TypeError, ValueError):
            pass
        return str(v).strip() == ""

    # Customer Type (after City) — restricted template list
    ct_raw = row.get("Customer Type")
    if blank(ct_raw):
        errs.append("Customer Type is required")
    else:
        c = _CUSTOMER_TEMPLATE_CUSTOMER_TYPE_BY_LOWER.get(str(ct_raw).strip().lower())
        if c is None:
            errs.append(
                "Customer Type must be one of: "
                f"{', '.join(_CUSTOMER_TEMPLATE_CUSTOMER_TYPE_ALLOWED)}"
            )

    ph_raw = row.get("phone")
    if blank(ph_raw):
        errs.append("phone is required")
    else:
        ph = re.sub(r"\D", "", str(ph_raw))
        if len(ph) != 10:
            errs.append("phone must be exactly 10 digits")

    for key in ("first_name", "last_name"):
        v = row.get(key)
        if blank(v):
            errs.append(f"{key} is required")
        else:
            s = str(v).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", s):
                errs.append(f"{key} must contain only letters and spaces")
            elif len(s) < 3:
                errs.append(f"{key} must be at least 3 characters")
            elif len(s) > 100:
                errs.append(f"{key} must not exceed 100 characters")

    if blank(row.get("sales_rep")):
        errs.append("sales_rep is required")
    elif len(str(row.get("sales_rep")).strip()) > 100:
        errs.append("sales_rep must not exceed 100 characters")

    for key in ("billing_address", "shipping_address"):
        v = row.get(key)
        if blank(v):
            errs.append(f"{key} is required")
        elif len(str(v).strip()) < 3:
            errs.append(f"{key} must be at least 3 characters")
        elif len(str(v).strip()) > 2000:
            errs.append(f"{key} is too long")

    st = row.get("street")
    if blank(st):
        errs.append("street is required")
    elif len(str(st).strip()) < 3:
        errs.append("street must be at least 3 characters")
    elif len(str(st).strip()) > 150:
        errs.append("street must not exceed 150 characters")

    stt = row.get("state")
    if blank(stt):
        errs.append("state is required")
    else:
        s = str(stt).strip()
        if not re.fullmatch(r"^[A-Za-z ]+$", s):
            errs.append("state must contain only letters and spaces")
        elif len(s) < 3:
            errs.append("state must be at least 3 characters")
        elif len(s) > 100:
            errs.append("state must not exceed 100 characters")

    z = row.get("zip_code")
    if blank(z):
        errs.append("zip_code is required")
    else:
        zs = re.sub(r"\D", "", str(z))
        if len(zs) != 6:
            errs.append("zip_code must be exactly 6 digits")

    ctry = row.get("country")
    if blank(ctry):
        errs.append("country is required")
    else:
        s = str(ctry).strip()
        if not re.fullmatch(r"^[A-Za-z ]+$", s):
            errs.append("country must contain only letters and spaces")
        elif len(s) < 3:
            errs.append("country must be at least 3 characters")
        elif len(s) > 50:
            errs.append("country must not exceed 50 characters")

    for key in ("payment_terms", "credit_term"):
        v = row.get(key)
        if blank(v):
            errs.append(f"{key} is required")
        elif len(str(v).strip()) > 50:
            errs.append(f"{key} must not exceed 50 characters")

    g = row.get("gstin")
    if blank(g):
        errs.append("gstin is required")
    else:
        gs = re.sub(r"\s", "", str(g).strip().upper())
        if not re.fullmatch(r"^[0-9A-Z]{15,20}$", gs):
            errs.append("gstin must be 15-20 alphanumeric characters")

    return errs


@app.route("/download-customer-template")
def download_customer_template():

    # 1. Empty dataframe (all schema fields except id, created_at; extras after City)
    df = pd.DataFrame(columns=_customer_import_template_column_headers())

    # 2. Save to memory
    output = io.BytesIO()
    df.to_excel(output, index=False, sheet_name="Customers")
    output.seek(0)

    # 3. Load workbook
    wb = load_workbook(output)
    ws = wb.active

    # ---------------- DROPDOWNS ----------------
    customer_type_validation = DataValidation(
        type="list",
        formula1='"Individual,Business,Organization,Corporate"',
        allow_blank=False
    )

    status_validation = DataValidation(
        type="list",
        formula1='"Active,Inactive"',
        allow_blank=False
    )

    # ---------------- CUSTOMER ID VALIDATION (NUMBERS ONLY) ----------------
    # Match the simple numeric rule used for Product ID:
    # Allow: Whole number, greater than 0
    customer_id_validation = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Invalid Customer ID",
        error="Customer ID must contain ONLY numbers greater than 0 (no letters or special characters)."
    )





    

    credit_limit_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Credit Limit",
        error="Credit Limit must be 0 or greater."
    )

    # ---------------- TEXT VALIDATIONS ----------------
    # Name validation (Customer template) – mirror the Product Name rule:
    # =AND(B2<>"",NOT(OR(ISNUMBER(SEARCH("0",B2)),...,ISNUMBER(SEARCH("9",B2)))))
    # NOTE: For openpyxl custom validations, formula1 should NOT start with "=".
    name_validation = DataValidation(
        type="custom",
        formula1=(
            'AND(B2<>"",'
            'NOT(OR('
            'ISNUMBER(SEARCH("0",B2)),'
            'ISNUMBER(SEARCH("1",B2)),'
            'ISNUMBER(SEARCH("2",B2)),'
            'ISNUMBER(SEARCH("3",B2)),'
            'ISNUMBER(SEARCH("4",B2)),'
            'ISNUMBER(SEARCH("5",B2)),'
            'ISNUMBER(SEARCH("6",B2)),'
            'ISNUMBER(SEARCH("7",B2)),'
            'ISNUMBER(SEARCH("8",B2)),'
            'ISNUMBER(SEARCH("9",B2))'
            '))'
            ')'
        ),
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Name",
        error="Name must contain only alphabets (letters) and spaces. Numbers and special characters are not allowed."
    )

    city_validation = DataValidation(
        type="custom",
        formula1='=AND(G2<>"",NOT(ISNUMBER(SEARCH("0",G2))))',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid City",
        error="City must contain only letters."
    )

    email_validation = DataValidation(
        type="custom",
        formula1='=AND(D2<>"",ISNUMBER(SEARCH("@",D2)),ISNUMBER(SEARCH(".",D2)))',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Email",
        error="Enter a valid email address."
    )

    # Resolve column letters by header name so template changes stay safe.
    header_to_col = {cell.value: cell.column_letter for cell in ws[1] if cell.value}
    customer_type_col = header_to_col.get("Customer Type")

    # Add validations
    ws.add_data_validation(customer_id_validation)
    ws.add_data_validation(name_validation)
    ws.add_data_validation(customer_type_validation)
    ws.add_data_validation(email_validation)
    ws.add_data_validation(status_validation)
    ws.add_data_validation(credit_limit_validation)
    ws.add_data_validation(city_validation)

    # Apply validations (Row 2–1000)
    customer_id_validation.add("A2:A1000")
    name_validation.add("B2:B1000")
    email_validation.add("D2:D1000")
    status_validation.add("E2:E1000")
    credit_limit_validation.add("F2:F1000")
    city_validation.add("G2:G1000")
    if customer_type_col:
        customer_type_validation.add(f"{customer_type_col}2:{customer_type_col}1000")

    # Save final file
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(
        final_output,
        as_attachment=True,
        download_name="Customer_Import_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/import", methods=["GET", "POST"])
def import_products():
    user_email = session.get("user")
    if not user_email:
        if request.method == "POST":
            return jsonify({
                "status": "error",
                "message": "Session expired. Please login first."
            }), 401
        return redirect(url_for("login", message="session_expired"))

    if request.method == "POST":
        file = request.files.get("file")

        if not file:
            return jsonify({
                "status": "error",
                "message": "No file uploaded"
            })

        rel_path = _upload_basename(file.filename, secure=True)
        _persist_module_upload(
            object_storage.MODULE_IMPORTS,
            IMPORT_UPLOAD_FOLDER,
            file,
            rel_path,
        )

        # Dummy response (replace later)
        return jsonify({
            "status": "success",
            "valid": 120,
            "invalid": 5,
            "skipped": 2,
            "errors": [
                'Missing "UOM" in Row 10,12,13',
                'Invalid GST Rate in Row 18'
            ]
        })

    # GET → open import page
    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "import-product.html",
        title="Import Products - Stackly",
        page="products",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


@app.get("/import-product")
def import_product_metadata():
    """Small JSON endpoint so /import page has a named Fetch/XHR entry."""
    user_email = session.get("user")
    if not user_email:
        return jsonify(
            {"success": False, "message": "Session expired. Please login first."}
        ), 401

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return jsonify(
        {
            "success": True,
            "page": "import-product",
            "current_user": {"email": user_email, "name": user_name},
        }
    ), 200


def _product_row_signature_from_excel(row) -> tuple:
    """
    Canonical 6-tuple for product duplicate detection (import + validation).
    Text fields compared case-insensitively; stock/price normalized so Excel 12 vs 12.0 match.
    """
    def _s(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        return str(v).strip()

    name = _s(row.get("Product Name"))
    t = _s(row.get("Type"))
    cat = _s(row.get("Category"))
    st = _s(row.get("Status"))
    stock_raw = row.get("Stock Level")
    price_raw = row.get("Price")
    try:
        stock = int(float(stock_raw)) if not pd.isna(stock_raw) else 0
    except (TypeError, ValueError):
        stock = 0
    try:
        price = float(price_raw) if not pd.isna(price_raw) else 0.0
    except (TypeError, ValueError):
        price = 0.0
    price_r = round(price + 1e-12, 2)
    return (name.lower(), t.lower(), cat.lower(), st.lower(), stock, price_r)


def _product_signature_from_stored(p: dict) -> tuple:
    """Same tuple as _product_row_signature_from_excel for products in product.json."""
    try:
        stock = int(float(p.get("stock_level", 0) or 0))
    except (TypeError, ValueError):
        stock = 0
    try:
        price = float(p.get("price", 0) or 0)
    except (TypeError, ValueError):
        price = 0.0
    price_r = round(price + 1e-12, 2)
    return (
        str(p.get("product_name") or "").strip().lower(),
        str(p.get("type") or "").strip().lower(),
        str(p.get("category") or "").strip().lower(),
        str(p.get("status") or "").strip().lower(),
        stock,
        price_r,
    )


def _product_import_filename_ok(filename):
    if not filename or not str(filename).strip():
        return False
    fn = str(filename).strip().lower()
    return fn.endswith(".csv") or fn.endswith(".xlsx")


def _read_product_import_dataframe(file):
    """Read uploaded product import file; only .csv and .xlsx are allowed."""
    if not _product_import_filename_ok(file.filename):
        return None, "Invalid file format. Only .csv and .xlsx files are allowed."
    try:
        file.seek(0)
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        return df, None
    except Exception:
        return None, "Invalid or corrupt file. Please upload a valid CSV or XLSX file."


@app.route("/upload", methods=["POST"])
def upload_file():
    file = request.files.get("file")

    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    df, read_err = _read_product_import_dataframe(file)
    if read_err:
        return jsonify({"error": read_err}), 400

    # Check if file is empty (no data rows, only headers)
    if df.empty or len(df) == 0:
        return jsonify({
            "error": "No data found",
            "message": "The uploaded file contains no data. Please ensure the file has at least one row of product data."
        }), 400

    # Ensure required columns exist in the uploaded template
    required_columns = [
        "Product ID",
        "Product Name",
        "Type",
        "Category",
        "Status",
        "Stock Level",
        "Price",
    ]
    missing_cols = [c for c in required_columns if c not in df.columns]
    if missing_cols:
        return (
            jsonify(
                {
                    "error": "Invalid template",
                    "message": "The uploaded Excel file does not match the required product import template.",
                    "missing_columns": missing_cols,
                }
            ),
            400,
        )

    valid_rows = 0
    invalid_rows = 0
    skipped_rows = 0
    error_details = []
    skipped_row_numbers = []  # Track which rows were skipped (completely blank)
    
    
    # Load existing products to check against database
    existing_products = load_products()
    existing_product_ids = {str(p.get("product_id", "")) for p in existing_products if p.get("product_id")}
    existing_signatures = {_product_signature_from_stored(p) for p in existing_products}
    
    # Track Product IDs for uniqueness validation within uploaded file
    seen_product_ids = {}  # key: Product ID (as string), value: first row number
    
    # Track seen row combinations for duplicate detection (excluding Product ID)
    seen_rows = {}  # signature -> first Excel row number

    # Helper to check blank (NaN or empty/whitespace)
    def is_blank(val):
        if pd.isna(val):
            return True
        return str(val).strip() == ""
    
    # Helper to normalize value for comparison
    def normalize_value(val):
        if pd.isna(val):
            return ""
        return str(val).strip()

    for index, row in df.iterrows():
        errors = []

        # Treat rows where ALL A–G columns are blank as skipped
        if all(
            is_blank(row.get(col_name))
            for col_name in ["Product ID", "Product Name", "Type", "Category", "Status", "Stock Level", "Price"]
        ):
            skipped_rows += 1
            skipped_row_numbers.append(index + 2)  # Store row number (Excel row, +2 for header)
            # Don't validate or count as valid/invalid; just skip
            continue

        # --- Product ID (optional - will be auto-generated if blank, must be unique if provided) ---
        pid_raw = row.get("Product ID")
        if not is_blank(pid_raw):
            # Only validate if Product ID is provided
            pid_str = str(pid_raw).strip()
            
            # Check if Product ID already has "P" prefix (e.g., "P124")
            if pid_str.upper().startswith("P") and len(pid_str) > 1:
                # Extract numeric part after "P"
                numeric_part = pid_str[1:].strip()
                try:
                    pid_num = float(numeric_part)
                    if not pid_num.is_integer() or int(pid_num) <= 0:
                        errors.append("Product ID must be a valid number after 'P' prefix")
                    else:
                        # Normalize to uppercase (P124)
                        pid_str = f"P{int(pid_num)}"
                except (ValueError, TypeError):
                    errors.append("Product ID must be a valid number after 'P' prefix")
            else:
                # Try to convert to number (handles both "12" and "12.0" from Excel)
                try:
                    pid_num = float(pid_str)
                    # Check if it's a whole number (no decimal part)
                    if not pid_num.is_integer():
                        errors.append("Product ID must be a whole number")
                    elif int(pid_num) <= 0:
                        errors.append("Product ID must be greater than 0")
                    else:
                        # Prepend "P" to numeric Product ID (e.g., 124 becomes P124)
                        pid_str = f"P{int(pid_num)}"
                except (ValueError, TypeError):
                    # If conversion fails, it's not a valid number
                    errors.append("Product ID must be a whole number")
                    pid_str = None
            
            # Validate uniqueness if Product ID is valid
            if pid_str and not any("Product ID" in err for err in errors):
                # Check if Product ID already exists in the uploaded file
                if pid_str in seen_product_ids:
                    first_row = seen_product_ids[pid_str]
                    errors.append(f"Duplicate Product ID: Product ID {pid_str} already exists in row {first_row}")
                # Check if Product ID already exists in the database
                elif pid_str in existing_product_ids:
                    errors.append(f"Duplicate Product ID: Product ID {pid_str} already exists in the system")
                else:
                    seen_product_ids[pid_str] = index + 2  # Store the row number (Excel row, +2 for header)
            # Note: If Product ID is blank, it will be auto-generated during import (P101, P102, etc.)

        # --- Product Name (mandatory, alphabets + spaces, min length 3) ---
        pname_raw = row.get("Product Name")
        if is_blank(pname_raw):
            errors.append("Product Name is required")
        else:
            product_name = str(pname_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", product_name):
                errors.append("Product Name must contain ONLY letters and spaces")
            elif len(product_name) < 3:
                errors.append("Product Name must be at least 3 characters")

        # --- Type (mandatory) ---
        type_raw = row.get("Type")
        if is_blank(type_raw):
            errors.append("Type is required")

        # --- Category (mandatory) ---
        category_raw = row.get("Category")
        if is_blank(category_raw):
            errors.append("Category is required")

        # --- Status (mandatory) ---
        status_raw = row.get("Status")
        if is_blank(status_raw):
            errors.append("Status is required")

        # --- Stock Level (mandatory, whole number >= 0) ---
        stock_raw = row.get("Stock Level")
        if is_blank(stock_raw):
            errors.append("Stock Level is required")
        else:
            try:
                stock_num = float(stock_raw)
                if not stock_num.is_integer():
                    errors.append("Stock Level must be a whole number")
                elif stock_num < 0:
                    errors.append("Stock Level must be 0 or greater")
            except Exception:
                errors.append("Stock Level must be a number")

        # --- Price (mandatory, number > 0) ---
        price_raw = row.get("Price")
        if is_blank(price_raw):
            errors.append("Price is required")
        else:
            try:
                price_num = float(price_raw)
                if price_num <= 0:
                    errors.append("Price must be greater than 0")
            except Exception:
                errors.append("Price must be a valid number")

        # --- Duplicate Row Check (same 6-field signature as save_product; normalized numbers) ---
        sig = _product_row_signature_from_excel(row)
        if sig in seen_rows:
            first_row = seen_rows[sig]
            errors.append(
                f"Duplicate row: This combination of Product Name, Type, Category, Status, Stock Level, and Price is identical to row {first_row}"
            )
        elif sig in existing_signatures:
            errors.append(
                "Duplicate product: this combination already exists in the system (same Name, Type, Category, Status, Stock Level, and Price)."
            )
        else:
            if any(sig[:4]) or sig[4] != 0 or sig[5] != 0.0:
                seen_rows[sig] = index + 2

        if errors:
            invalid_rows += 1
            error_details.append({
                "row": index + 2,  # +2 because header is row 1
                "errors": errors
            })
        else:
            valid_rows += 1

    return jsonify({
        "total_rows": len(df),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "skipped_rows": skipped_rows,
        "skipped_row_numbers": skipped_row_numbers,  # List of row numbers that were skipped
        "error_details": error_details
    })

@app.route("/import-products-validated", methods=["POST"])
def import_products_validated():
    file = request.files.get("file")

    if not file:
        return jsonify({"success": False, "message": "No file uploaded"}), 400

    df, read_err = _read_product_import_dataframe(file)
    if read_err:
        return jsonify({"success": False, "message": read_err}), 400

    products = load_products()
    
    # Get existing product IDs to determine next sequence number
    existing_ids = {str(p.get("product_id", "")).strip() for p in products if p.get("product_id")}
    existing_sigs = {_product_signature_from_stored(p) for p in products}
    batch_sigs = set()
    skipped_duplicates = 0
    
    # Find the maximum numeric value from existing product IDs
    max_num = 0
    for pid in existing_ids:
        # Extract numeric part from formats like "P101", "101", "P-101", etc.
        match = re.search(r"(\d+)$", pid)
        if match:
            max_num = max(max_num, int(match.group(1)))
    
    added = 0
    current_sequence = 0

    for _, row in df.iterrows():
        errors = []

        # Product Name validation (same rules as /upload)
        raw_name = row.get("Product Name")
        if pd.isna(raw_name):
            errors.append("Product Name is required")
        else:
            product_name = str(raw_name).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", product_name):
                errors.append("Product Name must contain ONLY letters and spaces")
            elif len(product_name) < 3:
                errors.append("Product Name must be at least 3 characters")

        if errors:
            # Skip invalid rows, we only import validated rows
            continue

        # Skip duplicate rows (same 6-field signature as /upload and save_product)
        sig = _product_row_signature_from_excel(row)
        if sig in batch_sigs or sig in existing_sigs:
            skipped_duplicates += 1
            continue

        # Basic extraction of remaining fields
        raw_id = row.get("Product ID")
        product_id = "" if pd.isna(raw_id) else str(raw_id).strip()

        # Auto-generate Product ID if blank
        if not product_id:
            # Generate next Product ID based on existing products + already added in this batch
            current_sequence += 1
            product_id = f"P{max_num + current_sequence}"
        else:
            # If Product ID is provided, normalize it to "P###" format
            # Check if Product ID already has "P" prefix (case-insensitive)
            if product_id.upper().startswith("P") and len(product_id) > 1:
                # Extract numeric part after "P"
                numeric_part = product_id[1:].strip()
                try:
                    pid_num = float(numeric_part)
                    if pid_num.is_integer() and int(pid_num) > 0:
                        # Normalize to "P###" format (e.g., "P124", "p124" -> "P124")
                        product_id = f"P{int(pid_num)}"
                        num_val = int(pid_num)
                        max_num = max(max_num, num_val)
                    else:
                        # Invalid format, keep as-is but try to extract numeric part for max_num
                        match = re.search(r"(\d+)$", product_id)
                        if match:
                            num_val = int(match.group(1))
                            max_num = max(max_num, num_val)
                except (ValueError, TypeError):
                    # Invalid format, keep as-is but try to extract numeric part for max_num
                    match = re.search(r"(\d+)$", product_id)
                    if match:
                        num_val = int(match.group(1))
                        max_num = max(max_num, num_val)
            else:
                # Product ID doesn't start with "P" - check if it's purely numeric
                try:
                    # Try to convert to number (handles both "124" and "124.0" from Excel)
                    pid_num = float(product_id)
                    if pid_num.is_integer() and int(pid_num) > 0:
                        # Prepend "P" to numeric Product ID (e.g., 124 becomes P124)
                        product_id = f"P{int(pid_num)}"
                        num_val = int(pid_num)
                        max_num = max(max_num, num_val)
                    else:
                        # Not a valid positive integer, keep as-is but try to extract numeric part
                        match = re.search(r"(\d+)$", product_id)
                        if match:
                            num_val = int(match.group(1))
                            max_num = max(max_num, num_val)
                except (ValueError, TypeError):
                    # If conversion fails, it might be alphanumeric, extract numeric part if present
                    match = re.search(r"(\d+)$", product_id)
                    if match:
                        num_val = int(match.group(1))
                        max_num = max(max_num, num_val)
                    # If no numeric part found, keep as-is

        type_val = "" if pd.isna(row.get("Type")) else str(row.get("Type")).strip()
        category_val = "" if pd.isna(row.get("Category")) else str(row.get("Category")).strip()
        status_val = "" if pd.isna(row.get("Status")) else str(row.get("Status")).strip() or "Active"

        stock_raw = row.get("Stock Level")
        price_raw = row.get("Price")

        try:
            stock_level = int(stock_raw) if not pd.isna(stock_raw) else 0
        except Exception:
            stock_level = 0

        try:
            price = float(price_raw) if not pd.isna(price_raw) else 0.0
        except Exception:
            price = 0.0

        item = {
            "product_id": product_id,
            "product_name": product_name,
            "type": type_val,
            "category": category_val,
            "status": status_val,
            "stock_level": stock_level,
            "price": price,
        }

        products.append(item)
        added += 1
        batch_sigs.add(sig)
        existing_sigs.add(sig)

    save_products(products)

    msg = f"Successfully imported {added} product(s)"
    if skipped_duplicates:
        msg += f" ({skipped_duplicates} duplicate row(s) skipped)"
    return jsonify(
        {
            "success": True,
            "added": added,
            "skipped_duplicates": skipped_duplicates,
            "message": msg,
        }
    )

  
@app.route("/save-product", methods=["POST"])
def save_product():
    try:
        product_id = generate_product_id()

        # ------- safe converters so empty / bad values won't crash ------
        def to_int(val, default=0):
            try:
                return int(val)
            except (TypeError, ValueError):
                return default

        def to_float(val, default=0.0):
            try:
                return float(val)
            except (TypeError, ValueError):
                return default

        # -------------------- IMAGE HANDLING ----------------------------
        image = request.files.get("product_image")
        image_filename = None

        if image and image.filename:
            rel_path = _upload_relative_path(str(product_id), image.filename, secure=True)
            stored, _ = _persist_module_upload(
                object_storage.MODULE_PRODUCT_IMAGES,
                PRODUCT_IMAGES_FOLDER,
                image,
                rel_path,
            )
            image_filename = stored

        # -------------------- PRODUCT OBJECT ----------------------------
        form = request.form

        product = {
            "product_id": str(product_id),
            "product_name": (form.get("product_name") or "").strip(),
            "type": (form.get("product_type") or "").strip(),
            "category": (form.get("category") or "").strip(),
            "status": (form.get("status") or "active").strip(),

            # fields used in table
            "stock_level": to_int(form.get("stock_level")),          
            "price": to_float(form.get("unit_price")),               

            # extra fields (for future use)
            "description": (form.get("description") or "").strip(),
            "sub_category": (form.get("sub_category") or "").strip(),
            "unit_price": (form.get("unit_price") or "").strip(),
            "discount": (form.get("discount") or "").strip(),
            "tax_code": (form.get("tax_code") or "").strip(),
            "quantity": (form.get("quantity") or "").strip(),
            "uom": (form.get("uom") or "").strip(),
            "reorder_level": (form.get("reorder_level") or "").strip(),
            "warehouse": (form.get("warehouse") or "").strip(),
            "size": (form.get("size") or "").strip(),
            "color": (form.get("color") or "").strip(),
            "weight": (form.get("weight") or "").strip(),
            "specifications": (form.get("specifications") or "").strip(),
            "related_products": (form.get("related_products") or "").strip(),
            "supplier": (form.get("supplier") or "").strip(),
            "product_usage": (form.get("product_usage") or "").strip(),
            "image": image_filename,
        }

        # -------------------- DUPLICATE VALIDATION -----------------------------
        products = load_products()
        
        # Normalize values for comparison (case-insensitive for text, exact for numbers)
        new_product_name = (product.get("product_name") or "").strip().lower()
        new_type = (product.get("type") or "").strip().lower()
        new_category = (product.get("category") or "").strip().lower()
        new_status = (product.get("status") or "").strip().lower()
        new_stock_level = product.get("stock_level", 0)
        new_price = product.get("price", 0.0)
        
        # Check 1: Duplicate product name (case-insensitive)
        for existing in products:
            existing_name = (existing.get("product_name") or "").strip().lower()
            if existing_name == new_product_name:
                return jsonify(
                    success=False, 
                    message=f"Product with name '{product.get('product_name')}' already exists. Please use a different product name."
                ), 409
        
        # Check 2: Duplicate combination (product_name + type + category + status + stock_level + price)
        for existing in products:
            existing_name = (existing.get("product_name") or "").strip().lower()
            existing_type = (existing.get("type") or "").strip().lower()
            existing_category = (existing.get("category") or "").strip().lower()
            existing_status = (existing.get("status") or "").strip().lower()
            existing_stock_level = existing.get("stock_level", 0)
            existing_price = existing.get("price", 0.0)
            
            # Compare all fields (case-insensitive for text, exact for numbers)
            if (existing_name == new_product_name and
                existing_type == new_type and
                existing_category == new_category and
                existing_status == new_status and
                existing_stock_level == new_stock_level and
                abs(existing_price - new_price) < 0.01):  # Float comparison with tolerance
                return jsonify(
                    success=False,
                    message=f"A product with the same combination (Name: '{product.get('product_name')}', Type: '{product.get('type')}', Category: '{product.get('category')}', Status: '{product.get('status')}', Stock Level: {new_stock_level}, Price: {new_price}) already exists."
                ), 409
        
        # -------------------- SAVE TO DB (single-row insert) -----------------------------
        tax_code = product.get("tax_code", "")
        tax_percent = None
        if tax_code:
            m = re.search(r"\((\d+(?:\.\d+)?)%\)", str(tax_code))
            if m:
                try:
                    tax_percent = float(m.group(1))
                except ValueError:
                    tax_percent = None

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                INSERT INTO products (
                    product_id, product_name, product_type, category_name,
                    unit_price, discount, description, sub_category,
                    quantity, stock_level, reorder_level,
                    weight, specifications, related_products,
                    status, product_usage, image,
                    tax_code, tax_percent, tax_description,
                    uom_name, uom_items, uom_description,
                    warehouse_name, warehouse_location, warehouse_manager,
                    warehouse_contact, warehouse_notes,
                    size, color,
                    supplier_name, supplier_contact, supplier_phone,
                    supplier_email, supplier_address
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s
                )
                """,
                (
                    product.get("product_id"),
                    product.get("product_name"),
                    product.get("type", ""),
                    product.get("category", ""),
                    to_float(form.get("unit_price")),
                    to_float(form.get("discount")),
                    product.get("description", ""),
                    product.get("sub_category", ""),
                    to_int(form.get("quantity")),
                    to_int(form.get("stock_level")),
                    to_int(form.get("reorder_level")),
                    product.get("weight", ""),
                    product.get("specifications", ""),
                    product.get("related_products", ""),
                    (product.get("status") or "Active").strip(),
                    product.get("product_usage", ""),
                    product.get("image", ""),
                    tax_code,
                    tax_percent,
                    "",
                    product.get("uom", ""),
                    0,
                    "",
                    product.get("warehouse", ""),
                    "",
                    "",
                    "",
                    "",
                    product.get("size", ""),
                    product.get("color", ""),
                    product.get("supplier", ""),
                    "",
                    "",
                    "",
                    "",
                ),
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return jsonify(success=True, product_id=product_id)

    except Exception as e:
        
        print("ERROR in /save-product:", e)
        return jsonify(success=False, message="Internal server error"), 500



# =========================================
# 6. MASTERS — Customer
# =========================================
def get_customers_from_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    SELECT 
        customer_id,
        name,
        first_name,
        last_name,
        company,
        customer_type,
        customer_status,
        email,
        phone,
        credit_limit,
        city,
        sales_rep,
        street,
        state,
        zip_code,
        country,
        payment_terms,
        credit_term,
        gstin,
        available_limit,
        billing_address,
        shipping_address
    FROM customers
""")

    rows = cur.fetchall()

    customers = []
    for r in rows:
        ct = r[5]
        customers.append({
        "customer_id": r[0],
        "name": r[1],
        "first_name": r[2],
        "last_name": r[3],
        "company": r[4],
        "customer_type": ct,
        "status": r[6],
        "email": r[7],
        "phone": r[8],
        "credit_limit": float(r[9] or 0),
        "city": r[10],
        "sales_rep": r[11],
        "company_type": ct,
        "street": r[12],
        "state": r[13],
        "zip_code": r[14],
        "country": r[15],
        "payment_terms": r[16],
        "credit_term": r[17],
        "gstin": r[18],
        "available_limit": float(r[19] or 0),
        "billing_address": r[20],
        "shipping_address": r[21]
    })

    cur.close()
    conn.close()
    return customers


@app.route("/customer")
def customer():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired"}), 401

    prof = get_current_user_profile() or {}
    user_name = prof.get("name") or "User"
    user_role = prof.get("role") or "User"

    # ✅ THIS PART CHANGE
    if request.args.get("format") == "json":
        customers_list = get_customers_from_db()
        return jsonify({
            "success": True,
            "data": {
                "items": customers_list
            }
        }), 200

    return render_template(
        "customer.html",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
    )

@app.route("/import-customer", methods=["GET", "POST"])
def import_customer():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    
    if request.method == "GET" and wants_json():
        return jsonify(
            {
                "success": True,
                "page": "import-customer",
                "current_user": {"email": user_email, "name": user_name},
            }
        ), 200

    return render_template(
        "import-customer.html",
        title="Import Customers - Stackly",
        page="customer",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )

_CUSTOMER_TYPE_ALLOWED = (
    "Retail",
    "Wholesale",
    "Corporate",
    "Online",
    "Distributor",
    "Individual",
    "Business",
    "Organization",
)
_CUSTOMER_TYPE_BY_LOWER = {t.lower(): t for t in _CUSTOMER_TYPE_ALLOWED}
_CUSTOMER_TEMPLATE_CUSTOMER_TYPE_ALLOWED = (
    "Individual",
    "Business",
    "Organization",
    "Corporate",
)
_CUSTOMER_TEMPLATE_CUSTOMER_TYPE_BY_LOWER = {
    t.lower(): t for t in _CUSTOMER_TEMPLATE_CUSTOMER_TYPE_ALLOWED
}


def normalize_customer_type(value):
    """Return canonical customer type, or None if empty/invalid. Accepts any casing (e.g. DISTRIBUTOR -> Distributor)."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    if not s:
        return None
    return _CUSTOMER_TYPE_BY_LOWER.get(s.lower())


def _customer_row_signature_from_excel(row) -> tuple:
    """
    Canonical 7-tuple for customer duplicate detection (same logical row as upload duplicate check).
    Excludes Customer ID. Credit limit normalized like product price (Excel 100 vs 100.0).
    """
    def _s(v):
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        return str(v).strip()

    name = _s(row.get("Name")).lower()
    company = _s(row.get("Company")).lower()
    ct_raw = row.get("Customer Type")
    if _s(ct_raw) == "":
        ctype_key = ""
    else:
        cn = normalize_customer_type(ct_raw)
        ctype_key = (cn or _s(ct_raw)).lower()

    email = _s(row.get("Email")).lower()
    status = _s(row.get("Status")).lower()
    cr = row.get("Credit Limit")
    try:
        credit = float(cr) if not pd.isna(cr) else 0.0
    except (TypeError, ValueError):
        credit = 0.0
    credit_r = round(credit + 1e-12, 2)
    city = _s(row.get("City")).lower()
    return (name, company, ctype_key, email, status, credit_r, city)


def _customer_signature_from_stored(c: dict) -> tuple:
    """Same tuple as _customer_row_signature_from_excel for rows in customer.json."""
    cr_raw = c.get("credit_limit", "")
    try:
        credit = float(str(cr_raw).replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        credit = 0.0
    credit_r = round(credit + 1e-12, 2)
    ct = normalize_customer_type(c.get("customer_type") or c.get("company_type"))
    if ct is None:
        ctype_key = str(c.get("customer_type") or c.get("company_type") or "").strip().lower()
    else:
        ctype_key = ct.lower()
    return (
        str(c.get("name") or "").strip().lower(),
        str(c.get("company") or "").strip().lower(),
        ctype_key,
        str(c.get("email") or "").strip().lower(),
        str(c.get("status") or "").strip().lower(),
        credit_r,
        str(c.get("city") or "").strip().lower(),
    )


@app.route("/import-customers-validated", methods=["POST"])
def import_customers_validated():
    user_email = session.get("user")
    if not user_email:
        return jsonify(success=False, message="Session expired. Please login."), 401

    file = request.files.get("file")
    if not file or file.filename.strip() == "":
        return jsonify(success=False, message="No file uploaded"), 400

    df, read_err = _read_customer_import_dataframe(file)
    if read_err:
        return jsonify(success=False, message=read_err), 400

    required_columns = _customer_import_all_required_columns()
    for col in required_columns:
        if col not in df.columns:
            return jsonify(success=False, message=f"Missing column: {col}"), 400

    customers = get_customers_from_db()
    existing_ids = {str(c.get("customer_id", "")).strip().lower() for c in customers if c.get("customer_id")}
    existing_emails = {
        str(c.get("email", "")).strip().lower()
        for c in customers
        if str(c.get("email", "")).strip()
    }
    existing_sigs = {_customer_signature_from_stored(c) for c in customers}
    batch_sigs = set()
    seen_emails = {}
    seen_customer_ids_in_batch = {}

    added = 0
    updated = 0
    skipped = 0
    skipped_duplicates = 0
    errors = []

    def is_blank(v):
        return pd.isna(v) or str(v).strip() == ""

    def is_valid_email(v):
        s = str(v or "").strip().lower()
        if not re.fullmatch(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", s):
            return False
        return s.endswith((".com", ".in", ".net", ".org"))

    max_num = 0
    for x in existing_ids:
        m = re.search(r"(\d+)$", str(x))
        if m:
            max_num = max(max_num, int(m.group(1)))
    current_sequence = 0

    def get_next_customer_id():
        nonlocal current_sequence
        current_sequence += 1
        return f"C{str(max_num + current_sequence).zfill(3)}"

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # _db_sync_customers_id_sequence(cur)

        for idx, row in df.iterrows():
            row_no = idx + 2
            try:
                if all(is_blank(row[col]) for col in required_columns):
                    skipped += 1
                    continue

                name = str(row.get("Name", "")).strip()
                company = str(row.get("Company", "")).strip()
                ctype_raw = row.get("Customer Type")
                email = str(row.get("Email", "")).strip().lower()
                status = str(row.get("Status", "")).strip()
                city = str(row.get("City", "")).strip()

                validation_errors = []

                if is_blank(name):
                    validation_errors.append("Name is required")
                if is_blank(company):
                    validation_errors.append("Company is required")

                ctype = normalize_customer_type(ctype_raw)

                if is_blank(email):
                    validation_errors.append("Email is required")
                elif not is_valid_email(email):
                    validation_errors.append("Invalid email format")
                elif email in seen_emails:
                    validation_errors.append(f"Duplicate Email in file (row {seen_emails[email]})")
                elif email in existing_emails:
                    validation_errors.append("Duplicate Email: already exists in system")
                else:
                    seen_emails[email] = row_no

                if status not in ["Active", "Inactive"]:
                    validation_errors.append("Status must be 'Active' or 'Inactive'")

                credit_val_raw = row.get("Credit Limit")
                if is_blank(credit_val_raw):
                    validation_errors.append("Credit Limit is required")
                    credit_limit_num = 0.0
                else:
                    try:
                        credit_limit_num = float(credit_val_raw)
                        if credit_limit_num < 0:
                            validation_errors.append("Credit Limit must be >= 0")
                    except (ValueError, TypeError):
                        validation_errors.append("Credit Limit must be a valid number")
                        credit_limit_num = 0.0

                if is_blank(city):
                    validation_errors.append("City is required")

                validation_errors.extend(_validate_customer_import_extra_fields(row))

                cid_raw = row.get("Customer ID")
                is_existing_id = False
                if is_blank(cid_raw):
                    customer_id = get_next_customer_id()
                else:
                    cid_str = str(cid_raw).strip()
                    if cid_str.upper().startswith("C"):
                        numeric = cid_str[1:].strip()
                    else:
                        numeric = cid_str
                    try:
                        cid_num = int(float(numeric))
                        if cid_num <= 0 or float(numeric) != float(cid_num):
                            raise ValueError("invalid")
                        customer_id = f"C{str(cid_num).zfill(3)}"
                    except Exception:
                        validation_errors.append("Customer ID must be a positive whole number")
                        customer_id = None

                if not customer_id:
                    validation_errors.append("Customer ID validation failed")
                    customer_id_lower = ""
                else:
                    customer_id_lower = customer_id.lower()
                    if customer_id_lower in seen_customer_ids_in_batch:
                        validation_errors.append(
                            f"Duplicate Customer ID in file (row {seen_customer_ids_in_batch[customer_id_lower]})"
                        )
                    else:
                        seen_customer_ids_in_batch[customer_id_lower] = row_no
                    is_existing_id = customer_id_lower in existing_ids

                if validation_errors:
                    skipped += 1
                    errors.append(f"Row {row_no}: " + ", ".join(validation_errors))
                    continue

                phone_val = re.sub(r"\D", "", str(row.get("phone") or ""))
                fn = str(row.get("first_name") or "").strip()
                ln = str(row.get("last_name") or "").strip()
                # available_limit is derived from Credit Limit for imported files.
                avail_lim = float(credit_limit_num or 0)
                sales_rep_v = str(row.get("sales_rep") or "").strip()
                bill_v = str(row.get("billing_address") or "").strip()
                ship_v = str(row.get("shipping_address") or "").strip()
                street_v = str(row.get("street") or "").strip()
                state_v = str(row.get("state") or "").strip()
                zip_v = re.sub(r"\D", "", str(row.get("zip_code") or ""))
                country_v = str(row.get("country") or "").strip()
                pay_v = str(row.get("payment_terms") or "").strip()
                cr_term_v = str(row.get("credit_term") or "").strip()
                gstin_v = re.sub(r"\s", "", str(row.get("gstin") or "").strip().upper())

                sig = _customer_row_signature_from_excel(row)
                if sig in batch_sigs or (not is_existing_id and sig in existing_sigs):
                    skipped += 1
                    skipped_duplicates += 1
                    continue

                if is_existing_id:
                    cur.execute("""
                        UPDATE customers
                        SET name=%s,
                            company=%s,
                            customer_type=%s,
                            customer_status=%s,
                            email=%s,
                            phone=%s,
                            credit_limit=%s,
                            city=%s,
                            first_name=%s,
                            last_name=%s,
                            available_limit=%s,
                            sales_rep=%s,
                            billing_address=%s,
                            shipping_address=%s,
                            street=%s,
                            state=%s,
                            zip_code=%s,
                            country=%s,
                            payment_terms=%s,
                            credit_term=%s,
                            gstin=%s
                        WHERE customer_id=%s
                    """, (
                        name,
                        company,
                        ctype,
                        status,
                        email,
                        phone_val,
                        credit_limit_num,
                        city,
                        fn,
                        ln,
                        avail_lim,
                        sales_rep_v,
                        bill_v,
                        ship_v,
                        street_v,
                        state_v,
                        zip_v,
                        country_v,
                        pay_v,
                        cr_term_v,
                        gstin_v,
                        customer_id,
                    ))
                    updated += 1
                else:
                    cur.execute("""
                        INSERT INTO customers (
                            customer_id,
                            name,
                            first_name,
                            last_name,
                            company,
                            customer_type,
                            customer_status,
                            email,
                            phone,
                            credit_limit,
                            city,
                            sales_rep,
                            street,
                            state,
                            zip_code,
                            country,
                            payment_terms,
                            credit_term,
                            gstin,
                            available_limit,
                            billing_address,
                            shipping_address
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                        )
                    """, (
                        customer_id,
                        name,
                        fn,
                        ln,
                        company,
                        ctype,
                        status,
                        email,
                        phone_val,
                        credit_limit_num,
                        city,
                        sales_rep_v,
                        street_v,
                        state_v,
                        zip_v,
                        country_v,
                        pay_v,
                        cr_term_v,
                        gstin_v,
                        avail_lim,
                        bill_v,
                        ship_v,
                    ))
                    added += 1

                existing_ids.add(customer_id_lower)
                existing_emails.add(email)
                existing_sigs.add(sig)
                batch_sigs.add(sig)

            except Exception as row_err:
                skipped += 1
                errors.append(f"Row {row_no}: {row_err}")

        conn.commit()
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify(success=False, message=f"Failed to import customers: {e}"), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    if added == 0 and updated == 0 and errors:
        return jsonify(
            success=False,
            message=f"Import failed: all rows were rejected. {errors[0]}",
            added=added,
            updated=updated,
            skipped=skipped,
            skipped_duplicates=skipped_duplicates,
            error_details=errors,
        ), 400

    return jsonify(
        success=True,
        added=added,
        updated=updated,
        skipped=skipped,
        skipped_duplicates=skipped_duplicates,
        error_details=errors,
    )

# =========================================
# ✅ API — Get All Customer(JSON)
# =========================================
@app.route("/api/customer", methods=["GET"])
def api_customer():
    """GET /api/customer — paginated + filtered at DB level."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401
    try:
        q = (request.args.get("q") or "").strip().lower()
        status = (request.args.get("status") or "").strip()
        ctype = (request.args.get("type") or "").strip()
        sales_rep = (request.args.get("sales_rep") or "").strip()

        try:
            page = max(1, int(request.args.get("page") or 1))
        except (TypeError, ValueError):
            page = 1
        try:
            page_size = int(request.args.get("page_size") or 10)
        except (TypeError, ValueError):
            page_size = 10
        page_size = min(max(page_size, 1), 100)

        where_parts = []
        where_params = []

        if q:
            like = f"%{q}%"
            where_parts.append("(LOWER(customer_id) LIKE %s OR LOWER(name) LIKE %s OR LOWER(company) LIKE %s)")
            where_params.extend([like, like, like])
        if status:
            where_parts.append("LOWER(customer_status) = %s")
            where_params.append(status.lower())
        if ctype:
            where_parts.append("LOWER(COALESCE(customer_type, '')) = %s")
            where_params.append(ctype.lower())
        if sales_rep:
            where_parts.append("LOWER(COALESCE(sales_rep, '')) = %s")
            where_params.append(sales_rep.lower())

        where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(f"SELECT COUNT(*) FROM customers {where_sql}", tuple(where_params))
        total_items = int((cur.fetchone() or [0])[0] or 0)
        total_pages = max(1, (total_items + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        offset = (page - 1) * page_size

        cur.execute(
            f"""
            SELECT customer_id, name, company, customer_type, customer_status, email, credit_limit, city, sales_rep
            FROM customers
            {where_sql}
            ORDER BY customer_id DESC
            LIMIT %s OFFSET %s
            """,
            tuple(where_params + [page_size, offset]),
        )
        rows = cur.fetchall()

        items = []
        for r in rows:
            ct = r[3]
            items.append({
                "customer_id": r[0],
                "name": r[1],
                "company": r[2],
                "customer_type": ct,
                "company_type": ct,
                "status": r[4],
                "email": r[5],
                "credit_limit": float(r[6] or 0),
                "city": r[7],
                "sales_rep": r[8] or "",
            })

        cur.execute("SELECT DISTINCT customer_status FROM customers WHERE customer_status IS NOT NULL AND customer_status <> '' ORDER BY customer_status")
        statuses = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT COALESCE(customer_type, '') AS ct FROM customers WHERE COALESCE(customer_type, '') <> '' ORDER BY ct")
        types = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT sales_rep FROM customers WHERE sales_rep IS NOT NULL AND sales_rep <> '' ORDER BY sales_rep")
        sales_reps = [r[0] for r in cur.fetchall()]

        cur.close()
        conn.close()

        response_data = {
            "success": True,
            "data": {
                "items": items,
                "page": page,
                "total_pages": total_pages,
                "total_items": total_items,
                "meta": {
                    "statuses": statuses,
                    "types": types,
                    "sales_reps": sales_reps
                }
            }
        }

        response = jsonify(response_data)
        response.headers['Content-Type'] = 'application/json; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        return response
        
    except Exception as e:
        print(f"Error in /api/customer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": str(e),
            "data": {
                "items": [],
                "page": 1,
                "total_pages": 1,
                "total_items": 0,
                "meta": {
                    "statuses": [],
                    "types": [],
                    "sales_reps": []
                }
            }
        }), 500


@app.route("/api/customer/<customer_id>", methods=["GET"])
def api_get_customer(customer_id):
    """
    GET /api/customer/<customer_id>
    
    Returns a single customer by ID )
    """
    try:
       
        customers = get_customers_from_db()
       
        customer = next((c for c in customers if str(c.get("customer_id")) == str(customer_id)), None)
        
        if not customer:
            return jsonify({
                "success": False,
                "message": "Customer not found"
            }), 404
        
        return jsonify({
            "success": True,
            "data": customer,
            "message": "Customer retrieved successfully"
        }), 200
        
    except Exception as e:
        print(f"Error in /api/customer/<customer_id>: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# =========================================
# ✅ API: UPDATE CUSTOMER (PUT /api/customer/<id>) — JSON for Postman
# =========================================
@app.route("/api/customer/<customer_id>", methods=["PUT"])
def api_update_customer(customer_id):
    """Update customer by ID. Requires JSON body. Same pattern as PUT /api/products/<id>."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    if not data:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    name = (data.get("name") or data.get("Name") or "").strip()
    if not name:
        return jsonify({"success": False, "message": "Customer name is required"}), 400

    try:
       
        conn = get_db_connection()
        cur = conn.cursor()

        # Check customer exists
        cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s", (customer_id,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "Customer not found"}), 404

        # Update customer
        cur.execute("""
            UPDATE customers
            SET name=%s,
                company=%s,
                customer_type=%s,
                email=%s,
                credit_limit=%s,
                customer_status=%s,
                city=%s
            WHERE customer_id=%s
        """, (
            (data.get("name") or "").strip(),
            (data.get("company") or "").strip(),
            (data.get("customer_type") or "").strip(),
            (data.get("email") or "").strip().lower(),
            float(data.get("credit_limit") or 0),
            (data.get("status") or "").strip(),
            (data.get("city") or "").strip(),
            customer_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Customer updated successfully"
        }), 200
   
    except Exception as e:
        print(f"Error in api_update_customer: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500


# =========================================
# ✅ API: DELETE CUSTOMER (DELETE /api/customer/<id>) — JSON for Postman
# =========================================
@app.route("/api/customer/<customer_id>", methods=["DELETE"])
def api_delete_customer(customer_id):
    """Delete customer by ID."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    
    conn = get_db_connection()
    cur = conn.cursor()

    # Check exists
    cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s", (customer_id,))
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Customer not found"}), 404

    # Delete
    cur.execute("DELETE FROM customers WHERE customer_id = %s", (customer_id,))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "success": True,
        "message": "Customer deleted successfully"
    }), 200


@app.route("/update-customer/<customer_id>", methods=["POST"])
def update_customer(customer_id):
    # Require login (same as Edit Product / api_update_customer)
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

   
    
    customers = get_customers_from_db()
  
    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    company = (data.get("company") or "").strip()
    # accept both keys, but prefer "customer_type"
    customer_type = (
        (data.get("customer_type") or "").strip()
        or (data.get("company_type") or "").strip()
    )
    email = (data.get("email") or "").strip().lower()
    credit_limit = (data.get("credit_limit") or "").strip()
    status = (data.get("status") or "").strip()
    city = (data.get("city") or "").strip()

 # ✅ DUPLICATE EMAIL CHECK (exclude current customer)
    for cust in customers:
        if str(cust.get("customer_id")) != str(customer_id):
            existing_email = (cust.get("email") or "").strip().lower()
            if email and existing_email == email:
                return jsonify({
                    "success": False,
                    "message": "Duplicate email! This email already exists."
                }), 409

    # find the matching customer
    found = False
    for cust in customers:
        if str(cust.get("customer_id")) == str(customer_id):
            cust["name"] = name
            cust["company"] = company
            # update both keys so data is consistent
            cust["customer_type"] = customer_type
            cust["company_type"] = customer_type
            cust["email"] = email
            cust["credit_limit"] = credit_limit
            cust["status"] = status
            cust["city"] = city
            found = True
            break

    if not found:
        return jsonify({"success": False, "message": "Customer not found"}), 404

    # save_customer(customer)
    return jsonify({"success": True, "message": "Customer updated"}), 200


# =========================================
# ✅ DELETE CUSTOMER (POST)
# =========================================
@app.route("/delete-customer/<cust_id>", methods=["POST"])
def delete_customer(cust_id):
   conn = get_db_connection()
   cur = conn.cursor()

   cur.execute("SELECT customer_id FROM customers WHERE customer_id=%s", (cust_id,))
   if not cur.fetchone():
       cur.close()
       conn.close()
       return jsonify({"ok": False, "message": "Customer not found"}), 404

   cur.execute("DELETE FROM customers WHERE customer_id=%s", (cust_id,))

   conn.commit()
   cur.close()
   conn.close()

   return jsonify({"ok": True, "message": "Customer deleted"})


# =========================================
# 6. MASTERS — Customer — Add New Customer
# =========================================
@app.route("/addnew-customer")
def addnew_customer():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "customer-addnew-customer.html",
        title="Add New Customer - Stackly",
        page="customer",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )


# =========================================
# ✅ API — Get All Customers (Add New Customer)
# =========================================
@app.route("/api/customers", methods=["GET"])
def get_customers():
   
    customers = get_customers_from_db()
    return jsonify(customers)


# =========================================
# ✅ API — Save Customer (Add New Customer)
# =========================================
@app.route("/api/customers", methods=["POST"])
def create_customer():
    data = request.get_json()

    if not data:
        return jsonify({"error": "No data received"}), 400

    gst_id = (data.get("gstNumber") or "").strip().upper()

    zip_digits = re.sub(r"\D", "", str(data.get("zipCode") or ""))
    if len(zip_digits) != 6:
        return jsonify({"error": "Zip code must be exactly 6 digits"}), 400

    try:
       
        customers = get_customers_from_db()

        if gst_id:
            for c in customers:
                if (c.get("gstNumber") or "").strip().upper() == gst_id:
                    return jsonify({
                        "error": "GST/Tax ID already exists"
                    }), 409  # Conflict

        # Generate customer ID in format C101, C102, etc.
        customer_id = generate_customer_id_for_master()
        
        # Transform form data to match customer.json format
        first_name = (data.get("firstName") or "").strip()
        last_name = (data.get("lastName") or "").strip()
        full_name = (first_name + " " + last_name).strip()
        
        # For company field: use name if customer type is Individual, otherwise leave empty
        # (The form doesn't have a company field, so we'll use name as company for consistency)
        company_name = full_name if data.get("customerType") == "Individual" else (data.get("company") or "")
        
        customer_data = {
            "customer_id": customer_id,
            "name": full_name or "Unknown",
            "company": company_name or full_name,
            "customer_type": data.get("customerType") or "",
            "status": data.get("customerStatus") or "Active",
            "email": (data.get("email") or "").strip().lower(),
            "credit_limit": str(data.get("creditLimit") or "0"),
            "city": data.get("city") or "",
            "sales_rep": (data.get("salesRep") if data.get("salesRep") != "custom" else data.get("salesRepCustom")) or "",
            "company_type": data.get("customerType") or "",
            "phone": data.get("phoneNumber") or "",
            "gstNumber": gst_id,
            "address": data.get("address") or "",
            "street": data.get("street") or "",
            "state": data.get("state") or "",
            "zipCode": zip_digits,
            "country": data.get("country") or "",
            "billingAddress": data.get("billingAddress") or "",
            "shippingAddress": data.get("shippingAddress") or "",
            "paymentTerms": (data.get("paymentTerms") if data.get("paymentTerms") != "custom" else data.get("paymentTermsCustom")) or "",
            "creditTerm": (data.get("creditTerm") if data.get("creditTerm") != "custom" else data.get("creditTermCustom")) or "",
            "availableLimit": str(data.get("availableLimit") or "0")
        }

       
        conn = get_db_connection()
        cur = conn.cursor()
        # _db_sync_customers_id_sequence(cur)

        cur.execute("""
        INSERT INTO customers (
            customer_id, name, first_name, last_name,
            company, customer_type, customer_status, email, phone,
            credit_limit, city, sales_rep,
            street, state, zip_code, country,
            payment_terms, credit_term, gstin,
            available_limit, billing_address, shipping_address
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        customer_id,
        customer_data["name"],
        data.get("firstName"),
        data.get("lastName"),
        customer_data["company"],
        customer_data["customer_type"],
        data.get("customerStatus"),
        customer_data["email"],
        data.get("phoneNumber"),
        float(customer_data["credit_limit"] or 0),
        customer_data["city"],
        customer_data["sales_rep"],
        data.get("street"),
        data.get("state"),
        zip_digits,
        data.get("country"),
        data.get("paymentTerms"),
        data.get("creditTerm"),
        data.get("gstNumber"),
        float(data.get("availableLimit") or 0),
        data.get("billingAddress"),
        data.get("shippingAddress")
    ))
        conn.commit()
        cur.close()
        conn.close()
        

        return jsonify({
            "message": "Customer saved successfully",
            "customerId": customer_id
        }), 201
    


    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error saving customer: {e}")
        return jsonify({"error": str(e)}), 500


# =========================================
# ✅ ID GENERATION
# =========================================
def generate_customer_id():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT customer_id FROM customers ORDER BY customer_id DESC LIMIT 1")
    row = cur.fetchone()

    cur.close()
    conn.close()

    if row and row[0]:
        try:
            num = int(row[0].replace("C", ""))
            return f"C{str(num + 1).zfill(3)}"   # ✅ zero padding
        except:
            return "C001"
    else:
        return "C001"


# =========================================
# ✅ ID GENERATION FOR CUSTOMER MASTER (C101, C102, etc.)
# =========================================
def generate_customer_id_for_master():
    customers = get_customers_from_db()

    ids = []
    for c in customers:
        cust_id = c.get("customer_id")   

        if cust_id and str(cust_id).startswith('C'):
            try:
                num = int(str(cust_id).replace("C", ""))
                ids.append(num)
            except:
                pass

    if ids:
        new_id = max(ids) + 1
    else:
        new_id = 1

    return f"C{str(new_id).zfill(3)}"


@app.route('/api/customers/new-id', methods=['GET'])
def get_new_customer_id():
    return jsonify({"customerId": generate_customer_id()})


@app.route('/api/customers/master-id', methods=['GET'])
def get_master_customer_id():
    return jsonify({"customerId": generate_customer_id_for_master()})


def _customer_import_filename_ok(filename):
    if not filename or not str(filename).strip():
        return False
    fn = str(filename).strip().lower()
    return fn.endswith(".csv") or fn.endswith(".xlsx")


def _read_customer_import_dataframe(file):
    if not _customer_import_filename_ok(file.filename):
        return None, "Invalid file format. Only .csv and .xlsx files are allowed."
    try:
        file.seek(0)
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        return df, None
    except Exception:
        return None, "Invalid or corrupt file. Please upload a valid CSV or XLSX file."


@app.route("/upload-customer", methods=["POST"])
def upload_customer_file():
    file = request.files.get("file")

    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    df, read_err = _read_customer_import_dataframe(file)
    if read_err:
        return jsonify({"error": read_err}), 400

    # Check if file is empty (no data rows, only headers)
    if df.empty or len(df) == 0:
        return jsonify({
            "error": "No data found",
            "message": "The uploaded file contains no data. Please ensure the file has at least one row of customer data."
        }), 400

    valid_rows = 0
    invalid_rows = 0
    skipped_rows = 0
    error_details = []
    skipped_row_numbers = []  
    
 
   
  
    existing_customers = get_customers_from_db()



    # Normalize Customer IDs for comparison 
    existing_customer_ids = {str(c.get("customer_id", "")).strip().lower() for c in existing_customers if c.get("customer_id")}
    
    # Track Customer IDs for uniqueness validation within uploaded file
    seen_customer_ids = {}  # key: Customer ID (as string, lowercase), value: first row number
    
    # Track Emails for uniqueness validation within uploaded file
    seen_emails = {}  # key: Email (as string, lowercase), value: first row number
    
    # Load existing emails from database for uniqueness check
    existing_emails = {
        str(c.get("email", "")).strip().lower()
        for c in existing_customers
        if str(c.get("email", "")).strip() != ""
    }
    existing_customer_signatures = {_customer_signature_from_stored(c) for c in existing_customers}
    
    # Track seen row combinations for duplicate detection (excluding Customer ID)
    seen_rows = {}  # signature -> first Excel row number

    # Helper to check blank (NaN or empty/whitespace)
    def is_blank(val):
        if pd.isna(val):
            return True
        return str(val).strip() == ""
    
    # Helper to normalize value for comparison
    def normalize_value(val):
        if pd.isna(val):
            return ""
        return str(val).strip()
    
    

    # Security validation helpers
    def has_script(v):
        if pd.isna(v):
            return False
        s = str(v).lower()
        patterns = [
            "<script", "</script", "javascript:", "onerror=", "onload=",
            "<img", "<svg", "<iframe"
        ]
        return any(p in s for p in patterns)
    
    def has_sql_injection(v):
        if pd.isna(v):
            return False
        s = str(v).lower()
        patterns = [
            r"\bor\s+1\s*=\s*1\b",      # or 1=1
            r"\bunion\s+select\b",      # union select
            r"\bdrop\s+table\b",        # drop table
            r"\bdelete\s+from\b",       # delete from
            r"\binsert\s+into\b",       # insert into
            r"\bupdate\s+\w+\s+set\b",  # update x set
            r"--",                      # SQL comment
            r";"                        # multiple statements
        ]
        return any(re.search(p, s) for p in patterns)

    # Required columns for validation (base + extra template columns)
    required_columns = _customer_import_all_required_columns()

    # Check if all required columns exist
    for col in required_columns:
        if col not in df.columns:
            return jsonify({
                "total_rows": len(df),
                "valid_rows": 0,
                "invalid_rows": len(df),
                "skipped_rows": 0,
                "skipped_row_numbers": [],
                "error_details": [{"row": 0, "errors": [f"Missing column: {col}"]}]
            }), 400

    # ---------------- ROW VALIDATION ----------------
    for index, row in df.iterrows():
        errors = []
        row_no = index + 2  # Excel row number (header is row 1)

        # Treat rows where ALL required columns are blank as skipped
        if all(
            is_blank(row.get(col_name))
            for col_name in required_columns
        ):
            skipped_rows += 1
            skipped_row_numbers.append(row_no)  # Store row number (Excel row, +2 for header)
            # Don't validate or count as valid/invalid; just skip
            continue

        # Block Script/HTML injection in ANY column
        if any(has_script(row.get(col)) for col in required_columns):
            invalid_rows += 1
            error_details.append({
                "row": row_no,
                "errors": ["Script/HTML content detected"]
            })
            continue
        
        # Block SQL injection patterns in ANY column
        if any(has_sql_injection(row.get(col)) for col in required_columns):
            invalid_rows += 1
            error_details.append({
                "row": row_no,
                "errors": ["SQL injection pattern detected"]
            })
            continue

                
        # ---------------- Customer ID (optional - will be auto-generated if blank, must be unique if provided) ----------------
        # Same validation pattern as Product ID in product import
        cid_raw = row.get("Customer ID")
        customer_id = None
        
        if not is_blank(cid_raw):
            # Only validate if Customer ID is provided
            cid_str = str(cid_raw).strip()
            
            # Check if Customer ID already has "C" prefix (e.g., "C124")
            if cid_str.upper().startswith("C") and len(cid_str) > 1:
                # Extract numeric part after "C"
                numeric_part = cid_str[1:].strip()
                try:
                    cid_num = float(numeric_part)
                    if not cid_num.is_integer() or int(cid_num) <= 0:
                        errors.append("Customer ID must be a valid number after 'C' prefix")
                    else:
                        # Normalize to uppercase with 3-digit sequence (e.g., C012)
                        cid_str = f"C{str(int(cid_num)).zfill(3)}"
                except (ValueError, TypeError):
                    errors.append("Customer ID must be a valid number after 'C' prefix")
            else:
                # Try to convert to number (handles both "12" and "12.0" from Excel)
                try:
                    cid_num = float(cid_str)
                    # Check if it's a whole number (no decimal part)
                    if not cid_num.is_integer():
                        errors.append("Customer ID must be a whole number")
                    elif int(cid_num) <= 0:
                        errors.append("Customer ID must be greater than 0")
                    else:
                        # Prepend "C" to numeric Customer ID with 3-digit sequence (e.g., 12 becomes C012)
                        cid_str = f"C{str(int(cid_num)).zfill(3)}"
                except (ValueError, TypeError):
                    # If conversion fails, it's not a valid number
                    errors.append("Customer ID must be a whole number")
                    cid_str = None
            
            # Validate uniqueness if Customer ID is valid
            if cid_str and not any("Customer ID" in err for err in errors):
                # Normalize to lowercase for comparison
                cid_normalized = cid_str.lower()
                
                # Check if Customer ID already exists in the uploaded file
                if cid_normalized in seen_customer_ids:
                    first_row = seen_customer_ids[cid_normalized]
                    errors.append(f"Duplicate Customer ID: Customer ID {cid_str} already exists in row {first_row}")
                # Check if Customer ID already exists in the database
                elif cid_normalized in existing_customer_ids:
                    errors.append(f"Duplicate Customer ID: Customer ID {cid_str} already exists in the system")
                else:
                    seen_customer_ids[cid_normalized] = row_no  # Store the row number (Excel row, +2 for header)
                    customer_id = cid_str
            # Note: If Customer ID is blank, it will be auto-generated during import (C101, C102, etc.)





        # ---------------- Name (mandatory, letters + spaces, min length 3) ----------------
        name_raw = row.get("Name")
        if is_blank(name_raw):
            errors.append("Name is required")
        else:
            name = str(name_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", name):
                errors.append("Name must contain ONLY letters and spaces")
            elif len(name) < 3:
                errors.append("Name must be at least 3 characters")
            elif len(name) > 40:
                errors.append("Name must not exceed 40 characters")

        # ---------------- Company (mandatory) ----------------
        company_raw = row.get("Company")
        if is_blank(company_raw):
            errors.append("Company is required")
        else:
            company = str(company_raw).strip()
            if len(company) < 3:
                errors.append("Company must be at least 3 characters")
            elif len(company) > 50:
                errors.append("Company must not exceed 50 characters")
            # Allow letters, numbers, spaces, and common business symbols
            if not re.fullmatch(r"^[A-Za-z0-9 &.,'()\/-]+$", company):
                errors.append("Company contains invalid characters")

        # ---------------- Email (mandatory, valid format, must be unique) ----------------
        email_raw = row.get("Email")
        if is_blank(email_raw):
            errors.append("Email is required")
        else:
            email_str = str(email_raw).strip().lower()
            
            # Basic email format check
            basic_ok = re.fullmatch(
                r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$",
                email_str
            )
            
            # Allowed TLDs (matching product import pattern)
            allowed_tlds = {"com", "in", "net", "org", "co.in"}
            tld = email_str.rsplit(".", 1)[-1] if "." in email_str else ""
            
            if not basic_ok:
                errors.append("Invalid email format")
            elif tld not in allowed_tlds:
                errors.append(f"Email domain must end with one of: {', '.join(allowed_tlds)}")
            elif len(email_str) > 50:
                errors.append("Email must not exceed 50 characters")
            else:
                # Check for duplicate Email in uploaded file 
                if email_str in seen_emails:
                    first_row = seen_emails[email_str]
                    errors.append(f"Duplicate Email: Email already exists in row {first_row}")
                # Check if Email already exists in the database
                elif email_str in existing_emails:
                    errors.append(f"Duplicate Email: Email already exists in the system")
                else:
                    seen_emails[email_str] = row_no  # Store the row number (Excel row, +2 for header)

        # ---------------- Status (mandatory) ----------------
        status_raw = row.get("Status")
        if is_blank(status_raw):
            errors.append("Status is required")
        else:
            status = str(status_raw).strip()
            if status not in ["Active", "Inactive"]:
                errors.append("Status must be 'Active' or 'Inactive'")

        # ---------------- Credit Limit (mandatory, number >= 0) ----------------
        credit_limit_raw = row.get("Credit Limit")
        if is_blank(credit_limit_raw):
            errors.append("Credit Limit is required")
        else:
            try:
                credit_limit_num = float(credit_limit_raw)
                if credit_limit_num < 0:
                    errors.append("Credit Limit must be 0 or greater")
                elif credit_limit_num > 10000000:
                    errors.append("Credit Limit must not exceed 10,000,000")
            except (ValueError, TypeError):
                errors.append("Credit Limit must be a valid number")

        # ---------------- City (mandatory, letters + spaces, min length 3) ----------------
        city_raw = row.get("City")
        if is_blank(city_raw):
            errors.append("City is required")
        else:
            city = str(city_raw).strip()
            if not re.fullmatch(r"^[A-Za-z ]+$", city):
                errors.append("City must contain ONLY letters and spaces")
            elif len(city) < 3:
                errors.append("City must be at least 3 characters")
            elif len(city) > 40:
                errors.append("City must not exceed 40 characters")

        # ---------------- Extra template columns (mandatory) ----------------
        errors.extend(_validate_customer_import_extra_fields(row))

        # --- Duplicate row / duplicate-in-system (normalized signature; credit limit like product price) ---
        sig = _customer_row_signature_from_excel(row)
        if sig in seen_rows:
            first_row = seen_rows[sig]
            errors.append(
                f"Duplicate row: This combination of Name, Company, Customer Type, Email, Status, Credit Limit, and City is identical to row {first_row}"
            )
        elif sig in existing_customer_signatures:
            errors.append(
                "Duplicate customer: this combination already exists in the system (same Name, Company, Type, Email, Status, Credit Limit, and City)."
            )
        else:
            if sig != ("", "", "", "", "", 0.0, ""):
                seen_rows[sig] = row_no

        if errors:
            invalid_rows += 1
            error_details.append({
                "row": row_no,  # +2 because header is row 1
                "errors": errors
            })
        else:
            valid_rows += 1

    return jsonify({
        "total_rows": len(df),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "skipped_rows": skipped_rows,
        "skipped_row_numbers": skipped_row_numbers,
        "error_details": error_details
    })

@app.route("/api/customers", methods=["GET"])
def api_get_customers():
    try:
        customers = get_customers_from_db()

        return jsonify({
            "success": True,
            "customers": customers
        })

    except Exception as e:
        print("API ERROR:", e)
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


   
# =========================================
# 6.Masters— Suppliers
# =========================================
@app.route("/suppliers")
def suppliers():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))
 
    # ✅ Get user name
    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break
 
    # ✅ FETCH FROM DATABASE (NOT JSON)
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
 
    cur.execute("""
        SELECT supplier_id, supplier_name, status, supplier_type, supplier_tier, created_at
        FROM suppliers
        ORDER BY created_at DESC
    """)
 
    rows = cur.fetchall()
 
    cur.close()
    conn.close()
 
    # ✅ MAP DATA FOR UI
    suppliers_rows = []
 
    for s in rows:
        suppliers_rows.append({
            "id": (s.get("supplier_id") or "").strip(),
            "name": (s.get("supplier_name") or "").strip(),
            "created_date": str(s.get("created_at", ""))[:10],
            "status": (s.get("status") or "").strip(),
            "type": (s.get("supplier_type") or "").strip(),
            "tier": (s.get("supplier_tier") or "").strip(),
        })
 
    return render_template(
        "suppliers.html",
        title="Supplier Master - Stackly",
        page="suppliers",
        section="inventory",
        user_email=user_email,
        user_name=user_name,
        suppliers=suppliers_rows,
    )
 
 
@app.route("/supplier-new")
def supplier_new():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))
 
    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break
 
    return render_template(
        "supplier-new.html",
        title="Create Supplier - Stackly",
        page="supplier_new",
        section="inventory",
        user_email=user_email,
        user_name=user_name,
    )
 
 
# =========================================
# Supplier Backend (Clean REST + PostgreSQL)
# =========================================
#---------intialize regex patterns for supplier validation
SUPPLIER_NAME_REGEX = re.compile(r"^[A-Za-z0-9 .,&()'/-]{3,100}$")
SUPPLIER_CONTACT_REGEX = re.compile(r"^[A-Za-z .'-]{2,80}$")
SUPPLIER_PHONE_REGEX = re.compile(r"^\+?[0-9][0-9\s-]{6,19}$")
SUPPLIER_EMAIL_REGEX = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
SUPPLIER_FIELDS = (
    "supplier_id",
    "supplier_name",
    "gstin",
    "company_registration_number",
    "legal_entity_name",
    "country_of_registration",
    "supplier_type",
    "supplier_tier",
    "status",
    "product_detail",
 
    "contact_first_name",
    "contact_last_name",
    "designation_role",
    "alternate_contact_no",
    "email",
    "phone_number",
    "website",
    "relationship_manager",
 
    "registered_office_address",
    "mailing_address",
    "warehouse_address",
    "billing_address",
    "registered_billing_address",
 
    "bank_name",
    "payment_method",
    "bank_account_no",
    "payment_terms",
    "iban_swift_code",
    "tax_withholding_setup",
 
    "currency",
    "categories_served",
    "inco_terms",
    "product_service_catalog",
    "freight_terms",
 
    "minimum_order_quantity",
    "return_replacement_policy",
    "average_delivery_time_days",
 
    "contract_references",
    "compliance_certifications",
    "risk_notes_flags",
    "compliance_status",
    "last_risk_assessment_date",
    "risk_ratings",
 
    "on_time_delivery_rate",
    "quality_ratings",
    "defect_return_rate",
 
    "last_evaluation_date",
    "contract_breaches",
    "improvement_plans",
    "complaints_registered",
 
    "external_key_contact",
    "visit_history_meeting_notes",
    "comments",
    "created_by"
)
 
# Columns that must not receive '' from the client (PostgreSQL date types reject empty string)
SUPPLIER_DATE_FIELDS = frozenset({
    "last_risk_assessment_date",
    "last_evaluation_date",
})
SUPPLIER_DATE_DISPLAY_FMT = "%d-%m-%Y"
 
 
def _supplier_date_display(val):
    """DB / API value → dd-mm-yyyy for supplier edit UI."""
    if val is None or val == "":
        return None
    if hasattr(val, "strftime"):
        return val.strftime(SUPPLIER_DATE_DISPLAY_FMT)
    s = str(val).strip()
    if not s:
        return None
    if re.match(r"^\d{2}-\d{2}-\d{4}$", s):
        return s
    iso = s.split("T")[0]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
        try:
            return datetime.strptime(iso, "%Y-%m-%d").date().strftime(SUPPLIER_DATE_DISPLAY_FMT)
        except ValueError:
            return s
    return s
 
 
def _supplier_date_parse_for_db(val):
    """Client dd-mm-yyyy (or ISO) → datetime.date for PostgreSQL DATE columns."""
    if val is None:
        return None
    if isinstance(val, str) and not val.strip():
        return None
    if hasattr(val, "strftime") and not isinstance(val, str):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if not s:
        return None
    if re.match(r"^\d{2}-\d{2}-\d{4}$", s):
        try:
            return datetime.strptime(s, SUPPLIER_DATE_DISPLAY_FMT).date()
        except ValueError:
            return None
    iso = s.split("T")[0]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", iso):
        try:
            return datetime.strptime(iso, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None
 
SUPPLIER_REQUIRED_FIELD_LABELS = (
    ("supplier_id", "Supplier ID"),
    ("supplier_name", "Supplier name"),
    ("gstin", "GSTIN"),
    ("company_registration_number", "Company registration number"),
    ("legal_entity_name", "Legal entity name"),
    ("country_of_registration", "Country of registration"),
    ("supplier_type", "Supplier type"),
    ("supplier_tier", "Supplier tier"),
    ("status", "Status"),
    ("contact_first_name", "Primary contact first name"),
    ("contact_last_name", "Last name"),
    ("email", "Email"),
    ("phone_number", "Phone number"),
    ("registered_office_address", "Registered office address"),
)
 
 
def _supplier_value_blank(v):
    return v is None or (isinstance(v, str) and not str(v).strip())
 
 
def validate_supplier_required(row):
    """Return an error message if required fields (per supplier form) are missing or invalid, else None."""
    row = row or {}
    for key, label in SUPPLIER_REQUIRED_FIELD_LABELS:
        if _supplier_value_blank(row.get(key)):
            return f"{label} is required."
    sid = (row.get("supplier_id") or "").strip().upper()
    if not re.match(r"^SUP-\d{3,}$", sid):
        return "Supplier ID must be in SUP-001 format."
    email = (row.get("email") or "").strip()
    if not SUPPLIER_EMAIL_REGEX.match(email):
        return "Enter a valid email address."
    phone = (row.get("phone_number") or "").strip()
    if not SUPPLIER_PHONE_REGEX.match(phone):
        return "Enter a valid phone number."
    fn = (row.get("contact_first_name") or "").strip()
    ln = (row.get("contact_last_name") or "").strip()
    if not SUPPLIER_CONTACT_REGEX.match(fn):
        return "Primary contact first name must be 2–80 letters (spaces, dots, apostrophes allowed)."
    if not SUPPLIER_CONTACT_REGEX.match(ln):
        return "Last name must be 2–80 letters (spaces, dots, apostrophes allowed)."
    sname = (row.get("supplier_name") or "").strip()
    if not SUPPLIER_NAME_REGEX.match(sname):
        return "Supplier name must be 3–100 characters and use allowed characters only."
    return None
#---------function to generate next supplier ID in SUP-NNN format
 
def generate_supplier_id():
    """Next SUP-NNN id based on existing rows (matches supplier-new.js /^SUP-\\d{3,}$/)."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT supplier_id FROM suppliers")
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
 
    best = 0
    for row in rows:
        sid = (row[0] or "").strip()
        if not sid:
            continue
        up = sid.upper()
        if not up.startswith("SUP-"):
            continue
        suffix = up[4:]
        if not suffix.isdigit():
            continue
        try:
            n = int(suffix)
            if n > best:
                best = n
        except ValueError:
            continue
    return f"SUP-{str(best + 1).zfill(3)}"
 
#---------API endpoint to get a new supplier ID (for supplier creation form)
@app.route("/api/suppliers/new-id", methods=["GET"])
def get_new_supplier_id():
    return jsonify({"supplierId": generate_supplier_id()})
 
 
def _normalize_supplier_row_from_request(data):
    """Build a dict with every SUPPLIER_FIELDS key for INSERT/UPSERT (missing JSON keys → None)."""
    data = dict(data or {})
    row = {}
    for col in SUPPLIER_FIELDS:
        v = data.get(col)
        if col in SUPPLIER_DATE_FIELDS:
            if v is None or (isinstance(v, str) and v.strip() == ""):
                v = None
            else:
                v = _supplier_date_parse_for_db(v)
        elif isinstance(v, str) and v.strip() == "":
            v = None
        row[col] = v
    return row
 
#----API endpoint to create a new supplier
@app.route("/api/suppliers", methods=["POST"])
def create_supplier():
    raw = dict(request.json or {})
    if not raw.get("supplier_id"):
        raw["supplier_id"] = generate_supplier_id()
 
    data = _normalize_supplier_row_from_request(raw)
    req_err = validate_supplier_required(data)
    if req_err:
        return jsonify({"success": False, "error": req_err}), 400
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
        cur.execute("""
            INSERT INTO suppliers (
                supplier_id, supplier_name, gstin,
                company_registration_number, legal_entity_name,
                country_of_registration, supplier_type, supplier_tier,
                status, product_detail,
 
                contact_first_name, contact_last_name,
                designation_role, alternate_contact_no,
                email, phone_number, website, relationship_manager,
 
                registered_office_address, mailing_address,
                warehouse_address, billing_address,
                registered_billing_address,
 
                bank_name, payment_method, bank_account_no,
                payment_terms, iban_swift_code, tax_withholding_setup,
 
                currency, categories_served, inco_terms,
                product_service_catalog, freight_terms,
 
                minimum_order_quantity, return_replacement_policy,
                average_delivery_time_days,
 
                contract_references, compliance_certifications,
                risk_notes_flags, compliance_status,
                last_risk_assessment_date, risk_ratings,
 
                on_time_delivery_rate, quality_ratings, defect_return_rate,
 
                last_evaluation_date, contract_breaches,
                improvement_plans, complaints_registered,
 
                external_key_contact, visit_history_meeting_notes,
                comments, created_by
            )
            VALUES (
                %(supplier_id)s, %(supplier_name)s, %(gstin)s,
                %(company_registration_number)s, %(legal_entity_name)s,
                %(country_of_registration)s, %(supplier_type)s, %(supplier_tier)s,
                %(status)s, %(product_detail)s,
 
                %(contact_first_name)s, %(contact_last_name)s,
                %(designation_role)s, %(alternate_contact_no)s,
                %(email)s, %(phone_number)s, %(website)s, %(relationship_manager)s,
 
                %(registered_office_address)s, %(mailing_address)s,
                %(warehouse_address)s, %(billing_address)s,
                %(registered_billing_address)s,
 
                %(bank_name)s, %(payment_method)s, %(bank_account_no)s,
                %(payment_terms)s, %(iban_swift_code)s, %(tax_withholding_setup)s,
 
                %(currency)s, %(categories_served)s, %(inco_terms)s,
                %(product_service_catalog)s, %(freight_terms)s,
 
                %(minimum_order_quantity)s, %(return_replacement_policy)s,
                %(average_delivery_time_days)s,
 
                %(contract_references)s, %(compliance_certifications)s,
                %(risk_notes_flags)s, %(compliance_status)s,
                %(last_risk_assessment_date)s, %(risk_ratings)s,
 
                %(on_time_delivery_rate)s, %(quality_ratings)s, %(defect_return_rate)s,
 
                %(last_evaluation_date)s, %(contract_breaches)s,
                %(improvement_plans)s, %(complaints_registered)s,
 
                %(external_key_contact)s, %(visit_history_meeting_notes)s,
                %(comments)s, %(created_by)s
            )
        """, data)
 
        conn.commit()
        sid = (data.get("supplier_id") or "").strip()
        return jsonify(
            {"success": True, "message": "Supplier created successfully", "supplier_id": sid}
        ), 201
 
    except Exception as e:
        conn.rollback()
        print("❌ create_supplier:", repr(e))
        return jsonify({"success": False, "error": str(e)}), 500
 
    finally:
        cur.close()
        conn.close()
 
#----API endpoint to get all suppliers (for supplier list UI)
 
@app.route("/api/suppliers", methods=["GET"])
def get_suppliers():
    """List suppliers for dropdowns (purchase order, filters, etc.)."""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT supplier_id, supplier_name, email
            FROM suppliers
            ORDER BY supplier_id ASC
        """)
        rows = cur.fetchall()
        suppliers = [
            {
                "id": (row[0] or "").strip(),
                "name": (row[1] or "").strip(),
                "email": (row[2] or "").strip(),
            }
            for row in rows
        ]
        return jsonify(suppliers)
    finally:
        cur.close()
        conn.close()
#----API endpoint to get a single supplier by ID (for supplier detail/edit UI)
@app.route("/api/suppliers/<supplier_id>", methods=["GET"])
def get_supplier(supplier_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
 
    cur.execute("SELECT * FROM suppliers WHERE supplier_id = %s", (supplier_id,))
    supplier = cur.fetchone()
 
    cur.close()
    conn.close()
 
    if not supplier:
        return jsonify({"success": False, "message": "Supplier not found"}), 404
 
    row = dict(supplier)
    for key, val in list(row.items()):
        if key in SUPPLIER_DATE_FIELDS:
            row[key] = _supplier_date_display(val)
        elif hasattr(val, "isoformat"):
            row[key] = val.isoformat()
        elif type(val).__name__ == "Decimal":
            row[key] = float(val) if val is not None else None
 
    return jsonify({"success": True, "data": row})
 
#----API endpoint to update a supplier by ID (for supplier detail/edit UI)
@app.route("/api/suppliers/<supplier_id>", methods=["PUT"])
def update_supplier(supplier_id):
    data = dict(request.json or {})
 
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
 
    try:
        cur.execute("SELECT * FROM suppliers WHERE supplier_id = %s", (supplier_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Supplier not found"}), 404
 
        row_dict = dict(row)
        merged = dict(row_dict)
        for key, val in data.items():
            if key == "supplier_id":
                continue
            if key not in SUPPLIER_FIELDS:
                continue
            if key in SUPPLIER_DATE_FIELDS:
                if val == "" or (isinstance(val, str) and not str(val).strip()):
                    val = None
                else:
                    val = _supplier_date_parse_for_db(val)
            merged[key] = val
 
        for dcol in SUPPLIER_DATE_FIELDS:
            if merged.get(dcol) == "":
                merged[dcol] = None
 
        merged_check = dict(merged)
        merged_check["supplier_id"] = supplier_id
        for col in SUPPLIER_FIELDS:
            if col not in merged_check:
                continue
            v = merged_check.get(col)
            if isinstance(v, str) and v.strip() == "":
                merged_check[col] = None
        req_err = validate_supplier_required(merged_check)
        if req_err:
            return jsonify({"success": False, "error": req_err}), 400
 
        # Only SET columns that exist on this database table (avoids 500 if schema lags code)
        update_columns = [
            c for c in SUPPLIER_FIELDS
            if c != "supplier_id" and c in row_dict
        ]
        if not update_columns:
            return jsonify({"success": False, "message": "No updatable columns on suppliers table"}), 500
 
        params = {col: merged.get(col) for col in update_columns}
        params["supplier_id"] = supplier_id
 
        set_clause = ", ".join(f"{col} = %({col})s" for col in update_columns)
        has_updated_at = "updated_at" in row_dict
        tail = ", updated_at = CURRENT_TIMESTAMP" if has_updated_at else ""
 
        cur.execute(
            f"""
            UPDATE suppliers
            SET {set_clause}{tail}
            WHERE supplier_id = %(supplier_id)s
            """,
            params,
        )
 
        conn.commit()
        return jsonify({"success": True, "message": "Supplier updated successfully"})
 
    except Exception as e:
        conn.rollback()
        print("❌ update_supplier:", repr(e))
        return jsonify({"success": False, "error": str(e)}), 500
 
    finally:
        cur.close()
        conn.close()
 
 
def _ensure_supplier_attachment_table(cur=None):
    """supplier_attachments table + uploads/supplier_attachments/ folder."""
    try:
        os.makedirs(SUPPLIER_ATTACHMENTS_FOLDER, exist_ok=True)
    except OSError:
        pass
    own_conn = None
    own_cur = None
    if cur is None:
        own_conn = get_db_connection()
        own_cur = own_conn.cursor()
        cur = own_cur
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS supplier_attachments (
                id SERIAL PRIMARY KEY,
                supplier_id TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                uploaded_at TIMESTAMP DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_supplier_attachments_supplier_id
            ON supplier_attachments (supplier_id)
            """
        )
        cur.execute(
            """
            ALTER TABLE supplier_attachments
            ADD COLUMN IF NOT EXISTS category TEXT DEFAULT ''
            """
        )
        # Commit DDL on the active connection (required when cur is caller-supplied).
        cur.connection.commit()
    finally:
        if own_conn:
            own_cur.close()
            own_conn.close()
 
 
DOC_UPLOAD_ALLOWED_EXTENSIONS = frozenset({"pdf", "jpg", "jpeg", "png"})
DOC_UPLOAD_MAX_BYTES = 10 * 1024 * 1024

SUPPLIER_UPLOAD_ALLOWED_EXTENSIONS = DOC_UPLOAD_ALLOWED_EXTENSIONS
SUPPLIER_UPLOAD_MAX_BYTES = DOC_UPLOAD_MAX_BYTES


_DOC_UPLOAD_BLOCKED_MIMES = frozenset({
    "application/zip",
    "application/x-zip-compressed",
    "application/x-rar-compressed",
    "application/vnd.ms-excel",
    "application/msword",
    "application/vnd.openxmlformats-officedocument",
})


def _doc_upload_filename_allowed(filename, mimetype=None):
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext not in DOC_UPLOAD_ALLOWED_EXTENSIONS:
        return False
    mime = (mimetype or "").strip().lower().split(";", 1)[0]
    if mime and any(mime.startswith(blocked) for blocked in _DOC_UPLOAD_BLOCKED_MIMES):
        return False
    return True


def _supplier_upload_filename_allowed(filename):
    return _doc_upload_filename_allowed(filename)


@app.route("/api/supplier-attachments", methods=["POST"])
def upload_supplier_attachment():
    """Save supplier file to S3 supplier_attachments/ or local uploads/supplier_attachments/."""
    try:
        supplier_id = (request.form.get("supplier_id") or "").strip().upper()
        category = (request.form.get("category") or "").strip().lower()
        file = request.files.get("file")
        if not supplier_id:
            return jsonify({"success": False, "message": "supplier_id is required"}), 400
        if not file or not file.filename:
            return jsonify({"success": False, "message": "No file uploaded"}), 400
        if not _doc_upload_filename_allowed(
            file.filename, file.mimetype or "application/octet-stream"
        ):
            return jsonify({
                "success": False,
                "message": "Invalid file format. Only PDF, JPEG, and PNG files are allowed.",
            }), 400
        file_size = _upload_file_size_bytes(file)
        if file_size > SUPPLIER_UPLOAD_MAX_BYTES:
            return jsonify({
                "success": False,
                "message": "File too large. Maximum size is 10MB.",
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute("SELECT 1 FROM suppliers WHERE supplier_id = %s", (supplier_id,))
            if not cur.fetchone():
                return jsonify({"success": False, "message": "Supplier not found"}), 404
        finally:
            cur.close()
            conn.close()

        display_name = file.filename or "attachment"
        rel_path = _supplier_attachment_relative_path(supplier_id, display_name, category)
        save_path, _file_size = _persist_module_upload(
            object_storage.MODULE_SUPPLIER_ATTACHMENTS,
            SUPPLIER_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            _ensure_supplier_attachment_table(cur)
            cur.execute(
                """
                INSERT INTO supplier_attachments (supplier_id, file_name, file_path, category)
                VALUES (%s, %s, %s, %s)
                RETURNING id, uploaded_at
                """,
                (supplier_id, display_name, save_path, category),
            )
            row = cur.fetchone()
            attachment_id = row[0] if row else None
            if attachment_id is not None:
                cur.execute(
                    """
                    DELETE FROM supplier_attachments
                    WHERE supplier_id = %s AND file_name = %s
                      AND COALESCE(category, '') = %s AND id != %s
                    RETURNING file_path
                    """,
                    (supplier_id, display_name, category, attachment_id),
                )
                for old in cur.fetchall() or []:
                    _remove_stored_upload(old[0], SUPPLIER_ATTACHMENTS_FOLDER)
            conn.commit()
            uploaded_at = row[1].strftime("%Y-%m-%d %H:%M:%S") if row and row[1] else ""
        finally:
            cur.close()
            conn.close()

        return jsonify({
            "success": True,
            "id": attachment_id,
            "file_name": display_name,
            "file_path": save_path,
            "category": category,
            "uploaded_at": uploaded_at,
        })
    except Exception as e:
        print("❌ upload_supplier_attachment:", e)
        return jsonify({"success": False, "error": str(e)}), 500
 
 
@app.route("/api/supplier-attachments/<supplier_id>", methods=["GET"])
def get_supplier_attachments(supplier_id):
    supplier_id = (supplier_id or "").strip().upper()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_supplier_attachment_table(cur)
        cur.execute(
            """
            SELECT id, file_name, file_path, uploaded_at, COALESCE(category, '')
            FROM supplier_attachments
            WHERE supplier_id = %s
            ORDER BY uploaded_at DESC
            """,
            (supplier_id,),
        )
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                "id": r[0],
                "file_name": r[1],
                "file_path": r[2],
                "uploaded_at": r[3].strftime("%Y-%m-%d %H:%M:%S") if r[3] else "",
                "category": r[4] or "",
            })
        return jsonify({"success": True, "attachments": result})
    finally:
        cur.close()
        conn.close()
 
 
@app.route("/api/supplier-attachments/<supplier_id>/<int:attachment_id>/view")
def view_supplier_attachment(supplier_id, attachment_id):
    supplier_id = (supplier_id or "").strip().upper()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_supplier_attachment_table(cur)
        cur.execute(
            """
            SELECT file_name, file_path FROM supplier_attachments
            WHERE id = %s AND supplier_id = %s
            """,
            (attachment_id, supplier_id),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404
        stored = row[1]
        if object_storage.is_remote_url(str(stored or "")):
            return redirect(stored)
        full_path = _resolve_stored_file_path(stored)
        if not full_path or not os.path.isfile(full_path):
            return jsonify({"success": False, "message": "File not found on server"}), 404
        return send_file(full_path, as_attachment=False, download_name=row[0] or "attachment")
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/supplier-attachments/<supplier_id>/<int:attachment_id>/download")
def download_supplier_attachment(supplier_id, attachment_id):
    supplier_id = (supplier_id or "").strip().upper()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_supplier_attachment_table(cur)
        cur.execute(
            """
            SELECT file_name, file_path FROM supplier_attachments
            WHERE id = %s AND supplier_id = %s
            """,
            (attachment_id, supplier_id),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404
        stored = row[1]
        if object_storage.is_remote_url(str(stored or "")):
            return redirect(stored)
        full_path = _resolve_stored_file_path(stored)
        if not full_path or not os.path.isfile(full_path):
            return jsonify({"success": False, "message": "File not found on server"}), 404
        return send_file(full_path, as_attachment=True, download_name=row[0] or "attachment")
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/supplier-attachments/<supplier_id>/<int:attachment_id>", methods=["DELETE"])
def delete_supplier_attachment(supplier_id, attachment_id):
    supplier_id = (supplier_id or "").strip().upper()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_supplier_attachment_table(cur)
        cur.execute(
            """
            SELECT file_path FROM supplier_attachments
            WHERE id = %s AND supplier_id = %s
            """,
            (attachment_id, supplier_id),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404
        _remove_stored_upload(row[0], SUPPLIER_ATTACHMENTS_FOLDER)
        cur.execute(
            "DELETE FROM supplier_attachments WHERE id = %s AND supplier_id = %s",
            (attachment_id, supplier_id),
        )
        conn.commit()
        return jsonify({"success": True, "message": "Attachment deleted"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


#----API endpoint to delete a supplier by ID (for supplier list UI)
@app.route("/api/suppliers/<supplier_id>", methods=["DELETE"])
def delete_supplier(supplier_id):
    supplier_id = (supplier_id or "").strip().upper()
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
        _ensure_supplier_attachment_table(cur)
        cur.execute(
            "SELECT file_path FROM supplier_attachments WHERE supplier_id = %s",
            (supplier_id,),
        )
        for (file_path,) in cur.fetchall():
            _remove_stored_upload(file_path, SUPPLIER_ATTACHMENTS_FOLDER)
        cur.execute("DELETE FROM supplier_attachments WHERE supplier_id = %s", (supplier_id,))
        cur.execute("DELETE FROM suppliers WHERE supplier_id = %s", (supplier_id,))
        conn.commit()
 
        return jsonify({"success": True, "message": "Supplier deleted successfully"})
 
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
 
    finally:
        cur.close()
        conn.close()
 
 
 
 

# =========================================
@app.route("/crm")
def crm():
    return render_template("crm.html", page="crm")

# =========================================

# =========================================
# 1. ROOT & AUTH — OTP APIs
# =========================================
@app.route("/send_otp", methods=["POST"])
def send_otp():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()

    if not email:
        return jsonify(success=False, message="Email is required"), 400

    if len(email) > MAX_EMAIL_LENGTH:
        return jsonify(success=False, message="Email is too long. Max 50 characters."), 400

    if not EMAIL_REGEX.match(email):
        return jsonify(success=False, message="Enter a valid email address like name@gmail.com or name@yahoo.com"), 400

    # =========================================
    # BUG_008 — Simple per‑email OTP rate limit
    # Return HTTP 429 if too many OTPs are requested in a short window.
    # =========================================
    now = time.time()
    history = OTP_SEND_COUNT.get(email, [])
    # Keep only recent timestamps inside the window
    history = [ts for ts in history if now - ts <= OTP_WINDOW_SECONDS]
    if len(history) >= MAX_OTP_SENDS:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Too many OTP requests. Please wait a few minutes before trying again.",
                }
            ),
            429,
        )
    history.append(now)
    OTP_SEND_COUNT[email] = history

    otp = generate_otp()
    print("DEBUG OTP for", email, "=", otp)

    save_otp_in_db(email, otp)

    try:
        send_otp_email(email, otp)
    except Exception as e:
        print("Error sending OTP:", e)
        return jsonify(success=False, message="Error sending OTP. Try again."), 500

    return jsonify(success=True, message="OTP sent successfully!")


@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip().lower()
    otp = (data.get("otp") or "").strip()

    if verify_otp_in_db(email, otp):
        return jsonify({"success": True, "message": "OTP verified successfully!"}), 200
    return jsonify({"success": False, "message": "Invalid or expired OTP"}), 400


# =========================================
# 1. ROOT & AUTH — Signup API
# =========================================
# =========================================
# 1. ROOT & AUTH — Signup API
# =========================================
@app.route("/signup", methods=["POST"])
def signup():
    try:
        data = request.get_json() or {}
 
        first_name = (data.get("first_name") or data.get("firstName") or "").strip()
        last_name = (data.get("last_name") or data.get("lastName") or "").strip()
        name = (data.get("name") or "").strip()
        if not name:
            name = f"{first_name} {last_name}".strip()
        phone = (data.get("phone") or "").strip()
        email = (data.get("email") or "").strip().lower()
        password = (data.get("password") or "").strip()
        country_code, contact_number = _infer_country_and_contact_from_phone(phone)
 
        # ========= VALIDATION =========
        missing = []
        if not first_name:
            missing.append("First Name")
        if not last_name:
            missing.append("Last Name")
        if not phone:
            missing.append("Phone number")
        if not email:
            missing.append("Email")
        if not password:
            missing.append("Password")
 
        if missing:
            if len(missing) == 1:
                msg = f"⚠️ {missing[0]} is required"
            else:
                msg = "⚠️ " + ", ".join(missing[:-1]) + f" and {missing[-1]} are required"
            return jsonify({"success": False, "message": msg}), 400
 
        if not NAME_REGEX.match(first_name):
            return jsonify({"success": False, "message": "⚠️ First name must be 3–20 letters only"}), 400

        if not re.match(r"^[A-Za-z\s]{1,30}$", last_name):
            return jsonify({"success": False, "message": "⚠️ Last name must be 1–30 letters only"}), 400
 
        if not re.match(r"^\+\d{8,15}$", phone):
            return jsonify({"success": False, "message": "Enter valid phone like +91XXXXXXXXXX"}), 400
 
        if len(email) > MAX_EMAIL_LENGTH:
            return jsonify({"success": False, "message": "⚠️ Email too long"}), 400
 
        if not EMAIL_REGEX.match(email):
            return jsonify({"success": False, "message": "⚠️ Invalid email"}), 400
 
        if not is_email_otp_verified(email):
            return jsonify({
                "success": False,
                "message": "⚠️ Please verify OTP before signup"
            }), 400
 
        # ========= DB CONNECTION =========
        conn = get_db_connection()
        cur = conn.cursor()
 
        # ========= CHECK USER EXISTS =========
        cur.execute("SELECT user_id FROM users WHERE LOWER(email) = LOWER(%s)", (email,))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "⚠️ User already exists"}), 409
 
        # ========= INSERT USER =========
        import hashlib
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        cur.execute("""
            INSERT INTO users (name, phone, email, password, role, first_name, last_name, country_code, contact_number)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            name,
            phone,
            email,
            hashed_password,
            "User",
            first_name,
            last_name,
            country_code,
            contact_number,
        ))
 
        # If signup OTP exists in cache, persist it in DB columns too.
        otps = load_otps()
        otp_entry = otps.get(email) if isinstance(otps, dict) else None
        if isinstance(otp_entry, dict):
            cached_otp = (otp_entry.get("otp") or "").strip()
            cached_ts = otp_entry.get("timestamp")
            otp_expiry_dt = None
            if cached_ts:
                try:
                    otp_expiry_dt = datetime.fromtimestamp(float(cached_ts) + 300)
                except Exception:
                    otp_expiry_dt = None
            if cached_otp:
                cur.execute(
                    """
                    UPDATE users
                    SET email_otp = %s,
                        otp_expiry = %s
                    WHERE LOWER(email) = LOWER(%s)
                    """,
                    (cached_otp, otp_expiry_dt, email),
                )

        conn.commit()
        cur.close()
        conn.close()
 
        # ========= SEND EMAIL =========
        try:
            send_email(email, "Welcome!", f"Hello {name}, your account has been created successfully!")
        except Exception as mail_err:
            print("Welcome email failed:", mail_err)
 
        return jsonify({"success": True, "message": "Signup successful!"}), 200
 
    except Exception as e:
        print("Signup error:", e)
        return jsonify({"success": False, "message": "Server error"}), 500
# =========================================
# 1. ROOT & AUTH — Login API
# =========================================
 
@app.route("/login", methods=["POST"])
def login_post():
    try:
        if not request.is_json:
            return jsonify({"success": False, "message": "Expected JSON"}), 400
 
        data = request.get_json(silent=True) or {}
        email = (data.get("email") or "").strip().lower()
        password = (data.get("password") or "").strip()
        remember_me = data.get("rememberMe", False)
 
        if not email:
            return jsonify({"success": False, "message": "Email required"}), 400
        if not password:
            return jsonify({"success": False, "message": "Password required"}), 400
 
        conn = get_db_connection()
        cur = conn.cursor()
 
        # ✅ Get user
        cur.execute("""
            SELECT user_id, name, role, password, branch, department
            FROM users
            WHERE LOWER(email) = LOWER(%s)
            LIMIT 1
        """, (email,))
        user = cur.fetchone()
 
        if not user:
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "User not found"}), 404
 
        (
            user_id,
            db_name,
            db_role,
            db_password,
            db_branch,
            db_department
        ) = user
 
        # ❌ Wrong password
        import hashlib
        hashed_input = hashlib.sha256(password.encode()).hexdigest()
        password_match = (db_password == hashed_input) or (db_password == password)
        if not password_match:
            cur.close()
            conn.close()
            return jsonify({"success": False, "message": "Wrong password"}), 401
 
        # ✅ Password correct — set session directly
        session.permanent = True
        session["user"] = email
        session["role"] = db_role
        session["branch"] = db_branch or "Main Branch"
        session["department"] = db_department or ""
        session["last_active"] = time.time()
        session["remember_me"] = remember_me

        # Keep DB login audit columns updated for normal password login too.
        cur.execute(
            """
            UPDATE users
            SET last_seen = NOW()
            WHERE user_id = %s
            """,
            (user_id,),
        )
 
        conn.commit()
        cur.close()
        conn.close()
 
        return jsonify({"success": True, "message": "Login successful"}), 200
 
    except Exception as e:
        import traceback
        traceback.print_exc()
        print("LOGIN ERROR DETAIL:", repr(e))
        return jsonify({"success": False, "message": f"Server error: {str(e)}"}), 500
    #---------otp verify---------
 
@app.route("/verify-login-otp", methods=["POST"])
def verify_login_otp():
    try:
        data = request.get_json()
        user_id = data.get("user_id")
        otp = data.get("otp")
 
        conn = get_db_connection()
        cur = conn.cursor()
 
        cur.execute("""
            SELECT email, email_otp, otp_expiry, role, branch, department
            FROM users
            WHERE user_id = %s
        """, (user_id,))
        user = cur.fetchone()
 
        if not user:
            return jsonify({"success": False, "message": "User not found"}), 404
 
        email, db_otp, expiry, role, branch, department = user
 
        if otp != db_otp:
            return jsonify({"success": False, "message": "Invalid OTP"}), 400
 
        if expiry:
            expiry_naive = expiry.replace(tzinfo=None) if expiry.tzinfo else expiry
            if datetime.now() > expiry_naive:
                return jsonify({"success": False, "message": "OTP expired"}), 400
 
        # ✅ SUCCESS LOGIN
        session.permanent = True
        session["user"] = email
        session["role"] = role
        session["branch"] = branch or "Main Branch"
        session["department"] = department or ""
        session["last_active"] = time.time()
        session["remember_me"] = data.get("rememberMe", False)
 
        # ✅ Update last_seen
        cur.execute("""
            UPDATE users
            SET last_seen = NOW()
            WHERE user_id = %s
        """, (user_id,))
 
        conn.commit()
        cur.close()
        conn.close()
 
        return jsonify({"success": True, "message": "Login successful"})
 
    except Exception as e:
        print("❌ OTP error:", e)
        return jsonify({"success": False, "message": "Server error"}), 500
 
 


# @app.route("/login", methods=["POST"])
# def login_post():
#     try:
#         data = request.get_json()
#         email = (data.get("email") or "").strip().lower()
#         password = (data.get("password") or "").strip()

#         conn = get_db_connection()
#         cur = conn.cursor()

#         # Get user
#         cur.execute("""
#             SELECT id, password, failed_attempts, is_locked
#             FROM users
#             WHERE LOWER(email) = LOWER(%s)
#         """, (email,))
#         user = cur.fetchone()

#         if not user:
#             return jsonify({"success": False, "message": "User not found"}), 404

#         user_id, db_password, failed_attempts, is_locked = user

#         # 🔒 Check if locked
#         if is_locked:
#             return jsonify({"success": False, "message": "Account locked. Try later"}), 403

#         # ❌ Wrong password
#         if password != db_password:
#             failed_attempts += 1

#             if failed_attempts >= 3:
#                 cur.execute("""
#                     UPDATE users
#                     SET failed_attempts = %s,
#                         is_locked = TRUE,
#                         lock_time = NOW()
#                     WHERE id = %s
#                 """, (failed_attempts, user_id))
#             else:
#                 cur.execute("""
#                     UPDATE users
#                     SET failed_attempts = %s
#                     WHERE id = %s
#                 """, (failed_attempts, user_id))

#             conn.commit()
#             return jsonify({"success": False, "message": f"Wrong password ({failed_attempts}/3)"}), 401

#         # ✅ Correct password → reset attempts
#         cur.execute("""
#             UPDATE users
#             SET failed_attempts = 0
#             WHERE id = %s
#         """, (user_id,))

#         # 🔐 Generate OTP
#         otp = str(random.randint(100000, 999999))
#         expiry = datetime.now() + timedelta(minutes=5)

#         cur.execute("""
#             UPDATE users
#             SET email_otp = %s,
#                 otp_expiry = %s
#             WHERE id = %s
#         """, (otp, expiry, user_id))

#         conn.commit()

#         # 📧 Send OTP
#         send_email(email, "Login OTP", f"Your OTP is {otp}")

#         return jsonify({
#             "success": True,
#             "message": "OTP sent to email",
#             "user_id": user_id
#         })

#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)}), 500


# @app.route("/verify-login-otp", methods=["POST"])
# def verify_login_otp():
#     try:
#         data = request.get_json()
#         user_id = data.get("user_id")
#         otp = data.get("otp")

#         conn = get_db_connection()
#         cur = conn.cursor()

#         cur.execute("""
#             SELECT email, email_otp, otp_expiry
#             FROM users
#             WHERE id = %s
#         """, (user_id,))
#         user = cur.fetchone()

#         if not user:
#             return jsonify({"success": False, "message": "User not found"}), 404

#         email, db_otp, expiry = user

#         # ❌ OTP wrong
#         if otp != db_otp:
#             return jsonify({"success": False, "message": "Invalid OTP"}), 400

#         # ⏰ OTP expired
#         if datetime.now() > expiry:
#             return jsonify({"success": False, "message": "OTP expired"}), 400

#         # ✅ SUCCESS LOGIN
#         session["user"] = email

#         cur.execute("""
#             UPDATE users
#             SET last_seen = NOW(),
#                 email_otp = NULL,
#                 otp_expiry = NULL
#             WHERE id = %s
#         """, (user_id,))

#         conn.commit()

#         return jsonify({"success": True, "message": "Login successful"})

#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)}), 500


# =========================
# API: LIST + FILTER + PAGINATION (Supports HTML & JSON)
# =========================
@app.route("/api/products", methods=["GET"])
def api_products():
    """
    GET /api/products
    Query Parameters:
        - q: Search query
        - type: Filter by product type
        - category: Filter by category
        - status: Filter by status
        - stock: Filter by stock level (out/low/ok)
        - page: Page number (default: 1)
        - page_size: Items per page (default: 10)
        - format: Force format (json/html)
    
    Returns JSON or HTML based on Accept header
    """
    # BUG_001 / BUG_006: Require login for product list API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    q = (request.args.get("q") or "").strip().lower()
    ptype = (request.args.get("type") or "").strip()
    cat = (request.args.get("category") or "").strip()
    status = (request.args.get("status") or "").strip()
    brand = (request.args.get("brand") or "").strip()
    stock = (request.args.get("stock") or "").strip()

    # BUG_002: Robust validation for pagination parameters
    raw_page = request.args.get("page", "1")
    raw_page_size = request.args.get("page_size", "10")
    try:
        page = int(raw_page)
        page_size = int(raw_page_size)
    except (TypeError, ValueError):
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page' and 'page_size' must be integers.",
                }
            ),
            400,
        )

    if page <= 0 or page_size <= 0:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page' and 'page_size' must be greater than 0.",
                }
            ),
            400,
        )

    # Optional upper bound to avoid huge pages
    if page_size > 1000:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Invalid pagination parameters. 'page_size' is too large.",
                }
            ),
            400,
        )

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        where_clauses = []
        params = []

        if q:
            where_clauses.append(
                """
                (
                    LOWER(COALESCE(product_id, '')) LIKE %s
                    OR LOWER(COALESCE(product_name, '')) LIKE %s
                    OR LOWER(COALESCE(product_type, '')) LIKE %s
                    OR LOWER(COALESCE(category_name, '')) LIKE %s
                    OR LOWER(COALESCE(status, '')) LIKE %s
                )
                """
            )
            like = f"%{q}%"
            params.extend([like, like, like, like, like])
        if ptype:
            where_clauses.append("COALESCE(product_type, '') = %s")
            params.append(ptype)
        if cat:
            where_clauses.append("COALESCE(category_name, '') = %s")
            params.append(cat)
        if status:
            where_clauses.append("COALESCE(status, '') = %s")
            params.append(status)
        if brand:
            where_clauses.append("COALESCE(supplier_name, '') = %s")
            params.append(brand)
        if stock == "out":
            where_clauses.append("COALESCE(stock_level, 0) = 0")
        elif stock == "low":
            where_clauses.append("COALESCE(stock_level, 0) BETWEEN 1 AND 5")
        elif stock == "ok":
            where_clauses.append("COALESCE(stock_level, 0) > 5")

        where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

        cur.execute(
            f"SELECT COUNT(*)::int AS c FROM products {where_sql}",
            tuple(params),
        )
        total_items = int((cur.fetchone() or {}).get("c") or 0)
        total_pages = max(1, (total_items + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        offset = (page - 1) * page_size

        cur.execute(
            f"""
            SELECT
                product_id,
                product_name,
                product_type,
                category_name,
                status,
                stock_level,
                unit_price,
                supplier_name
            FROM products
            {where_sql}
            ORDER BY
                CASE
                    WHEN product_id ~ '[0-9]+$'
                    THEN CAST(REGEXP_REPLACE(product_id, '^\\D+', '') AS INTEGER)
                    ELSE 0
                END DESC,
                product_id DESC
            LIMIT %s OFFSET %s
            """,
            tuple(params + [page_size, offset]),
        )
        rows = cur.fetchall() or []

        items = []
        for r in rows:
            d = dict(r)
            d["type"] = d.get("product_type") or ""
            d["category"] = d.get("category_name") or ""
            d["status"] = str(d.get("status") or "").strip().capitalize()
            d["stock_level"] = int(d.get("stock_level") or 0)
            d["price"] = float(d.get("unit_price") or 0.0)
            d["brand"] = d.get("supplier_name") or ""
            items.append(d)

        cur.execute("SELECT DISTINCT product_type FROM products WHERE COALESCE(product_type,'') <> '' ORDER BY product_type")
        types = [str(r.get("product_type") or "") for r in (cur.fetchall() or []) if str(r.get("product_type") or "").strip()]
        cur.execute("SELECT DISTINCT category_name FROM products WHERE COALESCE(category_name,'') <> '' ORDER BY category_name")
        categories = [str(r.get("category_name") or "") for r in (cur.fetchall() or []) if str(r.get("category_name") or "").strip()]
        cur.execute("SELECT DISTINCT status FROM products WHERE COALESCE(status,'') <> '' ORDER BY status")
        statuses = [str(r.get("status") or "") for r in (cur.fetchall() or []) if str(r.get("status") or "").strip()]
        cur.execute("SELECT DISTINCT supplier_name FROM products WHERE COALESCE(supplier_name,'') <> '' ORDER BY supplier_name")
        brands = [str(r.get("supplier_name") or "") for r in (cur.fetchall() or []) if str(r.get("supplier_name") or "").strip()]
    finally:
        cur.close()
        conn.close()

    response_data = {
        "success": True,
        "data": {
        "items": items,
        "page": page,
        "total_pages": total_pages,
        "total_items": total_items,
            "meta": {
                "types": types,
                "categories": categories,
                "statuses": statuses,
                "brands": brands
            }
        }
    }

    if wants_json():
        return jsonify(response_data), 200
    else:
        # For HTML, redirect to products page or return JSON anyway
        return jsonify(response_data), 200



# =========================
# API: DELETE PRODUCT (Supports HTML & JSON)
# =========================
@app.route("/api/products/<product_id>", methods=["DELETE"])
def api_delete_product(product_id):
    """
    DELETE /api/products/<product_id>
    
    Deletes a product by ID
    """
    # BUG_001 / BUG_006: Require login for delete API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    deleted_product = None
    try:
        cur.execute("SELECT * FROM products WHERE product_id = %s", (product_id,))
        deleted_product = cur.fetchone()
        if not deleted_product:
            error_response = {
                "success": False,
                "message": "Product not found",
                "error": f"Product with ID '{product_id}' does not exist"
            }
            return jsonify(error_response), 404

        cur.execute("DELETE FROM products WHERE product_id = %s", (product_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        msg = str(e)
        if "ForeignKeyViolation" in msg or "still referenced" in msg or "violates foreign key constraint" in msg:
            return jsonify({
                "success": False,
                "message": "Cannot delete product because it is used in other records.",
                "error": msg,
            }), 409
        return jsonify({
            "success": False,
            "message": "Delete failed",
            "error": msg,
        }), 500
    finally:
        cur.close()
        conn.close()

    if not deleted_product:
        error_response = {
            "success": False,
            "message": "Product not found",
            "error": f"Product with ID '{product_id}' does not exist"
        }
        return jsonify(error_response), 404
    
    response_data = {
        "success": True,
        "message": "Product deleted successfully",
        "data": dict(deleted_product)
    }

    return jsonify(response_data), 200
# =========================
# API: IMPORT CSV
# CSV columns: product_id,product_name,type,category,status,stock_level,price
# =========================
@app.route("/api/products/import", methods=["POST"])
def api_import_products():
    # BUG_001 / BUG_006: Require login for import API
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if "file" not in request.files:
        return jsonify({"message": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith(".csv"):
        return jsonify({"message": "Only CSV supported"}), 400

    products = load_products()
    existing_ids = {p.get("product_id") for p in products}

    decoded = file.stream.read().decode("utf-8", errors="ignore").splitlines()
    reader = csv.DictReader(decoded)

    added = 0
    for row in reader:
        pid = (row.get("product_id") or "").strip()
        if not pid or pid in existing_ids:
            continue

        item = {
            "product_id": pid,
            "product_name": (row.get("product_name") or "").strip(),
            "type": (row.get("type") or "").strip(),
            "category": (row.get("category") or "").strip(),
            "status": (row.get("status") or "Active").strip(),
            "stock_level": int((row.get("stock_level") or 0)),
            "price": float((row.get("price") or 0))
        }
        products.append(item)
        existing_ids.add(pid)
        added += 1

    save_products(products)
    return jsonify({"message": f"Import done ✅ Added {added} products"})


# =========================================
# ✅ SESSION TIMEOUT FUNCTION
# =========================================
def check_session_timeout():
    if "user" not in session:
        return False

    # If "Remember Me" is checked, skip inactivity timeout check
    remember_me = session.get("remember_me", False)
    if remember_me:
        # Still update last_active for tracking, but don't expire session
        session["last_active"] = time.time()
        return True

    # For normal sessions, check inactivity timeout
    last_active = session.get("last_active", 0)
    now = time.time()

    if now - last_active > INACTIVITY_TIMEOUT:
        session.clear()
        return False

    session["last_active"] = now
    return True


# =========================================
# 9. UTILITY — Logout
# =========================================
@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("last_active", None)
    session.pop("remember_me", None)
    return redirect(url_for("login", message="logged_out"))


# =========================================
# 9. UTILITY — Global Search
# =========================================
@app.route("/search")
def global_search():
    q = (request.args.get("q") or "").strip().lower()
    results = []

    if not q:
        return jsonify({"results": []})

    users = load_users()
    for u in users:
        if not isinstance(u, dict):
            continue

        name = (u.get("name") or "").strip()
        email = (u.get("email") or "").strip()
        phone = (u.get("phone") or "").strip()

        if (q in name.lower() or q in email.lower() or q in phone):
            results.append({
                "type": "User",
                "label": f"{name} - {email}",
                "url": "/manage-users",
            })

    menu_items = [
        ("Dashboard", "/dashboard"),
        ("Masters", "/manage-users"),
        ("Manage Users", "/manage-users"),
        ("Products", "/products"),
        ("Customer", "/customer"),
        ("Department Role", "/department-role"),
        ("CRM", "/crm"),
        ("Enquiry List", "/crm"),
        ("Quotation Module", "/crm"),
        ("Sales", "/crm"),
        ("Delivery Note Module", "/crm"),
        ("Invoice Module", "/crm"),
        ("Delivery Note Return", "/crm"),
        ("Invoice Return Module", "/crm"),
    ]

    for label, url in menu_items:
        if q in label.lower():
            results.append({
                "type": "Menu",
                "label": label,
                "url": url,
            })

    return jsonify({"results": results})


# =========================================
# 2. DASHBOARD — End of dashboard routes
# =========================================
@app.route("/api/top-products")
def top_products():
    try:
        query = """
            SELECT
                product_name AS name,
                SUM(qty)::float AS qty,
                ROUND(
                    (SUM(qty) * 100.0 / SUM(SUM(qty)) OVER()),
                    2
                )::float AS percentage
            FROM sales_order_items
            GROUP BY product_name
            ORDER BY qty DESC
            LIMIT 5;
        """
        rows = fetch_all(query)
        result = [
            {"name": str(r["name"] or ""), "qty": float(r["qty"] or 0), "percentage": float(r["percentage"] or 0)}
            for r in (rows or [])
        ]
        return jsonify(result)
    except Exception as e:
        print(f"top_products error: {e}")
        return jsonify([]), 200
@app.route("/api/monthly-sales")
def monthly_sales():
    try:
        query = """
            SELECT
                TO_CHAR(m.month, 'Mon') AS month,
                COALESCE(SUM(s.grand_total), 0) AS total
            FROM generate_series(
                DATE_TRUNC('year', CURRENT_DATE),
                DATE_TRUNC('year', CURRENT_DATE) + INTERVAL '11 months',
                INTERVAL '1 month'
            ) AS m(month)
            LEFT JOIN sales_orders s
                ON DATE_TRUNC('month', s.order_date) = m.month
            GROUP BY m.month
            ORDER BY m.month;
        """
        rows = fetch_all(query)
        result = [
            {"month": str(r["month"]), "total": float(r["total"] or 0)}
            for r in (rows or [])
        ]
        return jsonify(result)
    except Exception as e:
        print(f"monthly_sales error: {e}")
        return jsonify([]), 200
 
# =========================================
# 3. MASTERS — Manage Users — API
# =========================================
@app.route("/api/users", methods=["GET"])
def api_get_users():
    """Get all users as JSON - for Postman/API testing"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    users = _db_fetch_users_ordered(include_id=False)

    user_name = "User"
    user_role = "User"

    current_email = (user_email or "").strip().lower()

    for u in users:
        if not isinstance(u, dict):
            continue
        u_email = (u.get("email") or "").strip().lower()
        if u_email == current_email:
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    return jsonify({
        "success": True,
        "users": [user_public_dict(u) for u in users if isinstance(u, dict)],
        "total": len(users),
        "current_user": {
            "email": user_email,
            "name": user_name,
            "role": user_role
        }
    }), 200


# 3. MASTERS — Manage Users — API (continued)
@app.route("/api/users/<int:user_index>", methods=["GET"])
def api_get_user(user_index):
    """Get a single user by index as JSON - for Postman/API testing"""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if user_index < 0:
        return jsonify({"success": False, "message": "User index out of range"}), 404

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT user_id, name, email, phone, role, first_name, last_name, country_code,
                   contact_number, branch, department, reporting_to, available_branches, employee_id
            FROM users
            ORDER BY user_id DESC
            OFFSET %s LIMIT 1
            """,
            (user_index,),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "User index out of range"}), 404
        user = {
            "name": row[1],
            "email": row[2],
            "phone": row[3],
            "role": row[4],
            "first_name": row[5],
            "last_name": row[6],
            "country_code": row[7],
            "contact_number": row[8],
            "branch": row[9],
            "department": row[10],
            "reporting_to": row[11],
            "available_branches": str(row[12]) if row[12] is not None else "",
            "employee_id": row[13],
        }
    finally:
        cur.close()
        conn.close()

    return jsonify({
        "success": True,
        "user": user_public_dict(user),
        "index": user_index
    }), 200


# =========================================
# ✅ API: CREATE USER (POST /api/users) — JSON for Postman
# =========================================
@app.route("/api/users", methods=["POST"])
def api_create_user():
    """Create new user. Requires JSON body. Use ?format=json or Accept: application/json."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    prof = get_current_user_profile() or {}
    user_role = prof.get("role") or "User"

    if normalize_role(user_role) not in ["superadmin", "admin"]:
        return jsonify({"success": False, "message": "Only Super Admin/Admin can create users."}), 403

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    users = load_users()

    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    email = (data.get("email") or "").strip()
    country_code = (data.get("country_code") or "").strip()
    contact_number = (data.get("contact_number") or "").strip()
    branch = (data.get("branch") or "").strip()
    department = (data.get("department") or "").strip()
    role = (data.get("role") or "").strip()
    reporting_to = (data.get("reporting_to") or "").strip()
    available_branches = (data.get("available_branches") or "").strip()
    employee_id = (data.get("employee_id") or "").strip()

    errors = []
    if not first_name or len(first_name) < 3:
        errors.append("First Name required (min 3 chars)")
    if not last_name or len(last_name) < 3:
        errors.append("Last Name required (min 3 chars)")
    if not email:
        errors.append("Email required")
    elif not EMAIL_REGEX.match(email):
        errors.append("Invalid email format")
    if not branch:
        errors.append("Branch required")
    if not department:
        errors.append("Department required")
    if not role:
        errors.append("Role required")
    if not reporting_to:
        errors.append("Reporting To required")
    if not available_branches:
        errors.append("Available Branches required")
    if not employee_id:
        errors.append("Employee ID required")

    for u in users:
        if isinstance(u, dict):
            if (u.get("email") or "").strip().lower() == email.lower():
                errors.append("Email already exists")
                break
            if (u.get("employee_id") or "") == employee_id:
                errors.append("Employee ID already exists")
                break

    if errors:
        return jsonify({"success": False, "message": "; ".join(errors), "errors": errors}), 400

    full_name = (first_name + " " + last_name).strip()
    full_phone = f"{country_code}{contact_number}" if country_code and contact_number else contact_number
    api_password = (data.get("password") or "").strip() or DEFAULT_BRANCH_USER_PASSWORD

    new_user = {
        "name": full_name,
        "phone": full_phone,
        "first_name": first_name,
        "last_name": last_name,
        "email": email,
        "country_code": country_code,
        "contact_number": contact_number,
        "branch": branch,
        "department": department,
        "role": role,
        "reporting_to": reporting_to,
        "available_branches": available_branches,
        "employee_id": employee_id,
        "password": api_password,
    }
    users.append(new_user)
    save_users(users)

    safe_user = {k: v for k, v in new_user.items() if k != "password"}
    return jsonify({
        "success": True,
        "message": "User created successfully",
        "user": safe_user,
    }), 201


# =========================================
# ✅ API: UPDATE USER (PUT /api/users/<index>) — JSON for Postman
# =========================================
@app.route("/api/users/<int:user_index>", methods=["PUT"])
def api_update_user(user_index):
    """Update user by index. Requires JSON body."""
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    if user_index < 0:
        return jsonify({"success": False, "message": "User index out of range"}), 404

    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s) LIMIT 1",
            ((user_email or "").strip(),),
        )
        current = cur.fetchone()
        if not current or normalize_role(current[0]) not in ["superadmin", "admin"]:
            return jsonify({"success": False, "message": "Only Super Admin/Admin can edit users."}), 403

        cur.execute(
            """
            SELECT user_id, name, email, phone, role, department, branch
            FROM users
            ORDER BY user_id DESC
            OFFSET %s LIMIT 1
            """,
            (user_index,),
        )
        base = cur.fetchone()
        if not base:
            return jsonify({"success": False, "message": "User index out of range"}), 404

        user_id = base[0]
        new_name = str(data.get("name", base[1] or "")).strip()
        new_email = str(data.get("email", base[2] or "")).strip()
        new_phone = str(data.get("phone", base[3] or "")).strip()
        new_role = str(data.get("role", base[4] or "User")).strip() or "User"
        new_department = str(data.get("department", base[5] or "")).strip()
        new_branch = str(data.get("branch", base[6] or "")).strip()

        cur.execute(
            """
            UPDATE users
            SET name=%s, email=%s, phone=%s, role=%s, department=%s, branch=%s
            WHERE id=%s
            """,
            (new_name, new_email, new_phone, new_role, new_department, new_branch, user_id),
        )
        conn.commit()

        refreshed = {
            "name": new_name,
            "email": new_email,
            "phone": new_phone,
            "role": new_role,
            "department": new_department,
            "branch": new_branch,
        }
        return jsonify({
            "success": True,
            "message": "User updated",
            "user": user_public_dict(refreshed),
        }), 200
    finally:
        cur.close()
        conn.close()


# =========================================
# ✅ API: DELETE USER (DELETE /api/users/<index>) — JSON for Postman
# =========================================
@app.route("/delete-user/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Not logged in"}), 401

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute(
            "SELECT role FROM users WHERE LOWER(email)=LOWER(%s)",
            (user_email,),
        )
        row = cur.fetchone()

        if not row:
            return jsonify({"success": False, "message": "User not found"}), 403

        if normalize_role(row[0]) != "superadmin":
            return jsonify({"success": False, "message": "Only super admin can delete"}), 403

        cur.execute(
            "SELECT user_id, email FROM users WHERE user_id = %s",
            (user_id,),
        )
        target = cur.fetchone()

        if not target:
            return jsonify({"success": False, "message": "User not found"}), 404

        db_id, deleted_email = target

        cur.execute("DELETE FROM users WHERE user_id = %s", (db_id,))
        conn.commit()

        return jsonify({
            "success": True,
            "message": "User deleted successfully",
            "deleted_email": deleted_email
        })

    except Exception as e:
        conn.rollback()
        print("❌ Delete error:", e)
        return jsonify({"success": False, "message": "Delete failed"}), 500

    finally:
        cur.close()
        conn.close()

# =========================================
# 7. CRM — Enquiry List
# =========================================

def generate_enquiry_id():
    row = fetch_one("SELECT enquiry_id FROM enquiry ORDER BY id DESC LIMIT 1")
    if not row:
        return "ENQ-0001"
    last_id = row["enquiry_id"]
    try:
        num = int(str(last_id).split("-")[1])
    except (IndexError, ValueError):
        return "ENQ-0001"
    return f"ENQ-{num + 1:04d}"


def _get_current_user_role():
    """Get current user's role from session / DB profile."""
    user_email = session.get("user")
    if not user_email:
        return None
    sr = session.get("role")
    if sr:
        return (str(sr).strip()).replace(" ", "").replace("_", "").lower()
    prof = get_current_user_profile()
    if prof and prof.get("role"):
        return (prof.get("role") or "User").strip().replace(" ", "").replace("_", "").lower()
    return "user"


def _get_logged_in_user_name():
    """Display name for the logged-in session from PostgreSQL users row."""
    user_email = session.get("user")
    if not user_email:
        return "User"

    dbu = _db_get_user_by_email(user_email)
    if dbu:
        n = (dbu.get("name") or "").strip()
        if n:
            return n

    return "User"


@app.route("/enquiry-list")
def enquiry_list():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    rows = fetch_all(
        """
        SELECT enquiry_id, first_name, last_name, email, phone_number, status,
               (SELECT COUNT(*) FROM enquiry_product ep WHERE ep.enquiry_id = e.enquiry_id) AS items_count
        FROM enquiry e
        ORDER BY created_at DESC
        """
    )
    enquiries = []
    for row in rows:
        enquiries.append(
            {
                "enquiry_id": row["enquiry_id"],
                "first_name": row["first_name"] or "",
                "last_name": row["last_name"] or "",
                "email": row["email"] or "",
                "phone_number": row["phone_number"] or "",
                "status": row["status"] or "New",
                "items": {},
            }
        )

    if wants_json():
        return (
            jsonify(
                {
                    "success": True,
                    "data": enquiries,
                    "total": len(enquiries),
                    "current_user": {
                        "email": user_email,
                        "name": user_name,
                        "role": user_role,
                    },
                    "permissions": get_effective_permissions_for_session(),
                }
            ),
            200,
        )

    return render_template(
        "enquiry-list.html",
        title="Enquiry List - Stackly",
        page="enquiry_list",
        section="crm",
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        enquiries=enquiries,
    )





@app.route("/api/enquiry/<enquiry_id>")
def get_enquiry(enquiry_id):
    row = fetch_one(
        """
        SELECT enquiry_id, first_name, last_name, phone_number, email
        FROM enquiry WHERE enquiry_id = %s
        """,
        (enquiry_id,),
    )
    if not row:
        return jsonify(success=False, message="Enquiry not found"), 404

    return jsonify(
        success=True,
        data={
            "enquiry_id": row["enquiry_id"],
            "first_name": row["first_name"] or "",
            "last_name": row["last_name"] or "",
            "phone": row["phone_number"] or "",
            "email": row["email"] or "",
        },
    )






@app.route("/update-enquiry/<enquiry_id>", methods=["POST"])
def update_enquiry(enquiry_id):
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can edit enquiries."), 403

    req_data = request.get_json()
    if not req_data:
        return jsonify(success=False, message="Invalid request"), 400

    updates = {}
    if "first_name" in req_data:
        updates["first_name"] = req_data["first_name"]
    if "last_name" in req_data:
        updates["last_name"] = req_data["last_name"]
    if "phone_number" in req_data:
        updates["phone_number"] = req_data["phone_number"]
    if "email" in req_data:
        updates["email"] = req_data["email"]

    if not updates:
        return jsonify(success=False, message="No fields to update"), 400

    set_clause = ", ".join([f"{k} = %s" for k in updates.keys()])
    values = list(updates.values()) + [enquiry_id]
    execute_query(
        f"UPDATE enquiry SET {set_clause}, updated_at = NOW() WHERE enquiry_id = %s",
        values,
    )

    return jsonify(success=True, message="Enquiry updated successfully")





@app.route("/delete-enquiry/<enquiry_id>", methods=["DELETE"])
def delete_enquiry(enquiry_id):
    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify(success=False, message="Only Super Admin can delete enquiries."), 403

    execute_query("DELETE FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    return jsonify(success=True, message="Enquiry deleted successfully")


def _enquiry_items_dict_for_api(enquiry_id):
    """Items keyed by product_id, shape expected by enquiry-list.js."""
    rows = fetch_all(
        """
        SELECT product_id, product_name, description,
               unit_price, selling_price, quantity, total
        FROM enquiry_product
        WHERE enquiry_id = %s
        """,
        (enquiry_id,),
    )
    result = {}
    for item in rows:
        pid = item["product_id"]
        result[pid] = {
            "item_code": pid,
            "item_name": item["product_name"],
            "description": item["description"],
            "unit_price": str(item["unit_price"]),
            "selling_price": str(item["selling_price"]),
            "quantity": item["quantity"],
            "total": str(item["total"]),
        }
    return result


def _enquiry_row_to_api(enquiry_id: str, row: dict) -> dict:
    if not row:
        return {}
    fn = row.get("first_name") or ""
    ln = row.get("last_name") or ""
    em = row.get("email") or ""
    ph = row.get("phone_number") or ""
    st = row.get("status") or "New"
    items = _enquiry_items_dict_for_api(enquiry_id)
    details = {
        "first_name": fn,
        "last_name": ln,
        "email": em,
        "phone": ph,
        "phone_number": ph,
        "status": st,
        "street": row.get("street"),
        "unit": row.get("unit"),
        "city": row.get("city"),
        "state": row.get("state"),
        "zip": row.get("zip"),
        "country": row.get("country"),
        "enquiry_type": row.get("enquiry_type"),
        "enquiry_description": row.get("enquiry_description"),
        "enquiry_channel": row.get("enquiry_channel"),
        "source": row.get("source"),
        "heard_about": row.get("heard_about"),
        "urgency": row.get("urgency"),
        "priority": row.get("priority"),
    }
    return {
        "enquiry_id": enquiry_id,
        "enquiry_details": details,
        "first_name": fn,
        "last_name": ln,
        "email": em,
        "phone_number": ph,
        "phone": ph,
        "status": st,
        "items": items,
    }


@app.route("/api/enquiries", methods=["GET"])
def api_enquiries_list():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    q = (request.args.get("search") or request.args.get("q") or "").strip().lower()
    status_filter = (request.args.get("status") or "").strip()
    raw_page = request.args.get("page")
    raw_page_size = request.args.get("page_size")
    use_pagination = bool(raw_page or raw_page_size)
    try:
        page = max(1, int(raw_page or 1))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(raw_page_size or 10)
    except (TypeError, ValueError):
        page_size = 10
    page_size = min(max(page_size, 1), 100)

    users = load_users()
    user_name = "User"
    user_role = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            user_role = (u.get("role") or "User").strip()
            break

    rows = fetch_all(
        """
        SELECT enquiry_id, first_name, last_name, email, phone_number, status
        FROM enquiry
        ORDER BY created_at DESC
        """
    )
    enquiries = []
    for row in rows:
        eid = row["enquiry_id"]
        st = row.get("status") or "New"
        if status_filter and st != status_filter:
            continue
        r = {
            "enquiry_id": eid,
            "first_name": row.get("first_name") or "",
            "last_name": row.get("last_name") or "",
            "email": row.get("email") or "",
            "phone_number": row.get("phone_number") or "",
            "phone": row.get("phone_number") or "",
            "status": st,
        }
        if q:
            blob = " ".join(
                [
                    str(eid),
                    r.get("first_name") or "",
                    r.get("last_name") or "",
                    r.get("email") or "",
                    r.get("phone_number") or "",
                    st,
                ]
            ).lower()
            if q not in blob:
                continue
        enquiries.append(r)

    total = len(enquiries)
    if use_pagination:
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, total_pages))
        start = (page - 1) * page_size
        page_items = enquiries[start : start + page_size]
    else:
        total_pages = 1
        page = 1
        page_items = enquiries

    return jsonify(
        {
            "success": True,
            "enquiries": page_items,
            "total": total,
            "page": page,
            "total_pages": total_pages,
            "current_user": {"email": user_email, "name": user_name, "role": user_role},
        }
    ), 200


@app.route("/api/enquiries/<enquiry_id>", methods=["GET"])
def api_enquiries_get_one(enquiry_id):
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    row = fetch_one("SELECT * FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    if not row:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    return jsonify({"success": True, "enquiry": _enquiry_row_to_api(enquiry_id, dict(row))}), 200


@app.route("/api/enquiries", methods=["POST"])
def api_enquiries_create():
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify({"success": False, "message": "Only Admin or Super Admin can create enquiries."}), 403

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    enquiry_id = (body.get("enquiry_id") or "").strip() or generate_enquiry_id()
    details_in = body.get("enquiry_details") if isinstance(body.get("enquiry_details"), dict) else {}
    items = body.get("items") if isinstance(body.get("items"), dict) else {}

    exists = fetch_one("SELECT 1 FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    if exists:
        return jsonify({"success": False, "message": "Enquiry ID already exists. Omit enquiry_id to auto-generate."}), 409

    execute_query(
        """
        INSERT INTO enquiry (
            id, enquiry_id, phone_number, first_name, last_name, email,
            street, unit, city, state, zip, country,
            enquiry_type, enquiry_description, enquiry_channel,
            source, heard_about, urgency, status, priority
        ) VALUES (
            (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry ep),
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        """,
        (
            enquiry_id,
            details_in.get("phone_number") or details_in.get("phone"),
            details_in.get("first_name"),
            details_in.get("last_name"),
            details_in.get("email"),
            details_in.get("street"),
            details_in.get("unit"),
            details_in.get("city"),
            details_in.get("state"),
            details_in.get("zip"),
            details_in.get("country"),
            details_in.get("enquiry_type"),
            details_in.get("enquiry_description"),
            details_in.get("enquiry_channel"),
            details_in.get("source"),
            details_in.get("heard_about"),
            details_in.get("urgency"),
            details_in.get("status", "New"),
            details_in.get("priority"),
        ),
    )

    for product_id, item in items.items():
        execute_query(
            """
            INSERT INTO enquiry_product
                (id, enquiry_id, product_id, product_name, description, unit_price, selling_price, quantity)
            VALUES (
                (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry_product ep),
                %s, %s, %s, %s, %s, %s, %s
            )
            """,
            (
                enquiry_id,
                product_id,
                item.get("item_name", ""),
                item.get("description", ""),
                item.get("unit_price", 0),
                item.get("selling_price", 0),
                item.get("quantity", 1),
            ),
        )

    row = fetch_one("SELECT * FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    return jsonify(
        {
            "success": True,
            "message": "Enquiry created",
            "enquiry": _enquiry_row_to_api(enquiry_id, dict(row)),
        }
    ), 201


@app.route("/api/enquiries/<enquiry_id>", methods=["PUT"])
def api_enquiries_update(enquiry_id):
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify({"success": False, "message": "Only Admin or Super Admin can update enquiries."}), 403

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"success": False, "message": "JSON body required"}), 400

    row = fetch_one("SELECT * FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    if not row:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    if "enquiry_details" in body and isinstance(body["enquiry_details"], dict):
        merged = dict(body["enquiry_details"])
        if merged.get("phone_number") is None and merged.get("phone") is not None:
            merged["phone_number"] = merged.get("phone")
        execute_query(
            """
            UPDATE enquiry SET
                phone_number = COALESCE(%s, phone_number),
                first_name = COALESCE(%s, first_name),
                last_name = COALESCE(%s, last_name),
                email = COALESCE(%s, email),
                street = COALESCE(%s, street),
                unit = COALESCE(%s, unit),
                city = COALESCE(%s, city),
                state = COALESCE(%s, state),
                zip = COALESCE(%s, zip),
                country = COALESCE(%s, country),
                enquiry_type = COALESCE(%s, enquiry_type),
                enquiry_description = COALESCE(%s, enquiry_description),
                enquiry_channel = COALESCE(%s, enquiry_channel),
                source = COALESCE(%s, source),
                heard_about = COALESCE(%s, heard_about),
                urgency = COALESCE(%s, urgency),
                status = COALESCE(%s, status),
                priority = COALESCE(%s, priority),
                updated_at = NOW()
            WHERE enquiry_id = %s
            """,
            (
                merged.get("phone_number"),
                merged.get("first_name"),
                merged.get("last_name"),
                merged.get("email"),
                merged.get("street"),
                merged.get("unit"),
                merged.get("city"),
                merged.get("state"),
                merged.get("zip"),
                merged.get("country"),
                merged.get("enquiry_type"),
                merged.get("enquiry_description"),
                merged.get("enquiry_channel"),
                merged.get("source"),
                merged.get("heard_about"),
                merged.get("urgency"),
                merged.get("status"),
                merged.get("priority"),
                enquiry_id,
            ),
        )

    if "items" in body:
        if not isinstance(body["items"], dict):
            return jsonify({"success": False, "message": "items must be an object"}), 400
        old_items = _enquiry_items_dict_for_api(enquiry_id)
        old_items.update(body["items"])
        execute_query("DELETE FROM enquiry_product WHERE enquiry_id = %s", (enquiry_id,))
        for product_id, item in old_items.items():
            execute_query(
                """
                INSERT INTO enquiry_product
                    (id, enquiry_id, product_id, product_name, description, unit_price, selling_price, quantity)
                VALUES (
                    (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry_product ep),
                    %s, %s, %s, %s, %s, %s, %s
                )
                """,
                (
                    enquiry_id,
                    product_id,
                    item.get("item_name", ""),
                    item.get("description", ""),
                    item.get("unit_price", 0),
                    item.get("selling_price", 0),
                    item.get("quantity", 1),
                ),
            )

    row = fetch_one("SELECT * FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    return jsonify(
        {
            "success": True,
            "message": "Enquiry updated",
            "enquiry": _enquiry_row_to_api(enquiry_id, dict(row)),
        }
    ), 200


@app.route("/api/enquiries/<enquiry_id>", methods=["DELETE"])
def api_enquiries_delete(enquiry_id):
    user_email = session.get("user")
    if not user_email:
        return jsonify({"success": False, "message": "Session expired. Please login first."}), 401

    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify({"success": False, "message": "Only Super Admin can delete enquiries."}), 403

    exists = fetch_one("SELECT 1 FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    if not exists:
        return jsonify({"success": False, "message": "Enquiry not found"}), 404

    execute_query("DELETE FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    return jsonify({"success": True, "message": "Enquiry deleted", "deleted_id": enquiry_id}), 200


@app.route("/api/enquiry-items/<enquiry_id>")
def get_enquiry_items(enquiry_id):
    items = fetch_all(
        """
        SELECT product_id, product_name, description,
               unit_price, selling_price, quantity, total
        FROM enquiry_product
        WHERE enquiry_id = %s
        """,
        (enquiry_id,),
    )
    result = {}
    for item in items:
        result[item["product_id"]] = {
            "item_code": item["product_id"],
            "item_name": item["product_name"],
            "description": item["description"],
            "unit_price": str(item["unit_price"]),
            "selling_price": str(item["selling_price"]),
            "quantity": item["quantity"],
            "total": str(item["total"]),
        }
    return jsonify(success=True, data=result)


@app.route("/update-enquiry-items/<enquiry_id>", methods=["POST"])
def update_enquiry_items(enquiry_id):
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can edit enquiry items."), 403

    req_data = request.get_json()
    if not req_data or "items" not in req_data:
        return jsonify(success=False, message="Invalid request"), 400

    for product_id, item in req_data["items"].items():
        execute_query(
            "DELETE FROM enquiry_product WHERE enquiry_id = %s AND product_id = %s",
            (enquiry_id, product_id),
        )
        execute_query(
            """
            INSERT INTO enquiry_product
                (id, enquiry_id, product_id, product_name, description, unit_price, selling_price, quantity)
            VALUES (
                (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry_product ep),
                %s, %s, %s, %s, %s, %s, %s
            )
            """,
            (
                enquiry_id,
                product_id,
                item.get("item_name", ""),
                item.get("description", ""),
                item.get("unit_price", 0),
                item.get("selling_price", 0),
                item.get("quantity", 1),
            ),
        )

    return jsonify(success=True, message="Enquiry items updated successfully")


    


@app.route("/new-enquiry")
def new_enquiry():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return redirect(url_for("enquiry_list") + "?message=create_denied")

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "new-enquiry.html",
        title="New-Enquiry - Stackly",
        page="new_enquiry",
        section="masters",
        user_email=user_email,
        user_name=user_name,
    )



@app.route("/generate-enquiry-id")
def generate_id():
    return jsonify(enquiry_id=generate_enquiry_id())


@app.route("/save-enquiry", methods=["POST"])
def save_enquiry():
    role = _get_current_user_role()
    if role not in ("admin", "superadmin"):
        return jsonify(success=False, message="Only Admin or Super Admin can create or update enquiries."), 403

    payload = request.get_json(silent=True)
    if not payload:
        payload = dict(request.form) if request.form else {}
    if not payload:
        return jsonify(success=False, message="Invalid request"), 400

    enquiry_id = payload.get("enquiry_id")
    if not enquiry_id:
        return jsonify(success=False, message="Enquiry ID required"), 400

    details = payload.get("enquiry_details") or {}
    items = payload.get("items") or {}

    existing = fetch_one("SELECT 1 FROM enquiry WHERE enquiry_id = %s", (enquiry_id,))
    if existing:
        execute_query(
            """
            UPDATE enquiry SET
                phone_number = %s,
                first_name = %s,
                last_name = %s,
                email = %s,
                street = %s,
                unit = %s,
                city = %s,
                state = %s,
                zip = %s,
                country = %s,
                enquiry_type = %s,
                enquiry_description = %s,
                enquiry_channel = %s,
                source = %s,
                heard_about = %s,
                urgency = %s,
                status = %s,
                priority = %s,
                updated_at = NOW()
            WHERE enquiry_id = %s
            """,
            (
                details.get("phone_number"),
                details.get("first_name"),
                details.get("last_name"),
                details.get("email"),
                details.get("street"),
                details.get("unit"),
                details.get("city"),
                details.get("state"),
                details.get("zip"),
                details.get("country"),
                details.get("enquiry_type"),
                details.get("enquiry_description"),
                details.get("enquiry_channel"),
                details.get("source"),
                details.get("heard_about"),
                details.get("urgency"),
                details.get("status", "New"),
                details.get("priority"),
                enquiry_id,
            ),
        )
    else:
        execute_query(
            """
            INSERT INTO enquiry (
                id, enquiry_id, phone_number, first_name, last_name, email,
                street, unit, city, state, zip, country,
                enquiry_type, enquiry_description, enquiry_channel,
                source, heard_about, urgency, status, priority
            ) VALUES (
                (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry ep),
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
            """,
            (
                enquiry_id,
                details.get("phone_number"),
                details.get("first_name"),
                details.get("last_name"),
                details.get("email"),
                details.get("street"),
                details.get("unit"),
                details.get("city"),
                details.get("state"),
                details.get("zip"),
                details.get("country"),
                details.get("enquiry_type"),
                details.get("enquiry_description"),
                details.get("enquiry_channel"),
                details.get("source"),
                details.get("heard_about"),
                details.get("urgency"),
                details.get("status", "New"),
                details.get("priority"),
            ),
        )

    execute_query("DELETE FROM enquiry_product WHERE enquiry_id = %s", (enquiry_id,))
    for product_id, item in items.items():
        execute_query(
            """
            INSERT INTO enquiry_product
                (id, enquiry_id, product_id, product_name, description, unit_price, selling_price, quantity)
            VALUES (
                (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry_product ep),
                %s, %s, %s, %s, %s, %s, %s
            )
            """,
            (
                enquiry_id,
                product_id,
                item.get("item_name", ""),
                item.get("description", ""),
                item.get("unit_price", 0),
                item.get("selling_price", 0),
                item.get("quantity", 1),
            ),
        )

    return jsonify(success=True)


@app.route("/get-enquiry-add-items/<enquiry_id>")
def get_enquiry_add_items_(enquiry_id):
    try:
        items = fetch_all(
            """
            SELECT product_id, product_name, description,
                   unit_price, selling_price, quantity, total
            FROM enquiry_product
            WHERE enquiry_id = %s
            """,
            (enquiry_id,),
        )
        result = {}
        for item in items:
            result[item["product_id"]] = {
                "item_code": item["product_id"],
                "item_name": item["product_name"],
                "description": item["description"],
                "unit_price": str(item["unit_price"]),
                "selling_price": str(item["selling_price"]),
                "quantity": item["quantity"],
                "total": str(item["total"]),
            }
        return jsonify({"success": True, "items": result})
    except Exception as e:
        print(f"Error in get_enquiry_add_items_: {e}")
        return jsonify({"success": False, "items": {}})


@app.route("/add-item", methods=["POST"])
def add_item():
    data = request.get_json()
    if not data or "enquiry_id" not in data or "item" not in data:
        return jsonify(error="Invalid request"), 400
    visible_id = data["enquiry_id"]
    item = data["item"]

    exists = fetch_one("SELECT 1 FROM enquiry WHERE enquiry_id = %s", (visible_id,))
    if not exists:
        return jsonify(error="Enquiry not found"), 400

    execute_query(
        """
        INSERT INTO enquiry_product (id, enquiry_id, product_id, product_name, description, unit_price, selling_price, quantity)
        VALUES (
            (SELECT COALESCE(MAX(ep.id), 0) + 1 FROM enquiry_product ep),
            %s, %s, %s, %s, %s, %s, %s
        )
        """,
        (
            visible_id,
            item["item_code"],
            item.get("item_name", ""),
            item.get("description", ""),
            item.get("unit_price", 0),
            item.get("selling_price", 0),
            item.get("quantity", 1),
        ),
    )
    return jsonify(status="item added")


@app.route("/check-email-enquiry")
def check_email_enquiry():
    try:
        email = request.args.get("email", "").lower()
        row = fetch_one(
            """
            SELECT enquiry_id, first_name, last_name, phone_number, email
            FROM enquiry WHERE LOWER(email) = %s
            """,
            (email,),
        )
        if row:
            return jsonify(
                {
                    "exists": True,
                    "enquiry_id": row["enquiry_id"],
                    "customer": {
                        "first_name": row["first_name"] or "",
                        "last_name": row["last_name"] or "",
                        "phone": row["phone_number"] or "",
                        "email": row["email"] or "",
                    },
                }
            )
        return jsonify({"exists": False})
    except Exception as e:
        print("❌ CHECK EMAIL ERROR:", e)
        return jsonify({"exists": False, "error": str(e)}), 500


@app.route("/get-product/<product_id>")
def get_product(product_id):
    product = fetch_one(
        """
        SELECT product_id, product_name, description, unit_price
        FROM products
        WHERE product_id = %s
        """,
        (product_id,),
    )
    if product:
        product = dict(product)
        product["unit_price"] = float(product["unit_price"])
        return jsonify({"success": True, "product": product})
    return jsonify({"success": False, "message": "Product not found"}), 404


@app.route("/delete-enquiry-item/<enquiry_id>/<item_code>", methods=["DELETE"])
def delete_enquiry_item(enquiry_id, item_code):
    role = _get_current_user_role()
    if role != "superadmin":
        return jsonify(success=False, message="Only Super Admin can delete enquiry items."), 403
    try:
        execute_query(
            """
            DELETE FROM enquiry_product
            WHERE enquiry_id = %s AND product_id = %s
            """,
            (enquiry_id, item_code),
        )
        return jsonify(success=True, message="Item permanently deleted")
    except Exception as e:
        print(f"Error deleting item: {e}")
        return jsonify(success=False, message=f"Error: {str(e)}"), 500


@app.route("/get-product-config")
def get_product_config():
    rows = fetch_all("SELECT product_id FROM products ORDER BY product_id")
    if not rows:
        return jsonify(success=False, message="No products found"), 404
    product_ids = [row["product_id"] for row in rows]
    max_id_length = max(len(pid) for pid in product_ids) if product_ids else 4
    return jsonify(
        success=True,
        max_id_length=max_id_length,
        product_ids=product_ids,
    )

# =========================================
# ✅ Helper function for quotation (QUOTATION)
# =========================================
# In-memory cache to avoid repeated file reads (speeds up list + get-quotation + filters)
_quotation_cache = None
_quotation_cache_time = 0.0
QUOTATION_CACHE_TTL = 3  # seconds

# ==================== DATABASE HELPER FUNCTIONS ====================
def load_quotations_from_db(filters=None, page=1, per_page=7):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        base_where = "WHERE 1=1"
        params = []
        if filters:
            q = filters.get('q')
            if q:
                base_where += " AND (LOWER(quotation_id) LIKE %s OR LOWER(customer_name) LIKE %s)"
                like = f"%{q}%"
                params.extend([like, like])
            status = filters.get('status')
            if status:
                base_where += " AND LOWER(status) = %s"
                params.append(status)
            qtype = filters.get('type')
            if qtype:
                base_where += " AND LOWER(quotation_type) = %s"
                params.append(qtype)
            sales_rep = filters.get('sales_rep')
            if sales_rep:
                base_where += " AND LOWER(sales_rep) = %s"
                params.append(sales_rep)

        # Count
        cur.execute(f"SELECT COUNT(*) as total FROM quotations {base_where}", params)
        total = cur.fetchone()['total']

        # Paginated data – no customer_id, no grand_total
        offset = (page - 1) * per_page
        cur.execute(f"""
            SELECT 
                quotation_id, 
                quotation_type, 
                customer_name,
                sales_rep,
                quotation_date, 
                status,
                grand_total
            FROM quotations
            {base_where}
            ORDER BY quotation_date DESC
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        items = cur.fetchall()
        for item in items:
            if item['quotation_date']:
                item['quotation_date'] = item['quotation_date'].isoformat()

        # Distinct sales reps
        cur.execute("SELECT DISTINCT sales_rep FROM quotations WHERE sales_rep IS NOT NULL AND sales_rep != ''")
        reps = [r['sales_rep'] for r in cur.fetchall()]

        return items, total, reps
    finally:
        cur.close()
        conn.close()

def get_full_quotation_from_db(quotation_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        # Main quotation (no customer_id, no grand_total)
        cur.execute("""
            SELECT 
                q.quotation_id, q.quotation_type, q.quotation_date, q.expiry_date,
                q.customer_name, q.customer_po, q.sales_rep, q.currency,  q.payment_terms AS "paymentTerms",    q.expected_date, q.status, q.created_at, q.last_updated,
                COALESCE(t.subtotal, 0) as subtotal,
                COALESCE(t.global_discount_percent, 0) as global_discount_percent,
                COALESCE(t.tax_summary, 0) as tax_summary,
                COALESCE(t.shipping_charge, 0) as shipping_charge,
                COALESCE(t.rounding_adjustment, 0) as rounding_adjustment,
                COALESCE(t.grand_total, 0) as grand_total
            FROM quotations q
            LEFT JOIN quotation_totals t ON q.quotation_id = t.quotation_id
            WHERE q.quotation_id = %s
        """, (quotation_id,))
        quotation = cur.fetchone()
        if not quotation:
            return None

        # Convert dates to strings
        for key in ['quotation_date', 'expiry_date', 'expected_date', 'created_at', 'last_updated']:
            if quotation.get(key):
                quotation[key] = quotation[key].isoformat()

        # Items (no sl_no)
        cur.execute("""
            SELECT 
                item_id, product_name, product_id, quantity, uom,
                unit_price, tax_percent as tax, discount_percent as discount, total
            FROM quotation_items
            WHERE quotation_id = %s
            ORDER BY item_id
        """, (quotation_id,))
        items = cur.fetchall()
        for item in items:
            for num in ['quantity', 'unit_price', 'tax', 'discount', 'total']:
                if item.get(num) is not None:
                    item[num] = float(item[num])
        quotation['items'] = items

        # Taxes – if you don't have quotation_tax table, return empty list
        quotation['taxes'] = []

        # Comments
        cur.execute("""
            SELECT comment_id, comment, created_by as user, created_at as time
            FROM quotation_comments
            WHERE quotation_id = %s
            ORDER BY created_at DESC
        """, (quotation_id,))
        comments = cur.fetchall()
        for c in comments:
            if c.get('time'):
                c['time'] = c['time'].strftime('%Y-%m-%d %H:%M:%S')
        quotation['comments'] = comments

        # Attachments
        cur.execute("""
            SELECT attachment_id as id, file_name as original_filename,
                   file_size as size, uploaded_at as upload_date
            FROM quotation_attachments
            WHERE quotation_id = %s
        """, (quotation_id,))
        attachments = cur.fetchall()
        for a in attachments:
            if a.get('upload_date'):
                a['upload_date'] = a['upload_date'].strftime('%Y-%m-%d %H:%M:%S')
        quotation['attachments'] = attachments

        # Build a 'totals' object to match frontend expectation
        quotation['totals'] = {
            'subtotal': float(quotation.get('subtotal', 0)),
            'global_discount_percent': float(quotation.get('global_discount_percent', 0)),
            'tax_summary': float(quotation.get('tax_summary', 0)),
            'shipping_charge': float(quotation.get('shipping_charge', 0)),
            'rounding_adjustment': float(quotation.get('rounding_adjustment', 0)),
            'grand_total': float(quotation.get('grand_total', 0))
        }

        return quotation
    finally:
        cur.close()
        conn.close()

def save_quotation_to_db(data):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        quotation_id = data['quotation_id']
        now = datetime.now()

        # Check existence
        cur.execute("SELECT 1 FROM quotations WHERE quotation_id = %s", (quotation_id,))
        exists = cur.fetchone()

        # Main fields – using customer_name, now including grand_total
        main_fields = {
            'quotation_id': quotation_id,
            'quotation_type': data.get('quotation_type'),
            'quotation_date': data.get('quotation_date'),
            'expiry_date': data.get('expiry_date'),
            'customer_name': data.get('customer_name'),
            'customer_po': data.get('customer_po'),
            'sales_rep': data.get('sales_rep'),
            'currency': data.get('currency'),
            'payment_terms': data.get('paymentTerms'),
            'expected_date': data.get('expected_date'),
            'status': data.get('status', 'draft'),
            'last_updated': now,
            'grand_total': data.get('totals', {}).get('grand_total', 0)   # ✅ added comma
        }

        if exists:
            cur.execute("""
                UPDATE quotations SET
                    quotation_type = %(quotation_type)s,
                    quotation_date = %(quotation_date)s,
                    expiry_date = %(expiry_date)s,
                    customer_name = %(customer_name)s,
                    customer_po = %(customer_po)s,
                    sales_rep = %(sales_rep)s,
                    currency = %(currency)s,
                    payment_terms = %(payment_terms)s,
                    expected_date = %(expected_date)s,
                    status = %(status)s,
                    last_updated = %(last_updated)s,
                    grand_total = %(grand_total)s
                WHERE quotation_id = %(quotation_id)s
            """, main_fields)
        else:
            cur.execute("""
                INSERT INTO quotations (
                    quotation_id, quotation_type, quotation_date, expiry_date,
                    customer_name, customer_po, sales_rep,
                    currency, payment_terms, expected_date, status, last_updated, grand_total
                ) VALUES (
                    %(quotation_id)s, %(quotation_type)s, %(quotation_date)s, %(expiry_date)s,
                    %(customer_name)s, %(customer_po)s, %(sales_rep)s,
                    %(currency)s, %(payment_terms)s, %(expected_date)s, %(status)s, %(last_updated)s, %(grand_total)s
                )
            """, main_fields)

        # Replace items – no sl_no
        cur.execute("DELETE FROM quotation_items WHERE quotation_id = %s", (quotation_id,))
        for item in data.get('items', []):
            cur.execute("""
                INSERT INTO quotation_items (
                    quotation_id, product_name, product_id, quantity,
                    uom, unit_price, tax_percent, discount_percent, total
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                quotation_id,
                item.get('product_name'), item.get('product_id'),
                item.get('quantity'), item.get('uom'),
                item.get('unit_price'), item.get('tax', 0),
                item.get('discount', 0), item.get('total', 0)
            ))

        # Insert or update totals (quotation_totals table)
        totals = data.get('totals', {})
        cur.execute("""
            INSERT INTO quotation_totals (
                quotation_id, subtotal, global_discount_percent, tax_summary,
                shipping_charge, rounding_adjustment, grand_total, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (quotation_id) DO UPDATE SET
                subtotal = EXCLUDED.subtotal,
                global_discount_percent = EXCLUDED.global_discount_percent,
                tax_summary = EXCLUDED.tax_summary,
                shipping_charge = EXCLUDED.shipping_charge,
                rounding_adjustment = EXCLUDED.rounding_adjustment,
                grand_total = EXCLUDED.grand_total,
                updated_at = EXCLUDED.updated_at
        """, (
            quotation_id,
            totals.get('subtotal', 0),
            totals.get('global_discount_percent', 0),
            totals.get('tax_summary', 0),
            totals.get('shipping_charge', 0),
            totals.get('rounding_adjustment', 0),
            totals.get('grand_total', 0),
            now
        ))

        # Add comment if provided
        if data.get('comment_text'):
            cur.execute("""
                INSERT INTO quotation_comments (quotation_id, comment, created_by)
                VALUES (%s, %s, %s)
            """, (quotation_id, data['comment_text'], data.get('submitted_by', 'System')))

        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()

def check_duplicate_customer_po(quotation_id, customer_po):
    """Return True if duplicate exists (case-insensitive)"""
    if not customer_po:
        return False
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT 1 FROM quotations
            WHERE LOWER(customer_po) = %s AND quotation_id != %s
        """, (customer_po.lower(), quotation_id))
        return cur.fetchone() is not None
    finally:
        cur.close()
        conn.close()

def generate_quotation_id_db():
    """Generate QA-XXXX from existing DB records"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT quotation_id FROM quotations WHERE quotation_id ~ '^QA-[0-9]+$'")
        rows = cur.fetchall()
        max_num = 0
        for row in rows:
            try:
                num = int(row[0].split('-')[1])
                if num > max_num:
                    max_num = num
            except:
                continue
        new_num = max_num + 1
        return f"QA-{new_num:03d}"
    finally:
        cur.close()
        conn.close()

def update_quotation_status_in_db(quotation_id, new_status, status_date, rejection_reason, status_history_entry):
    """Update status and optionally add status history (as JSON field)"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Get current quotation to update status_history JSON (stored as text in DB)
        cur.execute("SELECT status_history FROM quotations WHERE quotation_id = %s", (quotation_id,))
        row = cur.fetchone()
        if not row:
            return False
        status_history = row[0] if row[0] else []
        if isinstance(status_history, str):
            status_history = json.loads(status_history)
        status_history.append(status_history_entry)
        cur.execute("""
            UPDATE quotations
            SET status = %s, status_date = %s, rejection_reason = %s,
                last_updated = %s, status_history = %s
            WHERE quotation_id = %s
        """, (new_status, status_date, rejection_reason, datetime.now(), json.dumps(status_history), quotation_id))
        conn.commit()
        return True
    finally:
        cur.close()
        conn.close()

def check_and_update_expired_quotations_db():
    """Update status to 'expired' for quotations with expiry_date < today and status in ('send','sent','submitted')"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        today = datetime.now().date()
        cur.execute("""
            SELECT quotation_id, expiry_date, status_history FROM quotations
            WHERE status IN ('send', 'sent', 'submitted')
              AND expiry_date < %s
        """, (today,))
        expired = cur.fetchall()
        for row in expired:
            qid = row[0]
            expiry_date = row[1]
            status_history = row[2] if row[2] else []
            if isinstance(status_history, str):
                status_history = json.loads(status_history)
            status_history.append({
                'status': 'expired',
                'date': today.isoformat(),
                'time': datetime.now().strftime('%H:%M:%S'),
                'reason': 'Auto-expired',
                'notes': f'Expired on {today} (valid until {expiry_date})'
            })
            cur.execute("""
                UPDATE quotations SET status = 'expired', status_history = %s, last_updated = %s
                WHERE quotation_id = %s
            """, (json.dumps(status_history), datetime.now(), qid))
        conn.commit()
        return [row[0] for row in expired]
    finally:
        cur.close()
        conn.close()

# ==================== ROUTES (unchanged names, but using DB) ====================

@app.route("/quotation")
def quotation():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))
    users = load_users()  # you must have load_users() defined elsewhere; keep as is
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break
    return render_template(
        "quotation.html",
        title="Quotation - Stackly",
        page="quotation",
        section="crm",
        user_email=user_email,
        user_name=user_name,
    )

@app.route("/api/quotations", methods=["GET"])
def api_quotations():
    if "user" not in session:
        return jsonify(success=False, message="Session expired"), 401

    q = (request.args.get("q") or "").strip().lower()
    status = (request.args.get("status") or "").strip().lower()
    qtype = (request.args.get("type") or "").strip().lower()
    sales_rep = (request.args.get("sales_rep") or "").strip().lower()
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 7))

    filters = {}
    if q:
        filters['q'] = q
    if status:
        filters['status'] = status
    if qtype:
        filters['type'] = qtype
    if sales_rep:
        filters['sales_rep'] = sales_rep

    items, total, reps = load_quotations_from_db(filters, page, per_page)
    total_pages = max(1, math.ceil(total / per_page))

    return jsonify(
        success=True,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        items=items,
        sales_reps=reps
    )

@app.route("/api/quotations", methods=["POST"])
def api_create_quotation():
    if "user" not in session:
        return jsonify(success=False, message="Session expired"), 401

    data = request.get_json(force=True) or {}
    new_id = generate_quotation_id_db()

    new_item = {
        "quotation_id": new_id,
        "quotation_type": (data.get("quotation_type") or "service").lower(),
        "customer_name": data.get("customer_name") or "",
        "sales_rep": data.get("sales_rep") or "",
        "quotation_date": data.get("quotation_date") or datetime.now().strftime("%Y-%m-%d"),
        "status": "draft",
        "grand_total": float(data.get("grand_total") or 0),
    }

    # Insert directly
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO quotations (quotation_id, quotation_type, customer_name, sales_rep, quotation_date, status, grand_total)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (new_id, new_item['quotation_type'], new_item['customer_name'], new_item['sales_rep'],
              new_item['quotation_date'], new_item['status'], new_item['grand_total']))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify(success=False, error=str(e)), 500
    finally:
        cur.close()
        conn.close()

    return jsonify(success=True, item=new_item)

@app.route("/add-new-quotation")
def add_new_quotation():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))
    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break
    return render_template(
        "add-new-quotation.html",
        title="Add-New-Quotation - Stackly",
        page="quotation",
        section="crm",
        user_email=user_email,
        user_name=user_name,
    )

@app.route("/get-customers-quotation")
def get_customers_quotation():
    try:
        # Use the correct connection function
        # If your app uses get_db_connection, keep it; otherwise change to get_connection
        conn = get_db_connection()  # or get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cur.execute("""
                SELECT 
                    customer_id,
                    name,
                    first_name,
                    last_name,
                    company,
                    email,
                    phone,
                    sales_rep,
                    payment_terms AS "paymentTerms",
                    credit_limit,
                    billing_address,
                    shipping_address,
                    customer_status AS status    -- ✅ fixed column name
                FROM customers
                ORDER BY name
            """)
            customers = cur.fetchall()
            return jsonify(customers)
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        print(f"❌ Error fetching customers: {e}")
        import traceback
        traceback.print_exc()
        # Return empty array so frontend doesn't break
        return jsonify([])


@app.route('/get-products')
def get_products():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cur.execute("""
                SELECT 
                    product_id,
                    product_name,
                    unit_price,
                    uom_name AS uom,          -- alias to match frontend
                    tax_percent AS tax,       -- alias to match frontend
                    discount,
                    quantity,
                    status
                FROM products
                WHERE status = 'Active'
                ORDER BY product_name
            """)
            products = cur.fetchall()
            for p in products:
                if p.get('unit_price') is not None:
                    p['unit_price'] = float(p['unit_price'])
                if p.get('tax') is not None:
                    p['tax'] = float(p['tax'])
                if p.get('discount') is not None:
                    p['discount'] = float(p['discount'])
                if p.get('quantity') is not None:
                    p['quantity'] = float(p['quantity'])
            return jsonify(products)
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        print(f"Error fetching products: {e}")
        return jsonify([]), 500



@app.route('/generate-quotation-id')
def generate_quotation_id_route():
    try:
        new_id = generate_quotation_id_db()
        return jsonify({'success': True, 'quotation_id': new_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/save-quotation', methods=['POST'])
def save_quotation():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        status = data.get('status', 'draft')

        # Duplicate PO check
        customer_po = (data.get('customer_po') or '').strip()
        if check_duplicate_customer_po(quotation_id, customer_po):
            return jsonify({
                'success': False,
                'error': 'Customer PO Reference already exists. Please use a unique value.',
                'duplicate_field': 'customer_po'
            }), 400

        # Add timestamps
        data['last_updated'] = datetime.now().isoformat()
        data['created_at'] = data.get('created_at', datetime.now().isoformat())

        # Status history (keep as list in data, will be stored as JSON in DB)
        if 'status_history' not in data:
            data['status_history'] = []
        status_entry = {
            'status': status,
            'date': data.get('status_date', datetime.now().isoformat()),
            'user': data.get('submitted_by', 'System'),
            'notes': f'Quotation {status}'
        }
        if status == 'rejected' and data.get('rejection_reason'):
            status_entry['notes'] = f'Quotation rejected: {data["rejection_reason"]}'
        data['status_history'].append(status_entry)

        save_quotation_to_db(data)
        return jsonify({
            'success': True,
            'message': f'Quotation {quotation_id} saved with status: {status}',
            'quotation_id': quotation_id,
            'status': status
        })
    except Exception as e:
        print(f"Error saving quotation: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/check-customer-po', methods=['GET'])
def check_customer_po():
    try:
        value = request.args.get('value', '').strip()
        exclude_id = request.args.get('exclude_quotation_id', '').strip()
        if not value:
            return jsonify({'success': True, 'duplicate': False})
        duplicate = check_duplicate_customer_po(exclude_id, value)
        return jsonify({'success': True, 'duplicate': duplicate})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-quotation/<quotation_id>')
def get_quotation(quotation_id):
    try:
        quotation = get_full_quotation_from_db(quotation_id)
        if not quotation:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        return jsonify({'success': True, 'quotation': quotation})
    except Exception as e:
        print(f"❌ Error in get_quotation: {e}")   # <-- add this
        import traceback
        traceback.print_exc()                      # <-- add this
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-quotations/<status>')
def get_quotations_by_status(status):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            if status and status != 'all':
                cur.execute("SELECT * FROM quotations WHERE status = %s", (status,))
            else:
                cur.execute("SELECT * FROM quotations")
            quotations = cur.fetchall()
            return jsonify({'success': True, 'quotations': quotations, 'count': len(quotations)})
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/add-comment', methods=['POST'])
def add_comment():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        comment = data.get('comment')
        user = data.get('user', 'Admin')
        
        if not quotation_id or not comment:
            return jsonify({'success': False, 'error': 'Missing quotation_id or comment'}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            # Check if quotation exists
            cur.execute("SELECT 1 FROM quotations WHERE quotation_id = %s", (quotation_id,))
            exists = cur.fetchone()
            
            if not exists:
                # Auto-create a minimal draft quotation
                now = datetime.now().date()
                cur.execute("""
                    INSERT INTO quotations (
                        quotation_id, quotation_type, quotation_date, expiry_date,
                        customer_name, status, created_at, last_updated
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    quotation_id,
                    'Standard',                    # default type
                    now,                           # quotation_date = today
                    (now + timedelta(days=15)),    # expiry_date = +15 days
                    'Auto-created',                # placeholder customer_name
                    'draft',
                    datetime.now(),
                    datetime.now()
                ))
                conn.commit()
                print(f"✅ Auto-created draft quotation {quotation_id} for comment")
            
            # Now insert the comment
            cur.execute("""
                INSERT INTO quotation_comments (quotation_id, comment, created_by)
                VALUES (%s, %s, %s)
            """, (quotation_id, comment, user))
            conn.commit()
            
            return jsonify({'success': True, 'message': 'Comment added successfully'})
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        print(f"Error adding comment: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-comments/<quotation_id>')
def get_comments(quotation_id):
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 5))
        offset = (page - 1) * per_page
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cur.execute("SELECT COUNT(*) as total FROM quotation_comments WHERE quotation_id = %s", (quotation_id,))
            total = cur.fetchone()['total']
            cur.execute("""
                SELECT comment, created_by as user, created_at as time
                FROM quotation_comments
                WHERE quotation_id = %s
                ORDER BY created_at DESC
                LIMIT %s OFFSET %s
            """, (quotation_id, per_page, offset))
            comments = cur.fetchall()
            for c in comments:
                c['time'] = c['time'].strftime('%Y-%m-%d %H:%M:%S') if c['time'] else ''
            return jsonify({
                'comments': comments,
                'total': total,
                'page': page,
                'per_page': per_page,
                'has_more': (offset + per_page) < total
            })
        finally:
            cur.close()
            conn.close()
    except Exception:
        return jsonify({'comments': [], 'total': 0, 'has_more': False})


ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png'}
MAX_FILE_SIZE_MB = 10
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload-attachment', methods=['POST'])
def upload_attachment():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400

        file = request.files['file']
        quotation_id = request.form.get('quotation_id')

        if not quotation_id:
            return jsonify({'success': False, 'error': 'quotation_id is required'}), 400

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        # Validate file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        if file_size > MAX_FILE_SIZE_BYTES:
            return jsonify({'success': False, 'error': f'File exceeds {MAX_FILE_SIZE_MB} MB'}), 400

        # Validate extension
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'File type not allowed'}), 400

        # ‼️ STEP 1: Make sure the quotation exists (auto‑create draft if missing)
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute("SELECT 1 FROM quotations WHERE quotation_id = %s", (quotation_id,))
            exists = cur.fetchone()
            if not exists:
                now = datetime.now().date()
                cur.execute("""
                    INSERT INTO quotations (
                        quotation_id, quotation_type, quotation_date, expiry_date,
                        customer_name, status, created_at, last_updated
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    quotation_id,
                    'Standard',                     # default type
                    now,                            # quotation_date = today
                    (now + timedelta(days=15)),     # expiry_date = +15 days
                    'Auto-created',                 # placeholder customer_name
                    'draft',
                    datetime.now(),
                    datetime.now()
                ))
                conn.commit()
                print(f"✅ Auto-created draft quotation {quotation_id} for attachment upload")
        finally:
            # Don't close the connection yet – we reuse it for the attachment insert
            pass

        rel_path = _upload_relative_path(quotation_id, file.filename)
        stored_path, file_size = _persist_module_upload(
            object_storage.MODULE_QUOTATION_ATTACHMENTS,
            QUOTATION_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )

        cur.execute("""
            INSERT INTO quotation_attachments (quotation_id, file_name, file_path, file_size)
            VALUES (%s, %s, %s, %s)
            RETURNING attachment_id
        """, (quotation_id, file.filename, stored_path, file_size))
        row = cur.fetchone()
        attachment_id = row[0] if row else None
        _purge_prior_same_name_files(
            cur,
            "quotation_attachments",
            "quotation_id",
            quotation_id,
            "file_name",
            file.filename,
            "attachment_id",
            attachment_id,
            "file_path",
            QUOTATION_ATTACHMENTS_FOLDER,
        )
        conn.commit()

        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'attachment': {
                'id': attachment_id,
                'original_filename': file.filename,
                'size': file_size,
                'upload_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        })

    except Exception as e:
        print(f"Upload error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
@app.route('/get-attachments/<quotation_id>')
def get_attachments(quotation_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT attachment_id as id, file_name as original_filename,
                   file_size as size, uploaded_at as upload_date
            FROM quotation_attachments
            WHERE quotation_id = %s
            ORDER BY uploaded_at DESC
        """, (quotation_id,))
        attachments = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'attachments': attachments})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/view-attachment/<attachment_id>')
def view_attachment(attachment_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT file_path, file_name FROM quotation_attachments WHERE attachment_id = %s", (attachment_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return jsonify({'error': 'Attachment not found'}), 404
        file_path = row[0]
        original_name = row[1]
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        full_path = _resolve_stored_file_path(file_path)
        if not full_path or not os.path.isfile(full_path):
            return jsonify({'error': 'File not found'}), 404
        return send_file(full_path, as_attachment=False, download_name=original_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download-attachment/<attachment_id>')
def download_attachment(attachment_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT file_path, file_name FROM quotation_attachments WHERE attachment_id = %s", (attachment_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if not row:
            return jsonify({'error': 'Attachment not found'}), 404
        file_path = row[0]
        original_name = row[1]
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        full_path = _resolve_stored_file_path(file_path)
        if not full_path or not os.path.isfile(full_path):
            return jsonify({'error': 'File not found'}), 404
        return send_file(full_path, as_attachment=True, download_name=original_name)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
@app.route('/delete-attachment/<attachment_id>', methods=['DELETE'])
def delete_attachment(attachment_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT file_path FROM quotation_attachments WHERE attachment_id = %s", (attachment_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Attachment not found'}), 404
        file_path = row[0]
        _remove_stored_upload(file_path, QUOTATION_ATTACHMENTS_FOLDER)
        cur.execute("DELETE FROM quotation_attachments WHERE attachment_id = %s", (attachment_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'success': True, 'message': 'Attachment deleted'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/check-quotation/<quotation_id>')
def check_quotation(quotation_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute("SELECT 1 FROM quotations WHERE quotation_id = %s", (quotation_id,))
            exists = cur.fetchone() is not None
            return jsonify({'success': True, 'exists': exists, 'quotation_id': quotation_id})
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/debug-quotations')
def debug_quotations():
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        try:
            cur.execute("SELECT quotation_id FROM quotations")
            ids = [r['quotation_id'] for r in cur.fetchall()]
            return jsonify({
                'success': True,
                'count': len(ids),
                'quotation_ids': ids,
                'file_path': 'database (PostgreSQL)',
                'file_exists': True
            })
        finally:
            cur.close()
            conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def delete_quotation_from_db(quotation_id):
    """Delete quotation (cascade deletes items, taxes, comments, attachments, totals)"""
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM quotations WHERE quotation_id = %s", (quotation_id,))
        conn.commit()
        return True
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        cur.close()
        conn.close()

@app.route('/delete-quotation/<quotation_id>', methods=['DELETE'])
def delete_quotation(quotation_id):
    try:
        # Check status
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute("SELECT status FROM quotations WHERE quotation_id = %s", (quotation_id,))
            row = cur.fetchone()
            if not row:
                return jsonify({'success': False, 'error': 'Quotation not found'}), 404
            if row[0] != 'draft':
                return jsonify({'success': False, 'error': 'Only draft quotations can be deleted'}), 403
        finally:
            cur.close()
            conn.close()
        delete_quotation_from_db(quotation_id)
        # Also delete physical attachment files (already handled by ON DELETE CASCADE in DB, but files remain)
        # Optionally clean up files – skip for brevity
        return jsonify({'success': True, 'message': f'Quotation {quotation_id} deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/update-quotation-status', methods=['POST'])
def update_quotation_status():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        new_status = data.get('status')
        status_date = data.get('status_date', datetime.now().isoformat())
        rejection_reason = data.get('rejection_reason', '')
        status_history = data.get('status_history', {})
        if not quotation_id or not new_status:
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        history_entry = {
            'status': new_status,
            'date': status_date,
            'user': status_history.get('user', 'System'),
            'notes': status_history.get('notes', f'Quotation {new_status}')
        }
        success = update_quotation_status_in_db(quotation_id, new_status, status_date, rejection_reason, history_entry)
        if not success:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        return jsonify({'success': True, 'message': f'Quotation {quotation_id} updated to {new_status}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


    

# ===================================================
# OTP GENERATION
# ===================================================

def generate_otp():
    """Generate 6-digit OTP"""
    return ''.join(random.choices(string.digits, k=6))

# ===================================================
# OTP RATE LIMITING FUNCTIONS
# ===================================================

def check_otp_limits(email, quotation_id):
    """
    Check OTP attempt limits
    Returns: (allowed, reason, attempts_left)
    """
    try:
        now = datetime.now()
        key = f"{email}:{quotation_id}"
        
        # Check if user is blocked
        if key in otp_blocked:
            block_info = otp_blocked[key]
            unblock_time = datetime.fromisoformat(block_info['unblock_time'])
            
            if now < unblock_time:
                wait_minutes = int((unblock_time - now).total_seconds() / 60)
                return False, f"Too many attempts. Try again in {wait_minutes} minutes.", 0
            else:
                del otp_blocked[key]
        
        # Get recent failed attempts
        recent_attempts = [a for a in otp_attempts.get(key, []) 
                          if (now - a['timestamp']) < timedelta(minutes=30)]
        
        failed_attempts = [a for a in recent_attempts if not a['success']]
        
        if len(failed_attempts) >= RATE_LIMIT_CONFIG['max_otp_attempts']:
            unblock_time = now + timedelta(minutes=RATE_LIMIT_CONFIG['otp_cooldown_minutes'])
            otp_blocked[key] = {'unblock_time': unblock_time.isoformat()}
            return False, f"Too many failed attempts. Try again after {RATE_LIMIT_CONFIG['otp_cooldown_minutes']} minutes.", 0
        
        attempts_left = RATE_LIMIT_CONFIG['max_otp_attempts'] - len(failed_attempts)
        return True, "Allowed", attempts_left
        
    except Exception as e:
        print(f"Error in check_otp_limits: {e}")
        return True, "Allowed", 5

def record_otp_attempt(email, quotation_id, success):
    """Record an OTP attempt"""
    key = f"{email}:{quotation_id}"
    otp_attempts[key].append({
        'timestamp': datetime.now(),
        'success': success
    })
    
    # Clean up old records
    cutoff = datetime.now() - timedelta(hours=24)
    otp_attempts[key] = [a for a in otp_attempts[key] if a['timestamp'] > cutoff]

def get_otp_attempts_left(email, quotation_id):
    """Get remaining OTP attempts"""
    key = f"{email}:{quotation_id}"
    now = datetime.now()
    
    if key in otp_blocked:
        return 0
    
    recent_attempts = [a for a in otp_attempts.get(key, []) 
                      if (now - a['timestamp']) < timedelta(minutes=30)]
    failed_attempts = [a for a in recent_attempts if not a['success']]
    return RATE_LIMIT_CONFIG['max_otp_attempts'] - len(failed_attempts)

# ===================================================
# OTP RESEND LIMITING
# ===================================================

def check_resend_limits(email, quotation_id):
    """Check if resend is allowed (max 5 per 24 hours)"""
    now = datetime.now()
    key = f"{email}:{quotation_id}"
    
    resends = [a for a in otp_resend_attempts.get(key, []) 
               if (now - a['timestamp']) < timedelta(hours=24)]
    
    if len(resends) >= 5:
        return False, "Maximum resend attempts reached. Try again after 24 hours.", 0
    
    attempts_left = 5 - len(resends)
    return True, "Allowed", attempts_left

def record_resend_attempt(email, quotation_id):
    """Record a resend attempt"""
    key = f"{email}:{quotation_id}"
    otp_resend_attempts[key].append({'timestamp': datetime.now()})
    
    # Clean up
    cutoff = datetime.now() - timedelta(hours=24)
    otp_resend_attempts[key] = [a for a in otp_resend_attempts[key] if a['timestamp'] > cutoff]
    
    return get_resend_attempts_left(email, quotation_id)

def get_resend_attempts_left(email, quotation_id):
    """Get remaining resend attempts"""
    key = f"{email}:{quotation_id}"
    now = datetime.now()
    resends = [a for a in otp_resend_attempts.get(key, []) 
               if (now - a['timestamp']) < timedelta(hours=24)]
    return 5 - len(resends)

# ===================================================
# EMAIL RATE LIMITING
# ===================================================

def check_email_limits(quotation_id, customer_email, recipient_email):
    """
    Check if email can be sent
    Returns: (allowed, reason, requires_approval)
    """
    now = datetime.now()
    key = f"{quotation_id}:{customer_email}"
    
    attempts = email_attempts.get(key, [])
    
    # Per-quotation limit
    quotation_emails = [a for a in attempts if a['quotation_id'] == quotation_id]
    
    if len(quotation_emails) >= RATE_LIMIT_CONFIG['max_emails_per_quotation']:
        return False, f"Maximum {RATE_LIMIT_CONFIG['max_emails_per_quotation']} emails reached", False
    
    # Per-recipient limit
    recipient_emails = [a for a in quotation_emails if a['recipient'] == recipient_email]
    
    if len(recipient_emails) >= RATE_LIMIT_CONFIG['max_emails_per_recipient']:
        return False, f"Already sent {RATE_LIMIT_CONFIG['max_emails_per_recipient']} emails to this recipient", False
    
    # Daily limit
    daily_emails = [a for a in attempts if a['timestamp'].date() == now.date()]
    
    if len(daily_emails) >= RATE_LIMIT_CONFIG['max_daily_emails_per_customer']:
        return False, f"Daily limit of {RATE_LIMIT_CONFIG['max_daily_emails_per_customer']} emails reached", False
    
    # Throttle
    if attempts:
        last_email = attempts[-1]
        time_diff = now - last_email['timestamp']
        min_wait = timedelta(minutes=RATE_LIMIT_CONFIG['min_time_between_emails_minutes'])
        
        if time_diff < min_wait:
            wait_minutes = RATE_LIMIT_CONFIG['min_time_between_emails_minutes'] - (time_diff.seconds // 60)
            return False, f"Please wait {wait_minutes} minutes between emails", False
    
    # Check if approval required
    requires_approval = len(quotation_emails) >= RATE_LIMIT_CONFIG['requires_approval_after']
    
    return True, "Allowed", requires_approval

def record_email_sent(quotation_id, customer_email, recipient_email, approved=False):
    """Record that an email was sent"""
    key = f"{quotation_id}:{customer_email}"
    
    email_attempts[key].append({
        'quotation_id': quotation_id,
        'recipient': recipient_email,
        'timestamp': datetime.now(),
        'approved': approved
    })
    
    # Clean up old records
    cutoff = datetime.now() - timedelta(days=30)
    email_attempts[key] = [a for a in email_attempts[key] if a['timestamp'] > cutoff]

def get_email_count(quotation_id, customer_email):
    """Get number of emails sent for this quotation"""
    key = f"{quotation_id}:{customer_email}"
    return len(email_attempts.get(key, []))

# ===================================================
# SEND OTP EMAIL
# ===================================================

def send_otp_email(email, otp, quotation_id=None):
    """Send OTP via email"""
    try:
        print(f"📧 Sending OTP to {email}")
        
        msg = MIMEMultipart()
        msg['Subject'] = f"Your OTP for Quotation {quotation_id}" if quotation_id else "Your OTP for Quotation"
        msg['From'] = SENDER_EMAIL
        msg['To'] = email
        
        body = f"""
        Your OTP for verification is: {otp}
        
        This OTP is valid for {OTP_EXPIRY_MINUTES} minutes.
        
        Please enter this OTP to complete your quotation request.
        
        If you didn't request this, please ignore this email.
        """
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Try port 587 with TLS
        try:
            server = smtplib.SMTP(SMTP_SERVER, 587, timeout=30)
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)
            server.quit()
            print(f"✅ OTP sent successfully")
            return True
        except Exception as e:
            print(f"❌ Port 587 failed: {e}")
            
            # Try port 465 with SSL
            try:
                context = ssl.create_default_context()
                server = smtplib.SMTP_SSL(SMTP_SERVER, 465, context=context, timeout=30)
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
                server.quit()
                print(f"✅ OTP sent via SSL")
                return True
            except Exception as e2:
                print(f"❌ Both ports failed: {e2}")
                # For development, just log the OTP
                print(f"📧 [DEV MODE] OTP for {email}: {otp}")
                return True
            
    except Exception as e:
        print(f"❌ Error sending OTP: {e}")
        # For development, just log the OTP
        print(f"📧 [DEV MODE] OTP for {email}: {otp}")
        return True


# ===================================================
# SINGLE SOURCE OF TRUTH - ONE PDF GENERATOR FOR ALL
# ===================================================
def generate_quotation_pdf(quotation, quotation_id=None):
    """Generate PDF from quotation dict (already fetched from DB) – INR only."""
    try:
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18,
            leftMargin=18,
            topMargin=16,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()

        # ---------- Custom styles (matching DNR) ----------
        company_style = ParagraphStyle(
            name="Quot_CompanyName",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#8c1f1f"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )

        company_info_style = ParagraphStyle(
            name="Quot_CompanyInfo",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=9,
            leading=12,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=1,
        )

        page_title_style = ParagraphStyle(
            name="Quot_PageTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.green,
            alignment=TA_CENTER,
            spaceBefore=12,
            spaceAfter=12,
        )

        section_style = ParagraphStyle(
            name="Quot_Section",
            parent=styles["Heading3"],
            fontName="DejaVuSans-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#8c1f1f"),
            spaceAfter=6,
            spaceBefore=10,
        )

        label_style = ParagraphStyle(
            name="Quot_Label",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#6b1a1a"),
        )

        value_style = ParagraphStyle(
            name="Quot_Value",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8.5,
            leading=11,
            textColor=colors.black,
        )

        header_small_style = ParagraphStyle(
            name="Quot_HeaderSmall",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )

        terms_style = ParagraphStyle(
            name="Quot_Terms",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8,
            leading=11,
            textColor=colors.black,
            leftIndent=8,
        )

        # ---------- Helper functions ----------
        def safe_str(val, default="-"):
            if val is None:
                return default
            s = str(val).strip()
            return s if s else default

        def safe_float(val, default=0.0):
            try:
                if val in (None, ""):
                    return default
                return float(val)
            except Exception:
                return default

        elements = []

        currency_symbol = "₹"
        currency_code = "IND"

        totals = quotation.get('totals', {})
        subtotal = safe_float(totals.get('subtotal', 0))
        total_tax = safe_float(totals.get('tax_summary', 0))
        shipping = safe_float(totals.get('shipping_charge', 0))
        grand_total = safe_float(totals.get('grand_total', 0))
        global_discount_pct = safe_float(totals.get('global_discount_percent', 0))
        rounding = safe_float(totals.get('rounding_adjustment', 0))
        global_discount_amt = subtotal * (global_discount_pct / 100) if global_discount_pct else 0

        # Calculate item-level discount total
        total_discount = 0.0
        for item in quotation.get('items', []):
            qty = safe_float(item.get('quantity', 0))
            price = safe_float(item.get('unit_price', 0))
            disc_pct = safe_float(item.get('discount', 0))
            if disc_pct > 0:
                total_discount += qty * price * (disc_pct / 100)

        # ---------- Company header (same as DNR) ----------
        elements.append(Paragraph("STACKLY", company_style))
        elements.append(Paragraph(
            "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
            company_info_style,
        ))
        elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
        elements.append(Paragraph("Email: info@stackly.com", company_info_style))
        elements.append(Spacer(1, 10))

        status = quotation.get('status', 'draft').upper()
        page_title_text = f"QUOTATION - {status}"
        elements.append(Paragraph(page_title_text, page_title_style))
        elements.append(Spacer(1, 2))

        # ---------- Watermark for rejected/expired (preserve original behavior) ----------
        if status.lower() in ['rejected', 'expired']:
            watermark_text = "⚠️ REJECTED - FOR REFERENCE ONLY ⚠️" if status.lower() == 'rejected' else "⚠️ EXPIRED - FOR REFERENCE ONLY ⚠️"
            watermark_color = colors.red if status.lower() == 'rejected' else colors.orange
            watermark_para = Paragraph(
                watermark_text,
                ParagraphStyle(
                    'Watermark',
                    parent=styles["Normal"],
                    fontName="DejaVuSans-Bold",
                    fontSize=12,
                    textColor=watermark_color,
                    alignment=TA_CENTER,
                    spaceAfter=10,
                    backColor=colors.HexColor("#f9f9f9"),
                )
            )
            elements.append(watermark_para)
            elements.append(Spacer(1, 4))

        # ---------- Quotation details table (label/value pairs, same style as DNR) ----------
        quot_date = safe_str(quotation.get('quotation_date', ''))
        expiry = safe_str(quotation.get('expiry_date', ''))
        sales_rep = safe_str(quotation.get('sales_rep', ''))
        po_ref = safe_str(quotation.get('customer_po', 'N/A'))
        payment_terms = safe_str(quotation.get('payment_terms', 'N/A'))

        details_data = [
            [
                Paragraph("<b>Quotation No:</b>", label_style),
                Paragraph(safe_str(quotation.get('quotation_id')), value_style),
                Paragraph("<b>Date:</b>", label_style),
                Paragraph(quot_date, value_style),
            ],
            [
                Paragraph("<b>Customer:</b>", label_style),
                Paragraph(safe_str(quotation.get('customer_name')), value_style),
                Paragraph("<b>Expiry Date:</b>", label_style),
                Paragraph(expiry, value_style),
            ],
            [
                Paragraph("<b>Sales Rep:</b>", label_style),
                Paragraph(sales_rep, value_style),
                Paragraph("<b>Currency:</b>", label_style),
                Paragraph(f"{currency_code} ({currency_symbol})", value_style),
            ],
            [
                Paragraph("<b>PO Reference:</b>", label_style),
                Paragraph(po_ref, value_style),
                Paragraph("<b>Payment Terms:</b>", label_style),
                Paragraph(payment_terms, value_style),
            ],
        ]

        details_table = Table(details_data, colWidths=[110, 170, 95, 145])
        details_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(details_table)
        elements.append(Spacer(1, 16))

        # ---------- Quotation items table ----------
        elements.append(Paragraph("QUOTATION ITEMS", section_style))
        elements.append(Spacer(1, 2))

        items = quotation.get('items', []) or []

        # Table header
        item_data = [
            [
                Paragraph("S.No", header_small_style),
                Paragraph("Product Name", header_small_style),
                Paragraph("Qty", header_small_style),
                Paragraph("UOM", header_small_style),
                Paragraph("Unit Price", header_small_style),
                Paragraph("Tax %", header_small_style),
                Paragraph("Disc %", header_small_style),
                Paragraph("Total", header_small_style),
            ]
        ]

        for idx, item in enumerate(items, start=1):
            product_name = safe_str(item.get("product_name"))
            qty = safe_float(item.get("quantity", 0))
            uom = safe_str(item.get("uom"))
            price = safe_float(item.get("unit_price", 0))
            tax = safe_float(item.get("tax", 0))
            disc = safe_float(item.get("discount", 0))
            line_total = safe_float(item.get("total", 0))
            if line_total == 0:
                line_subtotal = qty * price
                disc_amt = line_subtotal * (disc / 100) if disc > 0 else 0
                after_disc = line_subtotal - disc_amt
                tax_amt = after_disc * (tax / 100) if tax > 0 else 0
                line_total = after_disc + tax_amt

            item_data.append(
                [
                    Paragraph(str(idx), value_style),
                    Paragraph(product_name, value_style),
                    Paragraph(f"{qty:.2f}".rstrip('0').rstrip('.'), value_style),
                    Paragraph(uom, value_style),
                    Paragraph(f"{currency_symbol}{price:.2f}", value_style),
                    Paragraph(f"{tax:.1f}%" if tax > 0 else "-", value_style),
                    Paragraph(f"{disc:.1f}%" if disc > 0 else "-", value_style),
                    Paragraph(f"{currency_symbol}{line_total:.2f}", value_style),
                ]
            )

        if len(item_data) == 1:
            item_data.append(
                [
                    Paragraph("-", value_style),
                    Paragraph("No line items available", value_style),
                    Paragraph("-", value_style),
                    Paragraph("-", value_style),
                    Paragraph("-", value_style),
                    Paragraph("-", value_style),
                    Paragraph("-", value_style),
                    Paragraph("-", value_style),
                ]
            )

        items_table = Table(
            item_data,
            colWidths=[35, 140, 45, 40, 65, 45, 45, 75],
            repeatRows=1,
        )
        items_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (1, 1), (2, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(items_table)
        elements.append(Spacer(1, 16))

        # ---------- Tax and Totals Summary (styled as a table) ----------
        elements.append(Paragraph("TAX AND TOTALS SUMMARY", section_style))

        summary_data = []
        summary_data.append([Paragraph("Subtotal:", label_style), Paragraph(f"{currency_symbol}{subtotal:.2f}", value_style)])
        summary_data.append([Paragraph("Total Discount (Item Level):", label_style), Paragraph(f"{currency_symbol}{total_discount:.2f}", value_style)])
        summary_data.append([Paragraph("Total Tax:", label_style), Paragraph(f"{currency_symbol}{total_tax:.2f}", value_style)])
        if shipping >= 0:
            summary_data.append([Paragraph("Shipping Charge:", label_style), Paragraph(f"{currency_symbol}{shipping:.2f}", value_style)])
        if global_discount_pct >= 0:
            summary_data.append([Paragraph(f"Global Discount ({global_discount_pct:.1f}%):", label_style), Paragraph(f"-{currency_symbol}{global_discount_amt:.2f}", value_style)])
        if rounding != 0:
            sign = "+" if rounding > 0 else ""
            rounding_para = Paragraph(f"{sign}{currency_symbol}{abs(rounding):.2f}", value_style)
            if rounding > 0:
                rounding_para = Paragraph(f"{sign}{currency_symbol}{abs(rounding):.2f}", ParagraphStyle(name="RoundPos", parent=value_style, textColor=colors.green))
            elif rounding < 0:
                rounding_para = Paragraph(f"{sign}{currency_symbol}{abs(rounding):.2f}", ParagraphStyle(name="RoundNeg", parent=value_style, textColor=colors.red))
            summary_data.append([Paragraph("Rounding Adjustment:", label_style), rounding_para])
        summary_data.append([Paragraph("", label_style), Paragraph("", value_style)])  # separator row
        summary_data.append([Paragraph("<b>GRAND TOTAL:</b>", label_style), Paragraph(f"<b>{currency_symbol}{grand_total:.2f}</b>", value_style)])

        summary_table = Table(summary_data, colWidths=[200, 150])
        summary_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 0), (-1, -3), 9),
                    ("FONTSIZE", (0, -1), (-1, -1), 10),
                    ("ALIGN", (0, 0), (0, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LINEABOVE", (0, -2), (-1, -2), 0.5, colors.HexColor("#999999")),
                    ("LINEBELOW", (0, -2), (-1, -2), 0.5, colors.HexColor("#999999")),
                    ("LINEABOVE", (0, -1), (-1, -1), 1.5, colors.HexColor("#8c1f1f")),
                    ("LINEBELOW", (0, -1), (-1, -1), 1.5, colors.HexColor("#8c1f1f")),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#faf0f0")),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        # ---------- Terms and Conditions ----------
        elements.append(Paragraph("Terms and Conditions", section_style))
        terms_list = [
            "1. This quotation is valid until the expiry date mentioned above.",
            "2. Prices are subject to change without prior notice.",
            "3. Payment terms as agreed upon.",
            "4. Delivery charges extra if not specified.",
            "5. Goods once sold will not be taken back.",
            "6. All taxes and duties as applicable.",
        ]
        for term in terms_list:
            elements.append(Paragraph(term, terms_style))
        elements.append(Spacer(1, 12))

        # ---------- Generation footer ----------
        footer_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(footer_text, ParagraphStyle(name="Footer", parent=value_style, fontSize=7, alignment=TA_CENTER)))

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    except Exception as e:
        print(f"PDF generation error: {e}")
        import traceback
        traceback.print_exc()
        return None

@app.route('/generate-pdf/<quotation_id>')
def generate_pdf(quotation_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # fetch quotation header & totals
        cur.execute("""
            SELECT 
                q.quotation_id, q.status, q.quotation_date, q.expiry_date,
                q.customer_name, q.sales_rep, q.customer_po, q.payment_terms as payment_term,
                COALESCE(t.subtotal, 0) as subtotal,
                COALESCE(t.global_discount_percent, 0) as global_discount_percent,
                COALESCE(t.tax_summary, 0) as tax_summary,
                COALESCE(t.shipping_charge, 0) as shipping_charge,
                COALESCE(t.rounding_adjustment, 0) as rounding_adjustment,
                COALESCE(t.grand_total, 0) as grand_total
            FROM quotations q
            LEFT JOIN quotation_totals t ON q.quotation_id = t.quotation_id
            WHERE q.quotation_id = %s
        """, (quotation_id,))
        quotation = cur.fetchone()

        if not quotation:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404

        # fetch items
        cur.execute("""
            SELECT 
                item_id, product_name, quantity, uom, unit_price,
                tax_percent as tax, discount_percent as discount, total
            FROM quotation_items
            WHERE quotation_id = %s
            ORDER BY item_id
        """, (quotation_id,))
        items = cur.fetchall()

        # convert numeric fields to float
        for item in items:
            for key in ['quantity', 'unit_price', 'tax', 'discount', 'total']:
                if item.get(key) is not None:
                    item[key] = float(item[key])

        for idx, item in enumerate(items, 1):
            item['sl_no'] = idx

        quotation['items'] = items
        quotation['totals'] = {
            'subtotal': float(quotation.get('subtotal', 0)),
            'global_discount_percent': float(quotation.get('global_discount_percent', 0)),
            'tax_summary': float(quotation.get('tax_summary', 0)),
            'shipping_charge': float(quotation.get('shipping_charge', 0)),
            'rounding_adjustment': float(quotation.get('rounding_adjustment', 0)),
            'grand_total': float(quotation.get('grand_total', 0))
        }

        cur.close()
        conn.close()

        # generate PDF
        pdf_bytes = generate_quotation_pdf(quotation)

        from flask import make_response
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'

        status_text = quotation.get('status', 'draft').upper()
        if status_text == "DRAFT":
            return jsonify({'success': False, 'error': 'PDF not available for draft quotations'}), 403
        elif status_text in ["REJECTED", "EXPIRED"]:
            response.headers['Content-Disposition'] = f'inline; filename=quotation_{quotation_id}_REFERENCE.pdf'
        else:
            response.headers['Content-Disposition'] = f'attachment; filename=quotation_{quotation_id}.pdf'

        return response

    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


from email.message import EmailMessage

def get_quotation_from_db(quotation_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    try:
        cur.execute("""
            SELECT q.*, 
                   COALESCE(t.subtotal,0) as subtotal,
                   COALESCE(t.global_discount_percent,0) as global_discount_percent,
                   COALESCE(t.tax_summary,0) as tax_summary,
                   COALESCE(t.shipping_charge,0) as shipping_charge,
                   COALESCE(t.rounding_adjustment,0) as rounding_adjustment,
                   COALESCE(t.grand_total,0) as grand_total
            FROM quotations q
            LEFT JOIN quotation_totals t ON q.quotation_id = t.quotation_id
            WHERE q.quotation_id = %s
        """, (quotation_id,))
        quotation = cur.fetchone()
        if not quotation:
            return None
        # Fetch items
        cur.execute("SELECT * FROM quotation_items WHERE quotation_id = %s", (quotation_id,))
        items = cur.fetchall()
        for item in items:
            for col in ['quantity','unit_price','tax_percent','discount_percent','total']:
                if item.get(col) is not None:
                    item[col] = float(item[col])
        quotation['items'] = items
        quotation['totals'] = {
            'subtotal': float(quotation.get('subtotal',0)),
            'global_discount_percent': float(quotation.get('global_discount_percent',0)),
            'tax_summary': float(quotation.get('tax_summary',0)),
            'shipping_charge': float(quotation.get('shipping_charge',0)),
            'rounding_adjustment': float(quotation.get('rounding_adjustment',0)),
            'grand_total': float(quotation.get('grand_total',0))
        }
        return quotation
    finally:
        cur.close()
        conn.close()
@app.route("/send-quotation/<quotation_id>", methods=["POST"])
def send_quotation(quotation_id):
    quotation = get_quotation_from_db(quotation_id)
    if not quotation:
        return jsonify({"success": False, "message": "Quotation not found"}), 404

    pdf_bytes = generate_quotation_pdf(quotation)   # your INR‑only PDF generator
    if not pdf_bytes:
        return jsonify({"success": False, "message": "Error generating PDF"}), 500

    customer_email = quotation.get("customer_email")   # make sure 'customer_email' exists in DB or use customer_name lookup
    if not customer_email:
        return jsonify({"success": False, "message": "Customer email not found"}), 400


def send_otp_email(email, otp, quotation_id=None):
    """Send OTP via email (plain text)"""
    try:
        msg = MIMEMultipart()
        msg['Subject'] = f"Your OTP for Quotation {quotation_id}" if quotation_id else "Your OTP for Quotation"
        msg['From'] = SENDER_EMAIL
        msg['To'] = email
        body = f"Your OTP for verification is: {otp}\n\nThis OTP is valid for {OTP_EXPIRY_MINUTES} minutes."
        msg.attach(MIMEText(body, 'plain'))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)
        print(f"✅ OTP sent to {email}")
        return True
    except Exception as e:
        print(f"❌ OTP email error: {e}")
        return False

import ssl
def send_quotation_email_internal(quotation_id, recipient_email=None):
    """Send quotation PDF by email – simple style like DNR, no external template"""
    try:
        # 1. Fetch quotation from PostgreSQL
        quotation = get_full_quotation_from_db(quotation_id)
        if not quotation:
            return {'success': False, 'error': 'Quotation not found'}

        # If recipient email not provided, use from DB
        if not recipient_email:
            recipient_email = quotation.get('email')
        if not recipient_email:
            return {'success': False, 'error': 'No recipient email available'}

        # 2. Generate PDF (using your existing function)
        pdf_attachment = generate_quotation_pdf(quotation)
        if not pdf_attachment:
            return {'success': False, 'error': 'PDF generation failed'}

        customer_name = quotation.get('customer_name', 'Customer')
        now = datetime.now()

        # 3. Simple inline HTML (exactly like DNR email)
        html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family: Arial, sans-serif; color: #333;">
    <p>Dear {customer_name},</p>
    <p>Please find attached the quotation (<strong>{quotation_id}</strong>) for your requested items.</p>
    <p>Please let us know if you have any questions.</p>
    <br>
    <p>Regards,<br>Stackly Team</p>
</body>
</html>
"""

        text_body = f"""Dear {customer_name},

Please find attached the quotation ({quotation_id}) for your requested items.

Please let us know if you have any questions.

Regards,
Stackly Team"""

        # Build email
        msg = MIMEMultipart('mixed')
        msg['Subject'] = f"Quotation {quotation_id} from Stackly"
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email

        msg_alternative = MIMEMultipart('alternative')
        msg_alternative.attach(MIMEText(text_body, 'plain'))
        msg_alternative.attach(MIMEText(html_body, 'html'))
        msg.attach(msg_alternative)

        # Attach PDF
        attachment = MIMEApplication(pdf_attachment, _subtype="pdf")
        attachment.add_header('Content-Disposition', 'attachment',
                              filename=f"Quotation_{quotation_id}.pdf")
        msg.attach(attachment)

        # 4. Send with dual SMTP fallback
        try:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL(SMTP_SERVER, 465, context=context, timeout=30) as server:
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
        except Exception as ssl_err:
            print(f"SSL port 465 failed: {ssl_err}, trying STARTTLS on 587...")
            with smtplib.SMTP(SMTP_SERVER, 587, timeout=30) as server:
                server.starttls()
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)

        print(f"✅ Quotation {quotation_id} sent to {recipient_email}")
        return {'success': True, 'message': f'Quotation sent to {recipient_email}'}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}
# ===================================================
# API ROUTES
# ===================================================
@app.route('/api/request-approval', methods=['POST'])
def api_request_approval():
    """Request manager approval"""
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        recipient_email = data.get('recipient')
        reason = data.get('reason')
        
        print(f"📨 APPROVAL REQUESTED: Quotation {quotation_id}, Recipient {recipient_email}, Reason: {reason}")
        
        return jsonify({'success': True, 'message': 'Approval request sent'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/send-quotation-email', methods=['POST'])
def api_send_quotation_email():
    """Send quotation email directly (no OTP)"""
    try:
        data = request.get_json(silent=True) or {}
        quotation_id = data.get('quotation_id')
        email = (data.get('email') or '').strip()
        if not quotation_id or not email:
            return jsonify({'success': False, 'error': 'Missing quotation ID or email'}), 400
        if '@' not in email or '.' not in email.split('@')[-1]:
            return jsonify({'success': False, 'error': 'Invalid email address'}), 400

        customer_email = session.get('user_email', 'unknown@example.com')
        allowed, reason, requires_approval = check_email_limits(quotation_id, customer_email, email)
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 400

        result = send_quotation_email_internal(quotation_id, email)
        if result.get('success'):
            record_email_sent(quotation_id, customer_email, email, approved=not requires_approval)
            return jsonify({'success': True, 'message': 'Quotation sent successfully'})
        return jsonify({'success': False, 'error': result.get('error')}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/send', methods=['POST'])
def api_send_otp():
    try:
        data = request.json
        email = data.get('email')
        quotation_id = data.get('quotation_id')
        if not email or '@' not in email:
            return jsonify({'success': False, 'error': 'Invalid email'}), 400

        allowed, reason, attempts_left = check_otp_limits(email, quotation_id)
        if not allowed:
            return jsonify({'success': False, 'error': reason, 'attempts_left': 0}), 429

        otp = generate_otp()  # you have this function
        session[f'otp_{email}_{quotation_id}'] = {'otp': otp, 'created_at': datetime.now().isoformat()}
        if send_otp_email(email, otp, quotation_id):
            record_otp_attempt(email, quotation_id, True)
            return jsonify({'success': True, 'message': 'OTP sent', 'attempts_left': attempts_left})
        return jsonify({'success': False, 'error': 'Failed to send OTP'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/verify', methods=['POST'])
def api_verify_otp():
    try:
        data = request.json
        email = data.get('email')
        otp = data.get('otp')
        quotation_id = data.get('quotation_id')

        allowed, reason, _ = check_otp_limits(email, quotation_id)
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 429

        stored = session.get(f'otp_{email}_{quotation_id}')
        if not stored:
            record_otp_attempt(email, quotation_id, False)
            return jsonify({'success': False, 'error': 'OTP not found'}), 400

        created = datetime.fromisoformat(stored['created_at'])
        if datetime.now() - created > timedelta(minutes=OTP_EXPIRY_MINUTES):
            session.pop(f'otp_{email}_{quotation_id}', None)
            record_otp_attempt(email, quotation_id, False)
            return jsonify({'success': False, 'error': 'OTP expired'}), 400

        if stored['otp'] != otp:
            record_otp_attempt(email, quotation_id, False)
            left = get_otp_attempts_left(email, quotation_id)
            return jsonify({'success': False, 'error': f'Invalid OTP. {left} attempts left'}), 400

        # OTP verified
        record_otp_attempt(email, quotation_id, True)
        session.pop(f'otp_{email}_{quotation_id}', None)

        customer_email = session.get('user_email', 'unknown@example.com')
        allowed2, reason2, requires_approval = check_email_limits(quotation_id, customer_email, email)
        if not allowed2:
            return jsonify({'success': False, 'error': reason2}), 400

        result = send_quotation_email_internal(quotation_id, email)
        if result.get('success'):
            record_email_sent(quotation_id, customer_email, email, approved=not requires_approval)
            return jsonify({'success': True, 'message': 'Quotation sent successfully'})
        return jsonify({'success': False, 'error': result.get('error')}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/otp/resend', methods=['POST'])
def api_resend_otp(): 
    try:
        data = request.json
        email = data.get('email')
        quotation_id = data.get('quotation_id')

        allowed, reason, left = check_resend_limits(email, quotation_id)
        if not allowed:
            return jsonify({'success': False, 'error': reason}), 429

        otp = generate_otp()
        session[f'otp_{email}_{quotation_id}'] = {'otp': otp, 'created_at': datetime.now().isoformat()}
        if send_otp_email(email, otp, quotation_id):
            record_resend_attempt(email, quotation_id)
            return jsonify({'success': True, 'resend_attempts_left': left})
        return jsonify({'success': False, 'error': 'Failed to resend OTP'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/check-email-limit', methods=['POST'])
def api_check_email_limit():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        recipient_email = data.get('recipient')
        customer_email = session.get('user_email', 'unknown@example.com')
        allowed, reason, requires_approval = check_email_limits(quotation_id, customer_email, recipient_email)
        return jsonify({'success': True, 'allowed': allowed, 'reason': reason if not allowed else None, 'requires_approval': requires_approval})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get-email-count/<quotation_id>')
def get_email_count_route(quotation_id):
    try:
        customer_email = session.get('user_email', 'unknown@example.com')
        count = get_email_count(quotation_id, customer_email)
        return jsonify({'success': True, 'count': count, 'max': RATE_LIMIT_CONFIG['max_emails_per_quotation']})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
# ===================================================
# DIAGNOSTIC ROUTES
# ===================================================

@app.route('/test-smtp-connection')
def test_smtp_connection():
    """Test SMTP connection"""
    results = []
    
    try:
        import socket
        ip = socket.gethostbyname('smtp.gmail.com')
        results.append(f"✅ DNS resolved: {ip}")
    except Exception as e:
        results.append(f"❌ DNS failed: {e}")
    
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587, timeout=10)
        server.starttls()
        results.append("✅ Port 587 OK")
        server.quit()
    except Exception as e:
        results.append(f"❌ Port 587 failed: {e}")
    
    return "<br>".join(results)



# ===================================================
# API ENDPOINTS
# ===================================================


@app.route('/get-quotation/<quotation_id>', methods=['GET'])
def get_single_quotation(quotation_id):
    try:
        quotation = get_full_quotation_from_db(quotation_id)
        if not quotation:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        return jsonify({'success': True, 'quotation': quotation})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/check-expired-now', methods=['POST'])
def manual_expiry_check():
    try:
        expired_ids = check_and_update_expired_quotations_db()
        return jsonify({
            'success': True,
            'message': f'Expiry check completed',
            'expired_count': len(expired_ids),
            'total': None   # optional
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ===================================================
# OPTIONAL: Daily Scheduler (Run via cron)
# ===================================================
def daily_expiry_check():
    """
    Run this once per day via cron job
    Example cron: 0 0 * * * python3 /path/to/your/app.py
    """
    print(f"🕒 Daily expiry check started at {datetime.now()}")
    
    # Use the DB version – returns list of expired quotation IDs
    expired_ids = check_and_update_expired_quotations_db()
    
    # Optional: fetch counts from DB for summary
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM quotations")
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM quotations WHERE status IN ('send','sent','submitted')")
    sent = cur.fetchone()[0]
    cur.close()
    conn.close()
    
    print(f"📊 Summary:")
    print(f"   - Total quotations: {total}")
    print(f"   - Still valid (Sent): {sent}")
    print(f"   - Expired (just updated): {len(expired_ids)}")
    print(f"✅ Daily expiry check completed")


# ===================================================
# SOFT-C INTEGRATION
# ===================================================

@app.route('/sync-softc', methods=['POST'])
def sync_softc():
    try:
        data = request.json
        quotation_id = data.get('quotation_id')
        
        # Get quotation data
        with open(QUOTATION_FILE, 'r') as f:
            quotations = json.load(f)
        
        quotation = next((q for q in quotations if q['quotation_id'] == quotation_id), None)
        
        if not quotation:
            return jsonify({'success': False, 'error': 'Quotation not found'}), 404
        
        # 👇 DIRECT ASSIGNMENT - No .env file needed
        softc_api_url = "https://api.soft-c.com/sync"
        softc_api_key = "your-actual-api-key-here"  # Replace with your real key
        
        # Prepare data for Soft-C
        softc_data = {
            "quotation_number": quotation['quotation_id'],
            "customer_name": quotation.get('customer_name'),
            "date": quotation.get('quotation_date'),
            "expiry_date": quotation.get('expiry_date'),
            "currency": quotation.get('currency'),
            "items": quotation.get('items', []),
            "total": quotation.get('grand_total', '0.00'),
            "status": quotation.get('status')
        }
        
        # Send to Soft-C API
        headers = {
            'Authorization': f'Bearer {softc_api_key}',
            'Content-Type': 'application/json'
        }
        
        response = requests.post(softc_api_url, json=softc_data, headers=headers, timeout=10)
        
        if response.status_code == 200:
            return jsonify({'success': True})
        else:
            return jsonify({
                'success': False, 
                'error': f'Soft-C API error: {response.status_code}'
            }), 500
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
# ===================================================
# TEST PDF ROUTE
# ===================================================

@app.route('/test-pdf')
def test_pdf():
    """Simple test to verify PDF generation is working"""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.append(Paragraph("PDF Generation Test", styles['Title']))
        elements.append(Paragraph("If you can see this, ReportLab is working!", styles['Normal']))
        
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=test.pdf'
        
        return response
    except Exception as e:
        return f"Error: {str(e)}"


# =========================================
# BILL LOAD/SAVE HELPERS (DATABASE VERSION)
# =========================================
def fetch_one(query, params=None):
    conn = get_db_connection()
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            return cur.fetchone()
    finally:
        conn.close()

def fetch_all(query, params=None):
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, params)
            return cur.fetchall()
    finally:
        conn.close()

def execute_query(query, params=None):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(query, params)
            conn.commit()
    finally:
        conn.close()

# =========================================
# BILL LOAD/SAVE HELPERS (DATABASE VERSION)
# =========================================

def load_bills():
    """Load all bills from database with their items."""
    sql = """
        SELECT
            b.id,
            b.created_at,
            b.user_email AS user,
            b.payment_mode,
            b.invoice_total,
            COALESCE(
                (SELECT json_agg(
                    json_build_object(
                        'product_code', bi.product_code,
                        'product_name', bi.product_name,
                        'quantity', bi.quantity,
                        'price', bi.price,
                        'total', bi.total
                    )
                ) FROM bill_items bi WHERE bi.bill_id = b.id),
                '[]'::json
            ) AS items
        FROM quick_bills b
        ORDER BY b.id DESC
    """
    rows = fetch_all(sql)
    bills = []
    for row in rows:
        bills.append({
            "id": row["id"],
            "created_at": row["created_at"].isoformat(timespec="seconds") if row["created_at"] else "",
            "user": row["user"],
            "items": row["items"],
            "totals": {"invoice_total": float(row["invoice_total"]) if row["invoice_total"] else 0},
            "payment": {"mode": row["payment_mode"] or ""}
        })
    return bills
def save_bills(bills):
    """Replace all bills in database (full replace - for compatibility)."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM bill_items")
            cur.execute("DELETE FROM quick_bills")
            
            for b in bills:
                cur.execute("""
                    INSERT INTO quick_bills (id, created_at, user_email, payment_mode, invoice_total)
                    VALUES (%s, %s, %s, %s, %s)
                """, (b["id"], b.get("created_at"), b.get("user"), b.get("payment", {}).get("mode"), b.get("totals", {}).get("invoice_total")))
                
                for item in b.get("items", []):
                    quantity = item.get("quantity")
                    if quantity is None or quantity == "":
                        quantity = 1
                    price = item.get("price") or 0
                    total = item.get("total") or (quantity * price)
                    
                    cur.execute("""
                        INSERT INTO bill_items
                        (bill_id, product_code, product_name, quantity, price, total)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (b["id"], item.get("product_code"), item.get("product_name"), quantity, price, total))
            conn.commit()
    finally:
        conn.close()

def generate_bill_id(bills=None):
    """Get next bill ID from sequence."""
    row = fetch_one("SELECT nextval('quick_bills_id_seq') AS next_id")
    return row["next_id"]

# =========================================
# EMAIL CHECK
# =========================================

@app.route("/check-email", methods=["POST"])
def check_emails():
    data = request.get_json() or {}
    email = (data.get("email") or "").strip()
    if not email:
        return jsonify({"exists": False})

    row = fetch_one(
        "SELECT 1 FROM enquiry WHERE LOWER(TRIM(email)) = LOWER(TRIM(%s)) LIMIT 1",
        (email,),
    )
    return jsonify({"exists": bool(row)})


# =========================================
# QUICK BILLING PAGE
# =========================================

@app.route("/quick-billing")
def quick_billing():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    role = session.get("role", "")
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            role = (u.get("role") or role or "").strip()
            break

    return render_template(
        "quick-billing.html",
        page="quick_billing",
        title="Quick Billing - Stackly",
        user_email=user_email,
        user_name=user_name,
        role=role,
    )

# =========================================
# PRODUCTS ENDPOINT - UPDATED FOR YOUR SCHEMA
# =========================================

@app.route("/api/products/qb") 
def api_products_qb():
    """Fetch products from database using actual table schema."""
    try:
        rows = fetch_all("""
            SELECT 
                product_id,
                product_name,
                unit_price,
                COALESCE(discount, 0) as discount,
                tax_code,
                tax_percent,
                category_name,
                specifications
            FROM products 
            WHERE status = 'Active'
            ORDER BY product_id
        """)
        products = []
        for row in rows:
            pid = row["product_id"]
            specs = str(row.get("specifications") or "")
            barcode = ""
            for token in specs.replace(",", " ").split():
                digits = "".join(c for c in token if c.isdigit())
                if len(digits) in (8, 12, 13):
                    barcode = digits
                    break
            if not barcode and pid:
                pid_digits = "".join(c for c in str(pid) if c.isdigit())
                if len(pid_digits) >= 8:
                    barcode = pid_digits

            products.append({
                "product_id": pid,
                "code": pid,
                "name": row["product_name"],
                "price": float(row["unit_price"]) if row["unit_price"] else 0,
                "discount": float(row["discount"]) if row["discount"] else 0,
                "tax_code": row["tax_code"] or "NONE",
                "tax_percent": row.get("tax_percent"),
                "category": row.get("category_name") or "",
                "barcode": barcode,
            })
        return jsonify({"success": True, "products": products})
    except Exception as e:
        print(f"❌ Error fetching products: {e}")
        return jsonify({"success": False, "message": "Could not load products"}), 500

# =========================================
# DELETED ITEMS PAGE
# =========================================

@app.route("/quick-billing/deleted")
def quick_billing_deleted():
    """Legacy URL kept for compatibility."""
    return redirect(url_for("quick_removebilling"))


@app.route("/quick-removebilling")
def quick_removebilling():
    """Standalone Quick Remove Billing page (Removed Items view)."""
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    role = session.get("role", "")
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            role = (u.get("role") or role or "").strip()
            break

    return render_template(
        "quick-removebilling.html",
        page="quick-removebilling",
        role=role,
        title="Quick Remove Billing - Stackly",
        user_email=user_email,
        user_name=user_name,
    )

@app.get("/removed-items")
def removed_items_metadata():
    """Small JSON endpoint so /quick-billing/deleted page has a named Fetch/XHR entry."""
    user_email = session.get("user")
    if not user_email:
        return jsonify(
            {"success": False, "message": "Session expired. Please login first."}
        ), 401

    users = load_users()
    user_name = "User"
    role = session.get("role", "")
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            role = (u.get("role") or role or "").strip()
            break

    return jsonify(
        {
            "success": True,
            "page": "removed-items",
            "current_user": {"email": user_email, "name": user_name, "role": role},
        }
    ), 200

# =========================================
# HOLD BILL ENDPOINT (DATABASE VERSION)
# =========================================

@app.route("/api/hold-bill", methods=["GET", "POST", "DELETE"])
def handle_hold_bill():
    """Store temporary hold bill in database."""
    
    if request.method == "POST":
        data = request.get_json(silent=True) or {}
        try:
            execute_query("DELETE FROM hold_bill")
            execute_query(
                "INSERT INTO hold_bill (data) VALUES (%s)",
                (json.dumps(data),)
            )
            return jsonify({"status": "success"})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    if request.method == "GET":
        try:
            row = fetch_one("SELECT data FROM hold_bill LIMIT 1")
            if row:
                return jsonify({"held": True, "bill": row["data"]})
            return jsonify({"held": False})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    # DELETE
    try:
        execute_query("DELETE FROM hold_bill")
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# =========================================
# SAVE QUICK BILL ENDPOINT (DATABASE VERSION)
# =========================================
@app.route("/api/save-quick-bill", methods=["POST"])
def save_quick_bill():
    try:
        data = request.get_json(silent=True) or {}
        items = data.get("items") or []
        totals = data.get("totals") or {}
        payment = data.get("payment") or {}

        if not items:
            return jsonify({"success": False, "message": "No items to save"}), 400

        user_email = session.get("user") or ""

        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO quick_bills (created_at, user_email, payment_mode, invoice_total)
                    VALUES (NOW(), %s, %s, %s)
                    RETURNING id
                """, (user_email, payment.get("mode"), totals.get("invoice_total")))
                bill_id = cur.fetchone()[0]

                for item in items:   # note: there was a duplicate loop in your original, fix that too
                    quantity = item.get("quantity") or item.get("qty")
                    if quantity is None or quantity == "":
                        quantity = 1
                    price = item.get("price") or 0
                    total = item.get("total") or (quantity * price)

                    cur.execute("""
                        INSERT INTO bill_items
                        (bill_id, product_code, product_name, quantity, price, total)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (bill_id, item.get("product_code"), item.get("product_name"), quantity, price, total))
                conn.commit()
        finally:
            conn.close()

        return jsonify({"success": True, "billId": bill_id}), 201
    except Exception as e:
        print(f"❌ Unexpected error in save_quick_bill: {e}")
        return jsonify({"success": False, "message": "Server error while saving bill"}), 500
# =========================================
# QUICK BILLING REST API (DATABASE VERSION)
# =========================================

def _require_login_json():
    """Helper to check login for JSON endpoints."""
    user_email = session.get("user")
    if not user_email:
        return None, jsonify({"success": False, "message": "Session expired"}), 401
    return user_email, None, None

@app.route("/api/quick-billing", methods=["GET"])
def api_quick_billing_list():
    """List all quick bills with filters and pagination."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    conditions = []
    params = []
    
    q = (request.args.get("q") or "").strip()
    if q:
        conditions.append("(b.id::text ILIKE %s OR b.user_email ILIKE %s OR b.created_at::text ILIKE %s)")
        like = f"%{q}%"
        params.extend([like, like, like])
    
    user_filter = (request.args.get("user") or "").strip()
    if user_filter:
        conditions.append("b.user_email ILIKE %s")
        params.append(f"%{user_filter}%")
    
    date_from = (request.args.get("date_from") or "").strip()
    if date_from:
        conditions.append("b.created_at >= %s")
        params.append(date_from)
    
    date_to = (request.args.get("date_to") or "").strip()
    if date_to:
        conditions.append("b.created_at <= %s")
        params.append(date_to)
    
    where_clause = " AND ".join(conditions) if conditions else "1=1"
    
    count_sql = f"SELECT COUNT(*) FROM quick_bills b WHERE {where_clause}"
    total_items = fetch_one(count_sql, params)["count"]
    
    try:
        page = max(1, int(request.args.get("page") or 1))
        page_size = min(1000, max(1, int(request.args.get("page_size") or 10)))
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Invalid page or page_size"}), 400
    
    offset = (page - 1) * page_size
    total_pages = max(1, (total_items + page_size - 1) // page_size)
    page = min(page, total_pages)
    
    sql = f"""
        SELECT
            b.id, b.created_at, b.user_email AS user,
            b.payment_mode, b.invoice_total,
            COALESCE(
                (SELECT json_agg(
                    json_build_object(
                        'product_code', bi.product_code,
                        'product_name', bi.product_name,
                        'quantity', bi.quantity,
                        'price', bi.price,
                        'total', bi.total
                    )
                ) FROM bill_items bi WHERE bi.bill_id = b.id),
                '[]'::json
            ) AS items
        FROM quick_bills b
        WHERE {where_clause}
        ORDER BY b.id DESC
        LIMIT %s OFFSET %s
    """
    params_page = params + [page_size, offset]
    rows = fetch_all(sql, params_page)
    
    items = []
    for row in rows:
        items.append({
            "id": row["id"],
            "created_at": row["created_at"].isoformat(timespec="seconds") if row["created_at"] else "",
            "user": row["user"],
            "items": row["items"],
            "totals": {"invoice_total": float(row["invoice_total"]) if row["invoice_total"] else 0},
            "payment": {"mode": row["payment_mode"] or ""}
        })
    
    return jsonify({
        "success": True,
        "data": {
            "items": items,
            "page": page,
            "total_pages": total_pages,
            "total_items": total_items,
        }
    }), 200

@app.route("/api/quick-billing/<int:bill_id>", methods=["GET"])
def api_quick_billing_get(bill_id):
    """Return a single bill by id."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    sql = """
        SELECT
            b.id, b.created_at, b.user_email AS user,
            b.payment_mode, b.invoice_total,
            COALESCE(
                (SELECT json_agg(
                    json_build_object(
                        'product_code', bi.product_code,
                        'product_name', bi.product_name,
                        'quantity', bi.quantity,
                        'price', bi.price,
                        'total', bi.total
                    )
                ) FROM bill_items bi WHERE bi.bill_id = b.id),
                '[]'::json
            ) AS items
        FROM quick_bills b
        WHERE b.id = %s
    """
    row = fetch_one(sql, (bill_id,))
    if not row:
        return jsonify({"success": False, "message": "Bill not found"}), 404

    bill = {
        "id": row["id"],
        "created_at": row["created_at"].isoformat(timespec="seconds") if row["created_at"] else "",
        "user": row["user"],
        "items": row["items"],
        "totals": {"invoice_total": float(row["invoice_total"]) if row["invoice_total"] else 0},
        "payment": {"mode": row["payment_mode"] or ""}
    }
    return jsonify({"success": True, "data": bill}), 200

@app.route("/api/quick-billing", methods=["POST"])
def api_quick_billing_create():
    """Create a new quick bill."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    totals = data.get("totals") or {}
    payment = data.get("payment") or {}

    if not items:
        return jsonify({"success": False, "message": "At least one item is required"}), 400

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO quick_bills (created_at, user_email, payment_mode, invoice_total)
                    VALUES (NOW(), %s, %s, %s)
                    RETURNING id, created_at
                """, (session.get("user") or "", payment.get("mode"), totals.get("invoice_total")))
                bill_id, created_at = cur.fetchone()

                for item in items:
                    quantity = item.get("quantity") or 1
                    price = item.get("price") or 0
                    total = item.get("total") or (quantity * price)
                    
                    cur.execute("""
                        INSERT INTO bill_items
                        (bill_id, product_code, product_name, quantity, price, total)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        bill_id,
                        item.get("product_code"),
                        item.get("product_name"),
                        quantity,
                        price,
                        total
                    ))
                conn.commit()

        bill_entry = {
            "id": bill_id,
            "created_at": created_at.isoformat(timespec="seconds"),
            "user": session.get("user") or "",
            "items": items,
            "totals": totals,
            "payment": payment,
        }
        return jsonify({
            "success": True,
            "message": "Bill created successfully",
            "data": bill_entry,
        }), 201

    except Exception as e:
        print(f"❌ Error creating bill: {e}")
        return jsonify({"success": False, "message": "Could not save bill"}), 500

@app.route("/api/quick-billing/<int:bill_id>", methods=["PUT"])
def api_quick_billing_update(bill_id):
    """Update an existing bill."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    if not request.is_json:
        return jsonify({"success": False, "message": "Content-Type must be application/json"}), 400

    data = request.get_json(silent=True) or {}
    items = data.get("items")
    totals = data.get("totals")
    payment = data.get("payment")

    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1 FROM quick_bills WHERE id = %s", (bill_id,))
                if not cur.fetchone():
                    return jsonify({"success": False, "message": "Bill not found"}), 404

                if totals is not None or payment is not None:
                    cur.execute("""
                        UPDATE quick_bills
                        SET payment_mode = COALESCE(%s, payment_mode),
                            invoice_total = COALESCE(%s, invoice_total)
                        WHERE id = %s
                    """, (
                        payment.get("mode") if payment else None,
                        totals.get("invoice_total") if totals else None,
                        bill_id
                    ))

                if items is not None:
                    cur.execute("DELETE FROM bill_items WHERE bill_id = %s", (bill_id,))
                    for item in items:
                        quantity = item.get("quantity") or 1
                        price = item.get("price") or 0
                        total = item.get("total") or (quantity * price)
                        
                        cur.execute("""
                            INSERT INTO bill_items
                            (bill_id, product_code, product_name, quantity, price, total)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            bill_id,
                            item.get("product_code"),
                            item.get("product_name"),
                            quantity,
                            price,
                            total
                        ))
                conn.commit()

        return api_quick_billing_get(bill_id)

    except Exception as e:
        print(f"❌ Error updating bill {bill_id}: {e}")
        return jsonify({"success": False, "message": "Could not update bill"}), 500

@app.route("/api/quick-billing/<int:bill_id>", methods=["DELETE"])
def api_quick_billing_delete(bill_id):
    """Remove a bill."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status

    try:
        execute_query("DELETE FROM quick_bills WHERE id = %s", (bill_id,))
        return jsonify({"success": True, "message": "Bill deleted successfully"}), 200
    except Exception as e:
        print(f"❌ Error deleting bill {bill_id}: {e}")
        return jsonify({"success": False, "message": "Could not delete bill"}), 500

@app.route("/api/quick-billing/new-id", methods=["GET"])
def api_quick_billing_new_id():
    """Return the next bill id for UI use."""
    user_email, resp, status = _require_login_json()
    if resp is not None:
        return resp, status
    
    row = fetch_one("SELECT nextval('quick_bills_id_seq') AS next_id")
    return jsonify({"billId": row["next_id"]}), 200
#--------------------------------------------------------------------------------------------------------


 


# ---------- Helpers ----------
def load_sales_orders():
    """
    Load all sales orders from JSON storage.
    Creates the file if it does not exist.
    Returns an empty list if the file is invalid.
    """
    if not os.path.exists(SALES_ORDERS_FILE):
        with open(SALES_ORDERS_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
        return []

    with open(SALES_ORDERS_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
            return data if isinstance(data, list) else []
        except json.JSONDecodeError:
            return []

def save_sales_orders(data):
    """
    Save all sales orders back to JSON storage.
    """
    with open(SALES_ORDERS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def find_sales_order_by_id(so_id: str):
    """
    Find a sales order by SO ID.
    """
    orders = load_sales_orders()
    so_id = (so_id or "").strip()

    return next(
        (order for order in orders if str(order.get("so_id", "")).strip() == so_id),
        None
    )

def generate_sales_order_id():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT so_id FROM sales_orders
        ORDER BY so_id DESC
        LIMIT 1
    """)

    last = cur.fetchone()

    if last:
        last_id = int(last[0].split('-')[1])
        new_id = last_id + 1
    else:
        new_id = 1

    cur.close()
    conn.close()

    return f"SO-{str(new_id).zfill(3)}"

def upsert_sales_order(payload: dict, status_value: str):
    """
    Insert a new sales order or update an existing one.
    Supports older key names like order_id for backward compatibility.
    """
    orders = load_sales_orders()

    so_id = (payload.get("so_id") or payload.get("order_id") or "").strip()
    if not so_id:
        so_id = generate_sales_order_id()

    existing = next(
        (
            order for order in orders
            if str(order.get("so_id") or order.get("order_id") or "").strip() == so_id
        ),
        None
    )

    # Ignore placeholder customer names
    customer_name = (payload.get("customer_name") or "").strip()
    if customer_name.lower() in ["select customer", "—", "-", ""]:
        customer_name = ""

    # Autofill customer details from master
    customer = find_customer_by_name(customer_name) if customer_name else None
    if customer:
        payload["customer_id"] = customer.get("customer_id", "")
        payload["email"] = customer.get("email", "")
        payload["phone"] = customer.get("phone", "")
        payload["billing_address"] = customer.get("billingAddress", "")
        payload["shipping_address"] = customer.get("shippingAddress", "")

    now_iso = datetime.now().isoformat(timespec="seconds")

    base = {
        "so_id": so_id,
        "order_date": "",
        "sales_rep": "",
        "order_type": "",
        "status": status_value,
        "stock_status": "",

        "customer_name": "",
        "customer_id": "",
        "billing_address": "",
        "shipping_address": "",
        "email": "",
        "phone": "",

        "payment_method": "",
        "currency": "",
        "due_date": "",
        "terms": "",

        "shipping_method": "",
        "delivery_date": "",
        "tracking_number": "",
        "internal_notes": "",
        "customer_notes": "",

        "items": [],

        "subtotal": 0,
        "tax_total": 0,
        "rounding": 0,
        "global_discount": 0,
        "shipping_charges": 0,
        "grand_total": 0,

        "comments": [],
        "status_history": [],

        "cancel_reason": "",
        "cancelled_by": "",
        "cancelled_at": "",

        "created_at": existing.get("created_at") if existing else now_iso,
        "updated_at": now_iso,
    }

    # Merge in the correct order:
    # defaults -> existing data -> incoming payload -> forced system fields
    doc = {**base, **(existing or {}), **payload}
    doc["so_id"] = so_id
    doc["status"] = status_value
    doc["updated_at"] = now_iso

    if not doc.get("created_at"):
        doc["created_at"] = now_iso

    # Remove older key if present
    doc.pop("order_id", None)

    # Ensure expected list types
    if not isinstance(doc.get("items"), list):
        doc["items"] = []

    if not isinstance(doc.get("comments"), list):
        doc["comments"] = []

    if not isinstance(doc.get("status_history"), list):
        doc["status_history"] = []

    # Update existing or insert new
    if existing:
        idx = orders.index(existing)
        orders[idx] = doc
    else:
        orders.insert(0, doc)

    save_sales_orders(orders)
    return so_id


# =========================================
# SALES ORDER - PAGE ROUTES
# =========================================
@app.get("/sales-order")
def sales_order():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    user_name = "User"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
    SELECT name FROM users
    WHERE LOWER(email) = LOWER(%s)
    """, (user_email,))

    row = cur.fetchone()

    if row:
        user_name = row[0] or "User"

    cur.close()
    conn.close()

    return render_template(
        "sales-order.html",
        page="sales_order",
        title="Sales Order - Stackly",
        user_email=user_email,
        user_name=user_name,
    )


@app.get("/sales_order")
def sales_order_compat():
    return redirect("/sales-order", code=302)


@app.get("/sales-order/new")
def sales_order_new():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT name FROM users WHERE email=%s", (user_email,))
    row = cur.fetchone()
    user_name = row[0] if row else "User"

    so_id = generate_sales_order_id()

    cur.execute("""
        SELECT customer_id, name, sales_rep,
               email, phone, billing_address, shipping_address
        FROM customers
    """)
    rows = cur.fetchall()

    customers = []
    sales_reps_set = set()
    for r in rows:
        customers.append({
            "customer_id": r[0],
            "name": r[1],
            "sales_rep": r[2],
            "email": r[3],
            "phone": r[4],
            "billing_address": r[5],
            "shipping_address": r[6]
        })
        if r[2]:
            sales_reps_set.add(r[2])

    sales_reps = sorted(list(sales_reps_set))

    cur.close()
    conn.close()

    return render_template(
        "sales-new.html",
        so_id=so_id,
        customers=customers,
        sales_reps=sales_reps,
        user_name=user_name,
        page="sales_order"
    )


@app.get("/sales-order/edit/<so_id>")
def sales_order_edit(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT customer_id, name
        FROM customers
    """)
    rows = cur.fetchall()

    customers = []
    for r in rows:
        customers.append({
            "customer_id": r[0],
            "name": r[1]
        })

    cur.close()
    conn.close()

    sales_reps = sorted(
        {
            str(c.get("sales_rep", "")).strip()
            for c in customers
            if isinstance(c, dict) and str(c.get("sales_rep", "")).strip()
        }
    )

    return render_template(
        "sales-new.html",
        mode="edit",
        so_id=so_id,
        sales_reps=sales_reps,
        customers=customers,
        page="sales_order"
    )


# =========================================
# SALES ORDER - API ROUTES
# =========================================
@app.get("/api/sales-orders/next-id")
def api_sales_orders_next_id():
    return jsonify({
        "success": True,
        "so_id": generate_sales_order_id()
    })


@app.post("/api/sales-orders/<so_id>/comments")
def add_sales_order_comment(so_id):
    data = request.get_json()
    comment = (data.get("comment") or "").strip()
    user = (data.get("user") or "User").strip()

    if not comment:
        return jsonify({"success": False, "message": "Empty comment"}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 1
        FROM sales_orders
        WHERE so_id = %s
    """, (so_id,))
    exists = cur.fetchone()

    if not exists:
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Save draft first before adding comments."}), 400

    cur.execute("""
        INSERT INTO sales_order_comments (
            so_id,
            comment,
            created_by
        )
        VALUES (%s, %s, %s)
    """, (so_id, comment, user))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})



@app.get("/api/sales-orders/<so_id>/comments")
def get_sales_order_comments(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT comment, created_by, created_at
        FROM sales_order_comments
        WHERE so_id=%s
        ORDER BY created_at DESC
    """, (so_id,))

    rows = cur.fetchall()

    comments = []
    for r in rows:
        comments.append({
            "comment": r[0],
            "user": r[1],
            "time": r[2].strftime("%d/%m/%Y, %I:%M %p")
        })

    cur.close()
    conn.close()

    return jsonify({"success": True, "comments": comments})


@app.get("/api/sales-orders")
def api_sales_orders_list():
    """Sales Order list used by /sales-order.

    Returns ``{"orders": [...]}``. The previous implementation joined
    ``customers`` and only selected ``c.name``; on rows whose
    ``customer_id`` was missing (legacy data) the query worked but elsewhere
    a Postgres error caused the front-end to silently render "No Sales Orders
    Found". Mirror the working ``/api/sales-orders/all`` shape: use
    ``COALESCE(so.customer_name, c.name, '')`` so denormalised columns are
    honoured, defensively coerce values, and log the traceback if anything
    still fails.
    """
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # NOTE: We read ``order_date`` as text (``to_char``) instead of as a
        # ``date``. Legacy test data in the DB has at least one out-of-range
        # year (e.g. 222222), and psycopg blows up with
        # ``ValueError: year must be in 1..9999`` while materialising the row.
        # Casting in SQL bypasses that adapter path entirely.
        cur.execute(
            """
            SELECT
                so.so_id,
                so.order_type,
                COALESCE(so.customer_name, c.name, '') AS customer_name,
                so.sales_rep,
                CASE
                    WHEN so.order_date IS NULL THEN ''
                    ELSE to_char(so.order_date, 'YYYY-MM-DD')
                END AS order_date,
                so.status,
                so.stock_status,
                so.grand_total
            FROM sales_orders so
            LEFT JOIN customers c
              ON so.customer_id = c.customer_id
            ORDER BY so.created_at DESC NULLS LAST, so.so_id DESC
            """
        )

        rows = cur.fetchall() or []

        orders = []
        for r in rows:
            orders.append({
                "so_id": r[0] or "",
                "order_type": r[1] or "",
                "customer_name": r[2] or "",
                "sales_rep": r[3] or "",
                "order_date": r[4] or "",
                "status": r[5] or "",
                "stock_status": r[6] or "",
                "grand_total": float(r[7] or 0),
            })

        return jsonify({"orders": orders})

    except Exception as e:
        print("[/api/sales-orders] failed:", e)
        traceback.print_exc()
        return jsonify({"error": str(e), "orders": []}), 500

    finally:
        cur.close()
        conn.close()

@app.get("/api/sales-orders/available")
def api_sales_orders_available():
    """Sales Orders available for invoice creation.
    Excludes SOs already referenced by other invoices.
    If invoice_id is provided, keeps that invoice's SO available for edit mode.
    """
    invoice_id = (request.args.get("invoice_id") or "").strip() or None

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
                so.so_id,
                COALESCE(c.name, so.customer_name, '') AS customer_name
            FROM sales_orders so
            LEFT JOIN customers c
              ON so.customer_id = c.customer_id
            WHERE NOT EXISTS (
                SELECT 1
                FROM invoices i
                WHERE i.sale_order_ref = so.so_id
                  AND (%s IS NULL OR i.invoice_id <> %s)
            )
            ORDER BY so.created_at DESC
            """,
            (invoice_id, invoice_id),
        )
        rows = cur.fetchall()
        orders = [{"so_id": r[0], "customer_name": r[1]} for r in rows]
        return jsonify({"success": True, "orders": orders})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.post("/api/sales-orders")
def create_sales_order():
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO sales_orders (
            so_id, order_date, sales_rep, order_type, status,
            customer_id, customer_name, billing_address, shipping_address,
            email, phone,
            payment_method, currency, due_date, terms,
            shipping_method, delivery_date, tracking_number,internal_notes, customer_notes,
            subtotal, tax_total, global_discount, shipping_charges, grand_total
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["so_id"], data["order_date"], data["sales_rep"], data["order_type"], data["status"],
        data["customer_id"], data["customer_name"], data["billing_address"], data["shipping_address"],
        data["email"], data["phone"], data["payment_method"], data["currency"], data["due_date"],
        data["terms"], data["shipping_method"], data["delivery_date"], data["tracking_number"],
        data["internal_notes"], data["customer_notes"], data["subtotal"], data["tax_total"],
        data["global_discount"], data["shipping_charges"], data["grand_total"]
    ))

    for item in data["items"]:
        cur.execute("""
            INSERT INTO sales_order_items (
                so_id, product_id, product_name,
                qty, uom, price, tax_pct, disc_pct, line_total
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["so_id"], item["product_id"], item["product_name"], item["qty"],
            item["uom"], item["price"], item["tax_pct"], item["disc_pct"], item["line_total"]
        ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})


@app.get("/api/sales-orders/all")
def api_sales_orders_all():
    conn = get_db_connection()
    cur = conn.cursor()
    q = (request.args.get("q") or "").strip().lower()
    status = (request.args.get("status") or "").strip().lower()
    order_type = (request.args.get("order_type") or "").strip().lower()
    sales_rep = (request.args.get("sales_rep") or "").strip().lower()
    raw_page = request.args.get("page")
    raw_page_size = request.args.get("page_size")
    use_pagination = bool(raw_page or raw_page_size)
    try:
        page = max(1, int(raw_page or 1))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(raw_page_size or 10)
    except (TypeError, ValueError):
        page_size = 10
    page_size = min(max(page_size, 1), 100)

    where_parts = []
    params = []

    if q:
        where_parts.append(
            """
            (
                LOWER(so.so_id) LIKE %s OR
                LOWER(COALESCE(so.customer_name, c.name, '')) LIKE %s OR
                LOWER(COALESCE(so.sales_rep, '')) LIKE %s
            )
            """
        )
        like = f"%{q}%"
        params.extend([like, like, like])
    if status:
        where_parts.append("LOWER(COALESCE(so.status, '')) = %s")
        params.append(status)
    if order_type:
        where_parts.append("LOWER(COALESCE(so.order_type, '')) = %s")
        params.append(order_type)
    if sales_rep:
        where_parts.append("LOWER(COALESCE(so.sales_rep, '')) = %s")
        params.append(sales_rep)

    where_sql = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""

    cur.execute(
        f"""
        SELECT COUNT(*)
        FROM sales_orders so
        LEFT JOIN customers c ON c.customer_id = so.customer_id
        {where_sql}
        """,
        tuple(params),
    )
    total = int((cur.fetchone() or [0])[0] or 0)

    total_pages = max(1, (total + page_size - 1) // page_size) if use_pagination else 1
    page = max(1, min(page, total_pages))
    limit_offset_sql = ""
    query_params = list(params)
    if use_pagination:
        limit_offset_sql = "LIMIT %s OFFSET %s"
        query_params.extend([page_size, (page - 1) * page_size])

    cur.execute(
        f"""
        SELECT
            so.so_id,
            so.order_date,
            so.sales_rep,
            so.order_type,
            so.status,
            so.stock_status,
            so.customer_id,
            COALESCE(so.customer_name, c.name, '') AS customer_name,
            so.grand_total,
            so.created_at,
            so.updated_at
        FROM sales_orders so
        LEFT JOIN customers c ON c.customer_id = so.customer_id
        {where_sql}
        ORDER BY so.created_at DESC, so.so_id DESC
        {limit_offset_sql}
        """,
        tuple(query_params),
    )
    rows = cur.fetchall()
    page_items = []
    for r in rows:
        page_items.append(
            {
                "so_id": r[0],
                "order_date": str(r[1]) if r[1] else "",
                "sales_rep": r[2] or "",
                "order_type": r[3] or "",
                "status": r[4] or "",
                "stock_status": r[5] or "",
                "customer_id": r[6] or "",
                "customer_name": r[7] or "",
                "grand_total": float(r[8] or 0),
                "created_at": r[9].isoformat() if r[9] and hasattr(r[9], "isoformat") else "",
                "updated_at": r[10].isoformat() if r[10] and hasattr(r[10], "isoformat") else "",
            }
        )

    cur.execute("SELECT DISTINCT status FROM sales_orders WHERE status IS NOT NULL AND status <> '' ORDER BY status")
    statuses = [r[0] for r in (cur.fetchall() or []) if r and r[0]]
    cur.execute("SELECT DISTINCT order_type FROM sales_orders WHERE order_type IS NOT NULL AND order_type <> '' ORDER BY order_type")
    types = [r[0] for r in (cur.fetchall() or []) if r and r[0]]
    cur.execute("SELECT DISTINCT sales_rep FROM sales_orders WHERE sales_rep IS NOT NULL AND sales_rep <> '' ORDER BY sales_rep")
    reps = [r[0] for r in (cur.fetchall() or []) if r and r[0]]

    cur.close()
    conn.close()

    return jsonify({
        "orders": page_items,
        "total": total,
        "page": page,
        "total_pages": total_pages,
        "meta": {
            "statuses": statuses,
            "order_types": types,
            "sales_reps": reps,
        },
    })


@app.get("/api/sales-orders/<so_id>")
def get_one_sales_order(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Cast date columns to text (legacy rows may have years outside Python's 1..9999).
    cur.execute("""
        SELECT
            so_id, sales_rep, order_type, status,
            customer_id, customer_name, billing_address, shipping_address,
            email, phone,
            payment_method, currency, terms,
            shipping_method, tracking_number, internal_notes, customer_notes,
            subtotal, tax_total, global_discount, shipping_charges, grand_total,
            stock_status,
            CASE
                WHEN order_date IS NULL THEN ''
                ELSE to_char(order_date, 'YYYY-MM-DD')
            END AS order_date,
            CASE
                WHEN due_date IS NULL THEN ''
                ELSE to_char(due_date, 'YYYY-MM-DD')
            END AS due_date,
            CASE
                WHEN delivery_date IS NULL THEN ''
                ELSE to_char(delivery_date, 'YYYY-MM-DD')
            END AS delivery_date
        FROM sales_orders
        WHERE so_id=%s
    """, (so_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"success": False}), 404

    columns = [desc[0] for desc in cur.description]
    data = dict(zip(columns, row))

    data["order_date"] = str(data.get("order_date") or "")
    data["due_date"] = str(data.get("due_date") or "")
    data["delivery_date"] = str(data.get("delivery_date") or "")
    data["internal_notes"] = data.get("internal_notes") or ""
    data["customer_notes"] = data.get("customer_notes") or ""

    cur.execute("SELECT name FROM customers WHERE customer_id=%s", (data["customer_id"],))
    cust = cur.fetchone()
    data["customer_name"] = cust[0] if cust else ""

    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items_rows = cur.fetchall()
    items = []
    for i in items_rows:
        items.append({
            "product_id": i[2],
            "product_name": i[3],
            "qty": i[4],
            "uom": i[5],
            "price": i[6],
            "tax_pct": i[7],
            "disc_pct": i[8],
            "line_total": i[9]
        })
    data["items"] = items

    cur.execute("""
        SELECT comment, created_by, created_at
        FROM sales_order_comments
        WHERE so_id=%s
        ORDER BY created_at ASC
    """, (so_id,))
    rows = cur.fetchall()
    comments = []
    for r in rows:
        comments.append({
            "text": r[0],
            "user": r[1],
            "created_at": str(r[2])
        })
    data["comments"] = comments

    cur.close()
    conn.close()

    return jsonify({"success": True, "order": data})


@app.post("/api/sales-orders/save-draft")
def api_sales_orders_save_draft():
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT so_id FROM sales_orders WHERE so_id=%s",
        (data["so_id"],)
    )
    existing = cur.fetchone()

    if existing:
        cur.execute("""
            UPDATE sales_orders SET
                order_date=%s,
                sales_rep=%s,
                order_type=%s,
                status=%s,
                customer_id=%s,
                customer_name=%s,
                billing_address=%s,
                shipping_address=%s,
                email=%s,
                phone=%s,
                payment_method=%s,
                currency=%s,
                due_date=%s,
                terms=%s,
                shipping_method=%s,
                delivery_date=%s,
                tracking_number=%s,
                internal_notes=%s,
                customer_notes=%s,
                subtotal=%s,
                tax_total=%s,
                global_discount=%s,
                shipping_charges=%s,
                grand_total=%s
            WHERE so_id=%s
        """, (
            data["order_date"], data["sales_rep"], data["order_type"], "Draft",
            data["customer_id"], data["customer_name"],
            data["billing_address"], data["shipping_address"],
            data["email"], data["phone"],
            data["payment_method"], data["currency"], data["due_date"], data["terms"],
            data["shipping_method"], data["delivery_date"], data["tracking_number"],
            data.get("internal_notes", ""), data.get("customer_notes", ""),
            data["subtotal"], data["tax_total"], data["global_discount"],
            data["shipping_charges"], data["grand_total"], data["so_id"]
        ))
    else:
        cur.execute("""
            INSERT INTO sales_orders (
                so_id, order_date, sales_rep, order_type, status,
                customer_id, customer_name, billing_address, shipping_address,
                email, phone,
                payment_method, currency, due_date, terms,
                shipping_method, delivery_date, tracking_number,
                internal_notes, customer_notes,
                subtotal, tax_total, global_discount, shipping_charges, grand_total
            ) VALUES (
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                %s,%s,%s,%s,%s
            )
        """, (
            data["so_id"], data["order_date"], data["sales_rep"], data["order_type"], "Draft",
            data["customer_id"], data["customer_name"],
            data["billing_address"], data["shipping_address"],
            data["email"], data["phone"],
            data["payment_method"], data["currency"], data["due_date"], data["terms"],
            data["shipping_method"], data["delivery_date"], data["tracking_number"],
            data.get("internal_notes", ""), data.get("customer_notes", ""),
            data["subtotal"], data["tax_total"], data["global_discount"],
            data["shipping_charges"], data["grand_total"]
        ))

    cur.execute("DELETE FROM sales_order_items WHERE so_id=%s", (data["so_id"],))
    for item in data["items"]:
        cur.execute("""
            INSERT INTO sales_order_items (
                so_id, product_id, product_name, qty, uom, price, tax_pct, disc_pct, line_total
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["so_id"], item["product_id"], item["product_name"],
            item["qty"], item["uom"], item["price"], item["tax_pct"],
            item["disc_pct"], item["line_total"]
        ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})


@app.post("/api/sales-orders/submit")
def api_sales_orders_submit():
    conn = None
    try:
        data = request.get_json()

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT so_id FROM sales_orders WHERE so_id=%s", (data.get("so_id"),))
        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE sales_orders SET
                    order_date=%s, sales_rep=%s, order_type=%s, status=%s,
                    customer_id=%s, customer_name=%s, billing_address=%s, shipping_address=%s,
                    email=%s, phone=%s, payment_method=%s, currency=%s, due_date=%s, terms=%s,
                    shipping_method=%s, delivery_date=%s, tracking_number=%s,
                    internal_notes=%s, customer_notes=%s, subtotal=%s, tax_total=%s,
                    global_discount=%s, shipping_charges=%s, grand_total=%s
                WHERE so_id=%s
            """, (
                data["order_date"], data["sales_rep"], data["order_type"], "Submitted",
                data["customer_id"], data["customer_name"],
                data["billing_address"], data["shipping_address"],
                data["email"], data["phone"], data["payment_method"], data["currency"],
                data["due_date"], data["terms"], data["shipping_method"], data["delivery_date"],
                data["tracking_number"], data["internal_notes"], data["customer_notes"],
                data["subtotal"], data["tax_total"], data["global_discount"],
                data["shipping_charges"], data["grand_total"], data["so_id"]
            ))
            cur.execute("DELETE FROM sales_order_items WHERE so_id=%s", (data["so_id"],))
        else:
            cur.execute("""
                INSERT INTO sales_orders (
                    so_id, order_date, sales_rep, order_type, status, customer_id, customer_name,
                    billing_address, shipping_address, email, phone, payment_method, currency,
                    due_date, terms, shipping_method, delivery_date, tracking_number,
                    internal_notes, customer_notes, subtotal, tax_total, global_discount,
                    shipping_charges, grand_total
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data["so_id"], data["order_date"], data["sales_rep"], data["order_type"], "Submitted",
                data["customer_id"], data["customer_name"], data["billing_address"], data["shipping_address"],
                data["email"], data["phone"], data["payment_method"], data["currency"], data["due_date"],
                data["terms"], data["shipping_method"], data["delivery_date"], data["tracking_number"],
                data["internal_notes"], data["customer_notes"], data["subtotal"], data["tax_total"],
                data["global_discount"], data["shipping_charges"], data["grand_total"]
            ))

        for item in data.get("items", []):
            cur.execute("""
                INSERT INTO sales_order_items (
                    so_id, product_id, product_name, qty, uom, price, tax_pct, disc_pct, line_total
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                data["so_id"], item["product_id"], item["product_name"],
                item["qty"], item["uom"], item["price"], item["tax_pct"], item["disc_pct"], item["line_total"]
            ))

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500


@app.route("/api/sales-products", methods=["GET"])
def get_sales_products():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get available columns
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'products'
        """)
        available_cols = {r[0] for r in (cur.fetchall() or [])}

        # Dynamic column mapping (NO CHANGE)
        name_col = "product_name" if "product_name" in available_cols else ("name" if "name" in available_cols else None)

        price_col = (
            "price" if "price" in available_cols else
            ("unit_price" if "unit_price" in available_cols else
             ("selling_price" if "selling_price" in available_cols else None))
        )

        uom_col = "uom" if "uom" in available_cols else ("uom_name" if "uom_name" in available_cols else None)

        stock_col = (
            "stock_level" if "stock_level" in available_cols else
            ("available_stock" if "available_stock" in available_cols else
             ("quantity" if "quantity" in available_cols else
              ("stock" if "stock" in available_cols else
               ("qty" if "qty" in available_cols else
                ("opening_stock" if "opening_stock" in available_cols else None)))))
        )

        # 🟢 NEW: check tax & discount columns safely
        discount_col = "discount" if "discount" in available_cols else None
        tax_percent_col = "tax_percent" if "tax_percent" in available_cols else None

        if "product_id" not in available_cols:
            return jsonify({"success": True, "products": []})

        # Expressions
        name_expr = name_col if name_col else "''"
        price_expr = price_col if price_col else "0"
        uom_expr = uom_col if uom_col else "''"
        stock_expr = stock_col if stock_col else "0"

        # 🟢 NEW expressions
        discount_expr = f"COALESCE({discount_col}, 0)" if discount_col else "0"
        tax_expr = f"COALESCE({tax_percent_col}, 0)" if tax_percent_col else "0"

        # 🟢 UPDATED QUERY
        cur.execute(f"""
            SELECT 
                product_id,
                {name_expr} AS product_name,
                {price_expr} AS price,
                {uom_expr} AS uom,
                {stock_expr} AS stock_level,
                {discount_expr} AS discount,
                {tax_expr} AS tax_percent
            FROM products
            ORDER BY product_id
        """)

        rows = cur.fetchall()

        products = []
        for r in rows:
            pid = str(r[0] or "").strip()
            if not pid:
                continue

            products.append({
                "product_id": pid,
                "product_name": (r[1] or "").strip() if isinstance(r[1], str) else str(r[1] or "").strip(),
                "price": float(r[2] or 0),
                "uom": (r[3] or "").strip() if isinstance(r[3], str) else str(r[3] or "").strip(),
                "stock_level": float(r[4] or 0),

                # 🟢 NEW FIELDS
                "discount": float(r[5] or 0),
                "tax_percent": float(r[6] or 0),
            })

        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "products": products
        })

    except Exception as e:
        if conn:
            conn.rollback()
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# =========================================
# SALES ORDER - PDF
# =========================================
@app.get("/api/sales-orders/<so_id>/pdf")
def sales_order_pdf(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM sales_orders WHERE so_id=%s", (so_id,))
    order = cur.fetchone()
    order_cols = [desc[0] for desc in cur.description]

    if not order:
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Sales Order not found"}), 404

    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items = cur.fetchall()
    item_cols = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    so = dict(zip(order_cols, order))
    so["items"] = []
    for i in items:
        item = dict(zip(item_cols, i))
        so["items"].append({
            "product_id": item.get("product_id", ""),
            "product_name": item.get("product_name", ""),
            "qty": float(item.get("qty") or 0),
            "uom": item.get("uom", ""),
            "price": float(item.get("price") or 0),
            "tax_pct": float(item.get("tax_pct") or 0),
            "disc_pct": float(item.get("disc_pct") or 0),
            "line_total": float(item.get("line_total") or 0)
        })

    try:
        pdf_bytes = generate_sales_order_pdf_bytes(so)
        response = make_response(pdf_bytes)
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = f'inline; filename="{so_id}.pdf"'
        return response
    except Exception as e:
        print("Sales Order PDF error:", e)
        return jsonify({"success": False, "message": str(e)}), 500


def generate_sales_order_pdf_bytes(so):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=32
    )

    elements = []
    styles = getSampleStyleSheet()

    # =========================================
    # PDF STYLES
    # =========================================
    title_style = ParagraphStyle(
        "SOTitle",
        parent=styles["Heading1"],
        fontSize=24,
        leading=28,
        alignment=1,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=8,
        fontName="DejaVuSans-Bold"
    )

    company_style = ParagraphStyle(
        "SOCompany",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=1,
        spaceAfter=2,
        fontName="DejaVuSans"
    )

    status_style = ParagraphStyle(
        "SOStatus",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        alignment=1,
        textColor=colors.HexColor("#148a08"),
        spaceAfter=14,
        fontName="DejaVuSans-Bold"
    )

    heading_style = ParagraphStyle(
        "SOHeading",
        parent=styles["Heading2"],
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceBefore=8,
        spaceAfter=8,
        fontName="DejaVuSans-Bold"
    )

    table_cell_style = ParagraphStyle(
        "SOTableCell",
        parent=styles["Normal"],
        fontSize=7.6,
        leading=9,
        fontName="DejaVuSans",
        wordWrap="CJK"
    )

    table_label_style = ParagraphStyle(
        "SOTableLabel",
        parent=styles["Normal"],
        fontSize=7.6,
        leading=9,
        fontName="DejaVuSans-Bold",
        textColor=colors.HexColor("#5f2d2d"),
        wordWrap="CJK"
    )

    terms_heading_style = ParagraphStyle(
        "SOTermsHeading",
        parent=styles["Heading2"],
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        fontName="DejaVuSans-Bold"
    )

    terms_style = ParagraphStyle(
        "SOTerms",
        parent=styles["Normal"],
        fontSize=7.4,
        leading=10,
        fontName="DejaVuSans"
    )

    footer_style = ParagraphStyle(
        "SOFooter",
        parent=styles["Normal"],
        fontSize=7.5,
        textColor=colors.HexColor("#555555"),
        alignment=0
    )

    # =========================================
    # CURRENCY MAPPING
    # =========================================
    currency_code = so.get("currency", "INR")
    currency_map = {
        "USD": "$",
        "EUR": "€",
        "GBP": "£",
        "INR": "₹",
        "IND": "₹",
        "SGD": "S$"
    }
    currency_symbol = currency_map.get(currency_code, currency_code)

    # =========================================
    # PDF HEADER
    # =========================================
    elements.append(Paragraph("STACKLY", title_style))
    elements.append(Paragraph(
        "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
        company_style
    ))
    elements.append(Paragraph("Phone: +91 7010792745", company_style))
    elements.append(Paragraph("Email: info@stackly.com", company_style))
    elements.append(Spacer(1, 10))

    status_text = (so.get("status") or "Submitted").upper()
    elements.append(Paragraph(f"SALES ORDER - {status_text}", status_style))
    elements.append(Spacer(1, 4))

    # =========================================
    # SALES ORDER INFO TABLE
    # =========================================
    info_data = [
        [
            Paragraph("Sales Order Number:", table_label_style),
            Paragraph(str(so.get("so_id", "") or "-"), table_cell_style),
            Paragraph("Date:", table_label_style),
            Paragraph(str(so.get("order_date", "") or "-"), table_cell_style),
        ],
        [
            Paragraph("Customer:", table_label_style),
            Paragraph(str(so.get("customer_name", "") or "-"), table_cell_style),
            Paragraph("Delivery Date:", table_label_style),
            Paragraph(str(so.get("delivery_date", "") or "-"), table_cell_style),
        ],
        [
            Paragraph("Sales Rep:", table_label_style),
            Paragraph(str(so.get("sales_rep", "") or "-"), table_cell_style),
            Paragraph("Currency:", table_label_style),
            Paragraph(str(currency_code or "-"), table_cell_style),
        ],
        [
            Paragraph("Order Type:", table_label_style),
            Paragraph(str(so.get("order_type", "") or "-"), table_cell_style),
            Paragraph("Payment Terms:", table_label_style),
            Paragraph(str(so.get("terms", "N/A") or "N/A"), table_cell_style),
        ],
    ]

    info_table = Table(info_data, colWidths=[110, 145, 95, 130])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#efefef")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#efefef")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    # =========================================
    # ITEMS TABLE
    # =========================================
    elements.append(Paragraph("SALES ORDER ITEMS", heading_style))

    items = so.get("items", [])

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT product_id, product_name, unit_price
        FROM products
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    product_map = {
        str(r[0]).strip(): {
            "product_id": r[0],
            "product_name": r[1],
            "unit_price": r[2]
        }
        for r in rows
    }

    table_data = [[
        "S.No", "Product Name", "Qty", "UOM", "Unit Price", "Tax %", "Disc %", "Total"
    ]]

    subtotal_calc = 0.0
    total_tax_calc = 0.0
    total_discount_calc = 0.0

    for idx, item in enumerate(items, start=1):
        pid = str(item.get("product_id", "")).strip()
        product = product_map.get(pid, {})

        qty = float(item.get("qty", 0) or 0)
        uom = item.get("uom", "") or "Nos"

        unit_price = float(
            item.get("price")
            or item.get("unit_price")
            or product.get("unit_price")
            or product.get("price")
            or product.get("selling_price")
            or 0
        )

        tax_pct = float(item.get("tax_pct", 0) or 0)
        disc_pct = float(item.get("disc_pct", 0) or 0)

        line_subtotal = qty * unit_price
        discount_amt = line_subtotal * (disc_pct / 100)
        after_discount = line_subtotal - discount_amt
        tax_amt = after_discount * (tax_pct / 100)
        line_total = float(item.get("line_total", 0) or (after_discount + tax_amt))

        subtotal_calc += line_subtotal
        total_tax_calc += tax_amt
        total_discount_calc += discount_amt

        table_data.append([
            str(idx),
            str(item.get("product_name", "") or "-"),
            f"{qty:.2f}",
            str(uom),
            f"{currency_symbol}{unit_price:.2f}",
            f"{tax_pct:.1f}%",
            f"{disc_pct:.1f}%",
            f"{currency_symbol}{line_total:.2f}"
        ])

    items_table = Table(table_data, colWidths=[32, 170, 42, 40, 60, 44, 44, 58])
    items_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 14))

    # =========================================
    # TOTALS SUMMARY
    # =========================================
    elements.append(Paragraph("TAX AND TOTALS SUMMARY", heading_style))

    shipping_charge = float(so.get("shipping_charges", 0) or 0)
    global_discount = float(so.get("global_discount", 0) or 0)
    rounding = float(so.get("rounding", 0) or 0)

    grand_total = float(
        so.get("grand_total", 0)
        or (subtotal_calc - global_discount + total_tax_calc + shipping_charge + rounding)
    )

    summary_data = [
        ["Subtotal:", f"{currency_symbol}{subtotal_calc:.2f}"],
        ["Total Discount (Item Level):", f"{currency_symbol}{total_discount_calc:.2f}"],
        ["Total Tax:", f"{currency_symbol}{total_tax_calc:.2f}"],
        ["Shipping Charge:", f"{currency_symbol}{shipping_charge:.2f}"],
        ["Global Discount:", f"-{currency_symbol}{global_discount:.2f}"],
    ]

    if rounding != 0:
        sign = "+" if rounding > 0 else "-"
        summary_data.append(["Rounding Adjustment:", f"{sign}{currency_symbol}{abs(rounding):.2f}"])

    summary_data.append(["GRAND TOTAL:", f"{currency_symbol}{grand_total:.2f}"])

    summary_table = Table(summary_data, colWidths=[300, 200])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -2), "DejaVuSans"),
        ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
        ("FONTSIZE", (0, 0), (-1, -2), 8),
        ("FONTSIZE", (0, -1), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.whitesmoke),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#7f1f1f")),
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    # =========================================
    # TERMS AND CONDITIONS
    # =========================================
    elements.append(Paragraph("Terms and Conditions", terms_heading_style))

    terms_lines = [
        "1. This Sales Order is issued based on the confirmed order details.",
        "2. Delivery will be made as per the agreed schedule.",
        "3. Payment should be completed as per agreed terms.",
        "4. Shipping charges extra if applicable.",
        "5. Goods once sold will not be taken back.",
        "6. All taxes and duties as applicable.",
        f"7. Internal Notes: {so.get('internal_notes', '') or 'N/A'}",
        f"8. Customer Notes: {so.get('customer_notes', '') or 'N/A'}",
    ]

    for line in terms_lines:
        elements.append(Paragraph(line, terms_style))

    elements.append(Spacer(1, 18))

    # =========================================
    # FOOTER
    # =========================================
    generated_on = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated on: {generated_on}", footer_style))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# =========================================
# SALES ORDER - EMAIL API
# =========================================
@app.post("/api/sales-orders/<so_id>/email")
def sales_order_email(so_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM sales_orders WHERE so_id=%s", (so_id,))
    order = cur.fetchone()
    order_cols = [desc[0] for desc in cur.description]

    if not order:
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Sales Order not found"}), 404

    cur.execute("SELECT * FROM sales_order_items WHERE so_id=%s", (so_id,))
    items = cur.fetchall()
    item_cols = [desc[0] for desc in cur.description]
    cur.close()
    conn.close()

    so = dict(zip(order_cols, order))
    so["items"] = []
    for i in items:
        item = dict(zip(item_cols, i))
        so["items"].append({
            "product_id": item.get("product_id", ""),
            "product_name": item.get("product_name", ""),
            "qty": float(item.get("qty") or 0),
            "uom": item.get("uom", ""),
            "price": float(item.get("price") or 0),
            "tax_pct": float(item.get("tax_pct") or 0),
            "disc_pct": float(item.get("disc_pct") or 0),
            "line_total": float(item.get("line_total") or 0)
        })

    customer_email = (so.get("email") or "").strip()
    if not customer_email:
        return jsonify({"success": False, "message": "Customer email not found"}), 400

    try:
        pdf_bytes = generate_sales_order_pdf_bytes(so)
        customer_name = so.get("customer_name", "Customer")
        so_no = so.get("so_id", "")
        order_date = so.get("order_date", "")
        grand_total = so.get("grand_total", 0)
        currency = so.get("currency", "INR")

        subject = f"Sales Order {so_no} from Stackly"
        body = (
            f"Dear {customer_name},\n\n"
            f"Please find attached the sales order ({so_no}).\n"
            "The details have been prepared as per the information provided.\n\n"
            "Please let us know if you have any questions.\n\n"
            "Regards,\n"
            "Stackly Team"
        )

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = SENDER_EMAIL
        msg["To"] = customer_email
        msg.set_content(body)
        msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=f"{so_no}.pdf")

        context = ssl.create_default_context()
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls(context=context)
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.send_message(msg)

        return jsonify({"success": True})
    except Exception as e:
        print("Email error:", e)
        return jsonify({"success": False, "message": str(e)}), 500


# =========================================
# SALES ORDER - CANCEL API
# =========================================
@app.post("/api/sales-orders/<so_id>/cancel")
def cancel_sales_order(so_id):
    data = request.get_json()
    reason = data.get("reason")
    cancelled_by = data.get("cancelled_by")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            UPDATE sales_orders
            SET status = 'Cancelled'
            WHERE so_id = %s
        """, (so_id,))

        cursor.execute("""
            INSERT INTO sales_order_cancellation (
                so_id,
                reason,
                cancelled_by
            )
            VALUES (%s, %s, %s)
        """, (so_id, reason, cancelled_by))

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        print("ERROR:", e)
        conn.rollback()
        return jsonify({"success": False, "message": str(e)})
    finally:
        cursor.close()
        conn.close()


# =========================================
# DELIVERY NOTE - UTILITIES / HELPERS
# =========================================

# ==================
# DELIVERY NOTE -
# =================


def find_customer_by_name(name: str):
    if not name:
        return None

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT customer_id, name, email, phone, billing_address, shipping_address
        FROM customers
        WHERE LOWER(name) = LOWER(%s)
        LIMIT 1
    """, (name,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    if row:
        return {
            "customer_id": row[0],
            "name": row[1],
            "email": row[2],
            "phone": row[3],
            "billing_address": row[4],
            "shipping_address": row[5]
        }

    return None
# -----------------------------------------
# Delivery Notes JSON Storage
# -----------------------------------------
# def load_delivery_notes():
#     if not os.path.exists(DELIVERY_NOTE_FILE):
#         with open(DELIVERY_NOTE_FILE, "w", encoding="utf-8") as f:
#             json.dump([], f)
#         return []
#
#     with open(DELIVERY_NOTE_FILE, "r", encoding="utf-8") as f:
#         try:
#             return json.load(f)
#         except json.JSONDecodeError:
#             return []
#
#
# def save_delivery_notes(data):
#     with open(DELIVERY_NOTE_FILE, "w", encoding="utf-8") as f:
#         json.dump(data, f, indent=2)


# -----------------------------------------
# Next DN ID generator (DN-0001 format)
# -----------------------------------------
def next_dn_id(notes):
    max_num = 0
    for n in notes:
        dn = str(n.get("dn_id", ""))
        if dn.startswith("DN-"):
            try:
                num = int(dn.split("-")[1])
                max_num = max(max_num, num)
            except:
                pass
    return f"DN-{max_num+1:04d}"


def _dn_pod_allowed(filename):
    ext = (filename or "").rsplit(".", 1)[-1].lower() if "." in (filename or "") else ""
    return ext in _DN_POD_EXTENSIONS


def _ensure_delivery_note_ack_columns(cur):
    for col, typ in (
        ("received_by", "TEXT"),
        ("contact_number", "TEXT"),
        ("pod_file_name", "TEXT"),
        ("pod_file_path", "TEXT"),
    ):
        cur.execute(f"ALTER TABLE delivery_notes ADD COLUMN IF NOT EXISTS {col} {typ}")
    cur.connection.commit()


# -----------------------------------------
# Get DN by ID
# -----------------------------------------
def get_dn_by_id(dn_id: str):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM delivery_notes WHERE dn_id=%s", (dn_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return None

    columns = [desc[0] for desc in cur.description]
    dn = dict(zip(columns, row))

    # fetch items
    cur.execute("""
    SELECT product_id, product_name, qty, uom, serial_no
    FROM delivery_note_items
    WHERE dn_id=%s
    """, (dn_id,))

    items_rows = cur.fetchall()

    items = []
    for i in items_rows:
        items.append({
            "product_id": i[0],
            "product_name": i[1],
            "qty": float(i[2]),
            "uom": i[3],
            "serial_no": i[4]
        })

    dn["items"] = items

    cur.close()
    conn.close()

    return dn


# -----------------------------------------
# EMAIL ATTACHMENT HELPER
# -----------------------------------------
def send_email_with_attachments(to_email, subject, body, from_email, password, attachments=None):
    smtp_server = "smtp.gmail.com"
    port = 587
    attachments = attachments or []

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    msg.attach(MIMEText(body, "plain", "utf-8"))

    for a in attachments:
        filename = a.get("filename", "attachment")
        content = a.get("content_bytes", b"")
        part = MIMEApplication(content, _subtype="pdf")
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP(smtp_server, port) as server:
            server.starttls(context=context)
            server.login(from_email, password)
            server.sendmail(from_email, [to_email], msg.as_string())
        return True
    except Exception as e:
        print("❌ Email send error:", e)
        return False


# =========================================
# DELIVERY NOTE - FIRST PAGE (List Page)
# delivery-note.html + delivery-note.js
# =========================================

@app.route("/delivery_note")
def delivery_note():
    user_email = session.get("user")
    user_name = "User"

    if user_email:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
        SELECT name FROM users
        WHERE LOWER(email) = LOWER(%s)
        """, (user_email,))

        row = cur.fetchone()

        if row:
            user_name = row[0] or "User"

        cur.close()
        conn.close()

    return render_template(
        "delivery-note.html",
        page="delivery_note",
        user_email=user_email,
        user_name=user_name,
    )

# =========================================
# DELIVERY NOTE - SECOND PAGE (New/Edit/View)
# deliverynote-new.html + deliverynote-new.js
# =========================================

@app.route("/delivery_note/new")
def delivery_note_new():
    conn = get_db_connection()
    cur = conn.cursor()

    # FIXED SALES ORDERS
    cur.execute("""
        SELECT so_id, customer_name
        FROM sales_orders
        ORDER BY created_at DESC
    """)
    rows = cur.fetchall()

    sales_orders = []
    for r in rows:
        sales_orders.append({
            "so_id": r[0],
            "customer_name": r[1]
        })

    # NEXT ID FIX (important)
    cur.execute("""
        SELECT dn_id FROM delivery_notes
        ORDER BY dn_id DESC
        LIMIT 1
    """)
    last = cur.fetchone()

    if last:
        last_num = int(last[0].split("-")[1])
        next_id = f"DN-{last_num + 1:03d}"
    else:
        next_id = "DN-001"

    cur.close()
    conn.close()

    return render_template(
        "deliverynote-new.html",
        page="delivery_note",
        sales_orders=sales_orders,
        next_dn_id=next_id,
        so_id=request.args.get("so_id", "").strip(),
        user_email=session.get("user"),
        user_name=_get_logged_in_user_name(),
    )


@app.route("/delivery_note/form")
def delivery_note_form():
    dn_id = request.args.get("id", "")
    mode = request.args.get("mode", "edit")

    conn = get_db_connection()
    cur = conn.cursor()

    # FIXED SALES ORDERS
    cur.execute("""
        SELECT so_id, customer_name
        FROM sales_orders
        ORDER BY created_at DESC
    """)
    rows = cur.fetchall()

    sales_orders = []
    for r in rows:
        sales_orders.append({
            "so_id": r[0],
            "customer_name": r[1]
        })

    cur.close()
    conn.close()

    return render_template(
        "deliverynote-new.html",
        page="delivery_note",
        dn_id=dn_id,
        mode=mode,
        sales_orders=sales_orders,
        user_email=session.get("user"),
        user_name=_get_logged_in_user_name(),
    )
# =========================================
# DELIVERY NOTE - API (List + Create)
# =========================================

@app.route("/api/delivery-notes", methods=["GET", "POST"])
def api_delivery_notes():

    # GET: list all delivery notes (for first page table)
    if request.method == "GET":
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
        SELECT dn_id,
        CASE
            WHEN delivery_date IS NULL THEN ''
            ELSE to_char(delivery_date, 'YYYY-MM-DD')
        END AS delivery_date,
        so_id, customer_name,
        delivery_type, destination_address,
        delivery_status, status
        FROM delivery_notes
        ORDER BY created_at DESC
        """)

        rows = cur.fetchall()

        notes = []
        for r in rows:
            notes.append({
                "dn_id": r[0],
                "delivery_date": str(r[1]),
                "sale_order_ref": r[2],  # keep same key
                "customer_name": r[3],
                "delivery_type": r[4],
                "destination_address": r[5],
                "delivery_status": r[6],
                "status": r[7],
            })

        cur.close()
        conn.close()

        return jsonify({"success": True, "data": notes})

    # POST: create new delivery note (from second page submit/save draft)
    data = request.get_json(force=True) or {}

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT dn_id FROM delivery_notes")
    rows = cur.fetchall()

    notes = [{"dn_id": r[0]} for r in rows]

    new_id = (data.get("dn_id") or "").strip()
    if not new_id:
        new_id = next_dn_id(notes)

    record = {
        "dn_id": new_id,
        "delivery_date": data.get("delivery_date", ""),
        "sale_order_ref": data.get("sale_order_ref", ""),
        "customer_name": data.get("customer_name", ""),
        "delivery_type": data.get("delivery_type", ""),
        "destination_address": data.get("destination_address", ""),
        "delivery_by": data.get("delivery_by", ""),
        "delivery_status": data.get("delivery_status", "draft"),
        "vehicle_no": data.get("vehicle_no", ""),
        "tracking_id": data.get("tracking_id", ""),
        "delivery_notes": data.get("delivery_notes", ""),
        "status": data.get("status", "Draft"),
        "items": data.get("items", []),
    }

    # Auto-fetch customer fields for email/phone/address (based on customer_name)
    customer_name = (record.get("customer_name") or "").strip()
    customer = find_customer_by_name(customer_name)

    if customer:
        record["customer_id"] = customer.get("customer_id", "")
        record["email"] = customer.get("email", "")
        record["phone"] = customer.get("phone", "")
        record["billing_address"] = customer.get("billing_address", "")
        record["shipping_address"] = customer.get("shipping_address", "")
    else:
        record["customer_id"] = ""
        record["email"] = ""
        record["phone"] = ""
        record["billing_address"] = ""
        record["shipping_address"] = ""

    # INSERT header
    # Keep FK valid: empty SO ref should be stored as NULL, not "".
    so_ref = (record.get("sale_order_ref") or "").strip() or None
    _ensure_delivery_note_ack_columns(cur)
    pod_path = (data.get("pod_file_path") or "").strip()
    pod_name = (data.get("pod_file") or data.get("pod_file_name") or "").strip()
    if pod_path and not pod_name:
        pod_name = _upload_basename(pod_path)

    cur.execute("""
    INSERT INTO delivery_notes (
        dn_id, so_id, customer_name, destination_address,
        delivery_date, delivery_type,
        status, delivery_status,
        delivery_by, vehicle_number, tracking_id, delivery_notes,
        received_by, contact_number, pod_file_name, pod_file_path
    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        record["dn_id"],
        so_ref,   # IMPORTANT mapping
        record["customer_name"],
        record["destination_address"],
        record["delivery_date"],
        record["delivery_type"],
        record["status"],
        record["delivery_status"],
        record["delivery_by"],
        record["vehicle_no"],
        record["tracking_id"],
        record["delivery_notes"],
        (data.get("received_by") or "").strip(),
        (data.get("contact_number") or "").strip(),
        pod_name or None,
        pod_path or None,
    ))

    # INSERT items
    # Some historical SO rows can reference product IDs that no longer exist in products.
    # To avoid hard-failing the whole DN save on FK, store product_id as NULL in that case.
    cur.execute("SELECT product_id FROM products")
    valid_product_ids = {str(r[0]).strip() for r in cur.fetchall() if r and r[0]}

    for it in record["items"]:
        raw_pid = str(it.get("product_id") or "").strip()
        safe_pid = raw_pid if raw_pid and raw_pid in valid_product_ids else None
        cur.execute("""
        INSERT INTO delivery_note_items (
            dn_id, product_id, product_name, qty, uom, serial_no
        ) VALUES (%s,%s,%s,%s,%s,%s)
        """, (
            record["dn_id"],
            safe_pid,
            it.get("product_name"),
            it.get("qty"),
            it.get("uom"),
            it.get("serial_no", "")
        ))

    # DN delivery status → linked Sales Order status
    if so_ref:
        dn_ds = str(record.get("delivery_status") or "").strip().lower().replace(" ", "_")
        so_norm = "LOWER(REPLACE(COALESCE(status, ''), ' ', '_'))"
        if dn_ds == "delivered":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'returned', 'delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Delivered", so_ref),
            )
        elif dn_ds == "partially_delivered":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'returned', 'delivered', 'partially_delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Partially Delivered", so_ref),
            )
        elif dn_ds == "returned":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'delivered', 'partially_delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Partially Delivered", so_ref),
            )

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True, "message": "Delivery Note Saved", "dn_id": new_id})


# =========================================
# DELIVERY NOTE - API (Get One + Update One)
# =========================================

@app.route("/api/delivery-notes/<dn_id>", methods=["GET", "PUT"])
def api_delivery_note_one(dn_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # =========================
    # FETCH HEADER
    # =========================
    _ensure_delivery_note_ack_columns(cur)
    cur.execute("""
    SELECT
        dn_id,
        CASE
            WHEN delivery_date IS NULL THEN ''
            ELSE to_char(delivery_date, 'YYYY-MM-DD')
        END AS delivery_date,
        so_id, customer_name, destination_address,
        delivery_type, status, delivery_status,
        delivery_by, vehicle_number, tracking_id, delivery_notes,
        received_by, contact_number, pod_file_name, pod_file_path
    FROM delivery_notes
    WHERE dn_id=%s
    """, (dn_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "Delivery Note not found"}), 404

    columns = [desc[0] for desc in cur.description]
    dn = dict(zip(columns, row))

    # =========================
    # FETCH ITEMS
    # =========================
    cur.execute("""
    SELECT product_id, product_name, qty, uom, serial_no
    FROM delivery_note_items
    WHERE dn_id=%s
    """, (dn_id,))

    items_rows = cur.fetchall()

    items = []
    for i in items_rows:
        items.append({
            "product_id": i[0],
            "product_name": i[1],
            "qty": float(i[2]),
            "uom": i[3],
            "serial_no": i[4]
        })

    dn["items"] = items  # correct place

    # =========================
    # GET
    # =========================
    if request.method == "GET":
        dn.setdefault("delivery_type", "")
        dn.setdefault("destination_address", "")
        dn.setdefault("vehicle_no", dn.get("vehicle_number", ""))
        dn.setdefault("tracking_id", "")
        dn.setdefault("delivery_by", "")
        dn.setdefault("delivery_notes", "")
        dn.setdefault("delivery_status", "draft")
        dn.setdefault("received_by", "")
        dn.setdefault("contact_number", "")
        pod_name = dn.get("pod_file_name") or ""
        pod_path = dn.get("pod_file_path") or ""
        dn["pod_file"] = pod_name
        dn["ack_pod_file_name"] = pod_name

        cur.close()
        conn.close()
        return jsonify({"success": True, "data": dn})

    # =========================
    # PUT
    # =========================
    payload = request.get_json(force=True) or {}

    # UPDATE HEADER
    so_ref = (payload.get("sale_order_ref") or "").strip() or None
    pod_path = (payload.get("pod_file_path") or "").strip()
    pod_name = (payload.get("pod_file") or payload.get("pod_file_name") or "").strip()
    if not pod_path or not pod_name:
        cur.execute(
            "SELECT pod_file_path, pod_file_name FROM delivery_notes WHERE dn_id = %s",
            (dn_id,),
        )
        prev = cur.fetchone()
        if prev:
            if not pod_path:
                pod_path = (prev[0] or "").strip()
            if not pod_name:
                pod_name = (prev[1] or "").strip()
    if pod_path and not pod_name:
        pod_name = _upload_basename(pod_path)
    cur.execute("""
    UPDATE delivery_notes SET
    delivery_date=%s,
    so_id=%s,
    customer_name=%s,
    delivery_type=%s,
    destination_address=%s,
    delivery_by=%s,
    delivery_status=%s,
    vehicle_number=%s,
    tracking_id=%s,
    delivery_notes=%s,
    status=%s,
    received_by=%s,
    contact_number=%s,
    pod_file_name=%s,
    pod_file_path=%s,
    updated_at=NOW()
    WHERE dn_id=%s
    """, (
        payload.get("delivery_date"),
        so_ref,
        payload.get("customer_name"),
        payload.get("delivery_type"),
        payload.get("destination_address"),
        payload.get("delivery_by"),
        payload.get("delivery_status"),
        payload.get("vehicle_no"),
        payload.get("tracking_id"),
        payload.get("delivery_notes"),
        payload.get("status"),
        (payload.get("received_by") or "").strip(),
        (payload.get("contact_number") or "").strip(),
        pod_name or None,
        pod_path or None,
        dn_id
    ))

    # DELETE old items
    cur.execute("DELETE FROM delivery_note_items WHERE dn_id=%s", (dn_id,))

    # INSERT new items
    for it in payload.get("items", []):
        cur.execute("""
        INSERT INTO delivery_note_items (
            dn_id, product_id, product_name, qty, uom, serial_no
        ) VALUES (%s,%s,%s,%s,%s,%s)
        """, (
            dn_id,
            it.get("product_id"),
            it.get("product_name"),
            it.get("qty"),
            it.get("uom"),
            it.get("serial_no", "")
        ))

    # DN delivery status → linked Sales Order status
    if so_ref:
        dn_ds = str(payload.get("delivery_status") or "").strip().lower().replace(" ", "_")
        so_norm = "LOWER(REPLACE(COALESCE(status, ''), ' ', '_'))"
        if dn_ds == "delivered":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'returned', 'delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Delivered", so_ref),
            )
        elif dn_ds == "partially_delivered":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'returned', 'delivered', 'partially_delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Partially Delivered", so_ref),
            )
        elif dn_ds == "returned":
            cur.execute(
                f"""
                UPDATE sales_orders SET status = %s
                WHERE so_id = %s
                  AND {so_norm} NOT IN (
                    'cancelled', 'delivered', 'partially_delivered',
                    'draft', 'in_transit'
                  )
                """,
                ("Partially Delivered", so_ref),
            )

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True, "message": "Delivery Note Updated"})


# =========================================
# DELIVERY NOTE - POD / Customer acknowledgement
# =========================================

@app.route("/api/delivery-notes/<dn_id>/acknowledgement", methods=["POST"])
def delivery_note_acknowledgement(dn_id):
    """Upload POD to S3 deliverynote_attachments/{dn_id}/ and save acknowledgement fields."""
    dn_id = (dn_id or "").strip()
    if not dn_id:
        return jsonify({"success": False, "message": "dn_id is required"}), 400

    received_by = (request.form.get("received_by") or "").strip()
    contact_number = (request.form.get("contact_number") or "").strip()
    file = request.files.get("file")

    if not received_by:
        return jsonify({"success": False, "message": "Received By is required"}), 400
    if not contact_number:
        return jsonify({"success": False, "message": "Contact Number is required"}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_delivery_note_ack_columns(cur)
        cur.execute(
            "SELECT pod_file_path FROM delivery_notes WHERE dn_id = %s",
            (dn_id,),
        )
        row = cur.fetchone()
        dn_exists = row is not None
        old_path = (row[0] or "").strip() if row else ""

        pod_path = old_path or (request.form.get("pod_file_path") or "").strip()
        pod_name = (request.form.get("pod_file_name") or "").strip()
        if file and file.filename:
            if not _dn_pod_allowed(file.filename):
                return jsonify({
                    "success": False,
                    "message": "Invalid file format. Upload only PDF, JPG, or PNG.",
                }), 400
            try:
                file.stream.seek(0, os.SEEK_END)
                size = file.stream.tell()
                file.stream.seek(0)
            except Exception:
                size = 0
            if size > MAX_FILE_SIZE_BYTES:
                return jsonify({
                    "success": False,
                    "message": f"File size exceeds {MAX_FILE_SIZE_MB} MB",
                }), 400

            pod_name = _upload_basename(file.filename)
            rel_path = _upload_relative_path(dn_id, pod_name)
            pod_path, _ = _persist_module_upload(
                object_storage.MODULE_DELIVERY_NOTE_ATTACHMENTS,
                DELIVERY_NOTE_ATTACHMENTS_FOLDER,
                file,
                rel_path,
            )
            if old_path and old_path != pod_path:
                _remove_stored_upload(old_path, DELIVERY_NOTE_ATTACHMENTS_FOLDER)
        elif not pod_path:
            return jsonify({
                "success": False,
                "message": "Please upload a POD file (PDF, JPG, PNG).",
            }), 400
        elif not pod_name:
            if dn_exists:
                cur.execute(
                    "SELECT pod_file_name FROM delivery_notes WHERE dn_id = %s",
                    (dn_id,),
                )
                name_row = cur.fetchone()
                pod_name = (name_row[0] or "").strip() if name_row else ""
            if not pod_name:
                pod_name = _upload_basename(pod_path)

        if dn_exists:
            cur.execute(
                """
                UPDATE delivery_notes
                SET received_by = %s,
                    contact_number = %s,
                    pod_file_name = %s,
                    pod_file_path = %s,
                    updated_at = NOW()
                WHERE dn_id = %s
                """,
                (received_by, contact_number, pod_name or None, pod_path or None, dn_id),
            )
            conn.commit()

        return jsonify({
            "success": True,
            "message": "Acknowledgement saved",
            "pod_file_name": pod_name,
            "pod_file_path": pod_path,
            "pod_file": pod_name,
        })
    except Exception as e:
        conn.rollback()
        print(f"delivery_note_acknowledgement error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/delivery-notes/<dn_id>/pod/download", methods=["GET"])
def delivery_note_pod_download(dn_id):
    dn_id = (dn_id or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_delivery_note_ack_columns(cur)
        cur.execute(
            "SELECT pod_file_path, pod_file_name FROM delivery_notes WHERE dn_id = %s",
            (dn_id,),
        )
        row = cur.fetchone()
        if not row or not (row[0] or "").strip():
            return jsonify({"success": False, "message": "POD file not found"}), 404
        file_path = (row[0] or "").strip()
        original_name = (row[1] or "").strip() or _upload_basename(file_path)
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        full_path = _resolve_stored_file_path(file_path)
        if not full_path or not os.path.isfile(full_path):
            return jsonify({"success": False, "message": "File not found"}), 404
        return send_file(full_path, as_attachment=True, download_name=original_name)
    finally:
        cur.close()
        conn.close()


@app.route("/api/delivery-notes/<dn_id>/pod", methods=["DELETE"])
def delivery_note_pod_delete(dn_id):
    dn_id = (dn_id or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_delivery_note_ack_columns(cur)
        cur.execute(
            "SELECT pod_file_path FROM delivery_notes WHERE dn_id = %s",
            (dn_id,),
        )
        row = cur.fetchone()
        old_path = (row[0] or "").strip() if row else ""
        if not old_path:
            old_path = (request.args.get("path") or "").strip()
        if old_path:
            _remove_stored_upload(old_path, DELIVERY_NOTE_ATTACHMENTS_FOLDER)
        if row:
            cur.execute(
                """
                UPDATE delivery_notes
                SET pod_file_name = NULL, pod_file_path = NULL, updated_at = NOW()
                WHERE dn_id = %s
                """,
                (dn_id,),
            )
            conn.commit()
        elif not old_path:
            return jsonify({"success": False, "message": "POD file not found"}), 404
        return jsonify({"success": True, "message": "POD removed"})
    finally:
        cur.close()
        conn.close()


# =========================================
# DELIVERY NOTE - API (Cancel DN)
# =========================================

@app.put("/api/delivery-notes/<dn_id>/cancel")
def cancel_delivery_note(dn_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # check exists
    cur.execute("SELECT delivery_status, status FROM delivery_notes WHERE dn_id=%s", (dn_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return jsonify(success=False, message="Delivery Note not found"), 404

    current_status = (row[0] or row[1] or "").strip().lower().replace(" ", "_")

    if current_status == "cancelled":
        cur.close()
        conn.close()
        return jsonify(success=True, message="Already cancelled")

    payload = request.get_json(silent=True) or {}

    cur.execute("""
    UPDATE delivery_notes SET
        delivery_status=%s,
        status=%s,
        updated_at=NOW()
    WHERE dn_id=%s
    """, (
        "Cancelled",
        "Cancelled",
        dn_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify(success=True, message="Delivery Note cancelled successfully")


# =========================================
# DELIVERY NOTE - API (PDF) - REPORTLAB
# =========================================
@app.get("/api/delivery-notes/<dn_id>/pdf")
def delivery_note_pdf(dn_id):
    dn = get_dn_by_id(dn_id)

    if not dn:
        return jsonify({"success": False, "message": "Delivery Note not found"}), 404

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)

    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="{dn_id}.pdf"'
    return response


# =========================================
# DELIVERY NOTE - API (Email with PDF) - REPORTLAB
# =========================================
@app.post("/api/delivery-notes/<dn_id>/email")
def email_delivery_note(dn_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # FETCH HEADER
    cur.execute("SELECT * FROM delivery_notes WHERE dn_id=%s", (dn_id,))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return jsonify({"success": False, "message": "DN not found"}), 404

    columns = [desc[0] for desc in cur.description]
    dn = dict(zip(columns, row))

    # FETCH ITEMS
    cur.execute("""
    SELECT product_id, product_name, qty, uom, serial_no
    FROM delivery_note_items
    WHERE dn_id=%s
    """, (dn_id,))

    items_rows = cur.fetchall()

    items = []
    for i in items_rows:
        items.append({
            "product_id": i[0],
            "product_name": i[1],
            "qty": float(i[2]),
            "uom": i[3],
            "serial_no": i[4]
        })

    dn["items"] = items

    cur.close()
    conn.close()

    if not dn:
        return jsonify({"success": False, "message": "DN not found"}), 404

    # 1) email from DN
    customer_email = (dn.get("email") or "").strip()

    # 2) fallback customer master
    if not customer_email:
        customer = find_customer_by_name(dn.get("customer_name", ""))
        if customer:
            customer_email = (customer.get("email") or "").strip()

    # 3) fallback sales order
    if not customer_email:
        sale_order_ref = dn.get("sale_order_ref") or dn.get("so_id")
        if sale_order_ref:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute("""
                SELECT email
                FROM sales_orders
                WHERE so_id = %s
            """, (sale_order_ref,))

            row = cur.fetchone()

            cur.close()
            conn.close()

            if row:
                customer_email = (row[0] or "").strip()

    if not customer_email:
        return jsonify({"success": False, "message": "Customer email not available"}), 400

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)

    ok = send_email_with_attachments(
        to_email=customer_email,
        subject=f"Delivery Note {dn_id}",
        body=(
            f"Dear {dn.get('customer_name', 'Customer')},\n\n"
            f"Please find attached the delivery note ({dn_id}) issued against your sales order {dn.get('sale_order_ref') or dn.get('so_id') or ''}.\n"
            "The delivery details have been processed as per the information provided.\n\n"
            "Please let us know if you have any questions.\n\n"
            "Regards,\n"
            "Stackly Team"
        ),
        from_email=EMAIL_ADDRESS,
        password=EMAIL_PASSWORD,
        attachments=[
            {"filename": f"{dn_id}.pdf", "content_bytes": pdf_bytes}
        ],
    )

    if not ok:
        return jsonify({"success": False, "message": "Email failed. Check SMTP/App password/Spam."}), 500

    return jsonify({"success": True, "message": "Email sent"})


# =========================================
# DELIVERY NOTE - PRINT PAGE (Optional route)
# Uses same PDF generator and streams inline
# =========================================
@app.get("/delivery-note/<dn_id>/print")
def delivery_note_print(dn_id):
    dn = get_dn_by_id(dn_id)
    if not dn:
        return "DN not found", 404

    pdf_bytes = generate_delivery_note_pdf_bytes(dn)
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f'inline; filename="{dn_id}.pdf"'
    return response


# =========================================
# DELIVERY NOTE - PDF GENERATOR (REPORTLAB)
# =========================================
def generate_delivery_note_pdf_bytes(dn):
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()

    # ---------------------------------------------------
    # STYLES
    # ---------------------------------------------------
    company_style = ParagraphStyle(
        name="CompanyName",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#8c1f1f"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    company_info_style = ParagraphStyle(
        name="CompanyInfo",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=1,
    )

    page_title_style = ParagraphStyle(
        name="PageTitle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.green,
        alignment=TA_CENTER,
        spaceBefore=12,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        name="DNSection",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        spaceBefore=10,
    )

    label_style = ParagraphStyle(
        name="DNLabel",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#6b1a1a"),
    )

    value_style = ParagraphStyle(
        name="DNValue",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8.5,
        leading=11,
        textColor=colors.black,
    )

    summary_white_style = ParagraphStyle(
        name="DNSummaryWhite",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.white,
    )

    small_style = ParagraphStyle(
        name="DNSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#444444"),
    )

    header_small_style = ParagraphStyle(
        name="DNHeaderSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    terms_style = ParagraphStyle(
        name="TermsStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=11,
        textColor=colors.black,
        leftIndent=8,
    )

    elements = []

    # ---------------------------------------------------
    # SAFE HELPERS
    # ---------------------------------------------------
    def safe_str(val, default="-"):
        if val is None:
            return default
        s = str(val).strip()
        return s if s else default

    def safe_float(val, default=0.0):
        try:
            if val in (None, ""):
                return default
            return float(val)
        except Exception:
            return default

    # ---------------------------------------------------
    # COMPANY HEADER
    # ---------------------------------------------------
    elements.append(Paragraph("STACKLY", company_style))
    elements.append(Paragraph(
        "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
        company_info_style
    ))
    elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
    elements.append(Paragraph("Email: info@stackly.com", company_info_style))
    elements.append(Spacer(1, 10))

    # ---------------------------------------------------
    # PAGE TITLE
    # ---------------------------------------------------
    status_text = safe_str(dn.get("delivery_status") or dn.get("status") or "SUBMITTED").upper()
    elements.append(Paragraph(f"DELIVERY NOTE - {status_text}", page_title_style))
    elements.append(Spacer(1, 2))

    # ---------------------------------------------------
    # TOP DETAILS TABLE
    # ---------------------------------------------------
    dn_details_data = [
        [
            Paragraph("<b>Delivery Note Number:</b>", label_style),
            Paragraph(safe_str(dn.get("dn_id")), value_style),
            Paragraph("<b>Date:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_date")), value_style),
        ],
        [
            Paragraph("<b>Customer:</b>", label_style),
            Paragraph(safe_str(dn.get("customer_name")), value_style),
            Paragraph("<b>Sales Order Ref:</b>", label_style),
            Paragraph(safe_str(dn.get("sale_order_ref") or dn.get("so_id")), value_style),
        ],
        [
            Paragraph("<b>Delivery Type:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_type")), value_style),
            Paragraph("<b>Delivery By:</b>", label_style),
            Paragraph(safe_str(dn.get("delivery_by")), value_style),
        ],
        [
            Paragraph("<b>Vehicle Number:</b>", label_style),
            Paragraph(safe_str(dn.get("vehicle_no") or dn.get("vehicle_number")), value_style),
            Paragraph("<b>Tracking ID:</b>", label_style),
            Paragraph(safe_str(dn.get("tracking_id")), value_style),
        ],
        [
            Paragraph("<b>Destination Address:</b>", label_style),
            Paragraph(safe_str(dn.get("destination_address")), value_style),
            Paragraph("<b>Status:</b>", label_style),
            Paragraph(status_text, value_style),
        ],
    ]

    details_table = Table(dn_details_data, colWidths=[110, 170, 95, 145])
    details_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 16))

    # ---------------------------------------------------
    # DELIVERY NOTE ITEMS
    # ---------------------------------------------------
    elements.append(Paragraph("DELIVERY NOTE ITEMS", section_style))
    elements.append(Spacer(1, 2))

    items = dn.get("items", []) or []

    item_data = [[
        Paragraph("S.No", header_small_style),
        Paragraph("Product Name", header_small_style),
        Paragraph("Product ID", header_small_style),
        Paragraph("Qty", header_small_style),
        Paragraph("UOM", header_small_style),
        Paragraph("Serial No(s)", header_small_style),
    ]]

    for idx, item in enumerate(items, start=1):
        product_name = safe_str(item.get("product_name"))
        product_id = safe_str(item.get("product_id"))
        qty = safe_float(item.get("qty"), 0.0)
        uom = safe_str(item.get("uom"))
        serial_no = safe_str(
            item.get("serial_no")
            or item.get("serial_nos")
            or item.get("serial_numbers"),
            ""
        )

        item_data.append([
            Paragraph(str(idx), value_style),
            Paragraph(product_name, value_style),
            Paragraph(product_id, value_style),
            Paragraph(f"{qty:.2f}".rstrip("0").rstrip("."), value_style),
            Paragraph(uom, value_style),
            Paragraph(serial_no, value_style),
        ])

    if len(item_data) == 1:
        item_data.append(["-", "No line items available", "-", "-", "-", "-"])

    items_table = Table(
        item_data,
        colWidths=[35, 170, 72, 42, 50, 124],
        repeatRows=1
    )
    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 16))

    # ---------------------------------------------------
    # DELIVERY NOTES
    # ---------------------------------------------------
    delivery_notes = safe_str(dn.get("delivery_notes"), "").strip()
    if delivery_notes:
        elements.append(Paragraph("Delivery Notes", section_style))
        notes_table = Table([[Paragraph(delivery_notes, value_style)]], colWidths=[500])
        notes_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#999999")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(notes_table)
        elements.append(Spacer(1, 12))

    # ---------------------------------------------------
    # TERMS AND CONDITIONS
    # ---------------------------------------------------
    elements.append(Paragraph("Terms and Conditions", section_style))

    terms_list = [
        "1. This Delivery Note is issued based on the confirmed order details.",
        "2. Delivery will be made as per the agreed schedule.",
        "3. Kindly verify the delivered items at the time of receipt.",
        "4. Any shortage or damage should be reported immediately.",
        "5. Goods once delivered will be considered accepted unless otherwise notified.",
    ]

    for term in terms_list:
        elements.append(Paragraph(term, terms_style))

    # ---------------------------------------------------
    # BUILD PDF
    # ---------------------------------------------------
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

# ============================================
# INVOICE-LIST
# ============================================
@app.get("/invoice-list")
def invoice_list():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "invoice-list.html",
        page="invoice",
        title="Invoice  List - Stackly",
        user_email=user_email,
        user_name=user_name,
    )



@app.route("/get-invoice")
def get_invoice():
    # ✅ First update overdue invoices
    update_overdue_invoices()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            id,
            invoice_id,
            sale_order_ref,
            customer_name,
            invoice_date,
            due_date,
            payment_status,
            status
        FROM invoices
        ORDER BY id DESC
    """)
    
    rows = cursor.fetchall()
    
    data = []
    for r in rows:
        data.append({
            "id": r[0],
            "invoice_id": r[1],
            "sale_order_ref": r[2],
            "customer_name": r[3],
            "invoice_date": str(r[4]) if r[4] else "",
            "due_date": str(r[5]) if r[5] else "",
            "payment_status": r[6],
            "status": r[7]  # ← This will be 'Overdue' if updated
        })
    
    cursor.close()
    conn.close()
    
    return jsonify(data)


@app.route('/api/invoice/<invoice_id>', methods=['GET'])
def get_invoice_api(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # 1. Invoice header
        cur.execute("""
            SELECT
                invoice_id, sale_order_ref, invoice_date, due_date,
                invoice_status, payment_terms, customer_ref_no,
                customer_name, customer_id, billing_address,
                shipping_address, email, phone, contact_person,
                payment_method, currency, payment_ref_no,
                transaction_date, payment_status, amount_paid,
                status, invoice_tags, terms_conditions
            FROM invoices
            WHERE invoice_id = %s
        """, (invoice_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Invoice not found'}), 404

        def safe_date(value):
            if value is None:
                return ''
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d')
            return str(value)

        invoice = {
            'invoice_id': row[0],
            'sale_order_ref': row[1] or '',
            'invoice_date': safe_date(row[2]),
            'due_date': safe_date(row[3]),
            'invoice_status': row[4] or '',
            'payment_terms': row[5] or '',
            'customer_ref_no': row[6] or '',
            'customer_name': row[7] or '',
            'customer_id': row[8] or '',
            'billing_address': row[9] or '',
            'shipping_address': row[10] or '',
            'email': row[11] or '',
            'phone': row[12] or '',
            'contact_person': row[13] or '',
            'payment_method': row[14] or '',
            'currency': row[15] or '',
            'payment_ref_no': row[16] or '',
            'transaction_date': safe_date(row[17]),
            'payment_status': row[18] or '',
            'amount_paid': float(row[19]) if row[19] else 0,
            'status': row[20] or '',
            'invoice_tags': row[21] or '',
            'terms_conditions': row[22] or ''
        }

        # 2. Invoice Items
        items = []
        cur.execute("""
            SELECT product_name, product_id, quantity, uom,
                   unit_price, tax_pct, disc_pct
            FROM invoice_items
            WHERE invoice_id = %s
        """, (invoice_id,))
        for item in cur.fetchall():
            qty = float(item[2] or 0)
            price = float(item[4] or 0)
            tax = float(item[5] or 0)
            disc = float(item[6] or 0)
            total = qty * price * (1 - disc/100) * (1 + tax/100)
            items.append({
                'product_name': item[0] or '',
                'product_id': item[1] or '',
                'quantity': qty,
                'uom': item[3] or '',
                'unit_price': price,
                'tax_pct': tax,
                'disc_pct': disc,
                'total': total
            })

        # 3. Invoice Summary (FIXED – includes all columns)
        summary = {}
        cur.execute("""
            SELECT
                sub_total, tax_total, grand_total, amount_paid, balance_due,
                COALESCE(shipping_charges, 0),
                COALESCE(rounding_adjustment, 0),
                COALESCE(global_discount_pct, 0)
            FROM invoice_summary
            WHERE invoice_id = %s
        """, (invoice_id,))
        summary_row = cur.fetchone()
        if summary_row:
            summary = {
                'sub_total': float(summary_row[0] or 0),
                'tax_total': float(summary_row[1] or 0),
                'grand_total': float(summary_row[2] or 0),
                'amount_paid': float(summary_row[3] or 0),
                'balance_due': float(summary_row[4] or 0),
                'shipping_charges': float(summary_row[5] or 0),
                'rounding_adjustment': float(summary_row[6] or 0),
                'global_discount': float(summary_row[7] or 0)   # frontend expects 'global_discount'
            }
        else:
            summary = {
                'sub_total': 0, 'tax_total': 0, 'grand_total': 0,
                'amount_paid': 0, 'balance_due': 0,
                'shipping_charges': 0, 'rounding_adjustment': 0, 'global_discount': 0
            }

        # 4. Comments
        comments = []
        cur.execute("""
            SELECT text, created_at
            FROM invoice_comments
            WHERE invoice_id = %s
            ORDER BY created_at
        """, (invoice_id,))
        for comment in cur.fetchall():
            comments.append({
                'text': comment[0] or '',
                'date': comment[1].strftime('%Y-%m-%d %H:%M') if comment[1] else ''
            })

        # 5. Attachments
        attachments = []
        cur.execute("""
            SELECT id, filename, file_path, uploaded_at
            FROM invoice_attachments
            WHERE invoice_id = %s
        """, (invoice_id,))
        for att in cur.fetchall():
            attachments.append({
                'id': att[0],
                'name': att[1] or '',
                'path': att[2] or '',
                'date': att[3].strftime('%Y-%m-%d %H:%M') if att[3] else ''
            })

        cur.close()
        conn.close()

        return jsonify({
            'invoice': invoice,
            'items': items,
            'summary': summary,
            'comments': comments,
            'attachments': attachments
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

@app.route('/api/invoice/<invoice_id>/status', methods=['PUT'])
def update_invoice_status(invoice_id):
    data = request.json
    new_status = data.get('status')
    
    if not new_status or new_status not in ['Draft', 'Sent', 'Paid', 'Cancelled', 'Overdue']:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400
    
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Update status
        cur.execute("UPDATE invoices SET status = %s WHERE invoice_id = %s", (new_status, invoice_id))
        
        # If marking as paid, also update payment_status
        if new_status == 'Paid':
            cur.execute("UPDATE invoices SET payment_status = 'Paid' WHERE invoice_id = %s", (invoice_id,))
            # Also clear overdue if any
            cur.execute("UPDATE invoices SET status = 'Paid' WHERE invoice_id = %s AND status = 'Overdue'", (invoice_id,))
        
        conn.commit()
        return jsonify({'success': True, 'message': f'Invoice status: {new_status} successfully'})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()



@app.route('/update-invoice/<invoice_id>', methods=['PUT'])
def update_invoice(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Helper for optional text fields
        def none_if_empty(val):
            return val if val and val.strip() else None

        # Check if invoice exists
        cur.execute("SELECT 1 FROM invoices WHERE invoice_id = %s", (invoice_id,))
        if not cur.fetchone():
            return jsonify({"success": False, "error": "Invoice not found"}), 404

        # SAFE numeric conversion
        amount_paid = float(request.form.get('amount_paid') or 0)
        sub_total   = float(request.form.get('sub_total') or 0)
        tax_total   = float(request.form.get('tax_total') or 0)
        grand_total = float(request.form.get('grand_total') or 0)
        shipping    = float(request.form.get('shipping_charges') or 0)
        rounding    = float(request.form.get('rounding_adjustment') or 0)
        global_discount = float(request.form.get('global_discount') or 0)

        # Update invoices table
        cur.execute("""
            UPDATE invoices SET
                sale_order_ref = %s,
                invoice_date = %s,
                due_date = %s,
                invoice_status = %s,
                payment_terms = %s,
                customer_ref_no = %s,
                customer_name = %s,
                customer_id = %s,
                billing_address = %s,
                shipping_address = %s,
                email = %s,
                phone = %s,
                contact_person = %s,
                payment_method = %s,
                currency = %s,
                payment_ref_no = %s,
                transaction_date = %s,
                payment_status = %s,
                amount_paid = %s,
                status = %s,
                invoice_tags = %s,
                terms_conditions = %s
            WHERE invoice_id = %s
        """, (
            none_if_empty(request.form.get('sale_order_reference')),
            none_if_empty(request.form.get('invoice_date')),
            none_if_empty(request.form.get('due_date')),
            none_if_empty(request.form.get('invoice_status')),
            none_if_empty(request.form.get('payment_terms')),
            none_if_empty(request.form.get('customer_ref_no')),
            none_if_empty(request.form.get('customer_name')),
            none_if_empty(request.form.get('customer_id')),
            none_if_empty(request.form.get('billing_address')),
            none_if_empty(request.form.get('shipping_address')),
            none_if_empty(request.form.get('email')),
            none_if_empty(request.form.get('phone')),
            none_if_empty(request.form.get('contact_person')),
            none_if_empty(request.form.get('payment_method')),
            none_if_empty(request.form.get('currency')),
            none_if_empty(request.form.get('payment_ref_no')),
            none_if_empty(request.form.get('transaction_date')),
            none_if_empty(request.form.get('payment_status')),
            amount_paid,
            none_if_empty(request.form.get('status')),
            none_if_empty(request.form.get('invoice_tags')),
            none_if_empty(request.form.get('terms_conditions')),
            invoice_id
        ))

        # Replace items
        cur.execute("DELETE FROM invoice_items WHERE invoice_id = %s", (invoice_id,))
        items_json = request.form.get('itemsData')
        if items_json:
            items = json.loads(items_json)
            for item in items:
                cur.execute("""
                    INSERT INTO invoice_items (
                        invoice_id, product_name, product_id,
                        quantity, uom, unit_price, tax_pct, disc_pct
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    invoice_id,
                    item.get('product_name'),
                    item.get('product_id'),
                    int(float(item.get('quantity', 0))),
                    item.get('uom'),
                    float(item.get('unit_price', 0)),
                    float(item.get('tax_pct', 0)),
                    float(item.get('disc_pct', 0))
                ))

        # Replace summary (delete old, insert new with all columns)
        cur.execute("DELETE FROM invoice_summary WHERE invoice_id = %s", (invoice_id,))
        balance_due = grand_total - amount_paid
        cur.execute("""
            INSERT INTO invoice_summary (
                invoice_id, sub_total, tax_total, grand_total, amount_paid, balance_due,
                shipping_charges, rounding_adjustment, global_discount_pct
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            invoice_id,
            sub_total,
            tax_total,
            grand_total,
            amount_paid,
            balance_due,
            shipping,
            rounding,
            global_discount
        ))

        # Add history entry
        cur.execute("""
            INSERT INTO invoice_history (id, invoice_id, action, details, user_name, timestamp)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (
            str(uuid.uuid4()),
            invoice_id,
            "Invoice Updated",
            f"Invoice {invoice_id} updated",
            "Admin",
            datetime.now()
        ))

        conn.commit()
        return jsonify({"success": True, "message": "Invoice updated successfully"})

    except Exception as e:
        conn.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ---------------------------------------------------------------------
# PDF GENERATION FUNCTION
# ---------------------------------------------------------------------

def generate_invoice_pdf_bytes(invoice, items, summary):
    """
    Generate PDF bytes for an invoice using the same styling as Delivery Note Return PDF.
    invoice: dict with invoice header data
    items: list of item dicts
    summary: dict with summary totals
    """
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=18,
    )

    styles = getSampleStyleSheet()

    # ---------- custom styles (matching DNR PDF) ----------
    company_style = ParagraphStyle(
        name="Invoice_CompanyName",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#8c1f1f"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    company_info_style = ParagraphStyle(
        name="Invoice_CompanyInfo",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=1,
    )

    page_title_style = ParagraphStyle(
        name="Invoice_PageTitle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.green,
        alignment=TA_CENTER,
        spaceBefore=12,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        name="Invoice_Section",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        spaceBefore=10,
    )

    label_style = ParagraphStyle(
        name="Invoice_Label",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#6b1a1a"),
    )

    value_style = ParagraphStyle(
        name="Invoice_Value",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8.5,
        leading=11,
        textColor=colors.black,
    )

    header_small_style = ParagraphStyle(
        name="Invoice_HeaderSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    terms_style = ParagraphStyle(
        name="Invoice_Terms",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=11,
        textColor=colors.black,
        leftIndent=8,
    )

    # helpers (safe conversion)
    def safe_str(val, default="-"):
        if val is None:
            return default
        s = str(val).strip()
        return s if s else default

    def safe_float(val, default=0.0):
        try:
            if val in (None, ""):
                return default
            return float(val)
        except Exception:
            return default

    elements = []

    # company header (same as DNR)
    elements.append(Paragraph("STACKLY", company_style))
    elements.append(
        Paragraph(
            "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
            company_info_style,
        )
    )
    elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
    elements.append(Paragraph("Email: info@stackly.com", company_info_style))
    elements.append(Spacer(1, 10))

    # status & watermark
    status_text = safe_str(invoice.get("status") or "DRAFT").upper()
    elements.append(Paragraph(f"INVOICE - {status_text}", page_title_style))

    if status_text in ['CANCELLED', 'OVERDUE']:
        watermark_text = "CANCELLED - FOR REFERENCE ONLY" if status_text == 'CANCELLED' else "OVERDUE - PAYMENT REQUIRED"
        watermark_color = colors.red if status_text == 'CANCELLED' else colors.HexColor("#FFA500")
        watermark_style = ParagraphStyle(
            "Invoice_Watermark",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=14,
            textColor=watermark_color,
            alignment=TA_CENTER,
            backColor=colors.HexColor("#f9f2f2") if status_text == 'CANCELLED' else colors.HexColor("#fff5e6"),
            spaceAfter=12,
            spaceBefore=6,
            leading=18,
        )
        elements.append(Paragraph(f"⚠️ {watermark_text} ⚠️", watermark_style))
        elements.append(Spacer(1, 6))

    # ---------- 1. INVOICE INFORMATION ----------
    elements.append(Paragraph("INVOICE INFORMATION", section_style))
    inv_data = [
        [
            Paragraph("<b>Invoice Number:</b>", label_style),
            Paragraph(safe_str(invoice.get("invoice_id")), value_style),
            Paragraph("<b>Invoice Date:</b>", label_style),
            Paragraph(safe_str(invoice.get("invoice_date")), value_style),
        ],
        [
            Paragraph("<b>Sale Order Ref:</b>", label_style),
            Paragraph(safe_str(invoice.get("sale_order_ref")), value_style),
            Paragraph("<b>Due Date:</b>", label_style),
            Paragraph(safe_str(invoice.get("due_date")), value_style),
        ],
        [
            Paragraph("<b>Invoice Status:</b>", label_style),
            Paragraph(safe_str(invoice.get("status")), value_style),
            Paragraph("<b>Payment Terms:</b>", label_style),
            Paragraph(safe_str(invoice.get("payment_terms")), value_style),
        ],
        [
            Paragraph("<b>Customer Ref No:</b>", label_style),
            Paragraph(safe_str(invoice.get("customer_ref_no")), value_style),
            Paragraph("<b>Invoice Tags:</b>", label_style),
            Paragraph(safe_str(invoice.get("invoice_tags")), value_style),
        ],
    ]
    inv_table = Table(inv_data, colWidths=[110, 160, 95, 135])
    inv_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(inv_table)
    elements.append(Spacer(1, 12))

    # ---------- 2. CUSTOMER INFORMATION ----------
    elements.append(Paragraph("CUSTOMER INFORMATION", section_style))
    currency_code = invoice.get("currency", "USD")
    currency_map = {
        'USD': '$', 'EUR': '€', 'GBP': '£', 'JPY': '¥', 'IND': '₹', 'INR': '₹',
        'SGD': 'S$', 'CAD': 'C$', 'AUD': 'A$', 'CHF': 'Fr', 'CNY': '¥'
    }
    currency_symbol = currency_map.get(currency_code, currency_code)
    cust_data = [
        [
            Paragraph("<b>Customer Name:</b>", label_style),
            Paragraph(safe_str(invoice.get("customer_name")), value_style),
            Paragraph("<b>Customer ID:</b>", label_style),
            Paragraph(safe_str(invoice.get("customer_id")), value_style),
        ],
        [
            Paragraph("<b>Email:</b>", label_style),
            Paragraph(safe_str(invoice.get("email")), value_style),
            Paragraph("<b>Phone:</b>", label_style),
            Paragraph(safe_str(invoice.get("phone")), value_style),
        ],
        [
            Paragraph("<b>Contact Person:</b>", label_style),
            Paragraph(safe_str(invoice.get("contact_person")), value_style),
            Paragraph("<b>Currency:</b>", label_style),
            Paragraph(f"{currency_code} ({currency_symbol})", value_style),
        ],
    ]
    cust_table = Table(cust_data, colWidths=[110, 170, 95, 145])
    cust_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(cust_table)
    elements.append(Spacer(1, 12))

    # ---------- 3. ADDRESS INFORMATION ----------
    if invoice.get("billing_address") or invoice.get("shipping_address"):
        elements.append(Paragraph("ADDRESS INFORMATION", section_style))
        addr_data = [
            [
                Paragraph("<b>Billing Address:</b>", label_style),
                Paragraph(safe_str(invoice.get("billing_address")), value_style),
                Paragraph("<b>Shipping Address:</b>", label_style),
                Paragraph(safe_str(invoice.get("shipping_address")), value_style),
            ],
        ]
        addr_table = Table(addr_data, colWidths=[110, 170, 95, 145])
        addr_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
                    ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(addr_table)
        elements.append(Spacer(1, 12))

    # ---------- 4. PAYMENT INFORMATION ----------
    elements.append(Paragraph("PAYMENT INFORMATION", section_style))
    pay_data = [
        [
            Paragraph("<b>Payment Method:</b>", label_style),
            Paragraph(safe_str(invoice.get("payment_method")), value_style),
            Paragraph("<b>Payment Status:</b>", label_style),
            Paragraph(safe_str(invoice.get("payment_status")), value_style),
        ],
        [
            Paragraph("<b>Payment Ref No:</b>", label_style),
            Paragraph(safe_str(invoice.get("payment_ref_no")), value_style),
            Paragraph("<b>Transaction Date:</b>", label_style),
            Paragraph(safe_str(invoice.get("transaction_date")), value_style),
        ],
        [
            Paragraph("<b>Amount Paid:</b>", label_style),
            Paragraph(f"{currency_symbol}{safe_float(invoice.get('amount_paid')):.2f}", value_style),
            Paragraph("<b>Balance Due:</b>", label_style),
            Paragraph(f"{currency_symbol}{safe_float(summary.get('balance_due')):.2f}", value_style),
        ],
    ]
    pay_table = Table(pay_data, colWidths=[110, 170, 95, 145])
    pay_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(pay_table)
    elements.append(Spacer(1, 16))

    # ---------- 5. INVOICE ITEMS ----------
    if items:
        elements.append(Paragraph("INVOICE ITEMS", section_style))
        item_header = [
            Paragraph("S.No", header_small_style),
            Paragraph("Product Name", header_small_style),
            Paragraph("Product ID", header_small_style),
            Paragraph("Qty", header_small_style),
            Paragraph("UOM", header_small_style),
            Paragraph("Unit Price", header_small_style),
            Paragraph("Tax %", header_small_style),
            Paragraph("Disc %", header_small_style),
            Paragraph("Total", header_small_style),
        ]
        item_data = [item_header]

        for idx, it in enumerate(items, 1):
            qty = safe_float(it.get("quantity"))
            unit_price = safe_float(it.get("unit_price"))
            tax_pct = safe_float(it.get("tax_pct"))
            disc_pct = safe_float(it.get("disc_pct"))
            total = safe_float(it.get("total"))

            item_data.append([
                Paragraph(str(idx), value_style),
                Paragraph(safe_str(it.get("product_name")), value_style),
                Paragraph(safe_str(it.get("product_id")), value_style),
                Paragraph(f"{qty:.2f}".rstrip('0').rstrip('.'), value_style),
                Paragraph(safe_str(it.get("uom")), value_style),
                Paragraph(f"{currency_symbol}{unit_price:.2f}", value_style),
                Paragraph(f"{tax_pct:.1f}%" if tax_pct > 0 else "-", value_style),
                Paragraph(f"{disc_pct:.1f}%" if disc_pct > 0 else "-", value_style),
                Paragraph(f"{currency_symbol}{total:.2f}", value_style),
            ])

        col_widths = [35, 118, 58, 38, 38, 60, 40, 40, 65]
        items_table = Table(item_data, colWidths=col_widths, repeatRows=1)
        items_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("ALIGN", (1, 1), (2, -1), "LEFT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(items_table)
        elements.append(Spacer(1, 16))

    # ---------- 6. TAX AND TOTALS SUMMARY ----------
    elements.append(Paragraph("TAX AND TOTALS SUMMARY", section_style))

    # Compute item-level discount total
    total_discount = 0.0
    for it in items:
        line_sub = safe_float(it.get("quantity")) * safe_float(it.get("unit_price"))
        disc_pct = safe_float(it.get("disc_pct"))
        total_discount += line_sub * (disc_pct / 100)

    sub_total = safe_float(summary.get("sub_total"))
    tax_total = safe_float(summary.get("tax_total"))
    shipping = safe_float(summary.get("shipping_charges"))
    rounding = safe_float(summary.get("rounding_adjustment"))
    grand_total = safe_float(summary.get("grand_total"))
    amount_paid = safe_float(summary.get("amount_paid"))
    balance_due = safe_float(summary.get("balance_due"))
    global_disc_pct = safe_float(summary.get("global_discount_pct"))
    global_disc_amt = sub_total * (global_disc_pct / 100) if global_disc_pct > 0 else 0

    summary_rows = []
    summary_rows.append([Paragraph("<b>Subtotal:</b>", label_style), Paragraph(f"{currency_symbol}{sub_total:.2f}", value_style)])
    if total_discount > 0:
        summary_rows.append([Paragraph("<b>Item Level Discount:</b>", label_style), Paragraph(f"-{currency_symbol}{total_discount:.2f}", value_style)])
    if global_disc_pct > 0:
        summary_rows.append([Paragraph(f"<b>Global Discount ({global_disc_pct:.1f}%):</b>", label_style), Paragraph(f"-{currency_symbol}{global_disc_amt:.2f}", value_style)])
    summary_rows.append([Paragraph("<b>Total Tax:</b>", label_style), Paragraph(f"{currency_symbol}{tax_total:.2f}", value_style)])
    summary_rows.append([Paragraph("<b>Shipping Charge:</b>", label_style), Paragraph(f"{currency_symbol}{shipping:.2f}", value_style)])
    if rounding != 0:
        sign = '+' if rounding > 0 else ''
        summary_rows.append([Paragraph("<b>Rounding Adjustment:</b>", label_style), Paragraph(f"{sign}{currency_symbol}{abs(rounding):.2f}", value_style)])
    summary_rows.append([Paragraph("<b>GRAND TOTAL:</b>", label_style), Paragraph(f"{currency_symbol}{grand_total:.2f}", value_style)])
    summary_rows.append([Paragraph("<b>Amount Paid:</b>", label_style), Paragraph(f"{currency_symbol}{amount_paid:.2f}", value_style)])
    summary_rows.append([Paragraph("<b>BALANCE DUE:</b>", label_style), Paragraph(f"{currency_symbol}{balance_due:.2f}", value_style)])

    summary_table = Table(summary_rows, colWidths=[180, 120])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -4), colors.HexColor("#fafafa")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("FONTNAME", (0, 0), (-1, -4), "DejaVuSans"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, -3), (-1, -3), colors.HexColor("#e6e6e6")),
                ("FONTNAME", (0, -3), (-1, -3), "DejaVuSans-Bold"),
                ("BACKGROUND", (0, 0), (-1, -4), colors.HexColor("#fafafa")),
                ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
                ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    # ---------- 7. TERMS AND CONDITIONS ----------
    elements.append(Paragraph("Terms and Conditions", section_style))
    # terms_text = invoice.get("terms_conditions", "")
    # if terms_text:
        # for line in terms_text.split("\n"):
            # if line.strip():
                #  elements.append(Paragraph(line.strip(), terms_style))   # <-- no bullet

    # else:
    default_terms = [
            "1. This invoice is valid until the due date mentioned above.",
            "2. Payment terms as agreed upon.",
            "3. Goods once sold will not be taken back.",
            "4. All taxes and duties as applicable.",
            "5. Please quote invoice number when making payment.",
            "6. Late payment may incur additional charges.",
    ]
    for term in default_terms:
        elements.append(Paragraph(term, terms_style))
    elements.append(Spacer(1, 14))

    # ---------- 8. FOOTER ----------
    footer_style = ParagraphStyle(
        name="Invoice_Footer",
        parent=styles["Normal"],
        fontSize=7.5,
        textColor=colors.HexColor("#555555"),
        alignment=TA_CENTER,
    )
    generated_on = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elements.append(Paragraph(f"Generated on: {generated_on}", footer_style))
    elements.append(Paragraph("This is a system generated invoice - valid without signature", footer_style))

    # build PDF
    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# ---------------------------------------------------------------------
# FLASK ROUTE
# ---------------------------------------------------------------------
@app.route('/invoice/<invoice_id>/pdf')
def invoice_pdf(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # ----- 1. Invoice header with ALL fields -----
        cur.execute("""
            SELECT 
                invoice_id, sale_order_ref, invoice_date, due_date,
                customer_name, customer_id, email, phone, contact_person,
                payment_method, currency, payment_ref_no, transaction_date,
                payment_status, amount_paid, status, invoice_tags,
                billing_address, shipping_address, customer_ref_no,
                payment_terms, terms_conditions
            FROM invoices 
            WHERE invoice_id = %s
        """, (invoice_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'error': f'Invoice {invoice_id} not found.'}), 404

        invoice = {
            'invoice_id': row[0] or '',
            'sale_order_ref': row[1] or '',
            'invoice_date': row[2].strftime('%Y-%m-%d') if row[2] else '',
            'due_date': row[3].strftime('%Y-%m-%d') if row[3] else '',
            'customer_name': row[4] or '',
            'customer_id': row[5] or '',
            'email': row[6] or '',
            'phone': row[7] or '',
            'contact_person': row[8] or '',
            'payment_method': row[9] or '',
            'currency': row[10] or 'USD',
            'payment_ref_no': row[11] or '',
            'transaction_date': row[12].strftime('%Y-%m-%d') if row[12] else '',
            'payment_status': row[13] or '',
            'amount_paid': float(row[14] or 0),
            'status': row[15] or '',
            'invoice_tags': row[16] or '',
            'billing_address': row[17] or '',
            'shipping_address': row[18] or '',
            'customer_ref_no': row[19] or '',
            'payment_terms': row[20] or '',
            'terms_conditions': row[21] or ''
        }

        # ----- 2. Items -----
        cur.execute("""
            SELECT product_name, product_id, quantity, uom,
                   unit_price, tax_pct, disc_pct
            FROM invoice_items WHERE invoice_id = %s
        """, (invoice_id,))
        
        items = []
        for r in cur.fetchall():
            qty = float(r[2] or 0)
            price = float(r[4] or 0)
            tax = float(r[5] or 0)
            disc = float(r[6] or 0)
            total = qty * price * (1 - disc/100) * (1 + tax/100)
            items.append({
                'product_name': r[0] or '',
                'product_id': r[1] or '',
                'quantity': qty,
                'uom': r[3] or '',
                'unit_price': price,
                'tax_pct': tax,
                'disc_pct': disc,
                'total': total
            })

        # ----- 3. Summary -----
        cur.execute("""
            SELECT sub_total, tax_total, grand_total, amount_paid, balance_due,
                   COALESCE(shipping_charges,0), COALESCE(rounding_adjustment,0),
                   COALESCE(global_discount_pct,0)
            FROM invoice_summary WHERE invoice_id = %s
        """, (invoice_id,))
        
        summary_row = cur.fetchone()
        summary = {}
        if summary_row:
            summary = {
                'sub_total': float(summary_row[0] or 0),
                'tax_total': float(summary_row[1] or 0),
                'grand_total': float(summary_row[2] or 0),
                'amount_paid': float(summary_row[3] or 0),
                'balance_due': float(summary_row[4] or 0),
                'shipping_charges': float(summary_row[5] or 0),
                'rounding_adjustment': float(summary_row[6] or 0),
                'global_discount_pct': float(summary_row[7] or 0)
            }
        else:
            summary = {
                'sub_total': 0, 'tax_total': 0, 'grand_total': 0, 'amount_paid': 0,
                'balance_due': 0, 'shipping_charges': 0, 'rounding_adjustment': 0,
                'global_discount_pct': 0
            }

        cur.close()
        conn.close()

        # ----- 4. Generate PDF -----
        pdf_bytes = generate_invoice_pdf_bytes(invoice, items, summary)

        # ----- 5. Return PDF -----
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        
        if invoice['status'].lower() == 'draft':
            response.headers['Content-Disposition'] = 'inline; filename="invoice_preview.pdf"'
        else:
            response.headers['Content-Disposition'] = f'attachment; filename="invoice_{invoice_id}.pdf"'
        
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def send_invoice_email(recipient_email, invoice_data, items, summary, custom_message=""):
    """
    Send invoice email with PDF attachment – exactly like the DNR message style.
    No summary in body – just a simple message.
    """
    # Generate PDF using your existing function
    pdf_bytes = generate_invoice_pdf_bytes(invoice_data, items, summary)

    invoice_id = invoice_data.get('invoice_id', 'N/A')
    customer_name = invoice_data.get('customer_name', 'Customer')

    # Plain text body – exactly like DNR email (no numbers, no bullets)
    text_body = f"""
Dear {customer_name},

Please find attached the invoice ({invoice_id}) for your recent purchase.

Please let us know if you have any questions.

Regards,
Stackly Team
"""

    # Minimal HTML version – same simple message
    html_body = f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family: Arial, sans-serif; color: #333;">
    <p>Dear {customer_name},</p>
    <p>Please find attached the invoice (<strong>{invoice_id}</strong>) for your recent purchase.</p>
    <p>Please let us know if you have any questions.</p>
    <br>
    <p>Regards,<br>Stackly Team</p>
</body>
</html>
"""

    # Build email
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"Invoice {invoice_id} from Stackly"
    msg['From'] = os.getenv("EMAIL_ADDRESS")
    msg['To'] = recipient_email

    msg_alternative = MIMEMultipart('alternative')
    msg_alternative.attach(MIMEText(text_body, 'plain'))
    msg_alternative.attach(MIMEText(html_body, 'html'))
    msg.attach(msg_alternative)

    pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
    pdf_attachment.add_header('Content-Disposition', 'attachment', filename=f"Invoice_{invoice_id}.pdf")
    msg.attach(pdf_attachment)

    # Send
    try:
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD"))
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Invoice email send error: {e}")
        return False

@app.route('/api/invoice/<invoice_id>/send-email', methods=['POST'])
def send_invoice_email_api(invoice_id):
    data = request.get_json()
    recipient = data.get('email')
    custom_message = data.get('message', '')

    if not recipient:
        return jsonify({'success': False, 'error': 'Recipient email required'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Fetch invoice header with ALL fields
        cur.execute("""
            SELECT 
                invoice_id, sale_order_ref, invoice_date, due_date,
                customer_name, customer_id, email, phone, contact_person,
                payment_method, currency, payment_ref_no, transaction_date,
                payment_status, amount_paid, status, invoice_tags,
                billing_address, shipping_address, customer_ref_no,
                payment_terms, terms_conditions
            FROM invoices 
            WHERE invoice_id = %s
        """, (invoice_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Invoice not found'}), 404

        invoice = {
            'invoice_id': row[0] or '',
            'sale_order_ref': row[1] or '',
            'invoice_date': row[2].strftime('%Y-%m-%d') if row[2] else '',
            'due_date': row[3].strftime('%Y-%m-%d') if row[3] else '',
            'customer_name': row[4] or '',
            'customer_id': row[5] or '',
            'email': row[6] or '',
            'phone': row[7] or '',
            'contact_person': row[8] or '',
            'payment_method': row[9] or '',
            'currency': row[10] or 'USD',
            'payment_ref_no': row[11] or '',
            'transaction_date': row[12].strftime('%Y-%m-%d') if row[12] else '',
            'payment_status': row[13] or '',
            'amount_paid': float(row[14] or 0),
            'status': row[15] or '',
            'invoice_tags': row[16] or '',
            'billing_address': row[17] or '',
            'shipping_address': row[18] or '',
            'customer_ref_no': row[19] or '',
            'payment_terms': row[20] or '',
            'terms_conditions': row[21] or ''
        }

        # Items
        cur.execute("""
            SELECT product_name, product_id, quantity, uom, 
                   unit_price, tax_pct, disc_pct
            FROM invoice_items 
            WHERE invoice_id = %s
        """, (invoice_id,))
        
        items = []
        for r in cur.fetchall():
            qty = float(r[2] or 0)
            price = float(r[4] or 0)
            tax = float(r[5] or 0)
            disc = float(r[6] or 0)
            total = qty * price * (1 - disc/100) * (1 + tax/100)
            items.append({
                'product_name': r[0] or '',
                'product_id': r[1] or '',
                'quantity': qty,
                'uom': r[3] or '',
                'unit_price': price,
                'tax_pct': tax,
                'disc_pct': disc,
                'total': total
            })

        # Summary
        cur.execute("""
            SELECT sub_total, tax_total, grand_total, amount_paid, balance_due,
                   COALESCE(shipping_charges,0), COALESCE(rounding_adjustment,0),
                   COALESCE(global_discount_pct,0)
            FROM invoice_summary 
            WHERE invoice_id = %s
        """, (invoice_id,))
        
        summary_row = cur.fetchone()
        summary = {}
        if summary_row:
            summary = {
                'sub_total': float(summary_row[0] or 0),
                'tax_total': float(summary_row[1] or 0),
                'grand_total': float(summary_row[2] or 0),
                'amount_paid': float(summary_row[3] or 0),
                'balance_due': float(summary_row[4] or 0),
                'shipping_charges': float(summary_row[5] or 0),
                'rounding_adjustment': float(summary_row[6] or 0),
                'global_discount_pct': float(summary_row[7] or 0)
            }
        else:
            summary = {
                'sub_total': 0, 'tax_total': 0, 'grand_total': 0, 'amount_paid': 0,
                'balance_due': 0, 'shipping_charges': 0, 'rounding_adjustment': 0,
                'global_discount_pct': 0
            }

        cur.close()
        conn.close()

        # Send email
        success = send_invoice_email(recipient, invoice, items, summary, custom_message)
        
        if success:
            return jsonify({'success': True, 'message': 'Email sent successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to send email. Check SMTP credentials.'}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route('/generate-invoice-return/<invoice_id>')
def generate_invoice_return(invoice_id):
    """
    Called from invoice list when user clicks "generate invoice return".
    - If a return already exists for this invoice, redirect to its edit page.
    - Otherwise, create a new draft return and redirect to its edit page.
    """
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT invoice_return_id, status
        FROM invoice_return
        WHERE invoice_id = %s
        ORDER BY created_at DESC
        LIMIT 1
    """, (invoice_id,))
    existing = cur.fetchone()
    cur.close()
    conn.close()

    if existing:
        existing_id = existing[0]
        return redirect(f'/new-invoice-return?edit_id={existing_id}')
    else:
        new_id = generate_invoice_return_id()
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO invoice_return (invoice_return_id, invoice_id, status, created_at)
            VALUES (%s, %s, 'Draft', NOW())
        """, (new_id, invoice_id))
        conn.commit()
        cur.close()
        conn.close()
        return redirect(f'/new-invoice-return?edit_id={new_id}')


@app.route('/api/get-invoice-for-return/<invoice_id>')
def api_get_invoice_for_return(invoice_id):
    """API endpoint to fetch invoice details for return form"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                i.invoice_id,
                i.customer_name,
                i.customer_id,
                i.email,
                i.phone,
                i.contact_person,
                i.customer_ref_no,
                i.invoice_date,
                i.due_date,
                i.status,
                i.payment_status
            FROM invoices i
            WHERE i.invoice_id = %s
        """, (invoice_id,))

        invoice_row = cur.fetchone()

        if not invoice_row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Invoice not found'}), 404

        cur.execute("""
            SELECT
                COALESCE(grand_total, 0) as grand_total,
                COALESCE(global_discount_pct, 0) as global_discount_pct
            FROM invoice_summary
            WHERE invoice_id = %s
            ORDER BY created_at DESC
            LIMIT 1
        """, (invoice_id,))

        summary_row = cur.fetchone()
        grand_total = float(summary_row[0] or 0) if summary_row else 0
        global_discount_pct = float(summary_row[1] or 0) if summary_row else 0

        cur.execute("""
            SELECT
                id, product_id, product_name, quantity, uom,
                unit_price, tax_pct, disc_pct
            FROM invoice_items
            WHERE invoice_id = %s
            ORDER BY id
        """, (invoice_id,))

        items_rows = cur.fetchall()

        items = []
        for item_row in items_rows:
            product_id = item_row[1]
            quantity = float(item_row[3] or 0)
            unit_price = float(item_row[5] or 0)

            cur.execute("""
                SELECT COALESCE(SUM(return_quantity), 0)
                FROM invoice_return_items iri
                INNER JOIN invoice_return ir ON ir.invoice_return_id = iri.invoice_return_id
                WHERE ir.invoice_id = %s
                AND iri.product_id = %s
                AND LOWER(COALESCE(ir.status, '')) != 'cancelled'
            """, (invoice_id, product_id))

            returned_qty = float(cur.fetchone()[0] or 0)
            returnable_qty = quantity - returned_qty

            items.append({
                'id': item_row[0],
                'product_name': item_row[2],
                'product_id': product_id,
                'quantity': quantity,
                'returnable_quantity': returnable_qty if returnable_qty > 0 else 0,
                'unit_price': unit_price,
                'tax_pct': float(item_row[6] or 0),
                'disc_pct': float(item_row[7] or 0),
                'uom': item_row[4] or 'Nos'
            })

        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'customer_name': invoice_row[1] or '',
            'customer_id': invoice_row[2] or 'Auto Generate',
            'email': invoice_row[3] or '',
            'phone': invoice_row[4] or '',
            'contact_person': invoice_row[5] or '',
            'customer_ref_no': invoice_row[6] or '',
            'summary': {
                'original_total': grand_total,
                'global_discount_pct': global_discount_pct
            },
            'items': items
        })

    except Exception as e:
        print(f"Error fetching invoice details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/invoice-statuses')
def get_invoice_statuses():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT status FROM invoices WHERE status IS NOT NULL AND status != '' ORDER BY status")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    statuses = [row[0] for row in rows]
    return jsonify(statuses)


@app.route('/api/payment-statuses')
def get_payment_statuses():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT payment_status FROM invoices WHERE payment_status IS NOT NULL AND payment_status != '' ORDER BY payment_status")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    statuses = [row[0] for row in rows]
    return jsonify(statuses)


# =========================
# GENERATE INVOICE ID
# =========================
def generate_invoice_id():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT invoice_id FROM invoices ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return "INV-001"

    num = int(row[0].split("-")[1]) + 1
    return f"INV-{num:03d}"

# =========================
# NEW INVOICE PAGE
# =========================
@app.route("/new-invoice")
def new_invoice():
    current_invoice_id = (request.args.get("invoice_id") or "").strip() or None
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT
            so.so_id,
            COALESCE(c.name, so.customer_name, '') AS customer_name
        FROM sales_orders so
        LEFT JOIN customers c
          ON so.customer_id = c.customer_id
        WHERE NOT EXISTS (
            SELECT 1
            FROM invoices i
            WHERE i.sale_order_ref = so.so_id
              AND (%s IS NULL OR i.invoice_id <> %s)
        )
        ORDER BY so.so_id DESC
        """,
        (current_invoice_id, current_invoice_id),
    )

    rows = cur.fetchall()

    sales_orders = []
    for r in rows:
        sales_orders.append({
            "so_id": r[0],
            "customer_name": r[1]
        })

    cur.close()
    conn.close()

    invoice_id = generate_invoice_id()
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "new-invoice.html",
        page="invoice",
        title="New-Invoice-Stackly",
        user_email=user_email,
        user_name=user_name,
        invoice_id=invoice_id,
        sales_orders=sales_orders


    )


# =========================
# GET SALES ORDER (FROM DB)
# =========================
@app.get("/api/sales-orders/available-for-invoice")
def api_sales_orders_available_for_invoice():
    """Sales Orders available for invoice creation.
    Excludes Draft and Cancelled orders, and those already referenced by other invoices.
    If invoice_id is provided, keeps that invoice's SO available for edit mode.
    """
    invoice_id = (request.args.get("invoice_id") or "").strip() or None

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
                so.so_id,
                COALESCE(c.name, so.customer_name, '') AS customer_name
            FROM sales_orders so
            LEFT JOIN customers c
              ON so.customer_id = c.customer_id
            WHERE so.status NOT IN ('Draft', 'Cancelled')
              AND NOT EXISTS (
                SELECT 1
                FROM invoices i
                WHERE i.sale_order_ref = so.so_id
                  AND (%s IS NULL OR i.invoice_id <> %s)
            )
            ORDER BY so.created_at DESC
            """,
            (invoice_id, invoice_id),
        )
        rows = cur.fetchall()
        orders = [{"so_id": r[0], "customer_name": r[1]} for r in rows]
        return jsonify({"success": True, "orders": orders})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route('/get-sales-order/<so_id>')
def get_sales_order(so_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Get all sales order data (including summary fields)
        cur.execute("""
            SELECT customer_name, customer_id, billing_address,
                   shipping_address, email, phone,
                   payment_method, currency, due_date,
                   subtotal, tax_total, grand_total,
                   global_discount, shipping_charges
            FROM sales_orders
            WHERE so_id = %s
        """, (so_id,))
        sale = cur.fetchone()
        if not sale:
            return jsonify({})

        # Get items
        cur.execute("""
            SELECT product_name, product_id, qty, uom, price, tax_pct, disc_pct
            FROM sales_order_items
            WHERE so_id = %s
        """, (so_id,))
        items = cur.fetchall()

        return jsonify({
            "customer_name": sale[0],
            "customer_id": sale[1],
            "billing_address": sale[2],
            "shipping_address": sale[3],
            "email": sale[4],
            "phone": sale[5],
            "payment_method": sale[6],
            "currency": sale[7],
            "due_date": str(sale[8]) if sale[8] else None,
            "subtotal": float(sale[9] or 0),
            "tax_total": float(sale[10] or 0),
            "grand_total": float(sale[11] or 0),
            "global_discount": float(sale[12] or 0),
            "shipping_charges": float(sale[13] or 0),
            "rounding": 0.0,   # add column if needed, or set default
            "items": [
                {
                    "product_name": i[0],
                    "product_id": i[1],
                    "quantity": float(i[2]),
                    "uom": i[3],
                    "unit_price": float(i[4]),
                    "tax_pct": float(i[5]),
                    "disc_pct": float(i[6])
                } for i in items
            ]
        })
    finally:
        cur.close()
        conn.close()  

@app.route('/save-invoice', methods=['POST'])
def save_invoice():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        invoice_id = request.form.get('invoice_id') or generate_invoice_id()

        # Helper for optional text fields (convert empty strings to None)
        def none_if_empty(val):
            return val if val and val.strip() else None

        # SAFE numeric conversion (empty string → 0)
        amount_paid = float(request.form.get('amount_paid') or 0)
        sub_total   = float(request.form.get('sub_total') or 0)
        tax_total   = float(request.form.get('tax_total') or 0)
        grand_total = float(request.form.get('grand_total') or 0)
        shipping    = float(request.form.get('shipping_charges') or 0)
        rounding    = float(request.form.get('rounding_adjustment') or 0)
        global_discount = float(request.form.get('global_discount') or 0)

        # Insert into invoices
        cur.execute("""
            INSERT INTO invoices (
                invoice_id, sale_order_ref, invoice_date, due_date,
                invoice_status, payment_terms, customer_ref_no,
                customer_name, customer_id, billing_address,
                shipping_address, email, phone, contact_person,
                payment_method, currency, payment_ref_no,
                transaction_date, payment_status, amount_paid, status,
                invoice_tags, terms_conditions
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            invoice_id,
            none_if_empty(request.form.get('sale_order_reference')),
            none_if_empty(request.form.get('invoice_date')),
            none_if_empty(request.form.get('due_date')),
            none_if_empty(request.form.get('invoice_status')),
            none_if_empty(request.form.get('payment_terms')),
            none_if_empty(request.form.get('customer_ref_no')),
            none_if_empty(request.form.get('customer_name')),
            none_if_empty(request.form.get('customer_id')),
            none_if_empty(request.form.get('billing_address')),
            none_if_empty(request.form.get('shipping_address')),
            none_if_empty(request.form.get('email')),
            none_if_empty(request.form.get('phone')),
            none_if_empty(request.form.get('contact_person')),
            none_if_empty(request.form.get('payment_method')),
            none_if_empty(request.form.get('currency')),
            none_if_empty(request.form.get('payment_ref_no')),
            none_if_empty(request.form.get('transaction_date')),
            none_if_empty(request.form.get('payment_status')),
            amount_paid,
            none_if_empty(request.form.get('status')),
            none_if_empty(request.form.get('invoice_tags')),
            none_if_empty(request.form.get('terms_conditions'))
        ))

        # Insert items
        items_json = request.form.get('itemsData')
        if items_json:
            items = json.loads(items_json)
            for item in items:
                cur.execute("""
                    INSERT INTO invoice_items (
                        invoice_id, product_name, product_id,
                        quantity, uom, unit_price, tax_pct, disc_pct
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    invoice_id,
                    item.get('product_name'),
                    item.get('product_id'),
                    int(float(item.get('quantity', 0))),
                    item.get('uom'),
                    float(item.get('unit_price', 0)),
                    float(item.get('tax_pct', 0)),
                    float(item.get('disc_pct', 0))
                ))

        # Insert summary with all columns
        balance_due = grand_total - amount_paid
        cur.execute("""
            INSERT INTO invoice_summary (
                invoice_id, sub_total, tax_total, grand_total, amount_paid, balance_due,
                shipping_charges, rounding_adjustment, global_discount_pct
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            invoice_id,
            sub_total,
            tax_total,
            grand_total,
            amount_paid,
            balance_due,
            shipping,
            rounding,
            global_discount
        ))

        # Insert history
        cur.execute("""
            INSERT INTO invoice_history (id, invoice_id, action, details, user_name, timestamp)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (
            str(uuid.uuid4()),
            invoice_id,
            "Invoice Created",
            f"Invoice {invoice_id} created",
            "Admin",
            datetime.now()
        ))

        conn.commit()
        status = request.form.get('status', 'Draft')
        return jsonify({
            "success": True,
            "message": f"Invoice saved as {status} successfully"
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)})

    finally:
        cur.close()
        conn.close()

# =========================
# COMMENTS
# =========================
@app.route('/api/invoice/<invoice_id>/comments', methods=['GET'])
def get_comments_invoice(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, text, author, created_at 
        FROM invoice_comments
        WHERE invoice_id=%s
        ORDER BY created_at DESC
    """, (invoice_id,))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    comments = []
    for r in rows:
        comments.append({
            "id": r[0],
            "text": r[1],
            "author": r[2],
            "created_at": r[3]
        })

    return jsonify({"comments": comments})

@app.route('/api/invoice/<invoice_id>/comments', methods=['POST'])
def add_comment_invoice(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()

    data = request.get_json()

    cur.execute("""
        INSERT INTO invoice_comments (id, invoice_id, text, author, created_at)
        VALUES (%s,%s,%s,%s,%s)
    """, (
        str(uuid.uuid4()),
        invoice_id,
        data.get("comment_text"),
        "Admin",
        datetime.now()
    ))

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

# =========================
# ATTACHMENTS
# =========================
@app.route('/api/invoice/<invoice_id>/attachments', methods=['GET'])
def get_attachments_invoice(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, filename, file_path, size, uploaded_at
        FROM invoice_attachments
        WHERE invoice_id=%s
        ORDER BY uploaded_at DESC
    """, (invoice_id,))

    rows = cur.fetchall()
    cur.close()
    conn.close()

    data = []
    for r in rows:
        data.append({
            "id": r[0],
            "filename": r[1],
            "file_path": r[2],
            "size": r[3],
            "uploaded_at": r[4]
        })

    return jsonify({"success": True, "attachments": data})

@app.route('/api/invoice/<invoice_id>/attachments', methods=['POST'])
def upload_attachment_invoice(invoice_id):
    file = request.files['file']

    filename = file.filename
    rel_path = _upload_relative_path(invoice_id, filename)
    path, sz = _persist_module_upload(
        object_storage.MODULE_INVOICE_ATTACHMENTS,
        INVOICE_ATTACHMENTS_FOLDER,
        file,
        rel_path,
    )
    stored = _upload_basename(filename)

    conn = get_db_connection()
    cur = conn.cursor()
    att_id = str(uuid.uuid4())

    cur.execute("""
        INSERT INTO invoice_attachments 
        (id, invoice_id, filename, stored_name, file_path, size, uploaded_by, uploaded_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        att_id,
        invoice_id,
        filename,
        stored,
        path,
        sz,
        "Admin",
        datetime.now()
    ))
    _purge_prior_same_name_files(
        cur,
        "invoice_attachments",
        "invoice_id",
        invoice_id,
        "filename",
        filename,
        "id",
        att_id,
        "file_path",
        INVOICE_ATTACHMENTS_FOLDER,
    )

    conn.commit()
    cur.close()
    conn.close()

    return jsonify({"success": True})

import mimetypes

# Keep your existing download route (as_attachment=True)
@app.route('/api/invoice/<invoice_id>/attachments/<id>/download')
def download_attachment_invoice(invoice_id, id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT file_path, filename FROM invoice_attachments WHERE id=%s", (id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return "Not found", 404
    resolved = _resolve_stored_file_path(row[0])
    if object_storage.is_remote_url(str(resolved or "")):
        return redirect(resolved)
    if not resolved or not os.path.isfile(resolved):
        return "Not found", 404
    return send_file(resolved, as_attachment=True, download_name=row[1])

# NEW: View route (as_attachment=False)
@app.route('/api/invoice/<invoice_id>/attachments/<id>/view')
def view_attachment_invoice(invoice_id, id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT file_path, filename FROM invoice_attachments WHERE id=%s", (id,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return "Not found", 404
    resolved = _resolve_stored_file_path(row[0])
    if object_storage.is_remote_url(str(resolved or "")):
        return redirect(resolved)
    if not resolved or not os.path.isfile(resolved):
        return "Not found", 404
    # Serve inline (browser will display images, PDFs, etc.)
    return send_file(resolved, as_attachment=False, download_name=row[1])


@app.route('/api/invoice/<invoice_id>/attachments/<attachment_id>', methods=['DELETE'])
def delete_invoice_attachment(invoice_id, attachment_id):
    """Delete attachment from both database and filesystem"""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Get file path before deletion
        cur.execute("""
            SELECT file_path, filename 
            FROM invoice_attachments
            WHERE id = %s AND invoice_id = %s
        """, (attachment_id, invoice_id))
        
        row = cur.fetchone()
        
        if not row:
            cur.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Attachment not found'}), 404
        
        file_path = row[0]
        filename = row[1]
        
        # Delete from database
        cur.execute("""
            DELETE FROM invoice_attachments
            WHERE id = %s AND invoice_id = %s
        """, (attachment_id, invoice_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        _remove_stored_upload(file_path, INVOICE_ATTACHMENTS_FOLDER)
        
        return jsonify({
            'success': True, 
            'message': f'Attachment {filename} deleted successfully'
        })
        
    except Exception as e:
        print(f"❌ Error deleting attachment: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/payment-terms')
def get_payment_terms():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT DISTINCT payment_terms 
            FROM customers 
            WHERE payment_terms IS NOT NULL AND payment_terms != ''
            ORDER BY payment_terms
        """)

        rows = cur.fetchall()

        terms = [r[0] for r in rows]

        return jsonify({
            'success': True,
            'terms': terms
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()


@app.route('/api/customer-by-name/<name>')
def get_customer_by_name(name):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, name, email, phone, billing_address, shipping_address, payment_terms
            FROM customers
            WHERE LOWER(name) = LOWER(%s)
            LIMIT 1
        """, (name,))

        row = cur.fetchone()

        if not row:
            return jsonify({'success': True, 'customer': None})

        customer = {
            'id': row[0],
            'name': row[1],
            'email': row[2],
            'phone': row[3],
            'billing_address': row[4],
            'shipping_address': row[5],
            'paymentTerms': row[6]   # ⚠️ IMPORTANT (match your JS)
        }

        return jsonify({
            'success': True,
            'customer': customer
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    finally:
        cur.close()
        conn.close()


@app.route('/api/customer/<string:customer_id>/payment-term')
def get_customer_payment_term(customer_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Use customer_id column (e.g., 'CUST-001') instead of numeric id
        cur.execute("""
            SELECT payment_terms
            FROM customers
            WHERE customer_id = %s
        """, (customer_id,))
        row = cur.fetchone()
        if not row or row[0] is None:
            return jsonify({'success': True, 'payment_term': None})
        return jsonify({'success': True, 'payment_term': row[0]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route("/api/invoices", methods=["GET"])
def get_invoices():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT invoice_id, customer_name, invoice_date, total_amount, status
            FROM invoices
            ORDER BY created_at DESC
        """)

        rows = cur.fetchall()

        invoices = []
        for row in rows:
            invoices.append({
                "invoice_id": row[0],
                "customer_name": row[1],
                "invoice_date": str(row[2]),
                "total_amount": float(row[3]),
                "status": row[4]
            })

        cur.close()
        conn.close()

        return jsonify({"success": True, "invoices": invoices})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

def update_overdue_invoices():
    """Automatically update overdue invoices in database"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Update invoices that are overdue
        cur.execute("""
            UPDATE invoices
            SET status = 'Overdue'
            WHERE due_date < CURRENT_DATE
            AND payment_status != 'Paid'
            AND status NOT IN ('Paid', 'Cancelled', 'Overdue')
        """)
        
        updated_count = cur.rowcount
        conn.commit()
        
        if updated_count > 0:
            print(f"✅ Updated {updated_count} invoices to Overdue")
            
    except Exception as e:
        print(f"❌ Error updating overdue invoices: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()
    
    return updated_count




# =========================================
# INVOICE RETURN MODULE
# INVOICE-RETURN-LIST
# ========================================

@app.get("/invoice-return-list")
def invoice_return_list():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "invoice-return.html",
        page="inv_return",
        title="Invoice Return List - Stackly",
        user_email=user_email,
        user_name=user_name,
    )

@app.route('/api/invoice-returns')
def api_invoice_returns():
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        exclude_linked_for_dnr = (
            request.args.get("exclude_linked_for_dnr", "").strip() == "1"
        )
        if exclude_linked_for_dnr:
            cur.execute("""
                SELECT
                    ir.invoice_return_id,
                    ir.invoice_id,
                    ir.customer_name,
                    ir.return_date,
                    ir.status,
                    (
                        SELECT d.dnr_id
                        FROM deliverynote_returns d
                        WHERE TRIM(COALESCE(d.invoice_return_ref_id, ''))
                              = TRIM(ir.invoice_return_id)
                        LIMIT 1
                    ) AS linked_dnr_id
                FROM invoice_return ir
                ORDER BY ir.created_at DESC
            """)
        else:
            cur.execute("""
                SELECT 
                    invoice_return_id,
                    invoice_id,
                    customer_name,
                    return_date,
                    status
                FROM invoice_return
                ORDER BY created_at DESC
            """)
        rows = cur.fetchall()
        
        invoice_returns = []
        for row in rows:
            item = {
                "return_id": row[0],       # invoice_return_id
                "invoice_ref": row[1],     # invoice_id
                "customer_name": row[2],
                "return_date": row[3].strftime('%Y-%m-%d') if row[3] else "",
                "status": row[4]
            }
            if exclude_linked_for_dnr:
                item["linked_dnr_id"] = row[5] or ""
            invoice_returns.append(item)
        return jsonify(invoice_returns)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# =========================
# GENERATE NEW INVOICE RETURN ID
# =========================
def generate_invoice_return_id():
    conn = get_db_connection()
    cur = conn.cursor()
    # Use correct table name (singular)
    cur.execute("SELECT invoice_return_id FROM invoice_return ORDER BY created_at DESC LIMIT 1")
    row = cur.fetchone()
    cur.close()
    conn.close()
    if row and row[0]:
        last_id = row[0]
        # Assuming format IR-XXXX
        num = int(last_id.split('-')[1]) + 1
        return f"IR-{num:03d}"
    else:
        return "IR-001"

# =========================
# GET ALL INVOICES (for dropdown)
# =========================
def _invoice_has_any_returnable_qty(cur, invoice_id, exclude_return_id=None):
    """
    True if this invoice has at least one invoice_items line with remaining
    returnable quantity (same FIFO / pool rules as get_invoice_details).
    """
    if not invoice_id:
        return False
    cur.execute(
        """
        SELECT product_id, quantity, unit_price
        FROM invoice_items
        WHERE invoice_id = %s
        ORDER BY id
        """,
        (invoice_id,),
    )
    items_rows = cur.fetchall()
    if not items_rows:
        return False

    return_params = [invoice_id]
    ex_sql = ""
    if exclude_return_id:
        ex_sql = " AND ir.invoice_return_id <> %s "
        return_params.append(exclude_return_id)
    cur.execute(
        f"""
        SELECT iri.product_id, iri.unit_price, COALESCE(SUM(iri.return_quantity), 0)
        FROM invoice_return_items iri
        INNER JOIN invoice_return ir ON ir.invoice_return_id = iri.invoice_return_id
        WHERE ir.invoice_id = %s
          AND LOWER(COALESCE(ir.status, '')) <> 'cancelled'
          {ex_sql}
        GROUP BY iri.product_id, iri.unit_price
        """,
        tuple(return_params),
    )
    pool_left = defaultdict(float)
    for agg in cur.fetchall():
        pk = (agg[0] or "", round(float(agg[1] or 0), 6))
        pool_left[pk] += float(agg[2] or 0)

    def _pool_key(product_id, unit_price):
        return (product_id or "", round(float(unit_price or 0), 6))

    for row in items_rows:
        qf = float(row[1] or 0)
        pk = _pool_key(row[0], row[2])
        taken = min(qf, pool_left[pk])
        pool_left[pk] -= taken
        returnable = qf - taken
        if returnable > 0:
            return True
    return False


def get_all_invoices(existing_return_id=None):
    """
    Invoices eligible for a new return: at least one line still has returnable qty.
    Fully returned invoices are omitted.

    When editing/viewing an existing return, pass its invoice_return_id so return
    sums exclude this document, and the return's current invoice_id is always
    listed even if nothing else is returnable (edge: malformed data).
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        include_invoice_id = None
        if existing_return_id:
            cur.execute(
                "SELECT invoice_id FROM invoice_return WHERE invoice_return_id = %s",
                (existing_return_id,),
            )
            rr = cur.fetchone()
            if rr and rr[0]:
                include_invoice_id = rr[0]

        cur.execute(
            """
            SELECT invoice_id, customer_name
            FROM invoices
            ORDER BY created_at DESC
            """
        )
        rows = cur.fetchall()

        invoices = []
        for r in rows:
            inv_id = r[0]
            if include_invoice_id and inv_id == include_invoice_id:
                invoices.append({"invoice_id": inv_id, "customer_name": r[1] or ""})
                continue
            if _invoice_has_any_returnable_qty(
                cur, inv_id, exclude_return_id=existing_return_id
            ):
                invoices.append({"invoice_id": inv_id, "customer_name": r[1] or ""})
        return invoices
    finally:
        cur.close()
        conn.close()


@app.route("/new-invoice-return")
def new_invoice_return():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    # Read query parameters
    edit_id = request.args.get("edit_id")
    view_id = request.args.get("view_id")
    existing_id = edit_id or view_id

    # Determine invoice return ID for the hidden field
    invoice_return_id = existing_id if existing_id else generate_invoice_return_id()

    invoices = get_all_invoices(existing_return_id=existing_id)

    return render_template(
        "new-invoice-return.html",
        page="inv_return",
        title="Invoice Return",
        user_email=user_email,
        user_name=user_name,
        invoices=invoices,
        invoice_id=invoice_return_id
    )

# =========================
# FETCH INVOICE DETAILS (AJAX)
# =========================
@app.route("/get-invoice-details/<invoice_id>")
def get_invoice_details(invoice_id):
    conn = get_db_connection()
    cur = conn.cursor()

    # Optional: when editing an existing return, exclude its own lines from
    # "already returned" so remaining quantities are correct.
    exclude_return_id = (request.args.get("exclude_return_id") or "").strip() or None

    # =========================
    # 1. Invoice Main
    # =========================
    # Cast date columns to text (legacy rows may have years outside Python's 1..9999).
    cur.execute("""
        SELECT
            invoice_id,
            customer_name,
            customer_id,
            email,
            phone,
            contact_person,
            customer_ref_no,
            CASE
                WHEN invoice_date IS NULL THEN ''
                ELSE to_char(invoice_date, 'YYYY-MM-DD')
            END AS invoice_date,
            CASE
                WHEN due_date IS NULL THEN ''
                ELSE to_char(due_date, 'YYYY-MM-DD')
            END AS due_date
        FROM invoices
        WHERE invoice_id = %s
        ORDER BY created_at DESC LIMIT 1
    """, (invoice_id,))

    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"error": "Invoice not found"}), 404

    columns = [desc[0] for desc in cur.description]
    invoice = dict(zip(columns, row))

    # =========================
    # 2. Invoice Items — only lines with remaining returnable quantity
    # =========================
    cur.execute("""
        SELECT product_id, product_name, quantity, uom,
               unit_price, tax_pct, disc_pct
        FROM invoice_items
        WHERE invoice_id = %s
        ORDER BY id
    """, (invoice_id,))

    items_rows = cur.fetchall()

    # Pooled return quantities by (product_id, unit_price) across non-cancelled returns
    return_params = [invoice_id]
    ex_sql = ""
    if exclude_return_id:
        ex_sql = " AND ir.invoice_return_id <> %s "
        return_params.append(exclude_return_id)
    cur.execute(
        f"""
        SELECT iri.product_id, iri.unit_price, COALESCE(SUM(iri.return_quantity), 0)
        FROM invoice_return_items iri
        INNER JOIN invoice_return ir ON ir.invoice_return_id = iri.invoice_return_id
        WHERE ir.invoice_id = %s
          AND LOWER(COALESCE(ir.status, '')) <> 'cancelled'
          {ex_sql}
        GROUP BY iri.product_id, iri.unit_price
        """,
        tuple(return_params),
    )
    pool_left = defaultdict(float)
    for agg in cur.fetchall():
        pk = (agg[0] or "", round(float(agg[1] or 0), 6))
        pool_left[pk] += float(agg[2] or 0)

    def _pool_key(product_id, unit_price):
        return (product_id or "", round(float(unit_price or 0), 6))

    items = []
    for row in items_rows:
        qty = Decimal(row[2] or 0)
        price = Decimal(row[4] or 0)
        tax = Decimal(row[5] or 0)
        disc = Decimal(row[6] or 0)

        qf = float(qty)
        pk = _pool_key(row[0], row[4])
        taken = min(qf, pool_left[pk])
        pool_left[pk] -= taken
        returnable = qf - taken
        if returnable <= 0:
            continue

        # Calculate line total for display (full invoice line, not prorated)
        total = qty * price
        if disc > 0:
            total = total - (total * (disc / Decimal(100)))
        if tax > 0:
            total = total + (total * (tax / Decimal(100)))

        items.append({
            "product_id": row[0],
            "product_name": row[1],
            "quantity": qf,
            "returnable_quantity": float(returnable),
            "uom": row[3],
            "unit_price": float(price),
            "tax_pct": float(tax),
            "disc_pct": float(disc),
            "total": float(total)
        })

    all_line_items_returned = bool(items_rows) and not items

    # =========================
    # 3. Invoice Summary (simplified - only grand_total and global_discount_pct)
    # =========================
    cur.execute("""
        SELECT grand_total, global_discount_pct
        FROM invoice_summary
        WHERE invoice_id = %s
        ORDER BY created_at DESC LIMIT 1
    """, (invoice_id,))

    summary_row = cur.fetchone()

    if summary_row:
        grand_total = float(summary_row[0] or 0)
        global_discount_pct = float(summary_row[1] or 0)
        
        summary = {
            "original_total": grand_total,  # Original Grand Total from invoice
            "global_discount_pct": global_discount_pct,  # Global Discount % from invoice
            "subtotal": grand_total,  # Use grand_total as subtotal for now
            "discount_amount": 0,
            # "tax_amount":0,

            "rounding": 0,
            "refund_amount": 0
        }
    else:
        # Fallback if no summary found
        summary = {
            "original_total": 0,
            "global_discount_pct": 0,
            "subtotal": 0,
            "discount_amount": 0,
            # "tax_amount":0,
            "rounding": 0,
            "refund_amount": 0
        }

    # =========================
    # 4. Comments (optional - if needed)
    # =========================
    cur.execute("""
        SELECT text, created_at 
        FROM invoice_comments 
        WHERE invoice_id=%s 
        ORDER BY created_at DESC
    """, (invoice_id,))

    comments = []
    for c in cur.fetchall():
        comments.append({
            "text": c[0],
            "created_at": c[1].strftime("%Y-%m-%d %H:%M") if c[1] else ""
        })

    cur.close()
    conn.close()

    # =========================
    # 5. FINAL RESPONSE
    # =========================
    return jsonify({
        "invoice_id": invoice.get("invoice_id"),
        "customer_name": invoice.get("customer_name"),
        "customer_id": invoice.get("customer_id"),
        "email": invoice.get("email"),
        "phone": invoice.get("phone"),
        "contact_person": invoice.get("contact_person"),
        "customer_ref_no": invoice.get("customer_ref_no"),
        "items": items,
        "all_line_items_returned": all_line_items_returned,
        "summary": summary,
        "comments": comments
    })


def _invoice_grand_total_from_summary(cur, invoice_id):
    """
    Canonical 'Original Grand Total' for an invoice return: the linked invoice's
    grand_total from invoice_summary (latest row by created_at).

    invoice_return_summary.original_total can be wrong if an older client saved a
    bad payload; always prefer the source invoice total when present.
    """
    if not invoice_id:
        return None
    cur.execute(
        """
        SELECT grand_total
        FROM invoice_summary
        WHERE invoice_id = %s
        ORDER BY created_at DESC NULLS LAST
        LIMIT 1
        """,
        (invoice_id,),
    )
    row = cur.fetchone()
    if not row or row[0] is None:
        return None
    return float(row[0])


# =========================
# SAVE INVOICE RETURN
# =========================
from flask import request, jsonify
import traceback
@app.route("/save-invoice-return", methods=["POST"])
def save_invoice_return():
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        data = request.json
        invoice_return_id = data.get("invoice_return_id")

        # If no ID was sent, generate a new one
        if not invoice_return_id:
            invoice_return_id = generate_invoice_return_id()
            is_update = False
        else:
            # Check if the record already exists
            cur.execute("SELECT 1 FROM invoice_return WHERE invoice_return_id = %s", (invoice_return_id,))
            exists = cur.fetchone() is not None
            is_update = exists

        if is_update:
            # =========================
            # 1. UPDATE HEADER
            # =========================
            cur.execute("""
                UPDATE invoice_return
                SET invoice_id = %s,
                    customer_name = %s,
                    customer_id = %s,
                    email = %s,
                    phone = %s,
                    contact_person = %s,
                    customer_ref_no = %s,
                    return_date = %s,
                    refund_amount = %s,
                    status = %s
                WHERE invoice_return_id = %s
            """, (
                data.get("invoice_id"),
                data.get("customer_name"),
                data.get("customer_id"),
                data.get("email"),
                data.get("phone"),
                data.get("contact_person"),
                data.get("customer_ref_no"),
                data.get("return_date"),
                float(data.get("refund_amount", 0)),
                data.get("status"),
                invoice_return_id
            ))
            # Delete old items and summary (for update)
            cur.execute("DELETE FROM invoice_return_items WHERE invoice_return_id = %s", (invoice_return_id,))
            cur.execute("DELETE FROM invoice_return_summary WHERE invoice_return_id = %s", (invoice_return_id,))
        else:
            # =========================
            # 1. INSERT HEADER (for new record)
            # =========================
            cur.execute("""
                INSERT INTO invoice_return (
                    invoice_return_id,
                    invoice_id,
                    customer_name,
                    customer_id,
                    email,
                    phone,
                    contact_person,
                    customer_ref_no,
                    return_date,
                    refund_amount,
                    status
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                invoice_return_id,
                data.get("invoice_id"),
                data.get("customer_name"),
                data.get("customer_id"),
                data.get("email"),
                data.get("phone"),
                data.get("contact_person"),
                data.get("customer_ref_no"),
                data.get("return_date"),
                float(data.get("refund_amount", 0)),
                data.get("status")
            ))

        # =========================
        # 2. INSERT ITEMS (for both new and update)
        # =========================
        items = data.get("items", [])
        if not items:
            return jsonify({"error": "No items found"}), 400

        for item in items:
            cur.execute("""
                INSERT INTO invoice_return_items (
                    invoice_return_id,
                    product_id,
                    product_name,
                    invoice_quantity,
                    return_quantity,
                    serial_number,
                    return_reason,
                    uom,
                    unit_price,
                    tax_pct,
                    disc_pct,
                    total
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                invoice_return_id,
                item.get("product_id"),
                item.get("product_name"),
                float(item.get("invoice_quantity", 0)),
                float(item.get("quantity", 0)),
                item.get("serial_number", ""),
                item.get("return_reason", ""),
                item.get("uom"),
                float(item.get("unit_price", 0)),
                float(item.get("tax_pct", 0)),
                float(item.get("disc_pct", 0)),
                float(item.get("total", 0))
            ))

        # =========================
        # 3. INSERT SUMMARY (for both new and update)
        # =========================
        summary = data.get("summary", {})
        invoice_id_for_total = data.get("invoice_id")
        canonical_grand = _invoice_grand_total_from_summary(cur, invoice_id_for_total)
        original_total_val = (
            float(canonical_grand)
            if canonical_grand is not None
            else float(summary.get("original_total", 0))
        )
        cur.execute("""
            INSERT INTO invoice_return_summary (
                invoice_return_id,
                original_total,
                discount_pct,
                subtotal,
                discount_amount,
                rounding,
                refund_amount
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (
            invoice_return_id,
            original_total_val,
            float(summary.get("discount_pct", 0)),
            float(summary.get("subtotal", 0)),
            float(summary.get("discount_amount", 0)),
            float(summary.get("rounding", 0)),
            float(data.get("refund_amount", 0))
        ))

        conn.commit()

        return jsonify({
            "success": True,
            "invoice_return_id": invoice_return_id
        })

    except Exception as e:
        conn.rollback()
        print("❌ ERROR in save_invoice_return:", e)
        traceback.print_exc()
        return jsonify({
            "success": False,
            "message": "Failed to save invoice return",
            "error": str(e)
        }), 500

    finally:
        cur.close()
        conn.close()


def _ensure_invoice_return_comments_created_by_column(conn, cur):
    """Persist comment author display name; DDL committed separately so it survives read-only flows."""
    try:
        cur.execute(
            "ALTER TABLE invoice_return_comments ADD COLUMN IF NOT EXISTS created_by VARCHAR(255)"
        )
        conn.commit()
    except Exception as ex:
        conn.rollback()
        print("⚠️ invoice_return_comments.created_by ensure:", ex)


def _ensure_invoice_return_attachments_table(cur=None):
    """invoice_return_attachments + uploads/invoice_return_attachments/ (S3: invoice_return_attachments/)."""
    try:
        os.makedirs(INVOICE_RETURN_ATTACHMENTS_FOLDER, exist_ok=True)
    except OSError:
        pass
    own_conn = None
    own_cur = None
    if cur is None:
        own_conn = get_db_connection()
        own_cur = own_conn.cursor()
        cur = own_cur
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS invoice_return_attachments (
                id SERIAL PRIMARY KEY,
                invoice_return_id VARCHAR(50) NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_size BIGINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            ALTER TABLE invoice_return_attachments
            ADD COLUMN IF NOT EXISTS file_size BIGINT DEFAULT 0
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_invoice_return_attachments_ir_id
            ON invoice_return_attachments (invoice_return_id)
            """
        )
        cur.connection.commit()
    finally:
        if own_conn:
            own_cur.close()
            own_conn.close()


@app.route('/api/invoice-return/<invoice_return_id>/comments', methods=['GET'])
def get_comments_invoice_return(invoice_return_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        _ensure_invoice_return_comments_created_by_column(conn, cur)
        cur.execute("""
            SELECT id, comment, created_at, created_by
            FROM invoice_return_comments
            WHERE invoice_return_id=%s
            ORDER BY created_at DESC
        """, (invoice_return_id,))

        rows = cur.fetchall()

        comments = []
        for r in rows:
            created = r[2]
            author = (r[3] or "").strip() if len(r) > 3 else ""
            comments.append({
                "id": r[0],
                "text": r[1],
                "created_at": created.strftime("%Y-%m-%d %H:%M") if created else "",
                "author": author or "User",
            })

        return jsonify({"comments": comments})
    finally:
        cur.close()
        conn.close()


@app.route('/api/invoice-return/<invoice_return_id>/comments', methods=['POST'])
def add_comment_invoice_return(invoice_return_id):
    conn = get_db_connection()
    cur = conn.cursor()

    data = request.get_json()
    author_name = _get_logged_in_user_name()

    try:
        _ensure_invoice_return_comments_created_by_column(conn, cur)
        cur.execute("""
            INSERT INTO invoice_return_comments (invoice_return_id, comment, created_at, created_by)
            VALUES (%s,%s,%s,%s)
        """, (
            invoice_return_id,
            data.get("comment_text"),
            datetime.now(),
            author_name,
        ))

        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        print("❌ add_comment_invoice_return:", e)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@app.route('/api/invoice-return/<invoice_return_id>/attachments', methods=['GET'])
def get_attachments_invoice_return(invoice_return_id):
    invoice_return_id = (invoice_return_id or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_invoice_return_attachments_table(cur)
        cur.execute("""
            SELECT id, file_name, file_path, file_size, created_at
            FROM invoice_return_attachments
            WHERE invoice_return_id=%s
            ORDER BY created_at DESC
        """, (invoice_return_id,))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    data = []
    for r in rows:
        data.append({
            "id": r[0],
            "filename": r[1],
            "file_path": r[2],
            "size": r[3] if r[3] is not None else 0,
            "uploaded_at": r[4].strftime("%Y-%m-%d %H:%M") if r[4] else "",
        })
    return jsonify({"success": True, "attachments": data})


@app.route('/api/invoice-return/<invoice_return_id>/attachments', methods=['POST'])
def upload_attachment_invoice_return(invoice_return_id):
    conn = None
    cur = None
    try:
        invoice_return_id = (invoice_return_id or "").strip()
        if not invoice_return_id:
            return jsonify({"success": False, "message": "invoice_return_id is required"}), 400

        if 'file' not in request.files:
            return jsonify({"success": False, "message": "No file uploaded"}), 400

        file = request.files['file']
        if not file or file.filename == "":
            return jsonify({"success": False, "message": "Empty filename"}), 400

        original_name = _upload_basename(file.filename)
        file_type = file.mimetype or "application/octet-stream"
        file_size = _upload_file_size_bytes(file)

        if file_size <= 0:
            return jsonify({"success": False, "message": "Empty file cannot be uploaded"}), 400

        if file_size > DOC_UPLOAD_MAX_BYTES:
            return jsonify({
                "success": False,
                "message": f"File size exceeds {MAX_FILE_SIZE_MB} MB",
            }), 400

        if not _doc_upload_filename_allowed(original_name, file_type):
            return jsonify({
                "success": False,
                "message": "Invalid file format. Only PDF, JPEG, and PNG files are allowed.",
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_invoice_return_attachments_table(cur)

        cur.execute(
            "SELECT 1 FROM invoice_return WHERE invoice_return_id = %s",
            (invoice_return_id,),
        )
        if not cur.fetchone():
            conn.rollback()
            return jsonify({
                "success": False,
                "message": "Save the invoice return first, then attach files.",
            }), 400

        rel_path = _upload_relative_path(invoice_return_id, original_name)
        file_path, stored_size = _persist_module_upload(
            object_storage.MODULE_INVOICE_RETURN_ATTACHMENTS,
            INVOICE_RETURN_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )

        cur.execute("""
            INSERT INTO invoice_return_attachments (
                invoice_return_id, file_name, file_path, file_size, created_at
            ) VALUES (%s, %s, %s, %s, %s)
            RETURNING id, created_at
        """, (
            invoice_return_id,
            original_name,
            file_path,
            stored_size or file_size,
            datetime.now(),
        ))
        row = cur.fetchone()
        att_id = row[0] if row else None
        uploaded_at = row[1].strftime("%Y-%m-%d %H:%M") if row and row[1] else ""
        _purge_prior_same_name_files(
            cur,
            "invoice_return_attachments",
            "invoice_return_id",
            invoice_return_id,
            "file_name",
            original_name,
            "id",
            att_id,
            "file_path",
            INVOICE_RETURN_ATTACHMENTS_FOLDER,
        )

        conn.commit()

        return jsonify({
            "success": True,
            "message": "File uploaded successfully",
            "attachment": {
                "id": att_id,
                "filename": original_name,
                "file_path": file_path,
                "size": stored_size or file_size,
                "uploaded_at": uploaded_at,
            },
        })

    except Exception as e:
        if conn:
            conn.rollback()
        print("❌ upload_attachment_invoice_return:", e)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route('/api/invoice-return/<invoice_return_id>/attachments/<attachment_id>', methods=['DELETE'])
def delete_attachment_invoice_return(invoice_return_id, attachment_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_invoice_return_attachments_table(cur)

        cur.execute("""
            SELECT file_path 
            FROM invoice_return_attachments
            WHERE id=%s AND invoice_return_id=%s
        """, (attachment_id, invoice_return_id))

        row = cur.fetchone()

        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404

        file_path = row[0]

        _remove_stored_upload(file_path, INVOICE_RETURN_ATTACHMENTS_FOLDER)

        cur.execute("""
            DELETE FROM invoice_return_attachments
            WHERE id=%s AND invoice_return_id=%s
        """, (attachment_id, invoice_return_id))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Attachment deleted successfully"
        })

    except Exception as e:
        print("❌ Delete Error:", e)
        return jsonify({"success": False, "error": str(e)}), 500

# ===================================================
# DOWNLOAD (forces save to disk)
# ===================================================
@app.route('/api/invoice-return/<invoice_return_id>/attachments/<attachment_id>/download')
def download_invoice_return_attachment(invoice_return_id, attachment_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_invoice_return_attachments_table(cur)
        cur.execute("""
            SELECT file_path, file_name
            FROM invoice_return_attachments
            WHERE id = %s AND invoice_return_id = %s
        """, (attachment_id, invoice_return_id))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "Attachment not found"}), 404

        file_path = row[0]
        original_name = row[1]

        if object_storage.is_remote_url(str(file_path or "")):
            return redirect(file_path)
        resolved = _resolve_stored_file_path(file_path)
        if object_storage.is_remote_url(str(resolved or "")):
            return redirect(resolved)
        if not resolved or not os.path.exists(resolved):
            return jsonify({"error": "File not found on server"}), 404

        # as_attachment=True → forces download
        return send_file(resolved, as_attachment=True, download_name=original_name)

    except Exception as e:
        print("❌ Download error:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ===================================================
# VIEW (displays inline in browser)
# ===================================================
@app.route('/api/invoice-return/<invoice_return_id>/attachments/<attachment_id>/view')
def view_invoice_return_attachment(invoice_return_id, attachment_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_invoice_return_attachments_table(cur)
        cur.execute("""
            SELECT file_path, file_name
            FROM invoice_return_attachments
            WHERE id = %s AND invoice_return_id = %s
        """, (attachment_id, invoice_return_id))
        row = cur.fetchone()
        if not row:
            return jsonify({"error": "Attachment not found"}), 404

        file_path = row[0]
        original_name = row[1]

        if object_storage.is_remote_url(str(file_path or "")):
            return redirect(file_path)
        resolved = _resolve_stored_file_path(file_path)
        if object_storage.is_remote_url(str(resolved or "")):
            return redirect(resolved)
        if not resolved or not os.path.exists(resolved):
            return jsonify({"error": "File not found on server"}), 404

        # as_attachment=False → browser displays images/PDFs inline
        return send_file(resolved, as_attachment=False, download_name=original_name)

    except Exception as e:
        print("❌ View error:", e)
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()
        
@app.route("/api/invoice-return/<invoice_return_id>", methods=["GET"])
def get_invoice_return_data(invoice_return_id):
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        _ensure_invoice_return_comments_created_by_column(conn, cur)

        # 1. Header
        cur.execute("""
            SELECT 
                invoice_return_id,
                invoice_id,
                customer_name,
                customer_id,
                email,
                phone,
                contact_person,
                customer_ref_no,
                return_date,
                refund_amount,
                status
            FROM invoice_return
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Invoice return not found"}), 404
        
        invoice_return = {
            "invoice_return_id": row[0],
            "invoice_id": row[1],
            "customer_name": row[2],
            "customer_id": row[3],
            "email": row[4],
            "phone": row[5],
            "contact_person": row[6],
            "customer_ref_no": row[7],
            "return_date": row[8].strftime("%Y-%m-%d") if row[8] else "",
            "refund_amount": float(row[9]) if row[9] else 0,
            "status": row[10] or "Draft"
        }
        
        # 2. Items
        cur.execute("""
            SELECT 
                product_id,
                product_name,
                invoice_quantity,
                return_quantity,
                serial_number,
                return_reason,
                uom,
                unit_price,
                tax_pct,
                disc_pct,
                total
            FROM invoice_return_items
            WHERE invoice_return_id = %s
            ORDER BY id
        """, (invoice_return_id,))
        
        items = []
        for row in cur.fetchall():
            items.append({
                "product_id": row[0],
                "product_name": row[1],
                "invoice_quantity": float(row[2]) if row[2] else 0,
                "return_quantity": float(row[3]) if row[3] else 0,
                "serial_number": row[4] or "",
                "return_reason": row[5] or "",
                "uom": row[6] or "",
                "unit_price": float(row[7]) if row[7] else 0,
                "tax_pct": float(row[8]) if row[8] else 0,
                "disc_pct": float(row[9]) if row[9] else 0,
                "total": float(row[10]) if row[10] else 0
            })
        
        # 3. Comments
        cur.execute("""
            SELECT id, comment, created_at, created_by
            FROM invoice_return_comments
            WHERE invoice_return_id = %s
            ORDER BY created_at DESC
        """, (invoice_return_id,))
        
        comments = []
        for row in cur.fetchall():
            author = (row[3] or "").strip() if len(row) > 3 else ""
            comments.append({
                "id": row[0],
                "text": row[1],
                "created_at": row[2].strftime("%Y-%m-%d %H:%M:%S") if row[2] else "",
                "author": author or "User",
            })
        
        # 4. Attachments
        cur.execute("""
            SELECT id, file_name, file_path, created_at
            FROM invoice_return_attachments
            WHERE invoice_return_id = %s
            ORDER BY created_at DESC
        """, (invoice_return_id,))
        
        attachments = []
        for row in cur.fetchall():
            attachments.append({
                "id": row[0],
                "filename": row[1],
                "file_path": row[2],
                "uploaded_at": row[3].strftime("%Y-%m-%d %H:%M:%S") if row[3] else ""
            })
        
        # 5. Summary
        cur.execute("""
            SELECT original_total, discount_pct, subtotal, discount_amount, refund_amount
            FROM invoice_return_summary
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        
        summary_row = cur.fetchone()
        summary = {
            "original_total": float(summary_row[0]) if summary_row else 0,
            "discount_pct": float(summary_row[1]) if summary_row else 0,
            "subtotal": float(summary_row[2]) if summary_row else 0,
            "discount_amount": float(summary_row[3]) if summary_row else 0,
            "refund_amount": float(summary_row[4]) if summary_row else 0
        }
        inv_id = invoice_return.get("invoice_id")
        canonical = _invoice_grand_total_from_summary(cur, inv_id)
        if canonical is not None:
            summary["original_total"] = canonical
        
        return jsonify({
            "success": True,
            "invoice_return": invoice_return,
            "items": items,
            "comments": comments,
            "attachments": attachments,
            "summary": summary
        })
        
    except Exception as e:
        print("❌ Error fetching invoice return:", e)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route('/api/invoice-return/<return_id>/status', methods=['PUT'])
def update_invoice_return_status(return_id):
    data = request.json
    new_status = data.get('status')

    # Validate allowed statuses
    if not new_status or new_status not in ['draft', 'submitted', 'cancelled']:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Update the status
        cur.execute("UPDATE invoice_return SET status = %s WHERE invoice_return_id = %s", (new_status, return_id))

        # Check if any row was affected
        if cur.rowcount == 0:
            return jsonify({'success': False, 'error': 'Invoice return not found'}), 404

        conn.commit()
        return jsonify({'success': True, 'message': f'Status updated to {new_status}'})

    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()



# ==============================================
# PDF GENERATION (UPDATED)
# ==============================================
def generate_invoice_return_pdf_bytes(invoice_return, items, summary):
    """
    Generate PDF bytes for an invoice return with all columns.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=72, leftMargin=72,
                            topMargin=72, bottomMargin=72)

    elements = []
    styles = getSampleStyleSheet()

    # Currency symbol
    currency_symbol = '₹'

    # Custom styles (unchanged)
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2C3E50'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    company_style = ParagraphStyle(
        'Company',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=2,
        fontName='DejaVuSans'
    )
    status_style = ParagraphStyle(
        'Status',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        alignment=TA_CENTER,
        spaceAfter=14,
        fontName='DejaVuSans-Bold'
    )
    heading_style = ParagraphStyle(
        'Heading2',
        parent=styles['Heading2'],
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#2C3E50'),
        spaceBefore=8,
        spaceAfter=8,
        fontName='DejaVuSans-Bold'
    )
    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=7.6,
        leading=9,
        fontName='DejaVuSans',
        wordWrap='CJK'
    )
    terms_heading_style = ParagraphStyle(
        'TermsHeading',
        parent=styles['Heading2'],
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=6,
        fontName='DejaVuSans-Bold'
    )
    terms_style = ParagraphStyle(
        'Terms',
        parent=styles['Normal'],
        fontSize=7.4,
        leading=10,
        fontName='DejaVuSans'
    )
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=7.5,
        textColor=colors.HexColor('#555555'),
        alignment=TA_LEFT
    )

    # Company header
    elements.append(Paragraph("STACKLY", title_style))
    elements.append(Paragraph("MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008", company_style))
    elements.append(Paragraph("Phone: +91 7010792745", company_style))
    elements.append(Paragraph("Email: info@stackly.com", company_style))
    elements.append(Spacer(1, 10))

    # Status and watermark
    status_text = invoice_return.get('status', 'DRAFT').upper()
    status_color = {
        'DRAFT': colors.orange,
        'SUBMITTED': colors.blue,
        'CANCELLED': colors.red,
    }.get(status_text, colors.black)
    elements.append(Paragraph(f"INVOICE RETURN - {status_text}", status_style))

    if status_text == 'CANCELLED':
        elements.append(Paragraph(
            "⚠️ CANCELLED - FOR REFERENCE ONLY ⚠️",
            ParagraphStyle(
                'Watermark',
                parent=styles['Normal'],
                fontSize=16,
                textColor=colors.red,
                alignment=TA_CENTER,
                spaceAfter=20,
                spaceBefore=10,
                backColor=colors.lightgrey
            )
        ))
        elements.append(Spacer(1, 10))

    # ===========================================
    # SECTION 1: RETURN INFORMATION (unchanged)
    # =========================================
    elements.append(Paragraph("RETURN INFORMATION", heading_style))
    info_data = [
        ['Return Number:', invoice_return.get('invoice_return_id', '-'), 'Return Date:', invoice_return.get('return_date', '-')],
        ['Original Invoice:', invoice_return.get('invoice_id', '-'), 'Status:', invoice_return.get('status', '-')],
        ['Customer Ref No:', invoice_return.get('customer_ref_no', '-'), 'Refund Amount:', f"{currency_symbol}{invoice_return.get('refund_amount', 0):.2f}"],
    ]
    info_table = Table(info_data, colWidths=[120, 160, 100, 130])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,-1), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 15))

    # ===========================================
    # SECTION 2: CUSTOMER INFORMATION (unchanged)
    # =========================================
    elements.append(Paragraph("CUSTOMER INFORMATION", heading_style))
    customer_data = [
        ['Customer Name:', invoice_return.get('customer_name', '-'), 'Customer ID:', invoice_return.get('customer_id', '-')],
        ['Email:', invoice_return.get('email', '-'), 'Phone:', invoice_return.get('phone', '-')],
        ['Contact Person:', invoice_return.get('contact_person', '-'), 'Currency:', 'INR (₹)'],
    ]
    customer_table = Table(customer_data, colWidths=[120, 180, 100, 110])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,-1), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 1),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 15))

    # ===========================================
    # SECTION 3: RETURN ITEMS (UPDATED – added Invoice Qty column)
    # =========================================
    if items:
        elements.append(Paragraph("RETURN ITEMS", heading_style))

          # Define column widths that fit within ~450 points
        col_widths = [20, 60, 45, 40, 40, 35, 50, 45, 55, 55]
          # Sum = 20+60+45+40+40+35+50+45+55+55 = 445 (fits comfortably)

        table_data = [
        ['S.No', 'Product Name', 'Product ID', 'Invoice Qty', 'Return Qty',
         'UOM', 'Unit Price', 'Total', 'Serial Number', 'Return Reason']
        ]

    for idx, it in enumerate(items, 1):
        table_data.append([
            str(idx),
            it.get('product_name', '-'),
            it.get('product_id', '-'),
            f"{it.get('invoice_quantity', 0):.2f}",
            f"{it.get('return_quantity', 0):.2f}",
            it.get('uom', '-'),
            f"{currency_symbol}{it.get('unit_price', 0):.2f}",
            f"{currency_symbol}{it.get('total', 0):.2f}",
            it.get('serial_number', '-'),
            it.get('return_reason', '-')
        ])

    items_table = Table(table_data, colWidths=col_widths)
    items_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),                # slightly smaller font
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        # Center all header cells and most data cells
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # Right-align numeric columns (Unit Price, Total) for all rows
        ('ALIGN', (6, 1), (7, -1), 'RIGHT'),   # columns 6 and 7 (0-indexed)
        # Center other data columns (S.No, Product ID, Invoice Qty, Return Qty, UOM)
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (5, -1), 'CENTER'),
        # Left-align Product Name, Serial Number, Return Reason (they may wrap)
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (8, 1), (9, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 2),   # reduced padding from 4 to 2
        # Allow word wrapping for potentially long text columns
        ('WORDWRAP', (1, 1), (1, -1), True),    # Product Name
        ('WORDWRAP', (8, 1), (9, -1), True),    # Serial Number, Return Reason
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 20))

    # ===========================================
    # SECTION 4: SUMMARY (UPDATED – all columns)
    # =========================================
    elements.append(Paragraph("SUMMARY", heading_style))
    # 🆕 Include all summary fields (original_total, subtotal, discount_pct, discount_amount, rounding, refund_amount)
    summary_data = [
        ['Original Grand Total:', f"{currency_symbol}{summary.get('original_total', 0):.2f}"],
        ['Return Subtotal:', f"{currency_symbol}{summary.get('subtotal', 0):.2f}"],
        ['Global Discount (%) :', f"{summary.get('discount_pct', 0):.2f}%"],
        ['Global Discount Amount:', f"-{currency_symbol}{summary.get('discount_amount', 0):.2f}"],
        ['Rounding Adjustment:', f"{currency_symbol}{summary.get('rounding', 0):.2f}"],
        ['Amount to Refund:', f"{currency_symbol}{invoice_return.get('refund_amount', 0):.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,-1), (-1,-1), 'DejaVuSans-Bold'),
        ('FONTSIZE', (0,-1), (-1,-1), 11),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 5),
        ('LINEABOVE', (0,-1), (1,-1), 1, colors.black),
        ('BACKGROUND', (0,-1), (1,-1), colors.lightgrey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # ===========================================
    # SECTION 5: TERMS AND CONDITIONS (unchanged)
    # =========================================
    elements.append(Paragraph("Terms and Conditions", terms_heading_style))
    terms_text = invoice_return.get('terms_conditions', '')
    if terms_text:
        for line in terms_text.split('\n'):
            if line.strip():
                elements.append(Paragraph(f"• {line.strip()}", terms_style))
    else:
        default_terms = [
            "1. Return must be processed within 30 days of invoice date.",
            "2. All returned items must be in original condition with serial numbers intact.",
            "3. Refund will be processed after quality check.",
            "4. Please quote return number for any correspondence."
        ]
        for line in default_terms:
            elements.append(Paragraph(line, terms_style))
    elements.append(Spacer(1, 18))

    # Footer
    generated_on = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    elements.append(Paragraph(f"Generated on: {generated_on}", footer_style))
    elements.append(Paragraph("This is a system generated document - valid without signature", footer_style))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


# ==============================================
# EMAIL SENDING FUNCTION (UPDATED)
# ==============================================
def send_invoice_return_email(recipient_email, invoice_return, items, summary, custom_message=""):
    """
    Send invoice return email with HTML body and PDF attachment.
    HTML table now includes all columns (including Invoice Qty).
    """
    pdf_bytes = generate_invoice_return_pdf_bytes(invoice_return, items, summary)
    currency_symbol = '₹'

    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 0; padding: 20px; color: #333; }
            .container { max-width: 800px; margin: auto; background: #f9f9f9; border-radius: 8px; padding: 20px; }
            .header { text-align: center; border-bottom: 2px solid #a12828; padding-bottom: 10px; margin-bottom: 20px; }
            .logo { font-size: 28px; font-weight: bold; color: #a12828; }
            .success { background: #d4edda; color: #155724; padding: 10px; border-radius: 5px; margin: 15px 0; text-align: center; }
            .warning { background: #fff3cd; color: #856404; padding: 10px; border-radius: 5px; margin: 15px 0; text-align: center; }
            .info-section { margin: 20px 0; padding: 15px; background: white; border-radius: 5px; border: 1px solid #ddd; }
            .info-section h3 { margin-top: 0; color: #a12828; border-bottom: 1px solid #eee; padding-bottom: 8px; }
            .info-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
            .info-item { margin-bottom: 8px; }
            .info-label { font-weight: bold; color: #555; }
            .info-value { color: #333; }
            .items-table { width: 100%; border-collapse: collapse; margin: 15px 0; }
            .items-table th, .items-table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            .items-table th { background: #f2f2f2; }
            .summary { margin-top: 20px; text-align: right; }
            .footer { margin-top: 30px; font-size: 12px; text-align: center; color: #777; border-top: 1px solid #eee; padding-top: 15px; }
            .note { background: #fff3cd; padding: 10px; border-radius: 5px; margin: 15px 0; font-size: 13px; }
            .badge { display: inline-block; padding: 3px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; }
            .badge-draft { background: #e2e3e5; color: #383d41; }
            .badge-submitted { background: #d1ecf1; color: #0c5460; }
            .badge-cancelled { background: #f8d7da; color: #721c24; }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="logo">STACKLY</div>
                <div>MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008</div>
            </div>

            <div class="success">📄 Invoice Return Generated</div>

            <p>Hi {{ invoice_return.customer_name }},</p>
            <p>Your return request has been processed. Please find the attached PDF for details.</p>

            <!-- Return Information -->
            <div class="info-section">
                <h3>📦 Return Information</h3>
                <div class="info-grid">
                    <div class="info-item"><span class="info-label">Return Number:</span> <span class="info-value">{{ invoice_return.invoice_return_id }}</span></div>
                    <div class="info-item"><span class="info-label">Return Date:</span> <span class="info-value">{{ invoice_return.return_date }}</span></div>
                    <div class="info-item"><span class="info-label">Original Invoice:</span> <span class="info-value">{{ invoice_return.invoice_id }}</span></div>
                    <div class="info-item"><span class="info-label">Status:</span> <span class="info-value"><span class="badge badge-{{ invoice_return.status|lower }}">{{ invoice_return.status }}</span></span></div>
                    <div class="info-item"><span class="info-label">Customer Ref No:</span> <span class="info-value">{{ invoice_return.customer_ref_no or 'N/A' }}</span></div>
                    <div class="info-item"><span class="info-label">Refund Amount:</span> <span class="info-value">{{ currency_symbol }}{{ invoice_return.refund_amount|round(2) }}</span></div>
                </div>
            </div>

            <!-- Customer Information -->
            <div class="info-section">
                <h3>👤 Customer Information</h3>
                <div class="info-grid">
                    <div class="info-item"><span class="info-label">Name:</span> <span class="info-value">{{ invoice_return.customer_name }}</span></div>
                    <div class="info-item"><span class="info-label">Customer ID:</span> <span class="info-value">{{ invoice_return.customer_id or 'N/A' }}</span></div>
                    <div class="info-item"><span class="info-label">Email:</span> <span class="info-value">{{ invoice_return.email or 'N/A' }}</span></div>
                    <div class="info-item"><span class="info-label">Phone:</span> <span class="info-value">{{ invoice_return.phone or 'N/A' }}</span></div>
                    <div class="info-item"><span class="info-label">Contact Person:</span> <span class="info-value">{{ invoice_return.contact_person or 'N/A' }}</span></div>
                </div>
            </div>

            <!-- Return Items (UPDATED: added Invoice Qty column) -->
            <h3>🔄 Return Items</h3>
            <table class="items-table">
                <thead>
                    <tr>
                        <th>S.No</th>
                        <th>Product Name</th>
                        <th>Product ID</th>
                        <th>Invoice Qty</th>          <!-- 🆕 -->
                        <th>Return Qty</th>
                        <th>UOM</th>
                        <th>Unit Price</th>
                        <th>Total</th>
                        <th>Serial Number</th>
                        <th>Return Reason</th>
                    </tr>
                </thead>
                <tbody>
                    {% for it in items %}
                    <tr>
                        <td>{{ loop.index }}</td>
                        <td>{{ it.product_name }}</td>
                        <td>{{ it.product_id }}</td>
                        <td>{{ it.invoice_quantity|round(2) }}</td>      <!-- 🆕 -->
                        <td>{{ it.return_quantity|round(2) }}</td>
                        <td>{{ it.uom or '-' }}</td>
                        <td>{{ currency_symbol }}{{ it.unit_price|round(2) }}</td>
                        <td>{{ currency_symbol }}{{ it.total|round(2) }}</td>
                        <td>{{ it.serial_number or '-' }}</td>
                        <td>{{ it.return_reason or '-' }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>

            <!-- Summary (UPDATED: all fields) -->
            <div class="summary">
                <p><strong>Original Grand Total:</strong> {{ currency_symbol }}{{ summary.original_total|round(2) }}</p>
                <p><strong>Return Subtotal:</strong> {{ currency_symbol }}{{ summary.subtotal|round(2) }}</p>
                <p><strong>Global Discount ({{ summary.discount_pct }}%):</strong> -{{ currency_symbol }}{{ summary.discount_amount|round(2) }}</p>
                <p><strong>Rounding Adjustment:</strong> {{ currency_symbol }}{{ summary.rounding|round(2) }}</p>
                <p><strong>Amount to Refund:</strong> <span style="font-size: 18px; color: #a12828;">{{ currency_symbol }}{{ invoice_return.refund_amount|round(2) }}</span></p>
            </div>

            <div class="note">
                <strong>⚠️ Please Note:</strong><br>
                - The detailed return document is attached as a PDF.<br>
                - Return will be processed within 5-7 business days.<br>
                - Please quote return number for any inquiries.<br>
            </div>

            <div class="footer">
                <strong>STACKLY</strong><br>
                MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008<br>
                For any questions, contact us at <a href="mailto:support@stackly.com">support@stackly.com</a><br>
                Call: +91 7010792745<br>
                © 2028 Stackly. All rights reserved.
            </div>
        </div>
    </body>
    </html>
    """

    html_body = render_template_string(html_template,
                                       invoice_return=invoice_return,
                                       items=items,
                                       summary=summary,
                                       currency_symbol=currency_symbol,
                                       custom_message=custom_message)

    text_body = f"""
    Hi {invoice_return['customer_name']},

    Your return request {invoice_return['invoice_return_id']} has been processed.

    Return Details:
    - Return Number: {invoice_return['invoice_return_id']}
    - Original Invoice: {invoice_return['invoice_id']}
    - Return Date: {invoice_return['return_date']}
    - Status: {invoice_return['status']}
    - Refund Amount: {currency_symbol}{invoice_return['refund_amount']:.2f}

    Please find the attached PDF for complete details.

    Best regards,
    Stackly Team
    """

    # Build email
    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"Invoice Return {invoice_return['invoice_return_id']} from Stackly"
    msg['From'] = os.getenv("EMAIL_ADDRESS")
    msg['To'] = recipient_email

    msg_alternative = MIMEMultipart('alternative')
    msg_alternative.attach(MIMEText(text_body, 'plain'))
    msg_alternative.attach(MIMEText(html_body, 'html'))
    msg.attach(msg_alternative)

    pdf_attachment = MIMEApplication(pdf_bytes, _subtype='pdf')
    pdf_attachment.add_header('Content-Disposition', 'attachment', filename=f"Invoice_Return_{invoice_return['invoice_return_id']}.pdf")
    msg.attach(pdf_attachment)

    # Send email
    try:
        smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(os.getenv("EMAIL_ADDRESS"), os.getenv("EMAIL_PASSWORD"))
            server.send_message(msg)
        return True
    except Exception as e:
        print("Email send error:", e)
        return False


# ==============================================
# FLASK ROUTES (unchanged except for minor fixes)
# ==============================================
@app.route('/invoice-return/<invoice_return_id>/pdf')
def invoice_return_pdf(invoice_return_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Fetch return header
        cur.execute("""
            SELECT invoice_return_id, invoice_id, customer_name, customer_id,
                   email, phone, contact_person, customer_ref_no,
                   return_date, refund_amount, status
            FROM invoice_return
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'error': f'Invoice return {invoice_return_id} not found.'}), 404

        invoice_return = {
            'invoice_return_id': row[0] or '',
            'invoice_id': row[1] or '',
            'customer_name': row[2] or '',
            'customer_id': row[3] or '',
            'email': row[4] or '',
            'phone': row[5] or '',
            'contact_person': row[6] or '',
            'customer_ref_no': row[7] or '',
            'return_date': row[8].strftime('%Y-%m-%d') if row[8] else '',
            'refund_amount': float(row[9] or 0),
            'status': row[10] or '',
        }

        # Fetch items
        cur.execute("""
            SELECT product_id, product_name, invoice_quantity, return_quantity,
                   serial_number, return_reason, uom, unit_price, total
            FROM invoice_return_items
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        items = []
        for r in cur.fetchall():
            items.append({
                'product_id': r[0] or '',
                'product_name': r[1] or '',
                'invoice_quantity': float(r[2] or 0),
                'return_quantity': float(r[3] or 0),
                'serial_number': r[4] or '',
                'return_reason': r[5] or '',
                'uom': r[6] or '',
                'unit_price': float(r[7] or 0),
                'total': float(r[8] or 0),
            })

        # Fetch summary
        cur.execute("""
            SELECT original_total, discount_pct, subtotal, discount_amount,
                   rounding, refund_amount
            FROM invoice_return_summary
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        summary_row = cur.fetchone()
        summary = {}
        if summary_row:
            summary = {
                'original_total': float(summary_row[0] or 0),
                'discount_pct': float(summary_row[1] or 0),
                'subtotal': float(summary_row[2] or 0),
                'discount_amount': float(summary_row[3] or 0),
                'rounding': float(summary_row[4] or 0),
                'refund_amount': float(summary_row[5] or 0),
            }
        else:
            summary = {
                'original_total': 0,
                'discount_pct': 0,
                'subtotal': 0,
                'discount_amount': 0,
                'rounding': 0,
                'refund_amount': 0,
            }

        cur.close()
        conn.close()

        pdf_bytes = generate_invoice_return_pdf_bytes(invoice_return, items, summary)
        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename="invoice_return_{invoice_return_id}.pdf"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route('/api/invoice-return/<invoice_return_id>/email', methods=['POST'])
def send_invoice_return_email_api(invoice_return_id):
    data = request.get_json()
    recipient = data.get('email')
    custom_message = data.get('message', '')

    if not recipient:
        return jsonify({'success': False, 'error': 'Recipient email required'}), 400

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # Fetch return header (same as above)
        cur.execute("""
            SELECT invoice_return_id, invoice_id, customer_name, customer_id,
                   email, phone, contact_person, customer_ref_no,
                   return_date, refund_amount, status
            FROM invoice_return
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Return not found'}), 404

        invoice_return = {
            'invoice_return_id': row[0] or '',
            'invoice_id': row[1] or '',
            'customer_name': row[2] or '',
            'customer_id': row[3] or '',
            'email': row[4] or '',
            'phone': row[5] or '',
            'contact_person': row[6] or '',
            'customer_ref_no': row[7] or '',
            'return_date': row[8].strftime('%Y-%m-%d') if row[8] else '',
            'refund_amount': float(row[9] or 0),
            'status': row[10] or '',
        }

        # Fetch items
        cur.execute("""
            SELECT product_id, product_name, invoice_quantity, return_quantity,
                   serial_number, return_reason, uom, unit_price, total
            FROM invoice_return_items
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        items = []
        for r in cur.fetchall():
            items.append({
                'product_id': r[0] or '',
                'product_name': r[1] or '',
                'invoice_quantity': float(r[2] or 0),
                'return_quantity': float(r[3] or 0),
                'serial_number': r[4] or '',
                'return_reason': r[5] or '',
                'uom': r[6] or '',
                'unit_price': float(r[7] or 0),
                'total': float(r[8] or 0),
            })

        # Fetch summary
        cur.execute("""
            SELECT original_total, discount_pct, subtotal, discount_amount,
                   rounding, refund_amount
            FROM invoice_return_summary
            WHERE invoice_return_id = %s
        """, (invoice_return_id,))
        summary_row = cur.fetchone()
        summary = {}
        if summary_row:
            summary = {
                'original_total': float(summary_row[0] or 0),
                'discount_pct': float(summary_row[1] or 0),
                'subtotal': float(summary_row[2] or 0),
                'discount_amount': float(summary_row[3] or 0),
                'rounding': float(summary_row[4] or 0),
                'refund_amount': float(summary_row[5] or 0),
            }
        else:
            summary = {
                'original_total': 0,
                'discount_pct': 0,
                'subtotal': 0,
                'discount_amount': 0,
                'rounding': 0,
                'refund_amount': 0,
            }

        cur.close()
        conn.close()

        success = send_invoice_return_email(recipient, invoice_return, items, summary, custom_message)
        if success:
            return jsonify({'success': True, 'message': 'Email sent successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to send email'}), 500

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()



@app.route('/api/invoice-return/<return_id>/update-items-summary', methods=['PUT'])
def update_invoice_return_items_summary(return_id):
    """
    Update line items and summary for a draft invoice return.
    Expects JSON: { "items": [...], "summary": {...} }
    """
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        # 1. Check if return exists and status is 'draft'
        cur.execute("SELECT status FROM invoice_return WHERE invoice_return_id = %s", (return_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Invoice return not found'}), 404
        if row[0].lower() != 'draft':
            return jsonify({'success': False, 'error': 'Only draft returns can be updated'}), 403

        data = request.get_json()
        items = data.get('items', [])
        summary = data.get('summary', {})

        if not items:
            return jsonify({'success': False, 'error': 'No items provided'}), 400

        # 2. Replace items (delete old, insert new)
        cur.execute("DELETE FROM invoice_return_items WHERE invoice_return_id = %s", (return_id,))
        for item in items:
            cur.execute("""
                INSERT INTO invoice_return_items (
                    invoice_return_id,
                    product_id,
                    product_name,
                    invoice_quantity,
                    return_quantity,
                    serial_number,
                    return_reason,
                    uom,
                    unit_price,
                    tax_pct,
                    disc_pct,
                    total
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                return_id,
                item.get('product_id'),
                item.get('product_name'),
                float(item.get('invoice_quantity', 0)),
                float(item.get('quantity', 0)),
                item.get('serial_number', ''),
                item.get('return_reason', ''),
                item.get('uom'),
                float(item.get('unit_price', 0)),
                float(item.get('tax_pct', 0)),
                float(item.get('disc_pct', 0)),
                float(item.get('total', 0))
            ))

        # 3. Replace summary (delete old, insert new)
        cur.execute("DELETE FROM invoice_return_summary WHERE invoice_return_id = %s", (return_id,))
        cur.execute(
            "SELECT invoice_id FROM invoice_return WHERE invoice_return_id = %s",
            (return_id,),
        )
        inv_row = cur.fetchone()
        inv_id = inv_row[0] if inv_row else None
        canonical_grand = _invoice_grand_total_from_summary(cur, inv_id)
        original_total_val = (
            float(canonical_grand)
            if canonical_grand is not None
            else float(summary.get("original_total", 0))
        )
        cur.execute("""
            INSERT INTO invoice_return_summary (
                invoice_return_id,
                original_total,
                discount_pct,
                subtotal,
                discount_amount,
                rounding,
                refund_amount
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            return_id,
            original_total_val,
            float(summary.get("discount_pct", 0)),
            float(summary.get("subtotal", 0)),
            float(summary.get("discount_amount", 0)),
            float(summary.get("rounding", 0)),
            float(summary.get("refund_amount", 0))
        ))

        conn.commit()
        return jsonify({'success': True, 'message': 'Items and summary updated successfully'})

    except Exception as e:
        conn.rollback()
        print("❌ Error updating items/summary:", e)
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()


# =========================================
# DELIVERY NOTE RETURN LIST PAGE
# =========================================
@app.route("/deliverynote_return")
def dnr_page_list():

    return render_template(
        "deliverynote-return.html",
        page="deliverynote_return",
    )

# =========================================
# DELIVERY NOTE RETURN — schema
# =========================================

DNR_DATE_DISPLAY_FMT = "%d-%m-%Y"
DNR_DATE_INVALID_MSG = (
    "Invalid date. Use format DD-MM-YYYY (e.g. 31-05-2026)."
)
DNR_DATE_MUST_BE_TODAY_MSG = "DNR date must be today's date."


# =========================================
# DELIVERY NOTE RETURN LIST API
# =========================================
@app.route("/api/delivery-note-returns", methods=["GET"])
def get_delivery_note_returns():

    try:

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT *
            FROM deliverynote_returns
            ORDER BY
                dnr_date DESC NULLS LAST,
                CAST(
                    NULLIF(SPLIT_PART(dnr_id, '-', 2), '')
                    AS INTEGER
                ) DESC NULLS LAST
        """)

        rows = cur.fetchall()

        columns = [
            desc[0]
            for desc in cur.description
        ]

        result = []

        for row in rows:

            row_dict = dict(
                zip(columns, row)
            )

            dnr_date_raw = row_dict.get("dnr_date")
            if dnr_date_raw is None or dnr_date_raw == "":
                dnr_date_out = ""
            elif hasattr(dnr_date_raw, "strftime"):
                dnr_date_out = dnr_date_raw.strftime(DNR_DATE_DISPLAY_FMT)
            else:
                dnr_date_s = str(dnr_date_raw).strip()
                if not dnr_date_s:
                    dnr_date_out = ""
                elif re.match(r"^\d{2}-\d{2}-\d{4}$", dnr_date_s):
                    dnr_date_out = dnr_date_s
                else:
                    dnr_date_iso = dnr_date_s.split("T")[0]
                    if re.match(r"^\d{4}-\d{2}-\d{2}$", dnr_date_iso):
                        try:
                            dnr_date_out = (
                                datetime.strptime(dnr_date_iso, "%Y-%m-%d")
                                .date()
                                .strftime(DNR_DATE_DISPLAY_FMT)
                            )
                        except ValueError:
                            dnr_date_out = dnr_date_s
                    else:
                        dnr_date_out = dnr_date_s

            result.append({

                "dnr_id":
                    row_dict.get("dnr_id", ""),

                "invoice_return_ref":
                    row_dict.get("invoice_return_ref_id", ""),

                "customer_name":
                    row_dict.get("customer_name", ""),

                "dnr_date":
                    dnr_date_out,

                "status":
                    row_dict.get("status", "")

            })

        cur.close()
        conn.close()

        return jsonify(result)

    except Exception as e:

        print("DNR API ERROR:", e)

        return jsonify({
            "error": str(e)
        }), 500


# =========================================
# DELIVERY NOTE RETURN — single record
# =========================================

@app.route("/api/delivery-note-return/<dnr_id>", methods=["GET"])
def get_delivery_note_return_one(dnr_id):

    try:

        dnr_id = (dnr_id or "").strip()
        if not dnr_id:
            return jsonify({
                "success": False,
                "message": "DNR ID is required"
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            "SELECT * FROM deliverynote_returns WHERE dnr_id = %s",
            (dnr_id,),
        )
        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Delivery note return not found"
            }), 404

        columns = [desc[0] for desc in cur.description]
        header = dict(zip(columns, row))

        cur.execute("""
            SELECT
                product_id,
                product_name,
                uom,
                invoiced_qty,
                returned_qty,
                serial_no,
                return_reason
            FROM deliverynote_return_items
            WHERE dnr_id = %s
        """, (dnr_id,))

        items = []
        for item_row in cur.fetchall():
            items.append({
                "product_id": item_row[0] or "",
                "product_name": item_row[1] or "",
                "uom": item_row[2] or "",
                "invoiced_qty": item_row[3],
                "returned_qty": item_row[4],
                "serial_no": item_row[5] or "",
                "return_reason": item_row[6] or "",
            })

        comments = []
        try:
            cur.execute("""
                SELECT comment, created_by, created_at
                FROM deliverynote_return_comments
                WHERE dnr_id = %s
                ORDER BY created_at ASC
            """, (dnr_id,))
            for c_row in cur.fetchall():
                created_at = c_row[2]
                comments.append({
                    "comment": c_row[0] or "",
                    "created_by": (c_row[1] or "").strip() or "User",
                    "created_at": (
                        created_at.isoformat()
                        if hasattr(created_at, "isoformat")
                        else str(created_at or "")
                    ),
                })
        except Exception:
            pass

        history = []
        try:
            cur.execute("""
                SELECT action, description, created_by, created_at
                FROM deliverynote_return_history
                WHERE dnr_id = %s
                ORDER BY created_at DESC
            """, (dnr_id,))
            for h_row in cur.fetchall():
                created_at = h_row[3]
                history.append({
                    "action": h_row[0] or "",
                    "description": h_row[1] or "",
                    "created_by": (h_row[2] or "").strip() or "User",
                    "created_at": (
                        created_at.isoformat()
                        if hasattr(created_at, "isoformat")
                        else str(created_at or "")
                    ),
                })
        except Exception:
            pass

        attachments = []

        try:
            cur.execute(
                """
                SELECT
                    id,
                    file_name,
                    file_type,
                    file_size,
                    uploaded_at
                FROM deliverynote_return_attachments
                WHERE dnr_id = %s
                ORDER BY uploaded_at ASC NULLS LAST, id ASC
                """,
                (dnr_id,),
            )

            for att_row in cur.fetchall() or []:
                uploaded_at = att_row[4]

                if hasattr(uploaded_at, "strftime"):
                    uploaded_at_text = uploaded_at.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    uploaded_at_text = str(uploaded_at or "")

                attachments.append({
                    "id": att_row[0],
                    "filename": att_row[1] or "",
                    "file_type": att_row[2] or "",
                    "size": int(att_row[3] or 0),
                    "uploaded_at": uploaded_at_text,
                })
        except Exception:
            pass

        cur.close()
        conn.close()

        dnr_date_raw = header.get("dnr_date")
        if dnr_date_raw is None or dnr_date_raw == "":
            dnr_date_val = ""
        elif hasattr(dnr_date_raw, "strftime"):
            dnr_date_val = dnr_date_raw.strftime(DNR_DATE_DISPLAY_FMT)
        else:
            dnr_date_s = str(dnr_date_raw).strip()
            if not dnr_date_s:
                dnr_date_val = ""
            elif re.match(r"^\d{2}-\d{2}-\d{4}$", dnr_date_s):
                dnr_date_val = dnr_date_s
            else:
                dnr_date_iso = dnr_date_s.split("T")[0]
                if re.match(r"^\d{4}-\d{2}-\d{2}$", dnr_date_iso):
                    try:
                        dnr_date_val = (
                            datetime.strptime(dnr_date_iso, "%Y-%m-%d")
                            .date()
                            .strftime(DNR_DATE_DISPLAY_FMT)
                        )
                    except ValueError:
                        dnr_date_val = dnr_date_s
                else:
                    dnr_date_val = dnr_date_s

        return jsonify({
            "success": True,
            "data": {
                "dnr_id": header.get("dnr_id", ""),
                "status": header.get("status", ""),
                "invoice_return_ref_id": (
                    header.get("invoice_return_ref_id") or ""
                ),
                "customer_name": header.get("customer_name") or "",
                "customer_id": header.get("customer_id") or "",
                "email": header.get("email") or "",
                "phone": header.get("phone") or "",
                "contact_person": header.get("contact_person") or "",
                "customer_ref_no": header.get("customer_ref_no") or "",
                "cancel_reason": header.get("cancel_reason") or "",
                "dnr_date": dnr_date_val,
                "items": items,
                "comments": comments,
                "history": history,
                "attachments": attachments,
            },
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# =========================================
# DELIVERY NOTE RETURN NEW PAGE
# =========================================

@app.route("/deliverynote_return/new")
def dnr_new_page():

    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT COALESCE(MAX(
                CAST(NULLIF(SPLIT_PART(dnr_id, '-', 2), '') AS INTEGER)
            ), 0)
            FROM deliverynote_returns
            WHERE dnr_id ~ '^DNR-[0-9]+$'
            """
        )
        max_num = int((cur.fetchone() or [0])[0] or 0)
        dnr_id = f"DNR-{max_num + 1:03d}"
        conn.commit()
    finally:
        cur.close()
        conn.close()

    return render_template(
        "deliverynotereturn-new.html",
        page="deliverynote_return",
        dnr_id=dnr_id,
    )


@app.route("/api/delivery-note-return/next-id", methods=["GET"])
def api_next_dnr_id():

    try:

        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT COALESCE(MAX(
                    CAST(NULLIF(SPLIT_PART(dnr_id, '-', 2), '') AS INTEGER)
                ), 0)
                FROM deliverynote_returns
                WHERE dnr_id ~ '^DNR-[0-9]+$'
                """
            )
            max_num = int((cur.fetchone() or [0])[0] or 0)
            dnr_id = f"DNR-{max_num + 1:03d}"
            conn.commit()
        finally:
            cur.close()
            conn.close()

        return jsonify({
            "success": True,
            "dnr_id": dnr_id
        })

    except Exception as e:

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# =========================================
# DELIVERY NOTE RETURN — ATTACHMENTS (S3 / local uploads)
# =========================================

def _ensure_deliverynote_return_attachments_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS deliverynote_return_attachments (
            id SERIAL PRIMARY KEY,
            dnr_id TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_path TEXT,
            file_type TEXT,
            file_size BIGINT,
            file_content BYTEA,
            uploaded_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        ALTER TABLE deliverynote_return_attachments
        ADD COLUMN IF NOT EXISTS file_path TEXT
        """
    )
    cur.execute(
        """
        ALTER TABLE deliverynote_return_attachments
        ADD COLUMN IF NOT EXISTS file_type TEXT
        """
    )
    cur.execute(
        """
        ALTER TABLE deliverynote_return_attachments
        ADD COLUMN IF NOT EXISTS file_size BIGINT
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_dnr_attachments_dnr_id
        ON deliverynote_return_attachments (dnr_id)
        """
    )
    cur.connection.commit()


def _dnr_attachment_send(row, as_attachment):
    """Serve DNR attachment from S3/local path or legacy BYTEA in DB."""
    file_name = row[0] or "attachment"
    file_type = row[1] or "application/octet-stream"
    file_path = (row[2] or "").strip() if len(row) > 2 else ""
    file_data = row[3] if len(row) > 3 else None

    if file_path:
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        resolved = _resolve_stored_file_path(file_path)
        if resolved and os.path.isfile(resolved):
            return send_file(
                resolved,
                mimetype=file_type,
                download_name=file_name,
                as_attachment=as_attachment,
            )

    if file_data is None:
        return jsonify({
            "success": False,
            "message": "Attachment file not found",
        }), 404

    if isinstance(file_data, memoryview):
        file_bytes = file_data.tobytes()
    else:
        file_bytes = bytes(file_data)

    return send_file(
        BytesIO(file_bytes),
        mimetype=file_type,
        download_name=file_name,
        as_attachment=as_attachment,
    )


@app.post("/api/dnr-upload-attachment")
def dnr_upload_attachment():
    conn = None
    cur = None

    try:
        if "file" not in request.files:
            return jsonify({
                "success": False,
                "error": "No file provided"
            }), 400

        file = request.files["file"]
        dnr_id = (request.form.get("dnr_id") or "").strip()

        if not dnr_id:
            return jsonify({
                "success": False,
                "error": "dnr_id required"
            }), 400

        if file.filename == "":
            return jsonify({
                "success": False,
                "error": "No file selected"
            }), 400

        filename = _upload_basename(file.filename)
        file_type = file.mimetype or "application/octet-stream"

        file_size = _upload_file_size_bytes(file)

        if file_size <= 0:
            return jsonify({
                "success": False,
                "error": "Empty file cannot be uploaded"
            }), 400

        if file_size > DOC_UPLOAD_MAX_BYTES:
            return jsonify({
                "success": False,
                "error": f"File size exceeds {MAX_FILE_SIZE_MB} MB"
            }), 400

        if not _doc_upload_filename_allowed(filename, file_type):
            return jsonify({
                "success": False,
                "error": "Invalid file format. Only PDF, JPEG, and PNG files are allowed.",
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()

        _ensure_deliverynote_return_attachments_schema(cur)

        cur.execute("""
            SELECT status
            FROM deliverynote_returns
            WHERE dnr_id = %s
        """, (dnr_id,))

        dnr_row = cur.fetchone()

        if not dnr_row:
            conn.rollback()
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "Save the delivery note return first, then attach files."
            }), 400

        if str(dnr_row[0] or "").strip().lower() == "cancelled":
            conn.rollback()
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "Cannot attach files to a cancelled delivery note return."
            }), 400

        cur.execute("""
            SELECT COUNT(*)
            FROM deliverynote_return_attachments
            WHERE dnr_id = %s
        """, (dnr_id,))

        current_count = int((cur.fetchone() or [0])[0] or 0)

        if current_count >= 10:
            conn.rollback()
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "Maximum 10 files allowed per delivery note return"
            }), 400

        rel_path = _upload_relative_path(dnr_id, filename)
        stored_path, stored_size = _persist_module_upload(
            object_storage.MODULE_DELIVERY_NOTE_RETURN_ATTACHMENTS,
            DELIVERY_NOTE_RETURN_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )
        if stored_size and stored_size > 0:
            file_size = stored_size

        cur.execute("""
            INSERT INTO deliverynote_return_attachments (
                dnr_id,
                file_name,
                file_path,
                file_type,
                file_size,
                uploaded_at
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            dnr_id,
            filename,
            stored_path,
            file_type,
            file_size,
            datetime.now()
        ))

        new_row = cur.fetchone()
        conn.commit()

        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "attachment": {
                "id": new_row[0] if new_row else None,
                "filename": filename,
                "file_type": file_type,
                "size": file_size,
                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        })

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        if cur:
            try:
                cur.close()
            except Exception:
                pass

        if conn:
            try:
                conn.close()
            except Exception:
                pass

        print(traceback.format_exc())
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


@app.get("/api/dnr-attachments/<dnr_id>")
def dnr_get_attachments(dnr_id):
    conn = None
    cur = None

    try:
        dnr_id = (dnr_id or "").strip()

        if not dnr_id:
            return jsonify({
                "success": False,
                "attachments": []
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_deliverynote_return_attachments_schema(cur)

        cur.execute("""
            SELECT
                id,
                file_name,
                file_type,
                file_size,
                uploaded_at
            FROM deliverynote_return_attachments
            WHERE dnr_id = %s
            ORDER BY uploaded_at ASC NULLS LAST, id ASC
        """, (dnr_id,))

        attachments = []

        for row in cur.fetchall() or []:
            uploaded_at = row[4]

            if hasattr(uploaded_at, "strftime"):
                uploaded_at_text = uploaded_at.strftime("%Y-%m-%d %H:%M:%S")
            else:
                uploaded_at_text = str(uploaded_at or "")

            attachments.append({
                "id": row[0],
                "filename": row[1] or "",
                "file_type": row[2] or "",
                "size": int(row[3] or 0),
                "uploaded_at": uploaded_at_text
            })

        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "attachments": attachments
        })

    except Exception as e:
        if cur:
            try:
                cur.close()
            except Exception:
                pass

        if conn:
            try:
                conn.close()
            except Exception:
                pass

        return jsonify({
            "success": False,
            "attachments": [],
            "error": str(e)
        }), 500


@app.get("/api/dnr-attachment/<att_id>/view")
def dnr_view_attachment(att_id):
    conn = None
    cur = None

    try:
        att_id = int(str(att_id).strip())

        conn = get_db_connection()
        cur = conn.cursor()

        _ensure_deliverynote_return_attachments_schema(cur)
        cur.execute("""
            SELECT
                file_name,
                file_type,
                file_path,
                COALESCE(file_data, file_content)
            FROM deliverynote_return_attachments
            WHERE id = %s
        """, (att_id,))

        row = cur.fetchone()

        cur.close()
        conn.close()

        if not row:
            return jsonify({
                "success": False,
                "message": "Attachment not found"
            }), 404

        return _dnr_attachment_send(row, as_attachment=False)

    except Exception as e:
        if cur:
            try:
                cur.close()
            except Exception:
                pass

        if conn:
            try:
                conn.close()
            except Exception:
                pass

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.get("/api/dnr-attachment/<att_id>/download")
def dnr_download_attachment(att_id):
    conn = None
    cur = None

    try:
        att_id = int(str(att_id).strip())

        conn = get_db_connection()
        cur = conn.cursor()

        _ensure_deliverynote_return_attachments_schema(cur)
        cur.execute("""
            SELECT
                file_name,
                file_type,
                file_path,
                COALESCE(file_data, file_content)
            FROM deliverynote_return_attachments
            WHERE id = %s
        """, (att_id,))

        row = cur.fetchone()

        cur.close()
        conn.close()

        if not row:
            return jsonify({
                "success": False,
                "message": "Attachment not found"
            }), 404

        return _dnr_attachment_send(row, as_attachment=True)

    except Exception as e:
        if cur:
            try:
                cur.close()
            except Exception:
                pass

        if conn:
            try:
                conn.close()
            except Exception:
                pass

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.delete("/api/dnr-attachment/<att_id>")
def dnr_delete_attachment(att_id):
    conn = None
    cur = None

    try:
        att_id = int(str(att_id).strip())

        conn = get_db_connection()
        cur = conn.cursor()

        _ensure_deliverynote_return_attachments_schema(cur)
        cur.execute("""
            SELECT
                a.id,
                a.dnr_id,
                r.status,
                a.file_path
            FROM deliverynote_return_attachments a
            JOIN deliverynote_returns r
              ON r.dnr_id = a.dnr_id
            WHERE a.id = %s
        """, (att_id,))

        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "Attachment not found"
            }), 404

        if str(row[2] or "").strip().lower() == "cancelled":
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "Cannot delete attachments on a cancelled record."
            }), 400

        file_path = (row[3] or "").strip()
        if file_path:
            _remove_stored_upload(file_path, DELIVERY_NOTE_RETURN_ATTACHMENTS_FOLDER)

        cur.execute("""
            DELETE FROM deliverynote_return_attachments
            WHERE id = %s
        """, (att_id,))

        conn.commit()

        cur.close()
        conn.close()

        return jsonify({
            "success": True
        })

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        if cur:
            try:
                cur.close()
            except Exception:
                pass

        if conn:
            try:
                conn.close()
            except Exception:
                pass

        print(traceback.format_exc())
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

# =========================================
# SAVE DELIVERY NOTE RETURN
# =========================================

@app.post("/api/save-delivery-note-return")
def save_delivery_note_return():

    conn = None

    try:

        data = request.get_json() or {}

        dnr_id = (data.get("dnr_id") or "").strip()
        if not dnr_id:
            return jsonify({
                "success": False,
                "message": "DNR ID is required"
            }), 400

        status = (data.get("status") or "Draft").strip()
        if status not in ("Draft", "Submitted"):
            return jsonify({
                "success": False,
                "message": "Invalid status. Use Draft or Submitted."
            }), 400

        invoice_return_ref_id = (
            data.get("invoice_return_ref_id")
            or data.get("invoice_return_ref")
            or ""
        ).strip()

        if not invoice_return_ref_id:
            return jsonify({
                "success": False,
                "message": "Select Invoice Return Reference ID"
            }), 400

        

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT status
            FROM invoice_return
            WHERE invoice_return_id = %s
            """,
            (invoice_return_ref_id,),
        )
        ir_row = cur.fetchone()
        if not ir_row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Invoice return not found"
            }), 400

        cur.execute(
            """
            SELECT status, invoice_return_ref_id
            FROM deliverynote_returns
            WHERE dnr_id = %s
            """,
            (dnr_id,),
        )
        existing = cur.fetchone()
        exists = existing is not None
        existing_ir_ref = (
            str(existing[1] or "").strip()
            if exists
            else ""
        )

        cur.execute(
            """
            SELECT dnr_id
            FROM deliverynote_returns
            WHERE TRIM(COALESCE(invoice_return_ref_id, '')) = TRIM(%s)
              AND TRIM(dnr_id) <> TRIM(%s)
            LIMIT 1
            """,
            (invoice_return_ref_id, dnr_id),
        )
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": (
                    "This invoice return is already linked to a delivery note "
                    "return (including cancelled). Each invoice return can only "
                    "be used once."
                ),
            }), 400

        ir_is_cancelled = (
            str(ir_row[0] or "").strip().lower() == "cancelled"
        )
        reusing_same_ir = (
            exists
            and existing_ir_ref == invoice_return_ref_id
        )
        if ir_is_cancelled and not reusing_same_ir:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Cannot use a cancelled invoice return"
            }), 400

        if status == "Submitted":
            comments_in = data.get("comments") or []
            has_comment = any(
                (c.get("comment") or "").strip()
                for c in comments_in
            )
            if not has_comment:
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": "Please add at least one comment before submitting"
                }), 400
        customer_name = data.get("customer_name") or ""
        customer_id = data.get("customer_id") or ""
        email = data.get("email") or ""
        phone = data.get("phone") or ""
        contact_person = data.get("contact_person") or ""
        customer_ref_no = (data.get("customer_ref_no") or "").strip()
        dnr_date = data.get("dnr_date") or None

        if customer_ref_no:
            if len(customer_ref_no) < 3 or len(customer_ref_no) > 30:
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": (
                        "Customer Ref No must be between 3 and 30 characters."
                    ),
                }), 400

            if not re.match(r"^[A-Za-z0-9\-_/ ]+$", customer_ref_no):
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": (
                        "Only letters, numbers, hyphen (-), underscore (_), "
                        "and slash (/) are allowed in Customer Ref No."
                    ),
                }), 400

        if dnr_date:
            dnr_date_val = (dnr_date or "").strip()
            if (
                not dnr_date_val
                or not re.match(r"^\d{2}-\d{2}-\d{4}$", dnr_date_val)
            ):
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": DNR_DATE_INVALID_MSG,
                }), 400
            try:
                parsed_dnr_date = datetime.strptime(
                    dnr_date_val, DNR_DATE_DISPLAY_FMT
                ).date()
            except ValueError:
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": DNR_DATE_INVALID_MSG,
                }), 400
            if parsed_dnr_date.year < 1900 or parsed_dnr_date.year > 2100:
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": DNR_DATE_INVALID_MSG,
                }), 400
            if parsed_dnr_date != date.today():
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": DNR_DATE_MUST_BE_TODAY_MSG,
                }), 400
            dnr_date = parsed_dnr_date

        if exists:
            current_status = str(existing[0] or "").strip().lower()
            if current_status == "cancelled":
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": "Cannot update a cancelled delivery note return"
                }), 400
            new_status_lower = status.strip().lower()
            if current_status == "submitted" and new_status_lower == "draft":
                cur.close()
                conn.close()
                return jsonify({
                    "success": False,
                    "message": "Cannot change a submitted delivery note return back to draft"
                }), 400

        if exists:
            cur.execute("""
                UPDATE deliverynote_returns SET
                    status = %s,
                    invoice_return_ref_id = %s,
                    customer_name = %s,
                    customer_id = %s,
                    email = %s,
                    phone = %s,
                    contact_person = %s,
                    customer_ref_no = %s,
                    dnr_date = %s
                WHERE dnr_id = %s
            """, (
                status,
                invoice_return_ref_id,
                customer_name,
                customer_id,
                email,
                phone,
                contact_person,
                customer_ref_no,
                dnr_date,
                dnr_id,
            ))
        else:
            cur.execute("""
                INSERT INTO deliverynote_returns (
                    dnr_id,
                    status,
                    invoice_return_ref_id,
                    customer_name,
                    customer_id,
                    email,
                    phone,
                    contact_person,
                    customer_ref_no,
                    dnr_date
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                dnr_id,
                status,
                invoice_return_ref_id,
                customer_name,
                customer_id,
                email,
                phone,
                contact_person,
                customer_ref_no,
                dnr_date,
            ))

        cur.execute(
            "DELETE FROM deliverynote_return_items WHERE dnr_id = %s",
            (dnr_id,),
        )
        for it in data.get("items") or []:
            cur.execute("""
                INSERT INTO deliverynote_return_items (
                    dnr_id,
                    product_id,
                    product_name,
                    uom,
                    invoiced_qty,
                    returned_qty,
                    serial_no,
                    return_reason
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                dnr_id,
                it.get("product_id") or "",
                it.get("product_name") or "",
                it.get("uom") or "",
                it.get("invoiced_qty"),
                it.get("returned_qty"),
                it.get("serial_no") or "",
                it.get("return_reason") or "",
            ))

        cur.execute(
            "DELETE FROM deliverynote_return_comments WHERE dnr_id = %s",
            (dnr_id,),
        )

        for c in data.get("comments") or []:
            comment_text = (c.get("comment") or "").strip()
            if not comment_text:
                continue

            comment_author = (
                (c.get("created_by") or "").strip()
                or (c.get("author") or "").strip()
                or _get_logged_in_user_name()
            )

            created_at_raw = (
                (c.get("created_at") or "")
                or (c.get("raw_created_at") or "")
            )

            created_at_val = datetime.now()

            if created_at_raw:
                try:
                    created_at_val = datetime.fromisoformat(
                        str(created_at_raw).replace("Z", "+00:00")
                    )
                except Exception:
                    created_at_val = datetime.now()

            cur.execute("""
                INSERT INTO deliverynote_return_comments (
                    dnr_id,
                    comment,
                    created_by,
                    created_at
                )
                VALUES (%s, %s, %s, %s)
            """, (
                dnr_id,
                comment_text,
                comment_author,
                created_at_val,
            ))
        cur.execute("""
            INSERT INTO deliverynote_return_history (
                dnr_id,
                action,
                description,
                created_by,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            dnr_id,
            f"Saved as {status}",
            f"Delivery Note Return {dnr_id} saved as {status}",
            _get_logged_in_user_name(),
            datetime.now(),
        ))


                # =========================================
        # AUTO UPDATE DELIVERY NOTE STATUS
        # When Delivery Note Return is Submitted
        # =========================================
        if status == "Submitted":
            cur.execute("""
                SELECT invoice_id
                FROM invoice_return
                WHERE invoice_return_id = %s
                LIMIT 1
            """, (invoice_return_ref_id,))

            inv_ret_row = cur.fetchone()

            if inv_ret_row and inv_ret_row[0]:
                invoice_id = str(inv_ret_row[0]).strip()

                cur.execute("""
                    SELECT sale_order_ref
                    FROM invoices
                    WHERE invoice_id = %s
                    LIMIT 1
                """, (invoice_id,))

                inv_row = cur.fetchone()

                if inv_row and inv_row[0]:
                    sale_order_ref = str(inv_row[0]).strip()

                    cur.execute("""
                        UPDATE delivery_notes
                        SET
                            status = %s,
                            delivery_status = %s
                        WHERE TRIM(COALESCE(so_id, '')) = TRIM(%s)
                    """, (
                        "Returned",
                        "Returned",
                        sale_order_ref
                    ))





        conn.commit()

        cur.close()
        conn.close()

        status_label = str(status or "").strip()
        if status_label.lower() == "draft":
            save_message = "Delivery note return saved as Draft successfully"
        elif status_label.lower() == "submitted":
            save_message = "Delivery note return saved as Submitted successfully"
        else:
            save_message = f"Delivery note return saved as {status_label} successfully"

        return jsonify({
            "success": True,
            "message": save_message,
            "dnr_id": dnr_id,
            "status": status,
        })

    except Exception as e:

        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        if getattr(e, "pgcode", None) == "23505":
            return jsonify({
                "success": False,
                "message": (
                    "DNR ID already exists. Refresh the page to get a new ID."
                ),
            }), 409

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.post("/api/cancel-delivery-note-return")
def cancel_delivery_note_return():

    conn = None

    try:

        data = request.get_json() or {}
        dnr_id = (data.get("dnr_id") or "").strip()
        cancel_reason = (data.get("reason") or "").strip()

        if not dnr_id:
            return jsonify({
                "success": False,
                "message": "DNR ID is required"
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT status FROM deliverynote_returns WHERE dnr_id = %s",
            (dnr_id,),
        )
        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "No saved delivery note return found. Save Draft first."
            }), 404

        current_status = str(row[0] or "").strip().lower()

        if current_status == "cancelled":
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Delivery note return is already cancelled"
            }), 400

        if current_status != "submitted":
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Only submitted delivery note returns can be cancelled"
            }), 400

        cur.execute(
            """
            UPDATE deliverynote_returns
            SET status = %s,
                cancel_reason = %s
            WHERE dnr_id = %s
            """,
            ("Cancelled", cancel_reason or None, dnr_id),
        )

        cancel_description = (
            f"Delivery Note Return {dnr_id} cancelled"
        )
        if cancel_reason:
            cancel_description += f": {cancel_reason}"

        cur.execute("""
            INSERT INTO deliverynote_return_history (
                dnr_id,
                action,
                description,
                created_by,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            dnr_id,
            "Cancelled",
            cancel_description,
            _get_logged_in_user_name(),
            datetime.now(),
        ))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Delivery note return saved as Cancelled successfully",
            "dnr_id": dnr_id,
            "status": "Cancelled",
        })

    except Exception as e:

        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@app.route("/deliverynote_return/form")
def deliverynote_return_form():

    dnr_id_param = (request.args.get("id") or "").strip()

    if dnr_id_param:
        dnr_id = dnr_id_param
    else:
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT COALESCE(MAX(
                    CAST(NULLIF(SPLIT_PART(dnr_id, '-', 2), '') AS INTEGER)
                ), 0)
                FROM deliverynote_returns
                WHERE dnr_id ~ '^DNR-[0-9]+$'
                """
            )
            max_num = int((cur.fetchone() or [0])[0] or 0)
            dnr_id = f"DNR-{max_num + 1:03d}"
            conn.commit()
        finally:
            cur.close()
            conn.close()

    return render_template(
        "deliverynotereturn-new.html",
        page="deliverynote_return",
        dnr_id=dnr_id,
    )

# =========================================
# DELIVERY NOTE RETURN - API (PDF)
# =========================================
@app.get("/api/delivery-note-returns/<dnr_id>/pdf")
def delivery_note_return_pdf(dnr_id):
    dnr_id = (dnr_id or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM deliverynote_returns WHERE dnr_id=%s",
        (dnr_id,),
    )
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return jsonify({
            "success": False,
            "message": "Delivery Note Return not found"
        }), 404

    columns = [desc[0] for desc in cur.description]
    dnr = dict(zip(columns, row))

    dnr_date_raw = dnr.get("dnr_date")
    if dnr_date_raw is None or dnr_date_raw == "":
        dnr["dnr_date"] = ""
    elif hasattr(dnr_date_raw, "strftime"):
        dnr["dnr_date"] = dnr_date_raw.strftime(DNR_DATE_DISPLAY_FMT)
    else:
        dnr_date_s = str(dnr_date_raw).strip()
        if not dnr_date_s:
            dnr["dnr_date"] = ""
        elif re.match(r"^\d{2}-\d{2}-\d{4}$", dnr_date_s):
            dnr["dnr_date"] = dnr_date_s
        else:
            dnr_date_iso = dnr_date_s.split("T")[0]
            if re.match(r"^\d{4}-\d{2}-\d{2}$", dnr_date_iso):
                try:
                    dnr["dnr_date"] = (
                        datetime.strptime(dnr_date_iso, "%Y-%m-%d")
                        .date()
                        .strftime(DNR_DATE_DISPLAY_FMT)
                    )
                except ValueError:
                    dnr["dnr_date"] = dnr_date_s
            else:
                dnr["dnr_date"] = dnr_date_s

    cur.execute("""
    SELECT
        product_id,
        product_name,
        uom,
        invoiced_qty,
        returned_qty,
        serial_no,
        return_reason
    FROM deliverynote_return_items
    WHERE dnr_id=%s
    """, (dnr_id,))

    items_rows = cur.fetchall()

    items = []
    for i in items_rows:
        items.append({
            "product_id": i[0],
            "product_name": i[1],
            "uom": i[2],
            "invoiced_qty": float(i[3] or 0),
            "returned_qty": float(i[4] or 0),
            "serial_no": i[5] or "",
            "reason": i[6] or "",
        })

    dnr["items"] = items

    cur.close()
    conn.close()

    pdf_bytes = generate_delivery_note_return_pdf_bytes(dnr)

    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = (
        f'inline; filename="{dnr_id}.pdf"'
    )

    return response





# =========================================
# DELIVERY NOTE RETURN - EMAIL WITH PDF
# =========================================
@app.post("/api/delivery-note-returns/<dnr_id>/email")
def email_delivery_note_return(dnr_id):
    dnr_id = (dnr_id or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT customer_name, email
        FROM deliverynote_returns
        WHERE dnr_id = %s
        """,
        (dnr_id,),
    )
    row = cur.fetchone()
    cur.close()
    conn.close()

    if not row:
        return jsonify({
            "success": False,
            "message": "Delivery Note Return not found"
        }), 404

    customer_name = row[0] or ""
    customer_email = (row[1] or "").strip()

    if not customer_email:
        customer = find_customer_by_name(customer_name)

        if customer:
            customer_email = (
                customer.get("email") or ""
            ).strip()

    if not customer_email:
        return jsonify({
            "success": False,
            "message": "Customer email not available"
        }), 400

    with app.test_client() as client:
        pdf_resp = client.get(
            f"/api/delivery-note-returns/{dnr_id}/pdf"
        )

    if pdf_resp.status_code != 200:
        return jsonify({
            "success": False,
            "message": "Delivery Note Return not found"
        }), 404

    pdf_bytes = pdf_resp.data

    ok = send_email(
        customer_email,
        f"Delivery Note Return {dnr_id}",
        (
            f"Dear {customer_name or 'Customer'},\n\n"
            f"Please find attached the Delivery Note Return "
            f"document ({dnr_id}).\n\n"
            f"The returned item details have been recorded "
            f"successfully in our system.\n\n"
            f"If you have any questions, please contact us.\n\n"
            f"Regards,\n"
            f"Stackly Team"
        ),
        attachments=[
            {
                "filename": f"{dnr_id}.pdf",
                "content_bytes": pdf_bytes
            }
        ]
    )

    if not ok:
        return jsonify({
            "success": False,
            "message": "Email failed. Check SMTP/App password."
        }), 500

    return jsonify({
        "success": True,
        "message": "Email sent successfully"
    })





# =========================================
# DELIVERY NOTE RETURN - PDF GENERATOR
# =========================================
def generate_delivery_note_return_pdf_bytes(dnr):

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18,
        leftMargin=18,
        topMargin=16,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()

    company_style = ParagraphStyle(
        name="DNRReturnCompanyName",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#8c1f1f"),
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    company_info_style = ParagraphStyle(
        name="DNRReturnCompanyInfo",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=9,
        leading=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=1,
    )

    page_title_style = ParagraphStyle(
        name="DNRReturnPageTitle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=16,
        leading=20,
        textColor=colors.green,
        alignment=TA_CENTER,
        spaceBefore=12,
        spaceAfter=12,
    )

    section_style = ParagraphStyle(
        name="DNRSection",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#8c1f1f"),
        spaceAfter=6,
        spaceBefore=10,
    )

    label_style = ParagraphStyle(
        name="DNRLabel",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#6b1a1a"),
    )

    value_style = ParagraphStyle(
        name="DNRValue",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8.5,
        leading=11,
        textColor=colors.black,
    )

    header_small_style = ParagraphStyle(
        name="DNRHeaderSmall",
        parent=styles["Normal"],
        fontName="DejaVuSans-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    terms_style = ParagraphStyle(
        name="DNRTermsStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=8,
        leading=11,
        textColor=colors.black,
        leftIndent=8,
    )

    elements = []

    def safe_str(val, default="-"):
        if val is None:
            return default

        s = str(val).strip()

        return s if s else default

    def safe_float(val, default=0.0):
        try:
            if val in (None, ""):
                return default

            return float(val)

        except Exception:
            return default

    elements.append(Paragraph("STACKLY", company_style))

    elements.append(Paragraph(
        "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
        company_info_style
    ))

    elements.append(Paragraph(
        "Phone: +91 7010792745",
        company_info_style
    ))

    elements.append(Paragraph(
        "Email: info@stackly.com",
        company_info_style
    ))

    elements.append(Spacer(1, 10))

    status_text = safe_str(
        dnr.get("status") or "SUBMITTED"
    ).upper()

    elements.append(Paragraph(
        f"DELIVERY NOTE RETURN - {status_text}",
        page_title_style
    ))

    elements.append(Spacer(1, 2))

    dnr_date_str = safe_str(dnr.get("dnr_date"))

    dnr_details_data = [
        [
            Paragraph("<b>DNR Number:</b>", label_style),
            Paragraph(safe_str(dnr.get("dnr_id")), value_style),

            Paragraph("<b>Date:</b>", label_style),
            Paragraph(dnr_date_str, value_style),
        ],

        [
            Paragraph("<b>Invoice Return Ref:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("invoice_return_ref_id")),
                value_style
            ),

            Paragraph("<b>Customer Ref No:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("customer_ref_no")),
                value_style
            ),
        ],

        [
            Paragraph("<b>Customer Name:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("customer_name")),
                value_style
            ),

            Paragraph("<b>Customer ID:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("customer_id")),
                value_style
            ),
        ],

        [
            Paragraph("<b>Email:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("email")),
                value_style
            ),

            Paragraph("<b>Phone:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("phone_no") or dnr.get("phone")),
                value_style
            ),
        ],

        [
            Paragraph("<b>Contact Person:</b>", label_style),
            Paragraph(
                safe_str(dnr.get("contact_person")),
                value_style
            ),

            Paragraph("<b>Status:</b>", label_style),
            Paragraph(status_text, value_style),
        ],
    ]

    details_table = Table(
        dnr_details_data,
        colWidths=[110, 170, 95, 145]
    )

    details_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),

        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),

        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elements.append(details_table)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph(
        "DELIVERY NOTE RETURN ITEMS",
        section_style
    ))

    elements.append(Spacer(1, 2))

    items = dnr.get("items", []) or []

    item_data = [[
        Paragraph("S.No", header_small_style),
        Paragraph("Product Name", header_small_style),
        Paragraph("Product ID", header_small_style),
        Paragraph("UOM", header_small_style),
        Paragraph("Invoiced Qty", header_small_style),
        Paragraph("Returned Qty", header_small_style),
        Paragraph("Serial No(s)", header_small_style),
        Paragraph("Reason", header_small_style),
    ]]

    for idx, item in enumerate(items, start=1):

        item_data.append([
            Paragraph(str(idx), value_style),

            Paragraph(
                safe_str(item.get("product_name")),
                value_style
            ),

            Paragraph(
                safe_str(item.get("product_id")),
                value_style
            ),

            Paragraph(
                safe_str(item.get("uom")),
                value_style
            ),

            Paragraph(
                str(safe_float(item.get("invoiced_qty"))),
                value_style
            ),

            Paragraph(
                str(safe_float(item.get("returned_qty"))),
                value_style
            ),

            Paragraph(
                safe_str(item.get("serial_no"), ""),
                value_style
            ),

            Paragraph(
                safe_str(item.get("reason")),
                value_style
            ),
        ])

    if len(item_data) == 1:
        item_data.append([
            "-", "-", "-", "-", "-", "-", "-", "-"
        ])

    items_table = Table(
        item_data,
        colWidths=[35, 110, 65, 45, 65, 65, 90, 75],
        repeatRows=1
    )

    items_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#a12828")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "DejaVuSans"),

        ("FONTSIZE", (0, 0), (-1, -1), 7.5),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),

        ("ALIGN", (1, 1), (2, -1), "LEFT"),

        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#999999")),

        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.white,
            colors.HexColor("#f7f7f7")
        ]),

        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),

        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elements.append(items_table)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph(
        "Terms and Conditions",
        section_style
    ))

    terms_list = [
        "1. This Delivery Note Return is issued based on the returned item details.",
        "2. Returned items are subject to verification and approval.",
        "3. Damaged or incorrect products should be reported immediately.",
        "4. Serial numbers should match the originally delivered items.",
        "5. The company reserves the right to reject invalid return requests.",
    ]

    for term in terms_list:
        elements.append(Paragraph(term, terms_style))

    doc.build(elements)

    pdf_bytes = buffer.getvalue()

    buffer.close()

    return pdf_bytes


# ========================================
# Purchase order page
# ========================================
@app.route("/api/purchase-list", methods=["GET"])
def purchase_list_api():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
 
        cur.execute("""
            SELECT
                po_number,
                supplier_name,
                pdate,
                ddate,
                status,
                payment_terms,
                p_value
            FROM purchase_orders
            ORDER BY created_at DESC
        """)
 
        rows = cur.fetchall()
 
        orders = []
 
        for r in rows:
 
            orders.append({
                "po_number": r[0] or "",
                "supplier": r[1] or "",
                "pdate": str(r[2]) if r[2] else "",
                "ddate": str(r[3]) if r[3] else "",
                "status": r[4] or "Draft",
                "payment_terms": r[5] or "",
                "grand_total": float(r[6] or 0)
            })
 
        return jsonify(orders)
 
    except Exception as e:
 
        print("purchase_list_api error:", e)
 
        return jsonify([]), 500
 
    finally:
 
        cur.close()
        conn.close()
 
 
@app.route("/purchase")
def purchase_page():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
 
        cur.execute("""
            SELECT
                po_number,
                supplier_name,
                pdate,
                ddate,
                status,
                payment_terms,
                COALESCE(p_value, 0) AS p_value
            FROM purchase_orders
            ORDER BY created_at DESC
        """)
 
        rows = cur.fetchall()
 
        orders = []
 
        for r in rows:
 
            orders.append({
                "po_number": r[0],
                "supplier": r[1] or "-",
                "pdate": str(r[2]) if r[2] else "",
                "ddate": str(r[3]) if r[3] else "",
                "status": r[4] or "Draft",
                "payment_terms": r[5] or "-",
                "grand_total": float(r[6]) if r[6] is not None else 0
            })
 
        cur.execute("""
            SELECT
                supplier_id,
                supplier_name,
                email
            FROM suppliers
            ORDER BY supplier_name ASC
        """)
 
        supplier_rows = cur.fetchall()
 
        suppliers = []
 
        for s in supplier_rows:
 
            suppliers.append({
                "supplier_id": s[0],
                "supplier_name": s[1],
                "email": s[2]
            })
 
        return render_template(
            "purchase.html",
            orders=orders,
            suppliers=suppliers,
            page="purchase"
        )
 
    except Exception as e:
 
        print("PURCHASE PAGE ERROR:", e)
 
        return str(e), 500
 
    finally:
 
        cur.close()
        conn.close()
 
@app.route("/purchase-order")
def purchase_order():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    po_number = generate_po_number()
 
    today = date.today().isoformat()
 
    # SALES ORDERS
 
    cur.execute("""
        SELECT so_id
        FROM sales_orders
        ORDER BY so_id DESC
    """)
 
    sales_orders = [r[0] for r in cur.fetchall()]
 
    # SUPPLIERS
 
    cur.execute("""
        SELECT
            supplier_id,
            supplier_name,
            email
        FROM suppliers
        ORDER BY supplier_name
    """)
 
    suppliers = []
 
    for r in cur.fetchall():
 
        suppliers.append({
            "id": r[0],
            "name": r[1],
            "email": r[2]
        })
 
    cur.close()
    conn.close()
 
    return render_template(
        "purchase-order.html",
        po_number=po_number,
        today=today,
        sales_orders=sales_orders,
        suppliers=suppliers
    )
 
# ========================================
# GENERATE PO NUMBER
# ========================================
 
def generate_po_number():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    cur.execute("""
        SELECT po_number
        FROM purchase_orders
        ORDER BY created_at DESC
        LIMIT 1
    """)
 
    row = cur.fetchone()
 
    if row and row[0]:
 
        last_po = row[0]
 
        try:
            last_num = int(last_po.split("-")[1])
 
        except:
            last_num = 0
 
        new_num = last_num + 1
 
    else:
 
        new_num = 1
 
    cur.close()
    conn.close()
 
    return f"PO-{new_num:03d}"
 
# Sales Order id
 
@app.route("/api/sales-order-purchase/<so_id>")
def get_sales_order_purchase(so_id):
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
 
        cur.execute("""
            SELECT
                product_id,
                product_name,
                qty,
                uom,
                price,
                tax_pct,
                disc_pct
            FROM sales_order_items
            WHERE so_id=%s
        """, (so_id,))
 
        rows = cur.fetchall()
 
        items = []
 
        for r in rows:
 
            items.append({
                "product_id": r[0],
                "product_name": r[1],
                "qty": r[2],
                "uom": r[3],
                "price": r[4],
                "tax_pct": r[5],
                "disc_pct": r[6]
            })
 
        return jsonify({
            "items": items
        })
 
    except Exception as e:
 
        print("SO FETCH ERROR:", e)
 
        return jsonify({"items": []}), 500
 
    finally:
 
        cur.close()
        conn.close()
 
 
 
 
 
@app.route("/api/products-new")
def get_products_new():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    cur.execute("""
        SELECT
            product_id,
            product_name,
            unit_price,
            tax_percent,
            uom_name,
            stock_level
        FROM products
    """)
 
    rows = cur.fetchall()
 
    cur.close()
    conn.close()
 
    return jsonify([
 
        {
            "product_id": r[0],
            "product_name": r[1],
            "unit_price": float(r[2] or 0),
            "tax_percent": float(r[3] or 0),
            "uom_name": r[4],
            "stock_level": r[5]
        }
 
        for r in rows
 
    ])
 
# ========================================
# SALES ORDER IDS API
# ========================================
 
@app.get("/api/sales-orders/ids")
def get_sales_order_ids():
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
 
        cur.execute("""
            SELECT so_id
            FROM sales_orders
            ORDER BY so_id DESC
        """)
 
        rows = cur.fetchall()
 
        sales_order_ids = [r[0] for r in rows]
 
        return jsonify(sales_order_ids)
 
    except Exception as e:
 
        print("SALES ORDER LIST ERROR:", e)
 
        return jsonify([]), 500
 
    finally:
 
        cur.close()
        conn.close()
 
 
@app.route("/api/save-po-purchase", methods=["POST"])
def save_po_purchase():
 
    conn = None
    cur = None
 
    try:
 
        data = request.json
 
        conn = get_db_connection()
        cur = conn.cursor()
 
        po_number = str(data.get("po_number") or "").strip()
 
        if not po_number:
 
            po_number = generate_po_number()
 
        supplier_id = data.get("supplier_id")
        supplier_name = data.get("supplier_name")
        supplier_email = data.get("supplier_email")
 
        so_id = data.get("so_id")
 
        pdate = data.get("pdate")
        ddate = data.get("ddate")
 
        payment_terms = data.get("payment_terms")
        inco_terms = data.get("inco_terms")
 
        status = data.get("status", "Draft")
 
        items = data.get("items", [])
 
        # =================================
        # CALCULATE TOTAL
        # =================================
 
        total_value = float(data.get("grand_total", 0))
 
        # =================================
        # CHECK EXISTING
        # =================================
 
        cur.execute("""
            SELECT po_number
            FROM purchase_orders
            WHERE po_number=%s
        """, (po_number,))
 
        existing = cur.fetchone()
 
        # =================================
        # INSERT
        # =================================
 
        if not existing:
 
            cur.execute("""
                INSERT INTO purchase_orders
                (
                    po_number,
                    supplier_id,
                    supplier_name,
                    supplier_email,
                    so_id,
                    pdate,
                    ddate,
                    p_value,
                    status,
                    payment_terms,
                    inco_terms,
                    created_at
                )
                VALUES
                (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()
                )
            """, (
                po_number,
                supplier_id,
                supplier_name,
                supplier_email,
                so_id,
                pdate,
                ddate,
                round(total_value, 2),
                status,
                payment_terms,
                inco_terms
            ))
 
           
 
        # =================================
        # UPDATE
        # =================================
 
        else:
 
           
 
            cur.execute("""
                UPDATE purchase_orders
                SET
                    supplier_id=%s,
                    supplier_name=%s,
                    supplier_email=%s,
                    so_id=%s,
                    pdate=%s,
                    ddate=%s,
                    p_value=%s,
                    status=%s,
                    payment_terms=%s,
                    inco_terms=%s
                WHERE po_number=%s
            """, (
                supplier_id,
                supplier_name,
                supplier_email,
                so_id,
                pdate,
                ddate,
                round(total_value, 2),
                status,
                payment_terms,
                inco_terms,
                po_number
            ))
 
        # =================================
        # DELETE OLD ITEMS
        # =================================
 
        cur.execute("""
            DELETE FROM purchase_items
            WHERE po_number=%s
        """, (po_number,))
 
        # =================================
        # INSERT ITEMS
        # =================================
 
        for item in items:
 
            cur.execute("""
                INSERT INTO purchase_items
                (
                   
                    po_number,
                    product_id,
                    product_name,
                    qty,
                    price,
                    tax,
                    discount,
                    uom
                )
                VALUES
                (
                    %s,%s,%s,%s,%s,%s,%s,%s
                )
            """, (
 
           
                po_number,
                item.get("product_id"),
                item.get("product_name"),
                float(item.get("qty", 0)),
                float(item.get("price", 0)),
                float(item.get("tax", 0)),
                float(item.get("discount", 0)),
                item.get("uom")
 
            ))
 
        conn.commit()
 
        return jsonify({
 
            "success": True,
            "message": "Purchase Order Saved Successfully",
            "po_number": po_number,
            "status": status
 
        })
 
    except Exception as e:
 
        if conn:
            conn.rollback()
 
        print("SAVE PO ERROR:", e)
 
        return jsonify({
 
            "success": False,
            "error": str(e)
 
        }), 500
 
    finally:
 
        if cur:
            cur.close()
 
        if conn:
            conn.close()
 
# ========================================
# DELETE PURCHASE ORDER
# ========================================
 
@app.route("/delete_po/<string:po_number>", methods=["DELETE"])
def delete_po(po_number):
 
    conn = get_db_connection()
    cur = conn.cursor()
 
    try:
 
        cur.execute("""
            SELECT
                po_number,
                status
            FROM purchase_orders
            WHERE po_number=%s
        """, (po_number,))
 
        po = cur.fetchone()
 
        if not po:
 
            return jsonify({
                "error": "PO Not Found"
            }), 404
 
        po_number, status = po
 
        if status != "Draft":
 
            return jsonify({
                "error": "Only Draft PO Can Delete"
            }), 400
 
        cur.execute("""
            DELETE FROM purchase_items
            WHERE po_number=%s
        """, (po_number,))
 
        cur.execute("""
            DELETE FROM purchase_orders
            WHERE po_number=%s
        """, (po_number,))
 
        conn.commit()
 
        return jsonify({
            "success": True
        })
 
    except Exception as e:
 
        print("DELETE ERROR:", e)
 
        return jsonify({
            "error": str(e)
        }), 500
 
    finally:
 
        cur.close()
        conn.close()
 
@app.route("/generate-purchase-pdf", methods=["POST"])
def generate_purchase_pdf():
 
    data = request.json
 
    buffer = BytesIO()
 
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=20,
        bottomMargin=20
    )
 
    styles = getSampleStyleSheet()
 
    # =========================================
    # STYLES
    # =========================================
 
    company_style = ParagraphStyle(
        "CompanyStyle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=24,
        leading=28,
        alignment=TA_CENTER,
        textColor=colors.darkred,
        spaceAfter=5,
    )
 
    address_style = ParagraphStyle(
        "AddressStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=10,
        leading=14,
        alignment=TA_CENTER,
    )
 
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading2"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.green,
        spaceAfter=20,
    )
 
    section_style = ParagraphStyle(
        "SectionStyle",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=14,
        textColor=colors.darkred,
        spaceAfter=10,
    )
 
    normal_style = ParagraphStyle(
        "NormalStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=10,
    )
 
    elements = []
 
    # =========================================
    # COMPANY HEADER
    # =========================================
 
    elements.append(Paragraph("STACKLY", company_style))
 
    elements.append(Paragraph(
        "MMR Complex, Chinna Tirupathi, Salem, Tamil Nadu - 636008",
        address_style
    ))
 
    elements.append(Paragraph(
        "Phone: +91 7010792745",
        address_style
    ))
 
    elements.append(Paragraph(
        "Email: info@stackly.com",
        address_style
    ))
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # TITLE
    # =========================================
 
    elements.append(Paragraph("PURCHASE ORDER", title_style))
 
    # =========================================
    # INFO TABLE
    # =========================================
 
    info_data = [
        ["PO Number:", data.get("po_number"), "Date:", data.get("pdate")],
        ["Supplier:", data.get("supplier"), "Delivery Date:", data.get("ddate")],
        ["Status:", data.get("status"), "Payment Terms:", data.get("payment_terms")]
    ]
 
    info_table = Table(
        info_data,
        colWidths=[120, 170, 120, 140]
    )
 
    info_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (0, -1), "DejaVuSans-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "DejaVuSans-Bold"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),
 
        ("TEXTCOLOR", (0, 0), (0, -1), colors.darkred),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.darkred),
 
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
 
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    elements.append(info_table)
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # ITEMS TITLE
    # =========================================
 
    elements.append(Paragraph("PURCHASE ORDER ITEMS", section_style))
 
    # =========================================
    # ITEMS TABLE
    # =========================================
 
    table_data = [[
        "S.No",
        "Product Name",
        "Qty",
        "Price",
        "Tax %",
        "Disc %",
        "Total"
    ]]
 
    for i, item in enumerate(data.get("items", []), start=1):
 
        qty = float(item.get("qty", 0))
        price = float(item.get("price", 0))
 
        total = qty * price
 
        table_data.append([
            str(i),
            item.get("product_name"),
            str(item.get("qty")),
            f"₹ {price:.2f}",
            str(item.get("tax")),
            str(item.get("discount")),
            f"₹ {total:.2f}"
        ])
 
    item_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[45, 180, 55, 75, 55, 60, 90]
    )
 
    item_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
 
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
 
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
 
        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 8),
 
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
    ]))
 
    elements.append(item_table)
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # TOTALS TABLE
    # =========================================
 
    totals_data = [
        ["Subtotal", data.get("subtotal")],
        ["Tax", data.get("tax")],
        ["Rounding", data.get("rounding")],
        ["Grand Total", data.get("grand_total")],
    ]
 
    totals_table = Table(
        totals_data,
        colWidths=[450, 110]
    )
 
    totals_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
 
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
 
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
    ]))
 
    elements.append(totals_table)
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # NOTES
    # =========================================
 
    elements.append(Paragraph("Notes", section_style))
 
    notes = data.get("notes") or "Thank you for your business!"
 
    elements.append(Paragraph(notes, normal_style))
 
    # =========================================
    # BUILD PDF
    # =========================================
 
    doc.build(elements)
 
    buffer.seek(0)
 
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{data.get('po_number')}.pdf",
        mimetype="application/pdf"
    )
 
@app.route("/api/purchase-orders/<po_number>/email", methods=["POST"])
def email_purchase_order(po_number):
 
    try:
        data = request.json
        data["po_number"] = po_number
 
        supplier_email = data.get("supplier_email")
 
        if not supplier_email:
 
            return jsonify({
                "success": False,
                "message": "Supplier email missing"
            })
 
        # =========================================
        # GENERATE PDF
        # =========================================
        pdf_bytes = build_purchase_order_pdf(data)
        # =========================================
        # EMAIL CONTENT
        # =========================================
 
        subject = f"Purchase Order - {po_number}"
 
        body = f"""
Dear Supplier,
 
We hope this message finds you well.
 
Please find attached the official Purchase Order issued by our company for your processing.
 
Purchase Order Information:
- PO Number: {po_number}
 
We request you to acknowledge receipt of this order and proceed with the necessary arrangements.
 
Should you require any clarification, please do not hesitate to contact us.
 
Regards,
Procurement Department
Stackly
"""
 
        # =========================================
        # SEND EMAIL
        # =========================================
 
        success = send_email_with_pdf(
            to_email=supplier_email,
            subject=subject,
            body=body,
            pdf_bytes=pdf_bytes,
            pdf_filename=f"{po_number}.pdf"
        )
 
        if success:
            return jsonify({"success": True})
 
        else:
            return jsonify({
                "success": False,
                "message": "Email send failed"
            })
 
    except Exception as e:
 
        print("EMAIL ERROR:", str(e))
 
        return jsonify({
            "success": False,
            "message": str(e)
        })
 
def send_email_with_pdf(
    to_email,
    subject,
    body,
    pdf_bytes,
    pdf_filename="purchase_order.pdf"
):
 
    try:
 
        msg = MIMEMultipart()
 
        msg["From"] = EMAIL_ADDRESS
        msg["To"] = to_email
        msg["Subject"] = subject
 
        # ✅ BODY (this is your subject/body you mentioned)
        msg.attach(MIMEText(body, "plain"))
 
        # ✅ PDF ATTACHMENT
        part = MIMEApplication(pdf_bytes, _subtype="pdf")
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename=pdf_filename
        )
        msg.attach(part)
 
        # ✅ SMTP SEND
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
 
        server.send_message(msg)   # IMPORTANT
 
        server.quit()
 
        return True
 
    except Exception as e:
        print("Email send error:", e)
        return False
   
def build_purchase_order_pdf(data):
 
    buffer = BytesIO()
 
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=20,
        bottomMargin=20
    )
 
    styles = getSampleStyleSheet()
 
    # =========================================
    # STYLES
    # =========================================
 
    company_style = ParagraphStyle(
        "CompanyStyle",
        parent=styles["Heading1"],
        fontName="DejaVuSans-Bold",
        fontSize=24,
        leading=28,
        alignment=TA_CENTER,
        textColor=colors.darkred,
        spaceAfter=5,
    )
 
    address_style = ParagraphStyle(
        "AddressStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=10,
        leading=14,
        alignment=TA_CENTER,
    )
 
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading2"],
        fontName="DejaVuSans-Bold",
        fontSize=20,
        leading=24,
        alignment=TA_CENTER,
        textColor=colors.green,
        spaceAfter=20,
    )
 
    section_style = ParagraphStyle(
        "SectionStyle",
        parent=styles["Heading3"],
        fontName="DejaVuSans-Bold",
        fontSize=14,
        textColor=colors.darkred,
        spaceAfter=10,
    )
 
    normal_style = ParagraphStyle(
        "NormalStyle",
        parent=styles["Normal"],
        fontName="DejaVuSans",
        fontSize=10,
    )
 
    elements = []
 
    # =========================================
    # COMPANY HEADER
    # =========================================
 
    elements.append(Paragraph("STACKLY", company_style))
 
    elements.append(Paragraph(
        "MMR Complex, Chinna Tirupathi, Salem, Tamil Nadu - 636008",
        address_style
    ))
 
    elements.append(Paragraph(
        "Phone: +91 7010792745",
        address_style
    ))
 
    elements.append(Paragraph(
        "Email: info@stackly.com",
        address_style
    ))
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # TITLE
    # =========================================
 
    elements.append(Paragraph("PURCHASE ORDER", title_style))
 
    # =========================================
    # INFO TABLE
    # =========================================
 
    po_number = data.get("po_number", "")
    po_number = po_number.replace("PO ", "").strip()
 
    info_data = [
        ["PO Number:", po_number, "Date:", data.get("pdate")],
        ["Supplier:", data.get("supplier"), "Delivery Date:", data.get("ddate")],
        ["Status:", data.get("status"), "Payment Terms:", data.get("payment_terms")]
    ]
 
    info_table = Table(
        info_data,
        colWidths=[120, 170, 120, 140]
    )
 
    info_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (0, -1), "DejaVuSans-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "DejaVuSans-Bold"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
        ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),
 
        ("TEXTCOLOR", (0, 0), (0, -1), colors.darkred),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.darkred),
 
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
 
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
 
    elements.append(info_table)
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # ITEMS TITLE
    # =========================================
 
    elements.append(Paragraph("PURCHASE ORDER ITEMS", section_style))
 
    # =========================================
    # ITEMS TABLE
    # =========================================
 
    table_data = [[
        "S.No",
        "Product Name",
        "Qty",
        "Price",
        "Tax %",
        "Disc %",
        "Total"
    ]]
 
    for i, item in enumerate(data.get("items", []), start=1):
 
        qty = float(item.get("qty", 0))
        price = float(item.get("price", 0))
        tax = float(item.get("tax", 0))
        discount = float(item.get("discount", 0))
 
        total = qty * price
 
        table_data.append([
            str(i),
            item.get("product_name"),
            str(item.get("qty")),
            f"₹ {price:.2f}",
            f"{tax:.2f}",
            f"{discount:.2f}",
            f"₹ {total:.2f}"
        ])
 
    item_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[45, 180, 55, 75, 55, 60, 90]
    )
 
    item_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
 
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
 
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
 
        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 8),
 
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
    ]))
 
    elements.append(item_table)
 
    elements.append(Spacer(1, 25))
 
    # =========================================
    # ORDER SUMMARY / TOTALS
    # =========================================
 
 
 
    subtotal = float(data.get("subtotal") or 0)
    tax_total = float(data.get("tax") or 0)
    discount_total = float(data.get("discount_total") or 0)
    rounding = float(data.get("rounding") or 0)
    grand_total = float(data.get("grand_total") or 0)
 
    totals_data = [
        ["Subtotal", f"₹ {subtotal:.2f}"],
        ["Tax", f"₹ {tax_total:.2f}"],
        ["Discount", f"₹ {discount_total:.2f}"],
        ["Rounding", f"₹ {rounding:.2f}"],
        ["Grand Total", f"₹ {grand_total:.2f}"],
    ]
 
    totals_table = Table(
        totals_data,
        colWidths=[450, 110]
    )
 
    totals_table.setStyle(TableStyle([
 
        ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
        ("FONTNAME", (0, -1), (-1, -1), "DejaVuSans-Bold"),
 
        ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),
 
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
 
        ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
 
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
 
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.black),
    ]))
 
    elements.append(totals_table)
    elements.append(Spacer(1, 25))
 
    # =========================================
    # NOTES
    # =========================================
 
    elements.append(Paragraph("Notes", section_style))
 
    notes = data.get("notes") or "Thank you for your business!"
 
    elements.append(Paragraph(notes, normal_style))
 
    # =========================================
    # BUILD PDF
    # =========================================
 
    doc.build(elements)
 
    buffer.seek(0)
 
    return buffer.read()
 
 
@app.route("/purchase/view/<po_number>")
def view_po(po_number):
 
    from psycopg2.extras import RealDictCursor
 
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
 
    try:
 
        cur.execute("""
            SELECT po.*, s.email as supplier_email
            FROM purchase_orders po
            LEFT JOIN suppliers s
            ON po.supplier_id = s.supplier_id
            WHERE po.po_number=%s
        """, (po_number,))
 
        po = cur.fetchone()
 
        if not po:
            return "PO Not Found", 404
 
        cur.execute("""
            SELECT *
            FROM purchase_items
            WHERE po_number=%s
        """, (po_number,))
 
        items = cur.fetchall()
 
        po_dict = dict(po)
 
        po_dict["items"] = [dict(i) for i in items]
 
        po_dict["pdate"] = str(po_dict["pdate"]) if po_dict.get("pdate") else ""
 
        po_dict["ddate"] = str(po_dict["ddate"]) if po_dict.get("ddate") else ""
 
        cur.execute("""
            SELECT so_id
            FROM sales_orders
            ORDER BY so_id DESC
        """)
 
        sales_orders = [r["so_id"] for r in cur.fetchall()]
 
        return render_template(
            "purchase-order.html",
            po_number=po_dict["po_number"],
            today=str(po_dict.get("pdate", "")),
            sales_orders=sales_orders,
            po_data=po_dict,
            mode="view",
            page="purchase",
        )
 
    except Exception as e:
 
        print("view_po error:", e)
 
        return str(e), 500
 
    finally:
 
        cur.close()
        conn.close()
@app.route("/purchase/edit/<po_number>")
def edit_po(po_number):
 
    from psycopg2.extras import RealDictCursor
 
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
 
    try:
 
        # ====================================
        # GET PURCHASE ORDER
        # ====================================
 
        cur.execute("""
            SELECT
                po.*,
                s.email AS supplier_email
            FROM purchase_orders po
            LEFT JOIN suppliers s
            ON po.supplier_id = s.supplier_id
            WHERE po.po_number=%s
        """, (po_number,))
 
        po = cur.fetchone()
 
        if not po:
            return "PO Not Found", 404
 
        # ====================================
        # GET PURCHASE ITEMS
        # ====================================
 
        cur.execute("""
            SELECT *
            FROM purchase_items
            WHERE po_number=%s
        """, (po_number,))
 
        items = cur.fetchall()
 
        po_dict = dict(po)
 
        po_dict["items"] = [dict(i) for i in items]
 
        po_dict["pdate"] = (
            str(po_dict["pdate"])
            if po_dict.get("pdate")
            else ""
        )
 
        po_dict["ddate"] = (
            str(po_dict["ddate"])
            if po_dict.get("ddate")
            else ""
        )
 
        # ====================================
        # SALES ORDERS
        # ====================================
 
        try:
 
            cur.execute("""
                SELECT so_id
                FROM sales_orders
                ORDER BY so_id DESC
            """)
 
            sales_orders = [
                r["so_id"]
                for r in cur.fetchall()
            ]
 
        except Exception:
 
            sales_orders = []
 
        return render_template(
            "purchase-order.html",
            po_number=po_dict["po_number"],
            today=str(po_dict.get("pdate", "")),
            sales_orders=sales_orders,
            po_data=po_dict,
            mode="edit",
            page="purchase",
        )
 
    except Exception as e:
 
        print("edit_po error:", e)
 
        import traceback
        traceback.print_exc()
 
        return str(e), 500
 
    finally:
 
        cur.close()
        conn.close()
# ========================================
# Comments
# ========================================
def _ensure_purchase_aux_tables(cur):
    """Ensure purchase comments/attachments tables exist in this DB."""
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_comments (
            id SERIAL PRIMARY KEY,
            po_number TEXT NOT NULL,
            comment TEXT NOT NULL,
            created_by TEXT DEFAULT 'Admin',
            created_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS purchase_attachments (
            id SERIAL PRIMARY KEY,
            po_number TEXT NOT NULL,
            file_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            uploaded_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_purchase_comments_po_number
        ON purchase_comments (po_number)
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_purchase_attachments_po_number
        ON purchase_attachments (po_number)
        """
    )
    cur.connection.commit()


@app.route("/api/purchase-comments", methods=["POST"])
def add_purchase_comment():
    try:
        data = request.json
 
        po_number = data.get("po_number")
        comment = data.get("comment")
        created_by = data.get("created_by", "Admin")
 
        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_purchase_aux_tables(cur)
 
        cur.execute("""
            INSERT INTO purchase_comments (po_number, comment, created_by)
            VALUES (%s, %s, %s)
        """, (po_number, comment, created_by))
 
        conn.commit()
        cur.close()
        conn.close()
 
        return jsonify({"success": True, "message": "Comment added"})
 
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
 
 
# ========================================
# Attachments
# ========================================
 
@app.route("/api/purchase-attachments", methods=["POST"])
def upload_purchase_attachment():
    try:
        po_number = request.form.get("po_number")
        file = request.files.get("file")
 
        if not file:
            return jsonify({"success": False, "message": "No file uploaded"}), 400
 
        filename = _upload_basename(file.filename)
        rel_path = _upload_relative_path(po_number, filename)
        save_path, _ = _persist_module_upload(
            object_storage.MODULE_PURCHASE_ATTACHMENTS,
            PURCHASE_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )
 
        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_purchase_aux_tables(cur)
 
        cur.execute("""
            INSERT INTO purchase_attachments (po_number, file_name, file_path)
            VALUES (%s, %s, %s)
            RETURNING id
        """, (po_number, filename, save_path))
        row = cur.fetchone()
        attachment_id = row[0] if row else None
 
        conn.commit()
        cur.close()
        conn.close()
 
        return jsonify({
            "success": True,
            "id": attachment_id,
            "file_name": filename,
            "file_path": save_path
        })
 
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
 
@app.route("/api/purchase-comments/<po_number>")
def get_purchase_comments(po_number):
 
    conn = get_db_connection()
    cur = conn.cursor()
    _ensure_purchase_aux_tables(cur)
 
    cur.execute("""
        SELECT
            comment,
            created_by,
            created_at
        FROM purchase_comments
        WHERE po_number = %s
        ORDER BY created_at DESC
    """, (po_number,))
 
    rows = cur.fetchall()
 
    comments = []
 
    for r in rows:
 
        comments.append({
            "comment": r[0],
            "created_by": r[1],
            "created_at": str(r[2])
        })
 
    cur.close()
    conn.close()
 
    return jsonify(comments)
 
@app.route("/api/purchase-attachments/<po_number>")
def get_purchase_attachments(po_number):
    conn = get_db_connection()
    cur = conn.cursor()
    _ensure_purchase_aux_tables(cur)
 
    cur.execute("""
        SELECT id, file_name, file_path, uploaded_at
        FROM purchase_attachments
        WHERE po_number = %s
        ORDER BY uploaded_at DESC
    """, (po_number,))
 
    rows = cur.fetchall()
 
    cur.close()
    conn.close()
 
    data = []
    for r in rows:
        data.append({
            "id": r[0],
            "file_name": r[1],
            "file_path": r[2],
            "uploaded_at": str(r[3]) if len(r) > 3 and r[3] is not None else "",
        })

    return jsonify({"success": True, "attachments": data})


@app.route("/api/purchase-attachments", methods=["DELETE"])
def delete_purchase_attachment():
    """Delete purchase attachment record and remove file from S3/local storage."""
    conn = None
    cur = None
    try:
        data = request.get_json(silent=True) or {}
        po_number = (data.get("po_number") or "").strip()
        attachment_id = data.get("attachment_id")
        file_path = (data.get("file_path") or "").strip()
        file_name = (data.get("file_name") or "").strip()

        if not po_number:
            return jsonify({"success": False, "message": "po_number is required"}), 400
        if attachment_id is None and not file_path and not file_name:
            return jsonify({"success": False, "message": "attachment_id or file_path/file_name is required"}), 400

        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_purchase_aux_tables(cur)

        if attachment_id is not None:
            cur.execute(
                """
                SELECT file_path
                FROM purchase_attachments
                WHERE po_number = %s AND id = %s
                LIMIT 1
                """,
                (po_number, attachment_id),
            )
        elif file_path:
            cur.execute(
                """
                SELECT file_path
                FROM purchase_attachments
                WHERE po_number = %s AND file_path = %s
                ORDER BY uploaded_at DESC
                LIMIT 1
                """,
                (po_number, file_path),
            )
        else:
            cur.execute(
                """
                SELECT file_path
                FROM purchase_attachments
                WHERE po_number = %s AND file_name = %s
                ORDER BY uploaded_at DESC
                LIMIT 1
                """,
                (po_number, file_name),
            )

        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404

        stored_path = (row[0] or "").strip()
        if stored_path:
            _remove_stored_upload(stored_path, PURCHASE_ATTACHMENTS_FOLDER)

        if attachment_id is not None:
            cur.execute(
                "DELETE FROM purchase_attachments WHERE po_number = %s AND id = %s",
                (po_number, attachment_id),
            )
        elif file_path:
            cur.execute(
                "DELETE FROM purchase_attachments WHERE po_number = %s AND file_path = %s",
                (po_number, file_path),
            )
        else:
            cur.execute(
                """
                DELETE FROM purchase_attachments
                WHERE po_number = %s AND file_name = %s
                AND file_path = %s
                """,
                (po_number, file_name, stored_path),
            )

        conn.commit()
        return jsonify({"success": True, "message": "Attachment deleted"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route("/api/purchase-attachments/<int:attachment_id>/view")
def view_purchase_attachment(attachment_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_purchase_aux_tables(cur)
        cur.execute(
            """
            SELECT file_path, file_name
            FROM purchase_attachments
            WHERE id = %s
            """,
            (attachment_id,),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404
        file_path = row[0]
        original_name = row[1]
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        resolved = _resolve_stored_file_path(file_path)
        if not resolved or not os.path.isfile(resolved):
            return jsonify({"success": False, "message": "File not found"}), 404
        return send_file(resolved, as_attachment=False, download_name=original_name)
    finally:
        cur.close()
        conn.close()


@app.route("/api/purchase-attachments/<int:attachment_id>/download")
def download_purchase_attachment(attachment_id):
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_purchase_aux_tables(cur)
        cur.execute(
            """
            SELECT file_path, file_name
            FROM purchase_attachments
            WHERE id = %s
            """,
            (attachment_id,),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "message": "Attachment not found"}), 404
        file_path = row[0]
        original_name = row[1]
        if object_storage.is_remote_url(file_path):
            return redirect(file_path)
        resolved = _resolve_stored_file_path(file_path)
        if not resolved or not os.path.isfile(resolved):
            return jsonify({"success": False, "message": "File not found"}), 404
        return send_file(resolved, as_attachment=True, download_name=original_name)
    finally:
        cur.close()
        conn.close()
 

# ========================================
# Stock-reciept
# ========================================

@app.route("/stock-receipt")
def stock_receipt():

    return render_template(
        "stock-reciept.html",
        page="stock_receipt"
    )

@app.route("/stock-new")
def stock_new():

    return render_template(
        "stock-new.html",
        page="stock_new"
    )

@app.route("/api/generate-grn")
def generate_grn():

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT grn_number
        FROM stock_receipts
        ORDER BY grn_number DESC
        LIMIT 1
    """)

    last = cur.fetchone()

    if last and last[0]:

        try:
            last_no = int(last[0].split("-")[1])
            new_no = last_no + 1

        except:
            new_no = 1

    else:

        new_no = 1

    grn = f"GRN-{new_no:03d}"

    cur.close()
    conn.close()

    return jsonify({
        "grn_number": grn
    })

@app.route("/api/submitted-pos")
def submitted_pos():

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""

        SELECT
            po_number

        FROM purchase_orders

        WHERE LOWER(status) = 'submitted'

        ORDER BY po_number DESC

    """)

    rows = cur.fetchall()

    result = []

    for row in rows:

        result.append({
            "po_number": row[0]
        })

    cur.close()
    conn.close()

    return jsonify(result)


@app.route("/api/purchase/<po_number>")
def get_purchase(po_number):

    conn = get_db_connection()
    cur = conn.cursor()

    # HEADER
    cur.execute("""
        SELECT
            po_number,
            supplier_name,
            supplier_email
        FROM purchase_orders
        WHERE po_number = %s
    """, (po_number,))

    po = cur.fetchone()

    if not po:

        return jsonify({
            "error": "PO not found"
        }), 404

    # ITEMS
    cur.execute("""
        SELECT
            product_id,
            product_name,
            qty,
            price,
            tax,
            discount,
            uom
        FROM purchase_items
        WHERE po_number = %s
    """, (po_number,))

    items = cur.fetchall()

    item_list = []

    for item in items:

        item_list.append({

            "product_id": item[0],
            "product_name": item[1],
            "qty": float(item[2] or 0),
            "price": float(item[3] or 0),
            "tax_pct": float(item[4] or 0),
            "disc_pct": float(item[5] or 0),
            "uom": item[6]

        })

    cur.close()
    conn.close()

    return jsonify({

        "po_number": po[0],
        "supplier_name": po[1],
        "supplier_email": po[2],
        "items": item_list

    })

# ========================================
# SAVE STOCK RECEIPT
# ========================================

@app.route("/api/save-stock", methods=["POST"])
def save_stock():

    data = request.json

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =========================================
        # INSERT STOCK RECEIPT HEADER
        # =========================================
        cur.execute("""

            INSERT INTO stock_receipts (

                grn_number,
                po_number,
                supplier_id,
                supplier_name,
                supplier_email,
                received_date,
                supplier_dn_no,
                supplier_invoice_no,
                received_by,
                qc_done_by,
                grand_total,
                status

            )

            VALUES (

                %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s

            )

        """, (

            data.get("grn_number"),
            data.get("po_number"),
            data.get("supplier_id"),
            data.get("supplier_name"),
            data.get("supplier_email"),
            data.get("received_date"),
            data.get("supplier_dn_no"),
            data.get("supplier_invoice_no"),
            data.get("received_by"),
            data.get("qc_done_by"),
            data.get("grand_total", 0),
            data.get("status", "draft")

        ))

        # =========================================
        # INSERT STOCK RECEIPT ITEMS
        # =========================================
        for item in data.get("items", []):

            cur.execute("""

                INSERT INTO stock_receipt_items (

                    grn_number,
                    product_id,
                    product_name,
                    uom,
                    qty_ordered,
                    qty_received,
                    accepted_qty,
                    rejected_qty,
                    qty_returned,
                    stock_in,
                    warehouse,
                    unit_price,
                    tax_pct,
                    disc_pct,
                    total

                )

                VALUES (

                    %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s

                )

            """, (

                data.get("grn_number"),

                item.get("product_id"),
                item.get("product_name"),
                item.get("uom"),

                item.get("qty_ordered", 0),
                item.get("qty_received", 0),
                item.get("accepted_qty", 0),
                item.get("rejected_qty", 0),
                item.get("qty_returned", 0),

                item.get("stock_in"),
                item.get("warehouse"),

                item.get("unit_price", 0),
                item.get("tax_pct", 0),
                item.get("disc_pct", 0),
                item.get("total", 0)

            ))

        conn.commit()

        return jsonify({

            "success": True,
            "message": "Stock Receipt Saved Successfully",
            "grn_number": data.get("grn_number")

        })

    except Exception as e:

        conn.rollback()

        import traceback
        traceback.print_exc()

        return jsonify({

            "success": False,
            "error": str(e)

        }), 500

    finally:

        cur.close()
        conn.close()

@app.route("/api/stock-receipts", methods=["GET"])
def stock_receipts():

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        cur.execute("""

            SELECT

                grn_number,
                po_number,
                supplier_name,
                supplier_email,
                received_date,
                grand_total,
                status,
                received_by,
                qc_done_by

            FROM stock_receipts

            ORDER BY grn_number DESC

        """)

        rows = cur.fetchall()

        result = []

        for row in rows:

            result.append({

                "grn_number": row[0],
                "po_number": row[1],
                "supplier_name": row[2],
                "supplier_email": row[3],
                "received_date": str(row[4]) if row[4] else "",
                "grand_total": float(row[5] or 0),
                "status": row[6],
                "received_by": row[7],
                "qc_done_by": row[8]

            })

        return jsonify(result)

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cur.close()
        conn.close()

@app.route("/api/stock-receipt/<grn>", methods=["GET"])
def view_stock(grn):

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =========================================
        # GET HEADER
        # =========================================
        cur.execute("""

            SELECT

    
                grn_number,
                po_number,
                supplier_id,
                supplier_name,
                supplier_email,
                received_date,
                supplier_dn_no,
                supplier_invoice_no,
                received_by,
                qc_done_by,
                grand_total,
                status

            FROM stock_receipts

            WHERE TRIM(grn_number) = TRIM(%s)

        """, (grn,))

        stock = cur.fetchone()

        print("GRN PARAM:", grn)
        print("DB STOCK:", stock)

        if not stock:

            return jsonify({
                "error": "Stock Receipt Not Found"
            }), 404

        # =========================================
        # GET ITEMS
        # =========================================
        cur.execute("""

            SELECT

                product_id,
                product_name,
                uom,
                qty_ordered,
                qty_received,
                accepted_qty,
                rejected_qty,
                qty_returned,
                stock_in,
                warehouse,
                unit_price,
                tax_pct,
                disc_pct,
                total

            FROM stock_receipt_items

            WHERE TRIM(grn_number) = TRIM(%s)

        """, (grn,))

        items = cur.fetchall()

        item_list = []

        for item in items:

            item_list.append({

                "product_id": item[0],
                "product_name": item[1],
                "uom": item[2],
                "qty_ordered": float(item[3] or 0),
                "qty_received": float(item[4] or 0),
                "accepted_qty": float(item[5] or 0),
                "rejected_qty": float(item[6] or 0),
                "qty_returned": float(item[7] or 0),
                "stock_in": item[8],
                "warehouse": item[9],
                "price": float(item[10] or 0),
                "tax_pct": float(item[11] or 0),
                "disc_pct": float(item[12] or 0),
                "total": float(item[13] or 0)

            })

        return jsonify({



            "grn_number": stock[0],
            "po_number": stock[1],
            "supplier_id": stock[2],
            "supplier_name": stock[3],
            "supplier_email": stock[4],
            "received_date": str(stock[5]) if stock[5] else "",
            "supplier_dn_no": stock[6],
            "supplier_invoice_no": stock[7],
            "received_by": stock[8],
            "qc_done_by": stock[9],
            "grand_total": float(stock[10] or 0),
            "status": stock[11],
            "items": item_list

  

        })

    except Exception as e:

        print("VIEW STOCK ERROR:", e)

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cur.close()
        conn.close()

@app.route("/api/stock-receipt-pdf/<grn>")
def stock_receipt_pdf(grn):

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =========================================
        # HEADER
        # =========================================
        cur.execute("""

            SELECT
                grn_number,
                po_number,
                supplier_name,
                supplier_email,
                received_date,
                received_by,
                qc_done_by,
                grand_total,
                status

            FROM stock_receipts

            WHERE grn_number = %s

        """, (grn,))

        stock = cur.fetchone()

        if not stock:

            return jsonify({
                "error": "Stock Receipt Not Found"
            }), 404

        # =========================================
        # ITEMS
        # =========================================
        cur.execute("""
            SELECT
                product_id,
                product_name,
                qty_received,
                accepted_qty,
                rejected_qty,
                warehouse,
                total
            FROM stock_receipt_items
            WHERE grn_number = %s
        """, (grn,))

        items = cur.fetchall()

        # =========================================
        # PDF START
        # =========================================
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()

        # =========================================
        # STYLES
        # =========================================

        company_style = ParagraphStyle(
            "CompanyStyle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold", 
            fontSize=24,
            leading=28,
            alignment=TA_CENTER,
            textColor=colors.darkred,
            spaceAfter=5,
        )

        address_style = ParagraphStyle(
            "AddressStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=10,
            leading=14,
            alignment=TA_CENTER,
        )

        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Heading2"],
            fontName="DejaVuSans-Bold",
            fontSize=20,
            leading=24,
            alignment=TA_CENTER,
            textColor=colors.green,
            spaceAfter=20,
        )

        section_style = ParagraphStyle(
            "SectionStyle",
            parent=styles["Heading3"],
            fontName="DejaVuSans",
            fontSize=14,
            textColor=colors.darkred,
            spaceAfter=10,
        )

        normal_style = ParagraphStyle(
            "NormalStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=10,
        )

        elements = []

        # =========================================
        # COMPANY HEADER
        # =========================================

        elements.append(Paragraph("STACKLY", company_style))

        elements.append(Paragraph(
            "MMR Complex, Chinna Tirupathi, Salem, Tamil Nadu - 636008",
            address_style
        ))

        elements.append(Paragraph(
            "Phone: +91 7010792745",
            address_style
        ))

        elements.append(Paragraph(
            "Email: info@stackly.com",
            address_style
        ))

        elements.append(Spacer(1, 25))

        # =========================================
        # TITLE
        # =========================================

        elements.append(Paragraph("STOCK RECEIPT", title_style))

        # =========================================
        # INFO TABLE
        # =========================================

        info_data = [

            ["GRN Number:", stock[0], "Date:", str(stock[4])],
            ["PO Number:", stock[1], "Supplier:", stock[2]],
            ["Received By:", stock[5], "QC Done By:", stock[6]],
            ["Status:", stock[8], "Grand Total:", f"₹ {stock[7]}"]

        ]

        info_table = Table(
            info_data,
            colWidths=[120, 170, 120, 140]
        )

        info_table.setStyle(TableStyle([

            ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
            ("FONTNAME", (0, 0), (0, -1), "DejaVuSans-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "DejaVuSans-Bold"),

            ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),

            ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
            ("BACKGROUND", (2, 0), (2, -1), colors.whitesmoke),

            ("TEXTCOLOR", (0, 0), (0, -1), colors.darkred),
            ("TEXTCOLOR", (2, 0), (2, -1), colors.darkred),

            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 10),

            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ]))

        elements.append(info_table)

        elements.append(Spacer(1, 25))

        # =========================================
        # ITEMS TITLE
        # =========================================

        elements.append(Paragraph("STOCK RECEIPT ITEMS", section_style))

        # =========================================
        # ITEMS TABLE
        # =========================================

        table_data = [[

            "S.No",
            "Product ID",
            "Product Name",
            "Qty Received",
            "Accepted",
            "Rejected",
            "Warehouse",
            "Total"

        ]]

        for i, item in enumerate(items, start=1):

            table_data.append([

                str(i),
                item[0],
                item[1],
                str(item[2]),
                str(item[3]),
                str(item[4]),
                item[5],
                f"₹ {float(item[6]):.2f}"

            ])

        item_table = Table(
            table_data,
            repeatRows=1,
            colWidths=[35, 70, 130, 55, 55, 55, 70, 80]
        )

        item_table.setStyle(TableStyle([

            ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
            ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),

            ("FONTSIZE", (0, 0), (-1, -1), 8),

            ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            ("GRID", (0, 0), (-1, -1), 0.8, colors.grey),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 0), (-1, 0), 10),

            ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 8),

            ("BACKGROUND", (0, 1), (-1, -1), colors.white),

        ]))

        elements.append(item_table)

        elements.append(Spacer(1, 25))

        # =========================================
        # NOTES
        # =========================================

        elements.append(Paragraph("Notes", section_style))

        elements.append(Paragraph(
            "Stock Receipt Generated Successfully.",
            normal_style
        ))

        # =========================================
        # BUILD PDF
        # =========================================

        doc.build(elements)

        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{grn}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:

        return jsonify({
            "error": str(e)
        }), 500

    finally:

        cur.close()
        conn.close() 

@app.route("/api/stock-receipts/<grn>/email", methods=["POST"])
def email_stock_receipt(grn):

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        data = request.json
        po_number = data.get("po_number")

        # GET EMAIL
        cur.execute("""
            SELECT supplier_email
            FROM stock_receipts
            WHERE grn_number = %s
        """, (grn,))

        row = cur.fetchone()

        if not row:
            return jsonify({"success": False, "message": "Not found"})

        supplier_email = row[0]

        # 🔥 SAME PDF USED HERE
        pdf_bytes = build_stock_receipt_pdf(grn)

        subject = f"Stock Receipt - {grn}"

        body = f"""
Dear Supplier,

Please find attached Stock Receipt.

GRN: {grn}
PO: {po_number}

Regards,
Stackly
"""

        success = send_email_with_pdf(
            to_email=supplier_email,
            subject=subject,
            body=body,
            pdf_bytes=pdf_bytes,
            pdf_filename=f"{grn}.pdf"
        )

        return jsonify({"success": success})

    finally:
        cur.close()
        conn.close()

@app.route("/api/stock-receipt-pdf/<grn>")
def stock_receipt_pdf_email(grn):

    buffer = build_stock_receipt_pdf(grn, return_buffer=True)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{grn}.pdf",
        mimetype="application/pdf"
    )


def build_stock_receipt_pdf(grn, return_buffer=False):

    conn = get_db_connection()
    cur = conn.cursor()

    try:

        # =========================
        # HEADER
        # =========================
        cur.execute("""
            SELECT grn_number, po_number, supplier_name,
                   supplier_email, received_date,
                   received_by, qc_done_by,
                   grand_total, status
            FROM stock_receipts
            WHERE grn_number = %s
        """, (grn,))

        stock = cur.fetchone()

        if not stock:
            return None

        # =========================
        # ITEMS
        # =========================
        cur.execute("""
            SELECT
                product_id,
                product_name,
                qty_received,
                accepted_qty,
                rejected_qty,
                warehouse,
                total
            FROM stock_receipt_items
            WHERE grn_number = %s
        """, (grn,))

        items = cur.fetchall()

        # =========================
        # PDF SETUP
        # =========================
        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()

        # SAME STYLES AS YOUR PDF
        company_style = ParagraphStyle(
            "CompanyStyle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold",
            fontSize=24,
            alignment=TA_CENTER,
            textColor=colors.darkred
        )

        address_style = ParagraphStyle(
            "AddressStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=10,
            alignment=TA_CENTER,
        )

        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Heading2"],
            fontName="DejaVuSans-Bold",
            fontSize=20,
            alignment=TA_CENTER,
            textColor=colors.green,
        )

        normal_style = ParagraphStyle(
            "NormalStyle",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=10,
        )

        elements = []

        # =========================
        # HEADER
        # =========================
        elements.append(Paragraph("STACKLY", company_style))
        elements.append(Paragraph("MMR Complex, Salem", address_style))
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("STOCK RECEIPT", title_style))
        elements.append(Spacer(1, 15))

        # =========================
        # INFO TABLE
        # =========================
        info_data = [
            ["GRN", stock[0], "PO", stock[1]],
            ["Supplier", stock[2], "Date", str(stock[4])],
            ["Received By", stock[5], "QC", stock[6]],
            ["Status", stock[8], "Total", f"₹ {stock[7]}"],
        ]

        info_table = Table(info_data, colWidths=[80, 150, 80, 150])

        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 15))

        # =========================
        # ITEMS TABLE
        # =========================
        table_data = [[
            "S.No", "Product", "Qty", "Accepted",
            "Rejected", "Warehouse", "Total"
        ]]

        for i, item in enumerate(items, 1):
            table_data.append([
                str(i),
                item[1],
                str(item[2]),
                str(item[3]),
                str(item[4]),
                item[5],
                f"₹ {float(item[6]):.2f}"
            ])

        item_table = Table(table_data, colWidths=[30, 120, 50, 60, 60, 70, 70])

        item_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
            ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkred),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(item_table)

        # =========================
        # BUILD PDF
        # =========================
        doc.build(elements)

        buffer.seek(0)

        # 🔥 IMPORTANT: SAME OUTPUT FOR BOTH
        return buffer if return_buffer else buffer.getvalue()

    finally:
        cur.close()
        conn.close()

@app.route("/api/stock-comments", methods=["POST"])
def add_stock_comment():
    try:
        data = request.json

        grn_number = data.get("grn_number")
        comment = data.get("comment")
        created_by = data.get("created_by", "Admin")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO stock_comments (grn_number, comment, created_by)
            VALUES (%s, %s, %s)
        """, (grn_number, comment, created_by))

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/stock-comments/<grn_number>")
def get_stock_comments(grn_number):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT comment, created_by, created_at
        FROM stock_comments
        WHERE grn_number = %s
        ORDER BY created_at DESC
    """, (grn_number,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    result = []

    for r in rows:
        result.append({
            "comment": r[0],
            "created_by": r[1],
            "created_at": str(r[2])
        })

    return jsonify(result)


def _ensure_stock_attachments_table(cur=None):
    """stock_attachments + uploads/stock_attachments/ (S3 prefix: stock_attachments/)."""
    try:
        os.makedirs(STOCK_ATTACHMENTS_FOLDER, exist_ok=True)
    except OSError:
        pass
    own_conn = None
    own_cur = None
    if cur is None:
        own_conn = get_db_connection()
        own_cur = own_conn.cursor()
        cur = own_cur
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS stock_attachments (
                attachment_id SERIAL PRIMARY KEY,
                grn_number TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_path TEXT NOT NULL,
                uploaded_at TIMESTAMP DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_stock_attachments_grn
            ON stock_attachments (grn_number)
            """
        )
        cur.connection.commit()
    finally:
        if own_conn:
            own_cur.close()
            own_conn.close()


@app.route("/api/stock-attachments", methods=["POST"])
def upload_stock_attachment():
    conn = None
    cur = None
    try:
        grn_number = (request.form.get("grn_number") or "").strip()
        file = request.files.get("file")

        if not grn_number:
            return jsonify({"success": False, "message": "grn_number is required"}), 400
        if not file or not file.filename:
            return jsonify({"success": False, "message": "No file uploaded"}), 400

        filename = _upload_basename(file.filename)
        file_type = file.mimetype or "application/octet-stream"
        file_size = _upload_file_size_bytes(file)

        if file_size <= 0:
            return jsonify({"success": False, "message": "Empty file cannot be uploaded"}), 400
        if file_size > DOC_UPLOAD_MAX_BYTES:
            return jsonify({
                "success": False,
                "message": f"File size exceeds {MAX_FILE_SIZE_MB} MB",
            }), 400
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in {
            "pdf", "doc", "docx", "xls", "xlsx", "jpg", "jpeg", "png",
        }:
            return jsonify({
                "success": False,
                "message": "This file is not allowed",
            }), 400

        rel_path = _upload_relative_path(grn_number, filename)
        save_path, stored_size = _persist_module_upload(
            object_storage.MODULE_STOCK_ATTACHMENTS,
            STOCK_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )

        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_stock_attachments_table(cur)

        cur.execute(
            """
            INSERT INTO stock_attachments (grn_number, file_name, file_path)
            VALUES (%s, %s, %s)
            RETURNING attachment_id, uploaded_at
            """,
            (grn_number, filename, save_path),
        )
        row = cur.fetchone()
        attachment_id = row[0] if row else None
        uploaded_at = row[1].strftime("%Y-%m-%d %H:%M:%S") if row and row[1] else ""

        conn.commit()

        return jsonify({
            "success": True,
            "attachment_id": attachment_id,
            "file_name": filename,
            "file_path": save_path,
            "size": stored_size or file_size,
            "uploaded_at": uploaded_at,
        })

    except Exception as e:
        if conn:
            conn.rollback()
        print("❌ upload_stock_attachment:", e)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


@app.route("/api/stock-attachments/<grn_number>")
def get_stock_attachments(grn_number):
    grn_number = (grn_number or "").strip()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        _ensure_stock_attachments_table(cur)
        cur.execute(
            """
            SELECT attachment_id, file_name, file_path, uploaded_at
            FROM stock_attachments
            WHERE grn_number = %s
            ORDER BY uploaded_at DESC
            """,
            (grn_number,),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    result = []
    for r in rows:
        result.append({
            "attachment_id": r[0],
            "file_name": r[1],
            "file_path": r[2],
            "uploaded_at": str(r[3]) if r[3] is not None else "",
        })

    return jsonify({"success": True, "attachments": result})





# -------------------------------
# CREDIT NOTE LIST PAGE
# -------------------------------
@app.get("/credit-note")
def credit_note():
    user_email = session.get("user", "")
    users = load_users()

    user_name = "User"

    for u in users:
        if isinstance(u, dict):
            email = (u.get("email") or "").lower()

            if email == user_email.lower():
                user_name = u.get("name") or "User"
                break

    return render_template(
        "credit.html",
        page="credit_note",
        title="Credit Note - Stackly",
        user_email=user_email,
        user_name=user_name,
    )

# -------------------------------
# NEW / EDIT CREDIT NOTE PAGE
# -------------------------------
@app.get("/new-credit-note")
def new_credit_note():
    user_email = session.get("user", "")
    users = load_users()

    user_name = "User"

    for u in users:
        if isinstance(u, dict):
            email = (u.get("email") or "").lower()

            if email == user_email.lower():
                user_name = u.get("name") or "User"
                break

    raw_mode = request.args.get("mode")
    credit_param = (
        request.args.get("credit_note_id")
        or request.args.get("crn_id")
        or ""
    ).strip()

    if credit_param:
        credit_id = credit_param

        if raw_mode is None or str(raw_mode).strip() == "":
            mode = "view"
        else:
            mode = str(raw_mode).strip().lower()

    else:
        if raw_mode is None or str(raw_mode).strip() == "":
            mode = "new"
        else:
            mode = str(raw_mode).strip().lower()

        if mode == "new":
            conn = None
            cur = None
            last_num_row = None

            try:
                conn = get_db_connection()
                cur = conn.cursor()

                cur.execute(
                    """
                    SELECT COALESCE(
                        MAX(CAST(SUBSTRING(credit_note_id FROM 'CRN-(\\d+)$') AS INT)),
                        0
                    )
                    FROM credit_notes
                    """
                )

                last_num_row = cur.fetchone()

            except Exception as e:
                print("generate_credit_note_id error:", e)

            finally:
                try:
                    if cur:
                        cur.close()
                except Exception:
                    pass

                try:
                    if conn:
                        conn.close()
                except Exception:
                    pass

            try:
                new_num = int((last_num_row[0] if last_num_row else 0) or 0) + 1
            except Exception:
                new_num = 1

            credit_id = f"CRN-{str(new_num).zfill(3)}"

        else:
            credit_id = ""

    return render_template(
        "credit-new.html",
        page="credit_note",
        title="New Credit Note - Stackly",
        user_email=user_email,
        user_name=user_name,
        credit_id=credit_id,
        mode=mode,
    )


# -------------------------------
# GET INVOICE LIST FOR CREDIT NOTE
# -------------------------------
@app.get("/api/invoices-credit")
def get_invoices_credit():
    last_err = None

    for _ in range(2):
        conn = None
        cur = None

        try:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute(
                """
                SELECT invoice_id,
                       customer_name,
                       status
                FROM invoices
                WHERE LOWER(TRIM(COALESCE(status, ''))) NOT IN ('cancelled', 'draft')
                ORDER BY id DESC
                """
            )

            rows = cur.fetchall()
            invoice_list = [
                {
                    "invoice_id": r[0] or "",
                    "customer_name": r[1] or "",
                    "status": r[2] or "",
                }
                for r in rows
            ]

            return jsonify({
                "success": True,
                "invoices": invoice_list
            })

        except Exception as e:
            last_err = e

            if "SSL connection has been closed unexpectedly" not in str(e):
                break

        finally:
            try:
                if cur:
                    cur.close()
            except Exception:
                pass

            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    return jsonify({
        "invoices": [],
        "success": False,
        "message": str(last_err or "Failed to load invoices")
    }), 500


# -------------------------------
# GET INVOICE DETAILS FOR CREDIT NOTE
# -------------------------------
@app.get("/api/invoice-details-credit/<invoice_id>")
def get_invoice_details_credit(invoice_id):
    last_err = None

    for _ in range(2):
        conn = None
        cur = None

        try:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute(
                """
                SELECT i.invoice_id,
                       i.customer_name,
                       i.customer_id,
                       i.billing_address,
                       i.phone,
                       i.invoice_date,
                       i.due_date,
                       i.payment_terms,
                       i.status,
                       i.payment_status,
                       COALESCE(s.grand_total, 0),
                       COALESCE(s.amount_paid, 0),
                       COALESCE(s.balance_due, 0)
                FROM invoices i
                LEFT JOIN invoice_summary s
                    ON s.invoice_id = i.invoice_id
                WHERE i.invoice_id = %s
                LIMIT 1
                """,
                (invoice_id,),
            )

            row = cur.fetchone()

            if not row:
                return jsonify({
                    "success": False,
                    "message": "Invoice not found"
                }), 404

            cur.execute(
                """
                SELECT product_name,
                       product_id,
                       quantity,
                       uom,
                       unit_price,
                       tax_pct,
                       disc_pct
                FROM invoice_items
                WHERE invoice_id = %s
                """,
                (invoice_id,),
            )

            items_rows = cur.fetchall()
            items = []

            for r in (items_rows or []):
                qty = float(r[2] or 0)
                unit_price = float(r[4] or 0)
                tax_percent = float(r[5] or 0)
                discount = float(r[6] or 0)

                total = round(
                    (qty * unit_price)
                    * (1 + tax_percent / 100)
                    * (1 - discount / 100),
                    2
                )

                items.append({
                    "product_name": r[0] or "",
                    "product_id": r[1] or "",
                    "quantity": qty,
                    "uom": r[3] or "",
                    "unit_price": unit_price,
                    "tax_percent": tax_percent,
                    "discount": discount,
                    "total": total,
                })

            return jsonify({
                "success": True,
                "invoice": {
                    "invoice_id": row[0] or "",
                    "customer_name": row[1] or "",
                    "customer_id": row[2] or "",
                    "billing_address": row[3] or "",
                    "phone": row[4] or "",
                    "invoice_date": str(row[5]) if row[5] else "",
                    "due_date": str(row[6]) if row[6] else "",
                    "payment_terms": row[7] or "",
                    "status": row[8] or "",
                    "payment_status": row[9] or "",
                    "grand_total": float(row[10] or 0),
                    "amount_paid": float(row[11] or 0),
                    "balance_due": float(row[12] or 0),
                },
                "items": items
            })

        except Exception as e:
            last_err = e

            if "SSL connection has been closed unexpectedly" not in str(e):
                break

        finally:
            try:
                if cur:
                    cur.close()
            except Exception:
                pass

            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    print(f"get_invoice_details_credit error: {last_err}")

    return jsonify({
        "success": False,
        "message": str(last_err)
    }), 500


# -------------------------------
# GET INVOICE RETURN ITEMS FOR CREDIT NOTE
# -------------------------------
@app.get("/api/invoice-return-items/<invoice_id>")
def get_credit_return_items(invoice_id):
    rows = []
    last_err = None

    for _ in range(2):
        conn = None
        cur = None

        try:
            conn = get_db_connection()
            cur = conn.cursor()

            queries = [
                """
                SELECT iri.product_name,
                       iri.product_id,
                       iri.return_quantity,
                       iri.uom,
                       iri.return_reason,
                       iri.unit_price,
                       COALESCE(iri.tax_pct, 0) AS tax_percent,
                       COALESCE(iri.disc_pct, 0) AS discount,
                       iri.total
                FROM invoice_return_items iri
                INNER JOIN invoice_return ir
                    ON ir.invoice_return_id = iri.invoice_return_id
                WHERE ir.invoice_id = %s
                ORDER BY iri.id
                """,
                """
                SELECT product_name,
                       product_id,
                       return_qty,
                       uom,
                       reason,
                       unit_price,
                       tax_percent,
                       discount,
                       total
                FROM invoice_return_items
                WHERE invoice_id = %s
                """,
                """
                SELECT product_name,
                       product_id,
                       returned_qty,
                       uom,
                       reason,
                       unit_price,
                       tax_percent,
                       discount,
                       total
                FROM invoice_return_items
                WHERE invoice_id = %s
                """,
                """
                SELECT product_name,
                       product_id,
                       quantity,
                       uom,
                       reason,
                       unit_price,
                       tax_percent,
                       discount,
                       total
                FROM invoice_return_items
                WHERE invoice_id = %s
                """,
            ]

            for q in queries:
                try:
                    cur.execute(q, (invoice_id,))
                    rows = cur.fetchall()
                    break
                except Exception:
                    conn.rollback()

            return jsonify({
                "items": [
                    {
                        "product_name": r[0],
                        "product_id": r[1],
                        "return_qty": r[2],
                        "uom": r[3],
                        "reason": r[4],
                        "unit_price": r[5],
                        "tax_percent": r[6],
                        "discount": r[7],
                        "total": r[8],
                    }
                    for r in rows
                ]
            })

        except Exception as e:
            last_err = e

            if "SSL connection has been closed unexpectedly" not in str(e):
                break

        finally:
            try:
                if cur:
                    cur.close()
            except Exception:
                pass

            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    print(f"get_credit_return_items error: {last_err}")

    return jsonify({
        "items": [],
        "success": False,
        "message": str(last_err)
    }), 500



# -------------------------------
# GET CREDIT NOTE DETAILS BY ID
# -------------------------------
@app.get("/api/credit-notes/<credit_note_id>")
def get_credit_note_by_id(credit_note_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT
                credit_note_id, credit_note_date, invoice_ref_id, created_by, branch, currency,
                customer_name, customer_id, billing_address, phone, invoice_date, due_date,
                payment_terms, invoice_status, payment_status, invoice_total, amount_paid,
                balance_due, invoice_return_amount, balance_to_refund, refund_mode, refund_paid,
                refund_date, status, created_at, updated_at
            FROM credit_notes
            WHERE credit_note_id = %s
            LIMIT 1
            """,
            (credit_note_id,),
        )

        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Credit note not found"
            }), 404

        cur.execute(
            """
            SELECT
                product_name, product_id, returned_qty, uom, return_reason,
                unit_price, tax_percent, discount_percent, line_total
            FROM credit_note_items
            WHERE credit_note_id = %s
            ORDER BY item_id
            """,
            (credit_note_id,),
        )

        item_rows = cur.fetchall() or []

        items = []

        for ir in item_rows:
            items.append({
                "sno": None,
                "product_name": ir[0] or "",
                "product_id": ir[1] or "",
                "returned_qty": float(ir[2] or 0),
                "return_qty": float(ir[2] or 0),
                "uom": ir[3] or "",
                "reason": ir[4] or "",
                "return_reason": ir[4] or "",
                "unit_price": float(ir[5] or 0),
                "tax_percent": float(ir[6] or 0),
                "discount": float(ir[7] or 0),
                "discount_percent": float(ir[7] or 0),
                "total": float(ir[8] or 0),
                "line_total": float(ir[8] or 0),
            })

        cur.execute(
            """
            SELECT activity_type, message, created_by, created_at
            FROM credit_note_activity
            WHERE credit_note_id = %s
            ORDER BY created_at DESC, activity_id DESC
            """,
            (credit_note_id,),
        )

        comments_rows = cur.fetchall() or []

        comments = []

        for activity_type, message, created_by, created_at in comments_rows:
            msg = str(message or "").strip()

            if not msg:
                continue

            at_ms = None

            try:
                if created_at and hasattr(created_at, "timestamp"):
                    at_ms = int(created_at.timestamp() * 1000)
            except Exception:
                at_ms = None

            comments.append({
                "type": activity_type or "COMMENT",
                "user": str(created_by or "User").strip() or "User",
                "message": msg,
                "at": at_ms,
            })

        cur.close()
        conn.close()

        status_raw = str(row[23] or "").strip().lower()

        if status_raw == "submitted":
            status_norm = "Submitted"
        elif status_raw in ("cancelled", "canceled"):
            status_norm = "Cancelled"
        else:
            status_norm = "Draft"

        created_at_value = ""
        updated_at_value = ""

        if len(row) > 24 and row[24] is not None:
            if hasattr(row[24], "isoformat"):
                created_at_value = row[24].isoformat()
            else:
                created_at_value = str(row[24])

        if len(row) > 25 and row[25] is not None:
            if hasattr(row[25], "isoformat"):
                updated_at_value = row[25].isoformat()
            else:
                updated_at_value = str(row[25])

        return jsonify({
            "success": True,
            "item": {
                "credit_note_id": row[0] or "",
                "credit_note_date": str(row[1]) if row[1] else "",
                "invoice_ref_id": row[2] or "",
                "created_by": row[3] or "",
                "branch": row[4] or "",
                "currency": row[5] or "INR",
                "customer_name": row[6] or "",
                "customer_id": row[7] or "",
                "billing_address": row[8] or "",
                "phone": row[9] or "",
                "invoice_date": str(row[10]) if row[10] else "",
                "due_date": str(row[11]) if row[11] else "",
                "payment_terms": row[12] or "",
                "invoice_status": row[13] or "",
                "payment_status": row[14] or "",
                "invoice_total": float(row[15] or 0),
                "amount_paid": float(row[16] or 0),
                "balance_due": float(row[17] or 0),
                "invoice_return_amount": float(row[18] or 0),
                "balance_to_refund": float(row[19] or 0),
                "refund_mode": row[20] or "",
                "refund_paid": float(row[21] or 0),
                "refund_date": str(row[22]) if row[22] else "",
                "status": status_norm,
                "items": items,
                "comments": comments,
                "created_at": created_at_value,
                "updated_at": updated_at_value,
            }
        })

    except Exception as e:
        print(f"get_credit_note_by_id error: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


# -------------------------------
# GENERATE CREDIT NOTE PDF
# -------------------------------
@app.get("/api/credit-notes/<credit_note_id>/pdf")
def credit_note_pdf(credit_note_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT
                credit_note_id, credit_note_date, invoice_ref_id, created_by, branch, currency,
                customer_name, customer_id, billing_address, phone, invoice_date, due_date,
                payment_terms, invoice_status, payment_status, invoice_total, amount_paid,
                balance_due, invoice_return_amount, balance_to_refund, refund_mode, refund_paid,
                refund_date, status
            FROM credit_notes
            WHERE credit_note_id = %s
            LIMIT 1
            """,
            (credit_note_id,),
        )

        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Credit note not found"
            }), 404

        cur.execute(
            """
            SELECT
                product_name, product_id, returned_qty, uom, return_reason,
                unit_price, tax_percent, discount_percent, line_total
            FROM credit_note_items
            WHERE credit_note_id = %s
            ORDER BY item_id
            """,
            (credit_note_id,),
        )

        item_rows = cur.fetchall() or []

        items = []

        for ir in item_rows:
            items.append({
                "sno": None,
                "product_name": ir[0] or "",
                "product_id": ir[1] or "",
                "returned_qty": float(ir[2] or 0),
                "uom": ir[3] or "",
                "reason": ir[4] or "",
                "unit_price": float(ir[5] or 0),
                "tax_percent": float(ir[6] or 0),
                "discount": float(ir[7] or 0),
                "total": float(ir[8] or 0),
            })

        cur.close()
        conn.close()

        cn = {
            "credit_note_id": row[0] or "",
            "credit_note_date": str(row[1]) if row[1] else "",
            "invoice_ref_id": row[2] or "",
            "created_by": row[3] or "",
            "branch": row[4] or "",
            "currency": row[5] or "INR",
            "customer_name": row[6] or "",
            "customer_id": row[7] or "",
            "billing_address": row[8] or "",
            "phone": row[9] or "",
            "invoice_date": str(row[10]) if row[10] else "",
            "due_date": str(row[11]) if row[11] else "",
            "payment_terms": row[12] or "",
            "invoice_status": row[13] or "",
            "payment_status": row[14] or "",
            "invoice_total": float(row[15] or 0),
            "amount_paid": float(row[16] or 0),
            "balance_due": float(row[17] or 0),
            "invoice_return_amount": float(row[18] or 0),
            "balance_to_refund": float(row[19] or 0),
            "refund_mode": row[20] or "",
            "refund_paid": float(row[21] or 0),
            "refund_date": str(row[22]) if row[22] else "",
            "status": row[23] or "Draft",
            "items": items,
        }

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18,
            leftMargin=18,
            topMargin=16,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()

        company_style = ParagraphStyle(
            name="CN_CompanyName",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#8c1f1f"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )

        company_info_style = ParagraphStyle(
            name="CN_CompanyInfo",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=9,
            leading=12,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=1,
        )

        page_title_style = ParagraphStyle(
            name="CN_PageTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.green,
            alignment=TA_CENTER,
            spaceBefore=12,
            spaceAfter=12,
        )

        section_style = ParagraphStyle(
            name="CN_Section",
            parent=styles["Heading3"],
            fontName="DejaVuSans-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#8c1f1f"),
            spaceAfter=6,
            spaceBefore=10,
        )

        label_style = ParagraphStyle(
            name="CN_Label",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#6b1a1a"),
        )

        value_style = ParagraphStyle(
            name="CN_Value",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8.5,
            leading=11,
            textColor=colors.black,
        )

        header_small_style = ParagraphStyle(
            name="CN_HeaderSmall",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )

        elements = []

        elements.append(Paragraph("STACKLY", company_style))
        elements.append(Paragraph(
            "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
            company_info_style,
        ))
        elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
        elements.append(Paragraph("Email: info@stackly.com", company_info_style))
        elements.append(Spacer(1, 10))

        status_text = str(cn.get("status") or "DRAFT").strip().upper()
        elements.append(Paragraph(f"CREDIT NOTE - {status_text}", page_title_style))
        elements.append(Spacer(1, 2))

        details_data = [
            [
                Paragraph("<b>Credit Note No:</b>", label_style),
                Paragraph(str(cn.get("credit_note_id") or "-"), value_style),
                Paragraph("<b>Date:</b>", label_style),
                Paragraph(str(cn.get("credit_note_date") or "-"), value_style),
            ],
            [
                Paragraph("<b>Invoice Ref ID:</b>", label_style),
                Paragraph(str(cn.get("invoice_ref_id") or "-"), value_style),
                Paragraph("<b>Customer:</b>", label_style),
                Paragraph(str(cn.get("customer_name") or "-"), value_style),
            ],
            [
                Paragraph("<b>Customer ID:</b>", label_style),
                Paragraph(str(cn.get("customer_id") or "-"), value_style),
                Paragraph("<b>Phone:</b>", label_style),
                Paragraph(str(cn.get("phone") or "-"), value_style),
            ],
            [
                Paragraph("<b>Branch:</b>", label_style),
                Paragraph(str(cn.get("branch") or "-"), value_style),
                Paragraph("<b>Currency:</b>", label_style),
                Paragraph(str(cn.get("currency") or "-"), value_style),
            ],
            [
                Paragraph("<b>Payment Status:</b>", label_style),
                Paragraph(str(cn.get("payment_status") or "-"), value_style),
                Paragraph("<b>Invoice Status:</b>", label_style),
                Paragraph(str(cn.get("invoice_status") or "-"), value_style),
            ],
        ]

        details_table = Table(details_data, colWidths=[110, 170, 95, 145])
        details_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        elements.append(details_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("RETURNED LINE ITEMS", section_style))
        elements.append(Spacer(1, 2))

        item_data = [[
            Paragraph("S.No", header_small_style),
            Paragraph("Product Name", header_small_style),
            Paragraph("Product ID", header_small_style),
            Paragraph("Qty", header_small_style),
            Paragraph("UOM", header_small_style),
            Paragraph("Reason", header_small_style),
            Paragraph("Unit Price", header_small_style),
            Paragraph("Tax %", header_small_style),
            Paragraph("Discount %", header_small_style),
            Paragraph("Total", header_small_style),
        ]]

        for idx, item in enumerate(cn.get("items", []) or [], start=1):
            item_data.append([
                Paragraph(str(item.get("sno") or idx), value_style),
                Paragraph(str(item.get("product_name") or "-"), value_style),
                Paragraph(str(item.get("product_id") or "-"), value_style),
                Paragraph(str(item.get("returned_qty") or "0"), value_style),
                Paragraph(str(item.get("uom") or "-"), value_style),
                Paragraph(str(item.get("reason") or "-"), value_style),
                Paragraph(str(item.get("unit_price") or "0"), value_style),
                Paragraph(str(item.get("tax_percent") or "0"), value_style),
                Paragraph(str(item.get("discount") or "0"), value_style),
                Paragraph(str(item.get("total") or "0"), value_style),
            ])

        if len(item_data) == 1:
            item_data.append([
                Paragraph("1", value_style),
                Paragraph("-", value_style),
                Paragraph("-", value_style),
                Paragraph("0", value_style),
                Paragraph("-", value_style),
                Paragraph("-", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
            ])

        items_table = Table(
            item_data,
            colWidths=[30, 95, 55, 35, 35, 80, 55, 40, 55, 55],
            repeatRows=1,
        )

        items_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8c1f1f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#8a8a8a")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ffffff")),
        ]))

        elements.append(items_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("REFUND / ADJUSTMENT SUMMARY", section_style))

        summary_data = [
            [Paragraph("<b>Invoice Total:</b>", label_style), Paragraph(str(cn.get("invoice_total") or "0"), value_style)],
            [Paragraph("<b>Amount Paid:</b>", label_style), Paragraph(str(cn.get("amount_paid") or "0"), value_style)],
            [Paragraph("<b>Balance Due:</b>", label_style), Paragraph(str(cn.get("balance_due") or "0"), value_style)],
            [Paragraph("<b>Invoice Return Amount:</b>", label_style), Paragraph(str(cn.get("invoice_return_amount") or "0"), value_style)],
            [Paragraph("<b>Balance to Refund:</b>", label_style), Paragraph(str(cn.get("balance_to_refund") or "0"), value_style)],
            [Paragraph("<b>Refund Mode:</b>", label_style), Paragraph(str(cn.get("refund_mode") or "-"), value_style)],
            [Paragraph("<b>Refund Paid:</b>", label_style), Paragraph(str(cn.get("refund_paid") or "0"), value_style)],
            [Paragraph("<b>Refund Date:</b>", label_style), Paragraph(str(cn.get("refund_date") or "-"), value_style)],
        ]

        summary_table = Table(summary_data, colWidths=[190, 120])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#999999")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7c7c7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        elements.append(summary_table)

        doc.build(elements)

        pdf_bytes = buffer.getvalue()
        buffer.close()

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            download_name=f"{credit_note_id}.pdf",
            as_attachment=False,
        )

    except Exception as e:
        print(f"credit_note_pdf error: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# -------------------------------
# SEND CREDIT NOTE EMAIL
# -------------------------------
@app.post("/api/credit-notes/<credit_note_id>/email")
def credit_note_email(credit_note_id):
    try:
        payload = request.get_json(silent=True) or {}
        recipient = (payload.get("email") or "").strip()

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT
                credit_note_id, credit_note_date, invoice_ref_id, created_by, branch, currency,
                customer_name, customer_id, billing_address, phone, invoice_date, due_date,
                payment_terms, invoice_status, payment_status, invoice_total, amount_paid,
                balance_due, invoice_return_amount, balance_to_refund, refund_mode, refund_paid,
                refund_date, status
            FROM credit_notes
            WHERE credit_note_id = %s
            LIMIT 1
            """,
            (credit_note_id,),
        )

        row = cur.fetchone()

        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Credit note not found"
            }), 404

        invoice_ref_id = (row[2] or "").strip()
        customer_id = (row[7] or "").strip()

        # Auto fetch customer email if email not given from frontend
        if not recipient:
            customer_row = None

            if invoice_ref_id:
                customer_row = fetch_one(
                    """
                    SELECT i.email
                    FROM invoices i
                    WHERE i.invoice_id = %s
                      AND COALESCE(TRIM(i.email), '') <> ''
                    LIMIT 1
                    """,
                    (invoice_ref_id,),
                )

            if (not customer_row) and invoice_ref_id:
                customer_row = fetch_one(
                    """
                    SELECT c.email
                    FROM invoices i
                    JOIN customers c ON c.customer_id = i.customer_id
                    WHERE i.invoice_id = %s
                      AND COALESCE(TRIM(c.email), '') <> ''
                    LIMIT 1
                    """,
                    (invoice_ref_id,),
                )

            if (not customer_row) and customer_id:
                customer_row = fetch_one(
                    """
                    SELECT email
                    FROM customers
                    WHERE customer_id = %s
                      AND COALESCE(TRIM(email), '') <> ''
                    LIMIT 1
                    """,
                    (customer_id,),
                )

            recipient = ((customer_row or {}).get("email") or "").strip()

        if not recipient:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Customer email not found for this invoice/customer. Update email in invoice or customer master."
            }), 400

        cur.execute(
            """
            SELECT
                product_name, product_id, returned_qty, uom, return_reason,
                unit_price, tax_percent, discount_percent, line_total
            FROM credit_note_items
            WHERE credit_note_id = %s
            ORDER BY item_id
            """,
            (credit_note_id,),
        )

        item_rows = cur.fetchall() or []

        items = []

        for ir in item_rows:
            items.append({
                "sno": None,
                "product_name": ir[0] or "",
                "product_id": ir[1] or "",
                "returned_qty": float(ir[2] or 0),
                "uom": ir[3] or "",
                "reason": ir[4] or "",
                "unit_price": float(ir[5] or 0),
                "tax_percent": float(ir[6] or 0),
                "discount": float(ir[7] or 0),
                "total": float(ir[8] or 0),
            })

        cur.close()
        conn.close()

        cn = {
            "credit_note_id": row[0] or "",
            "credit_note_date": str(row[1]) if row[1] else "",
            "invoice_ref_id": row[2] or "",
            "created_by": row[3] or "",
            "branch": row[4] or "",
            "currency": row[5] or "INR",
            "customer_name": row[6] or "",
            "customer_id": row[7] or "",
            "billing_address": row[8] or "",
            "phone": row[9] or "",
            "invoice_date": str(row[10]) if row[10] else "",
            "due_date": str(row[11]) if row[11] else "",
            "payment_terms": row[12] or "",
            "invoice_status": row[13] or "",
            "payment_status": row[14] or "",
            "invoice_total": float(row[15] or 0),
            "amount_paid": float(row[16] or 0),
            "balance_due": float(row[17] or 0),
            "invoice_return_amount": float(row[18] or 0),
            "balance_to_refund": float(row[19] or 0),
            "refund_mode": row[20] or "",
            "refund_paid": float(row[21] or 0),
            "refund_date": str(row[22]) if row[22] else "",
            "status": row[23] or "Draft",
            "items": items,
        }

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18,
            leftMargin=18,
            topMargin=16,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()

        company_style = ParagraphStyle(
            name="CN_Email_CompanyName",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#8c1f1f"),
            alignment=TA_CENTER,
            spaceAfter=4,
        )

        company_info_style = ParagraphStyle(
            name="CN_Email_CompanyInfo",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=9,
            leading=12,
            textColor=colors.black,
            alignment=TA_CENTER,
            spaceAfter=1,
        )

        page_title_style = ParagraphStyle(
            name="CN_Email_PageTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold",
            fontSize=16,
            leading=20,
            textColor=colors.green,
            alignment=TA_CENTER,
            spaceBefore=12,
            spaceAfter=12,
        )

        section_style = ParagraphStyle(
            name="CN_Email_Section",
            parent=styles["Heading3"],
            fontName="DejaVuSans-Bold",
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#8c1f1f"),
            spaceAfter=6,
            spaceBefore=10,
        )

        label_style = ParagraphStyle(
            name="CN_Email_Label",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#6b1a1a"),
        )

        value_style = ParagraphStyle(
            name="CN_Email_Value",
            parent=styles["Normal"],
            fontName="DejaVuSans",
            fontSize=8.5,
            leading=11,
            textColor=colors.black,
        )

        header_small_style = ParagraphStyle(
            name="CN_Email_HeaderSmall",
            parent=styles["Normal"],
            fontName="DejaVuSans-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=TA_CENTER,
        )

        elements = []

        elements.append(Paragraph("STACKLY", company_style))
        elements.append(Paragraph(
            "MMR Complex, Chinna Thirupathi, near Chinna Muniyappan Kovil, Salem, Tamil Nadu - 636008",
            company_info_style,
        ))
        elements.append(Paragraph("Phone: +91 7010792745", company_info_style))
        elements.append(Paragraph("Email: info@stackly.com", company_info_style))
        elements.append(Spacer(1, 10))

        status_text = str(cn.get("status") or "DRAFT").strip().upper()
        elements.append(Paragraph(f"CREDIT NOTE - {status_text}", page_title_style))
        elements.append(Spacer(1, 2))

        details_data = [
            [
                Paragraph("<b>Credit Note No:</b>", label_style),
                Paragraph(str(cn.get("credit_note_id") or "-"), value_style),
                Paragraph("<b>Date:</b>", label_style),
                Paragraph(str(cn.get("credit_note_date") or "-"), value_style),
            ],
            [
                Paragraph("<b>Invoice Ref ID:</b>", label_style),
                Paragraph(str(cn.get("invoice_ref_id") or "-"), value_style),
                Paragraph("<b>Customer:</b>", label_style),
                Paragraph(str(cn.get("customer_name") or "-"), value_style),
            ],
            [
                Paragraph("<b>Customer ID:</b>", label_style),
                Paragraph(str(cn.get("customer_id") or "-"), value_style),
                Paragraph("<b>Phone:</b>", label_style),
                Paragraph(str(cn.get("phone") or "-"), value_style),
            ],
            [
                Paragraph("<b>Branch:</b>", label_style),
                Paragraph(str(cn.get("branch") or "-"), value_style),
                Paragraph("<b>Currency:</b>", label_style),
                Paragraph(str(cn.get("currency") or "-"), value_style),
            ],
            [
                Paragraph("<b>Payment Status:</b>", label_style),
                Paragraph(str(cn.get("payment_status") or "-"), value_style),
                Paragraph("<b>Invoice Status:</b>", label_style),
                Paragraph(str(cn.get("invoice_status") or "-"), value_style),
            ],
        ]

        details_table = Table(details_data, colWidths=[110, 170, 95, 145])
        details_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f3f3")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8a8a8a")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#a5a5a5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        elements.append(details_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("RETURNED LINE ITEMS", section_style))
        elements.append(Spacer(1, 2))

        item_data = [[
            Paragraph("S.No", header_small_style),
            Paragraph("Product Name", header_small_style),
            Paragraph("Product ID", header_small_style),
            Paragraph("Qty", header_small_style),
            Paragraph("UOM", header_small_style),
            Paragraph("Reason", header_small_style),
            Paragraph("Unit Price", header_small_style),
            Paragraph("Tax %", header_small_style),
            Paragraph("Discount %", header_small_style),
            Paragraph("Total", header_small_style),
        ]]

        for idx, item in enumerate(cn.get("items", []) or [], start=1):
            item_data.append([
                Paragraph(str(item.get("sno") or idx), value_style),
                Paragraph(str(item.get("product_name") or "-"), value_style),
                Paragraph(str(item.get("product_id") or "-"), value_style),
                Paragraph(str(item.get("returned_qty") or "0"), value_style),
                Paragraph(str(item.get("uom") or "-"), value_style),
                Paragraph(str(item.get("reason") or "-"), value_style),
                Paragraph(str(item.get("unit_price") or "0"), value_style),
                Paragraph(str(item.get("tax_percent") or "0"), value_style),
                Paragraph(str(item.get("discount") or "0"), value_style),
                Paragraph(str(item.get("total") or "0"), value_style),
            ])

        if len(item_data) == 1:
            item_data.append([
                Paragraph("1", value_style),
                Paragraph("-", value_style),
                Paragraph("-", value_style),
                Paragraph("0", value_style),
                Paragraph("-", value_style),
                Paragraph("-", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
                Paragraph("0", value_style),
            ])

        items_table = Table(
            item_data,
            colWidths=[30, 95, 55, 35, 35, 80, 55, 40, 55, 55],
            repeatRows=1,
        )

        items_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8c1f1f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#8a8a8a")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ffffff")),
        ]))

        elements.append(items_table)
        elements.append(Spacer(1, 16))

        elements.append(Paragraph("REFUND / ADJUSTMENT SUMMARY", section_style))

        summary_data = [
            [Paragraph("<b>Invoice Total:</b>", label_style), Paragraph(str(cn.get("invoice_total") or "0"), value_style)],
            [Paragraph("<b>Amount Paid:</b>", label_style), Paragraph(str(cn.get("amount_paid") or "0"), value_style)],
            [Paragraph("<b>Balance Due:</b>", label_style), Paragraph(str(cn.get("balance_due") or "0"), value_style)],
            [Paragraph("<b>Invoice Return Amount:</b>", label_style), Paragraph(str(cn.get("invoice_return_amount") or "0"), value_style)],
            [Paragraph("<b>Balance to Refund:</b>", label_style), Paragraph(str(cn.get("balance_to_refund") or "0"), value_style)],
            [Paragraph("<b>Refund Mode:</b>", label_style), Paragraph(str(cn.get("refund_mode") or "-"), value_style)],
            [Paragraph("<b>Refund Paid:</b>", label_style), Paragraph(str(cn.get("refund_paid") or "0"), value_style)],
            [Paragraph("<b>Refund Date:</b>", label_style), Paragraph(str(cn.get("refund_date") or "-"), value_style)],
        ]

        summary_table = Table(summary_data, colWidths=[190, 120])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fafafa")),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#999999")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7c7c7")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        elements.append(summary_table)

        doc.build(elements)

        pdf_bytes = buffer.getvalue()
        buffer.close()

        ok = send_email_with_attachments(
            to_email=recipient,
            subject=f"Credit Note {credit_note_id}",
            body=(
                f"Dear {cn.get('customer_name') or 'Customer'},\n\n"
                f"Please find attached the credit note ({credit_note_id}) issued against your invoice {cn.get('invoice_ref_id') or ''}.\n"
                "The refund/adjustment has been processed as per the details mentioned.\n\n"
                "Please let us know if you have any questions.\n\n"
                "Regards,\n"
                "Stackly Team"
            ),
            from_email=EMAIL_ADDRESS,
            password=EMAIL_PASSWORD,
            attachments=[
                {
                    "filename": f"{credit_note_id}.pdf",
                    "content_bytes": pdf_bytes,
                }
            ],
        )

        if not ok:
            return jsonify({
                "success": False,
                "message": "Email sending failed"
            }), 500

        return jsonify({
            "success": True,
            "message": f"Email sent successfully to {recipient}"
        })

    except Exception as e:
        print(f"credit_note_email error: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

def _ensure_credit_note_tables(cur):
    """Create credit note tables if missing (list GET must work before first save)."""
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS credit_notes (
            credit_note_id TEXT PRIMARY KEY,
            credit_note_date DATE,
            invoice_ref_id TEXT,
            created_by TEXT,
            branch TEXT,
            currency TEXT,
            customer_name TEXT,
            customer_id TEXT,
            billing_address TEXT,
            phone TEXT,
            invoice_date DATE,
            due_date DATE,
            payment_terms TEXT,
            invoice_status TEXT,
            payment_status TEXT,
            invoice_total NUMERIC,
            amount_paid NUMERIC,
            balance_due NUMERIC,
            invoice_return_amount NUMERIC,
            balance_to_refund NUMERIC,
            refund_mode TEXT,
            refund_paid NUMERIC,
            refund_date DATE,
            status TEXT DEFAULT 'Draft',
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    cur.execute("ALTER TABLE credit_notes ADD COLUMN IF NOT EXISTS email VARCHAR(100)")
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS credit_note_items (
            item_id SERIAL PRIMARY KEY,
            credit_note_id VARCHAR(50) NOT NULL,
            product_id VARCHAR(100),
            product_name VARCHAR(200),
            returned_qty NUMERIC(12,2) DEFAULT 0,
            uom VARCHAR(50),
            unit_price NUMERIC(12,2) DEFAULT 0,
            discount_percent NUMERIC(5,2) DEFAULT 0,
            tax_percent NUMERIC(5,2) DEFAULT 0,
            line_total NUMERIC(12,2) DEFAULT 0,
            return_reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS credit_note_activity (
            activity_id SERIAL PRIMARY KEY,
            credit_note_id VARCHAR(50) NOT NULL,
            activity_type VARCHAR(50),
            message TEXT,
            created_by VARCHAR(200),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


# -------------------------------
# GET / SAVE CREDIT NOTE LIST
# -------------------------------
@app.route("/api/credit-notes", methods=["GET", "POST"])
def get_credit_notes():
    if request.method == "POST":
        try:
            data = request.get_json(silent=True) or {}
            if not isinstance(data, dict):
                data = {}

            credit_note_id = (data.get("credit_note_id") or "").strip()
            if not credit_note_id:
                return jsonify({"success": False, "message": "credit_note_id is required"}), 400

            created_by_raw = str(data.get("created_by") or "").strip()
            created_by = re.sub(r"[^A-Za-z ]+", "", created_by_raw)
            created_by = re.sub(r"\s+", " ", created_by).strip()
            if (
                not created_by
                or len(created_by) < 3
                or len(created_by) > 30
                or not re.fullmatch(r"[A-Za-z ]+", created_by)
            ):
                return jsonify({
                    "success": False,
                    "message": "Created By must be 3 to 30 characters (letters and spaces only).",
                }), 400
            data["created_by"] = created_by

            refund_mode = str(data.get("refund_mode") or "").strip().lower()
            if refund_mode == "bank":
                refund_mode = "bank_transfer"
            if refund_mode not in {"", "cash", "bank_transfer", "upi", "cheque"}:
                return jsonify({
                    "success": False,
                    "message": "Invalid refund mode. Allowed: cash, bank_transfer, upi, cheque.",
                }), 400
            data["refund_mode"] = refund_mode

            status_raw = str(data.get("status") or "").strip().lower()
            if status_raw in ("", "draft"):
                status_norm = "Draft"
            elif status_raw == "submitted":
                status_norm = "Submitted"
            elif status_raw in ("cancelled", "canceled"):
                status_norm = "Cancelled"
            else:
                return jsonify({
                    "success": False,
                    "message": "Invalid status. Allowed values are Draft, Submitted, Cancelled.",
                }), 400
            data["status"] = status_norm

            try:
                invoice_total = max(float(str(data.get("invoice_total") or "").strip() or 0), 0.0)
            except Exception:
                invoice_total = 0.0
            try:
                amount_paid = max(float(str(data.get("amount_paid") or "").strip() or 0), 0.0)
            except Exception:
                amount_paid = 0.0
            if amount_paid > invoice_total:
                amount_paid = invoice_total
            try:
                invoice_return_amount = max(
                    float(str(data.get("invoice_return_amount") or "").strip() or 0), 0.0
                )
            except Exception:
                invoice_return_amount = 0.0
            if invoice_return_amount > invoice_total:
                invoice_return_amount = invoice_total
            try:
                refund_paid = max(float(str(data.get("refund_paid") or "").strip() or 0), 0.0)
            except Exception:
                refund_paid = 0.0
            refundable_base = min(invoice_return_amount, amount_paid)
            if refund_paid > refundable_base:
                refund_paid = refundable_base
            balance_due = max(invoice_total - amount_paid, 0.0)
            balance_to_refund = max(refundable_base - refund_paid, 0.0)

            comments_norm = []
            raw_comments = data.get("comments")
            if isinstance(raw_comments, list):
                now_ms = int(time.time() * 1000)
                for c in raw_comments[:200]:
                    if not isinstance(c, dict):
                        continue
                    u = str(c.get("user") or c.get("author") or "").strip() or "User"
                    m = str(c.get("message") or c.get("text") or c.get("comment") or "").strip()
                    if not m:
                        continue
                    if len(m) > 4000:
                        m = m[:4000]
                    at = c.get("at")
                    try:
                        at_i = int(float(at)) if at is not None else None
                    except (TypeError, ValueError):
                        at_i = None
                    if at_i is None or at_i <= 0:
                        at_i = now_ms
                    comments_norm.append({"user": u, "message": m, "at": at_i})

            raw_items = data.get("items")
            items_norm = raw_items if isinstance(raw_items, list) else []

            invoice_ref_id = (data.get("invoice_ref_id") or "").strip()
            customer_id = (data.get("customer_id") or "").strip()
            email_value = (data.get("email") or "").strip()

            conn = get_db_connection()
            cur = conn.cursor()
            _ensure_credit_note_tables(cur)

            if not email_value and invoice_ref_id:
                cur.execute(
                    """
                    SELECT i.email
                    FROM invoices i
                    WHERE i.invoice_id = %s
                      AND COALESCE(TRIM(i.email), '') <> ''
                    LIMIT 1
                    """,
                    (invoice_ref_id,),
                )
                inv_email_row = cur.fetchone()
                if inv_email_row:
                    email_value = (inv_email_row[0] or "").strip()
            if not email_value and invoice_ref_id:
                cur.execute(
                    """
                    SELECT c.email
                    FROM invoices i
                    JOIN customers c ON c.customer_id = i.customer_id
                    WHERE i.invoice_id = %s
                      AND COALESCE(TRIM(c.email), '') <> ''
                    LIMIT 1
                    """,
                    (invoice_ref_id,),
                )
                cust_email_row = cur.fetchone()
                if cust_email_row:
                    email_value = (cust_email_row[0] or "").strip()
            if not email_value and customer_id:
                cur.execute(
                    """
                    SELECT email
                    FROM customers
                    WHERE customer_id = %s
                      AND COALESCE(TRIM(email), '') <> ''
                    LIMIT 1
                    """,
                    (customer_id,),
                )
                cust_only_row = cur.fetchone()
                if cust_only_row:
                    email_value = (cust_only_row[0] or "").strip()

            cur.execute(
                """
                INSERT INTO credit_notes (
                    credit_note_id, credit_note_date, invoice_ref_id, created_by, branch, currency,
                    customer_name, customer_id, email, billing_address, phone, invoice_date, due_date,
                    payment_terms, invoice_status, payment_status, invoice_total, amount_paid,
                    balance_due, invoice_return_amount, balance_to_refund, refund_mode, refund_paid,
                    refund_date, status, updated_at
                ) VALUES (
                    %s, NULLIF(%s, '')::date, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, NULLIF(%s, '')::date, NULLIF(%s, '')::date,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    NULLIF(%s, '')::date, %s, NOW()
                )
                ON CONFLICT (credit_note_id) DO UPDATE SET
                    credit_note_date = EXCLUDED.credit_note_date,
                    invoice_ref_id = EXCLUDED.invoice_ref_id,
                    created_by = EXCLUDED.created_by,
                    branch = EXCLUDED.branch,
                    currency = EXCLUDED.currency,
                    customer_name = EXCLUDED.customer_name,
                    customer_id = EXCLUDED.customer_id,
                    email = EXCLUDED.email,
                    billing_address = EXCLUDED.billing_address,
                    phone = EXCLUDED.phone,
                    invoice_date = EXCLUDED.invoice_date,
                    due_date = EXCLUDED.due_date,
                    payment_terms = EXCLUDED.payment_terms,
                    invoice_status = EXCLUDED.invoice_status,
                    payment_status = EXCLUDED.payment_status,
                    invoice_total = EXCLUDED.invoice_total,
                    amount_paid = EXCLUDED.amount_paid,
                    balance_due = EXCLUDED.balance_due,
                    invoice_return_amount = EXCLUDED.invoice_return_amount,
                    balance_to_refund = EXCLUDED.balance_to_refund,
                    refund_mode = EXCLUDED.refund_mode,
                    refund_paid = EXCLUDED.refund_paid,
                    refund_date = EXCLUDED.refund_date,
                    status = EXCLUDED.status,
                    updated_at = NOW()
                """,
                (
                    credit_note_id,
                    data.get("credit_note_date") or "",
                    data.get("invoice_ref_id") or "",
                    data.get("created_by") or "",
                    data.get("branch") or "",
                    data.get("currency") or "INR",
                    data.get("customer_name") or "",
                    data.get("customer_id") or "",
                    email_value,
                    data.get("billing_address") or "",
                    data.get("phone") or "",
                    data.get("invoice_date") or "",
                    data.get("due_date") or "",
                    data.get("payment_terms") or "",
                    data.get("invoice_status") or "",
                    data.get("payment_status") or "Unpaid",
                    invoice_total,
                    amount_paid,
                    balance_due,
                    invoice_return_amount,
                    balance_to_refund,
                    data.get("refund_mode") or "",
                    refund_paid,
                    data.get("refund_date") or "",
                    data.get("status") or "Draft",
                ),
            )

            cur.execute("DELETE FROM credit_note_items WHERE credit_note_id = %s", (credit_note_id,))
            for idx, item in enumerate(items_norm, start=1):
                if not isinstance(item, dict):
                    continue
                cur.execute(
                    """
                    INSERT INTO credit_note_items (
                        credit_note_id, product_id, product_name,
                        returned_qty, uom, unit_price, discount_percent, tax_percent, line_total, return_reason
                    ) VALUES (
                        %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    (
                        credit_note_id,
                        str(item.get("product_id") or "").strip(),
                        str(item.get("product_name") or "").strip(),
                        max(
                            float(
                                str(
                                    item.get("returned_qty")
                                    if item.get("returned_qty") is not None
                                    else item.get("return_qty") or 0
                                ).strip()
                                or 0
                            ),
                            0.0,
                        ),
                        str(item.get("uom") or "").strip(),
                        max(float(str(item.get("unit_price") or "").strip() or 0), 0.0),
                        max(
                            float(
                                str(
                                    item.get("discount")
                                    if item.get("discount") is not None
                                    else item.get("discount_percent") or 0
                                ).strip()
                                or 0
                            ),
                            0.0,
                        ),
                        max(float(str(item.get("tax_percent") or "").strip() or 0), 0.0),
                        max(
                            float(
                                str(
                                    item.get("total")
                                    if item.get("total") is not None
                                    else item.get("line_total") or 0
                                ).strip()
                                or 0
                            ),
                            0.0,
                        ),
                        str(
                            item.get("reason")
                            if item.get("reason") is not None
                            else item.get("return_reason") or ""
                        ).strip(),
                    ),
                )

            cur.execute(
                "DELETE FROM credit_note_activity WHERE credit_note_id = %s AND activity_type IN ('COMMENT','HISTORY')",
                (credit_note_id,),
            )
            for c in comments_norm:
                user = str(c.get("user") or "User").strip() or "User"
                msg = str(c.get("message") or "").strip()
                if not msg:
                    continue
                cur.execute(
                    """
                    INSERT INTO credit_note_activity (credit_note_id, activity_type, message, created_by, created_at)
                    VALUES (%s, 'COMMENT', %s, %s, NOW())
                    """,
                    (credit_note_id, msg, user),
                )

            conn.commit()
            cur.close()
            conn.close()
            return jsonify({"success": True, "credit_note_id": credit_note_id})
        except Exception as e:
            print(f"save_credit_note error: {e}")
            return jsonify({"success": False, "message": str(e)}), 500

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        _ensure_credit_note_tables(cur)
        conn.commit()

        cur.execute(
            """
            SELECT
                credit_note_id,
                invoice_ref_id,
                customer_name,
                credit_note_date,
                status,
                COALESCE(payment_status, 'Unpaid')
            FROM credit_notes
            ORDER BY updated_at DESC, credit_note_id DESC
            """
        )

        rows = cur.fetchall()

        cur.close()
        conn.close()

        items = []

        for r in rows:
            status_raw = str(r[4] or "").strip().lower()

            if status_raw == "submitted":
                status_norm = "Submitted"
            elif status_raw in ("cancelled", "canceled"):
                status_norm = "Cancelled"
            else:
                status_norm = "Draft"

            items.append({
                "crn_id": r[0] or "",
                "invoice_ref_id": r[1] or "",
                "customer_name": r[2] or "",
                "credit_note_date": str(r[3]) if r[3] else "",
                "status": status_norm,
                "payment_status": r[5] or "Unpaid",
            })

        return jsonify({
            "success": True,
            "items": items
        })

    except Exception as e:
        print(f"get_credit_notes error: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

# -------------------------------
# MARK CREDIT NOTE AS PAID
# -------------------------------
@app.post("/api/credit-notes/<credit_note_id>/mark-paid")
def mark_credit_note_paid(credit_note_id):
    try:
        payload = request.get_json(silent=True) or {}

        try:
            invoice_total = float(payload.get("invoice_total") or 0)
        except Exception:
            invoice_total = 0

        try:
            amount_paid = float(payload.get("amount_paid") or 0)
        except Exception:
            amount_paid = 0

        try:
            invoice_return_amount = float(payload.get("invoice_return_amount") or 0)
        except Exception:
            invoice_return_amount = 0

        try:
            refund_paid = float(payload.get("refund_paid") or 0)
        except Exception:
            refund_paid = 0

        invoice_total = max(invoice_total, 0)
        amount_paid = max(amount_paid, 0)
        invoice_return_amount = max(invoice_return_amount, 0)
        refund_paid = max(refund_paid, 0)

        if amount_paid > invoice_total:
            amount_paid = invoice_total

        if invoice_return_amount > invoice_total:
            invoice_return_amount = invoice_total

        refundable_base = min(invoice_return_amount, amount_paid)

        refund_paid = refundable_base
        balance_due = max(invoice_total - amount_paid, 0)
        balance_to_refund = 0

        refund_mode = str(payload.get("refund_mode") or "").strip().lower()

        if refund_mode == "bank":
            refund_mode = "bank_transfer"

        if refund_mode not in {"", "cash", "bank_transfer", "upi", "cheque"}:
            return jsonify({
                "success": False,
                "message": "Invalid refund mode. Allowed: cash, bank_transfer, upi, cheque."
            }), 400

        refund_date = str(payload.get("refund_date") or "").strip()

        if not refund_date:
            refund_date = datetime.now().strftime("%Y-%m-%d")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS credit_note_activity (
                activity_id SERIAL PRIMARY KEY,
                credit_note_id VARCHAR(50) NOT NULL,
                activity_type VARCHAR(50),
                message TEXT,
                created_by VARCHAR(200),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """
        )

        cur.execute(
            """
            UPDATE credit_notes
            SET
                payment_status = 'Paid',
                refund_paid = %s,
                balance_due = %s,
                balance_to_refund = %s,
                refund_mode = %s,
                refund_date = NULLIF(%s, '')::date,
                updated_at = NOW()
            WHERE credit_note_id = %s
            """,
            (
                refund_paid,
                balance_due,
                balance_to_refund,
                refund_mode,
                refund_date,
                credit_note_id,
            ),
        )

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()

            return jsonify({
                "success": False,
                "message": "No saved credit note with this ID. Save Draft first, then mark as paid.",
            }), 404

        cur.execute(
            """
            UPDATE credit_notes
            SET status = 'Submitted', updated_at = NOW()
            WHERE credit_note_id = %s
            """,
            (credit_note_id,),
        )

        cur.execute(
            """
            INSERT INTO credit_note_activity (
                credit_note_id, activity_type, message, created_by, created_at
            ) VALUES (
                %s, 'HISTORY', %s, %s, NOW()
            )
            """,
            (
                credit_note_id,
                "Marked as paid",
                session.get("user", "User"),
            ),
        )

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            "success": True,
            "message": "Credit note marked as paid",
            "payment_status": "Paid",
            "refund_paid": refund_paid,
            "balance_due": balance_due,
            "balance_to_refund": balance_to_refund,
            "refund_mode": refund_mode,
            "refund_date": refund_date,
        })

    except Exception as e:
        print(f"mark_credit_note_paid error: {e}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

        
@app.post("/api/cn-upload-attachment")
def cn_upload_attachment():
    """Upload attachment for a saved credit note (same rules as delivery note return)."""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "No file provided"}), 400
        file = request.files["file"]
        credit_note_id = (request.form.get("credit_note_id") or "").strip()
        if not credit_note_id:
            return jsonify({"success": False, "error": "credit_note_id required"}), 400
        if file.filename == "":
            return jsonify({"success": False, "error": "No file selected"}), 400

        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        file.seek(0)
        if file_length > MAX_FILE_SIZE_BYTES:
            return jsonify(
                {"success": False, "error": f"File size exceeds {MAX_FILE_SIZE_MB} MB"}
            ), 400
        if not allowed_file(file.filename):
            return jsonify(
                {
                    "success": False,
                    "error": f'File type not allowed. Allowed: {", ".join(ALLOWED_EXTENSIONS)}',
                }
            ), 400

        

        hdr = fetch_one(
            "SELECT 1 AS ok FROM credit_notes WHERE credit_note_id = %s",
            (credit_note_id,),
        )
        if not hdr:
            return jsonify(
                {
                    "success": False,
                    "error": "Save the credit note as draft first, then attach files.",
                }
            ), 400

        cnt_row = fetch_one(
            """
            SELECT COUNT(*)::int AS c
            FROM credit_note_attachments
            WHERE credit_note_id = %s
            """,
            (credit_note_id,),
        )
        current_count = int((cnt_row or {}).get("c") or 0)
        if current_count >= 10:
            return jsonify(
                {
                    "success": False,
                    "error": "Maximum 10 files allowed per credit note",
                }
            ), 400

        original_filename = _upload_basename(file.filename)
        rel_path = _upload_relative_path(credit_note_id, original_filename)
        stored_path, _stored_size = _persist_module_upload(
            object_storage.MODULE_CREDIT_NOTE_ATTACHMENTS,
            CREDIT_NOTE_ATTACHMENTS_FOLDER,
            file,
            rel_path,
        )

        row = None
        conn = get_db_connection()
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                try:
                    os.makedirs(CREDIT_NOTE_ATTACHMENTS_FOLDER, exist_ok=True)
                except OSError:
                    pass
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS credit_note_attachments (
                        id SERIAL PRIMARY KEY,
                        credit_note_id TEXT NOT NULL,
                        file_name TEXT,
                        file_path TEXT,
                        uploaded_at TIMESTAMP DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = current_schema()
                      AND table_name = 'credit_note_attachments'
                    """
                )
                existing_cols = {str(r.get("column_name") or "").strip() for r in (cur.fetchall() or [])}

                insert_values = {
                    "credit_note_id": credit_note_id,
                    "uploaded_at": datetime.now(),
                }
                # Canonical schema
                if "file_name" in existing_cols:
                    insert_values["file_name"] = original_filename
                if "file_path" in existing_cols:
                    insert_values["file_path"] = stored_path
                # Legacy schema variants
                if "original_filename" in existing_cols:
                    insert_values["original_filename"] = original_filename
                if "filename" in existing_cols:
                    insert_values["filename"] = original_filename
                if "stored_filename" in existing_cols:
                    insert_values["stored_filename"] = _upload_basename(stored_path)
                if "size" in existing_cols:
                    insert_values["size"] = int(file_length)
                if "file_size" in existing_cols:
                    insert_values["file_size"] = int(file_length)
                if "uploaded_by" in existing_cols:
                    insert_values["uploaded_by"] = (session.get("user") or "system")

                cols = list(insert_values.keys())
                placeholders = ", ".join(["%s"] * len(cols))
                sql_cols = ", ".join(cols)
                cur.execute(
                    f"INSERT INTO credit_note_attachments ({sql_cols}) VALUES ({placeholders}) RETURNING id",
                    tuple(insert_values[c] for c in cols),
                )
                row = cur.fetchone()
            conn.commit()
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            _remove_stored_upload(stored_path, CREDIT_NOTE_ATTACHMENTS_FOLDER)
            raise
        finally:
            conn.close()

        new_id = row["id"] if row else None
        return jsonify(
            {
                "success": True,
                "attachment": {
                    "id": new_id,
                    "original_filename": original_filename,
                    "stored_filename": _upload_basename(stored_path),
                    "size": file_length,
                    "upload_date": datetime.now().isoformat(),
                },
            }
        )
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.get("/api/cn-attachments/<credit_note_id>")
def cn_get_attachments(credit_note_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            os.makedirs(CREDIT_NOTE_ATTACHMENTS_FOLDER, exist_ok=True)
        except OSError:
            pass
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS credit_note_attachments (
                id SERIAL PRIMARY KEY,
                credit_note_id TEXT NOT NULL,
                file_name TEXT,
                file_path TEXT,
                uploaded_at TIMESTAMP DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            SELECT id, file_name, file_path, uploaded_at
            FROM credit_note_attachments
            WHERE credit_note_id = %s
            ORDER BY id ASC
            """,
            (credit_note_id,),
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"cn_get_attachments: {e}")
        return jsonify({"success": False, "attachments": [], "error": str(e)}), 500

    attachments = []
    for r in rows or []:
        rid, fname, fpath, ts = r[0], r[1], r[2], r[3]
        tstr = ""
        if ts is not None:
            if hasattr(ts, "strftime"):
                tstr = ts.strftime("%Y-%m-%d %H:%M:%S")
            else:
                tstr = str(ts)
        raw_path = str(fpath or "")
        sz = 0
        ap = _resolve_stored_file_path(raw_path)
        if ap and os.path.isfile(ap) and not object_storage.is_remote_url(ap):
            try:
                sz = os.path.getsize(ap)
            except OSError:
                sz = 0
        attachments.append(
            {
                "id": rid,
                "original_filename": fname or "",
                "size": sz,
                "upload_date": tstr or "—",
            }
        )
    return jsonify({"success": True, "attachments": attachments})


@app.get("/api/cn-attachment/<att_id>/view")
def cn_view_attachment(att_id):
    try:
        att_id = int(str(att_id).strip())
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Invalid attachment id"}), 400
    try:
        os.makedirs(CREDIT_NOTE_ATTACHMENTS_FOLDER, exist_ok=True)
    except OSError:
        pass
    row = fetch_one(
        """
        SELECT id, credit_note_id, file_name, file_path
        FROM credit_note_attachments
        WHERE id = %s
        """,
        (att_id,),
    )
    if not row:
        return jsonify({"success": False, "message": "Attachment not found"}), 404
    row = dict(row)
    raw_fp = str(row.get("file_path") or "")
    if object_storage.is_remote_url(raw_fp):
        return redirect(raw_fp)
    abs_path = _resolve_stored_file_path(raw_fp)
    if (not abs_path) or object_storage.is_remote_url(abs_path) or (not os.path.isfile(abs_path)):
        return jsonify({"success": False, "message": "Attachment not found"}), 404
    return send_file(
        abs_path,
        download_name=row.get("file_name") or "file",
        as_attachment=False,
    )


@app.get("/api/cn-attachment/<att_id>/download")
def cn_download_attachment(att_id):
    try:
        att_id = int(str(att_id).strip())
    except (TypeError, ValueError):
        return jsonify({"success": False, "message": "Invalid attachment id"}), 400
    try:
        os.makedirs(CREDIT_NOTE_ATTACHMENTS_FOLDER, exist_ok=True)
    except OSError:
        pass
    row = fetch_one(
        """
        SELECT id, credit_note_id, file_name, file_path
        FROM credit_note_attachments
        WHERE id = %s
        """,
        (att_id,),
    )
    if not row:
        return jsonify({"success": False, "message": "Attachment not found"}), 404
    row = dict(row)
    raw_fp = str(row.get("file_path") or "")
    if object_storage.is_remote_url(raw_fp):
        return redirect(raw_fp)
    abs_path = _resolve_stored_file_path(raw_fp)
    if (not abs_path) or object_storage.is_remote_url(abs_path) or (not os.path.isfile(abs_path)):
        return jsonify({"success": False, "message": "Attachment not found"}), 404
    return send_file(
        abs_path,
        download_name=row.get("file_name") or "file",
        as_attachment=True,
    )


@app.delete("/api/cn-attachment/<att_id>")
def cn_delete_attachment(att_id):
    try:
        try:
            att_id = int(str(att_id).strip())
        except (TypeError, ValueError):
            return jsonify({"success": False, "error": "Invalid attachment id"}), 400
        
        row = fetch_one(
            """
            SELECT id, file_path
            FROM credit_note_attachments
            WHERE id = %s
            """,
            (att_id,),
        )
        if not row:
            return jsonify({"success": False, "error": "Attachment not found"}), 404
        row = dict(row)
        raw_file_path = str(row.get("file_path") or "").strip()

        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM credit_note_attachments WHERE id = %s", (att_id,))
            conn.commit()
        finally:
            conn.close()

        _remove_stored_upload(raw_file_path, CREDIT_NOTE_ATTACHMENTS_FOLDER)
        return jsonify({"success": True})
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"success": False, "error": str(e)}), 500


@app.post("/api/credit-notes/<credit_note_id>/cancel")
def cancel_credit_note(credit_note_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT COALESCE(status, ''), COALESCE(payment_status, '')
            FROM credit_notes
            WHERE credit_note_id = %s
            LIMIT 1
            """,
            (credit_note_id,),
        )
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "No saved credit note with this ID. Save Draft first, then cancel it.",
            })
        current_status = str(row[0] or "").strip().lower()
        current_payment_status = str(row[1] or "").strip().lower()
        if current_status == "submitted" and current_payment_status == "paid":
            cur.close()
            conn.close()
            return jsonify({
                "success": False,
                "message": "Submitted + Paid credit note cannot be cancelled.",
            }), 400
        cur.execute(
            """
            UPDATE credit_notes
            SET status = 'Cancelled', updated_at = NOW()
            WHERE credit_note_id = %s
            """,
            (credit_note_id,),
        )
        conn.commit()
        updated = cur.rowcount > 0
        cur.close()
        conn.close()
        if updated:
            return jsonify({"success": True})
        return jsonify({
            "success": False,
            "message": "No saved credit note with this ID. Save Draft first, then cancel it.",
        })
    except Exception as e:
        print(f"cancel_credit_note error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.delete("/api/credit-notes/<credit_note_id>")
def delete_credit_note(credit_note_id):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT file_path FROM credit_note_attachments WHERE credit_note_id = %s",
            (credit_note_id,),
        )
        for pr in cur.fetchall() or []:
            _remove_stored_upload(str(pr[0] or ""), CREDIT_NOTE_ATTACHMENTS_FOLDER)
        cur.execute("DELETE FROM credit_note_attachments WHERE credit_note_id = %s", (credit_note_id,))
        cur.execute("DELETE FROM credit_notes WHERE credit_note_id = %s", (credit_note_id,))
        conn.commit()
        deleted = cur.rowcount > 0
        cur.close()
        conn.close()
        if deleted:
            return jsonify({"success": True})
        return jsonify({
            "success": False,
            "message": "No saved credit note with this ID. Save Draft first.",
        })
    except Exception as e:
        print(f"delete_credit_note error: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# -------------------------------
# CREATE PAYMENT PAGE
# -------------------------------
@app.route("/create-payment")
def create_payment():
    user_email = session.get("user")
    if not user_email:
        return redirect(url_for("login", message="session_expired"))

    users = load_users()
    user_name = "User"
    for u in users:
        if isinstance(u, dict) and (u.get("email") or "").lower() == user_email.lower():
            user_name = u.get("name") or "User"
            break

    return render_template(
        "create-payment.html",
        title="Create Payment - Stackly",
        page="create_payment",
        section="finance",
        user_email=user_email,
        user_name=user_name,
    )


@app.route("/api/invoices_payments")
def get_invoices_payments():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                i.invoice_id,
                i.customer_name,
                COALESCE(iss.balance_due, 0) AS balance_due,
                i.payment_method
            FROM invoices i
            LEFT JOIN invoice_summary iss ON i.invoice_id = iss.invoice_id
            WHERE i.payment_status != 'Paid'
              AND i.status NOT IN ('Draft', 'Cancelled')
            ORDER BY i.created_at DESC
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        invoices = []
        for row in rows:
            invoices.append({
                "id": row[0],
                "customer_name": row[1],
                "balance_due": float(row[2]) if row[2] is not None else 0.0,
                "payment_method": row[3] if row[3] is not None else "",
            })
        return jsonify(invoices)
    except Exception as e:
        print("Error in /api/invoices_payments:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route("/api/payments", methods=["POST"])
def save_payment():
    try:
        data = request.get_json()
        invoice_id = data["invoiceId"]
        amount = Decimal(str(data["amount"]))
        customer_name = data["customerName"]
        payment_method = data["paymentMethod"]
        transaction_id = data["transactionId"]
        payment_date = data["date"]
        notes = data["notes"]

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO payments
            (invoice_id, customer_name, payment_method, transaction_id, amount, payment_date, notes)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (invoice_id, customer_name, payment_method, transaction_id, amount, payment_date, notes))

        cur.execute("SELECT balance_due FROM invoice_summary WHERE invoice_id = %s", (invoice_id,))
        row = cur.fetchone()

        if row:
            current_balance = row[0]
            new_balance = current_balance - amount
            if new_balance < 0:
                new_balance = Decimal("0")
            cur.execute("""
                UPDATE invoice_summary
                SET balance_due = %s
                WHERE invoice_id = %s
            """, (new_balance, invoice_id))
        else:
            cur.execute("SELECT invoice_total FROM invoices WHERE invoice_id = %s", (invoice_id,))
            inv_row = cur.fetchone()
            if inv_row:
                total_amount = inv_row[0] or Decimal("0")
                new_balance = Decimal(str(total_amount)) - amount
                if new_balance < 0:
                    new_balance = Decimal("0")
                cur.execute("""
                    INSERT INTO invoice_summary (invoice_id, balance_due)
                    VALUES (%s, %s)
                """, (invoice_id, new_balance))
            else:
                conn.rollback()
                return jsonify({"success": False, "error": "Invoice not found"}), 404

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({"success": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/invoice-summary/<invoice_id>/add-payment", methods=["PUT"])
def add_payment_to_invoice_summary(invoice_id):
    try:
        data = request.get_json()
        amount = float(data.get("amount", 0))
        payment_ref_no = data.get("payment_ref_no")
        transaction_date = data.get("transaction_date")

        if amount <= 0:
            return jsonify({"success": False, "error": "Amount must be positive"}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT amount_paid, grand_total
            FROM invoice_summary
            WHERE invoice_id = %s
        """, (invoice_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "error": "Invoice summary not found"}), 404

        current_paid = float(row[0]) if row[0] else 0.0
        grand_total = float(row[1]) if row[1] else 0.0
        new_paid = current_paid + amount
        new_balance_due = grand_total - new_paid

        cur.execute("""
            UPDATE invoice_summary
            SET amount_paid = %s, balance_due = %s
            WHERE invoice_id = %s
        """, (new_paid, new_balance_due, invoice_id))

        if payment_ref_no and transaction_date:
            cur.execute("""
                UPDATE invoices
                SET payment_ref_no = %s, transaction_date = %s
                WHERE invoice_id = %s
            """, (payment_ref_no, transaction_date, invoice_id))
        elif payment_ref_no:
            cur.execute("""
                UPDATE invoices
                SET payment_ref_no = %s
                WHERE invoice_id = %s
            """, (payment_ref_no, invoice_id))
        elif transaction_date:
            cur.execute("""
                UPDATE invoices
                SET transaction_date = %s
                WHERE invoice_id = %s
            """, (transaction_date, invoice_id))

        if new_balance_due <= 0:
            cur.execute("""
                UPDATE invoices
                SET payment_status = 'Paid'
                WHERE invoice_id = %s
            """, (invoice_id,))
        elif new_paid > 0 and new_balance_due > 0:
            cur.execute("""
                UPDATE invoices
                SET payment_status = 'Partial'
                WHERE invoice_id = %s
            """, (invoice_id,))

        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"success": False, "error": str(e)}), 500


# =========================================
# ✅ RUN APP
# =========================================
if __name__ == "__main__":
    import sys

    print("Application is running successfully")
    if object_storage.is_enabled():
        print(f"S3 uploads enabled (bucket: {object_storage._s3_bucket_id_from_env()})")
    else:
        print(
            "S3 uploads disabled — files save under uploads/ locally. "
            "Set S3_ACCESS_KEY_ID and S3_SECRET_ACCESS_KEY in .env for AWS S3."
        )
    # On Windows, debug auto-reload can log "Exception in thread Thread-1 (serve_forever)".
    # Set FLASK_USE_RELOADER=1 in .env to turn reload back on.
    use_reloader = os.environ.get("FLASK_USE_RELOADER", "1" if sys.platform != "win32" else "0")
    use_reloader = str(use_reloader).strip().lower() in ("1", "true", "yes", "on")
    app.run(debug=True, use_reloader=use_reloader)
    


